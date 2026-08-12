#!/opt/homebrew/opt/python@3.12/libexec/bin/python3
#
# Filename: fm.py
# Project: File Management
# Version: 2.00
# Description: File Manager — a unified interactive + CLI tool that combines
#              the compare, display, eject, find, monitor, remove, sync, zip,
#              optimize media, permissions, mount shares, local scripts, AWS
#              S3, and clean-up file-management scripts into a single
#              CB9Lib-based program.
# Maintainer: Cloud Box 9 Inc.
# Last Modified Date: 2026-08-12
#
# Combines the functionality of:
#   compareFilesRecursive.sh, compareSubFolders.sh, folderSizes.sh,
#   findLargeFiles.sh, findDuplicates.sh, findDuplicatesExt.sh,
#   removeFiles.sh, zipSubFolders.sh, zipView.py
#
# Usage (interactive):
#   fm.py
#
# Usage (CLI):
#   fm.py compare-2files   FILE_A FILE_B
#   fm.py convert FILE csv|json|xlsx|sql   (convert a .csv/.json/.xlsx data file;
#                                           output written beside it, same name,
#                                           new extension, collision-safe)
#   fm.py uuid [N]                         (generate N random UUIDs, one per line)
#   fm.py compare-contents A B [--recursive] [--by name|size|both] [--case-sensitive]
#   fm.py sizes [FOLDER] [--sort alpha|size]
#   fm.py drives                          (size and free space of all mounted drives)
#   fm.py drives-in-use                   (external drives + whether each is being read/written, macOS)
#   fm.py find folder PATTERN [ROOT]
#   fm.py find name   PATTERN [ROOT]
#   fm.py find ext    EXT     [ROOT]
#   fm.py find over   N       [ROOT]      (N = megabytes)
#   fm.py find under  N       [ROOT]      (N = megabytes)
#   fm.py find-files [ROOT] [--name PAT] [--ext E] [--over N] [--under N]   (combined AND)
#   fm.py find-dups   FOLDER...           (duplicate filenames, size table per folder)
#   fm.py find-fuzzy-dups FOLDER...       (close-name + close-size duplicates,
#                                          grouped with KEEP/DELETE markers)
#   fm.py find-missing A B [--in first|second|either] [--size]   (filenames in only one
#                                          folder; --size also matches the file size)
#   fm.py find-replace ROOT SEARCH REPLACE [--ext E] [--apply] [--bak] [--yes]
#                                          (find & replace text in files; dry run
#                                           unless --apply)
#   fm.py find-rename ROOT (--prepend TEXT | --append TEXT | --replace FIND NEW)
#                     [--ext E] [--apply] [--yes]
#                                          (find files and rename them; dry run
#                                           unless --apply)
#   fm.py remove folder      PATH        [--delete] [--yes]
#   fm.py remove name        PATTERN ROOT [--delete] [--yes]   (files by name)
#   fm.py remove folder-name PATTERN ROOT [--delete] [--yes]   (folders by name)
#   fm.py remove dup-name    FOLDER...    [--delete] [--yes]
#   fm.py remove dup-hash    FOLDER...    [--delete] [--yes]
#   fm.py remove dup-fuzzy   FOLDER...    [--delete] [--yes]   (close name + close size)
#   fm.py remove zero-size   FOLDER...    [--delete] [--yes]   (empty 0-byte files)
#   fm.py eject [--list] [--force] [--yes]   (eject all external drives — macOS)
#   fm.py monitor FOLDER [--no-recursive] [--ext jpg,png] [--csv]   (Ctrl-C stops)
#   fm.py monitor --profile NAME          (run a monitorProfiles entry from fmConfig.json)
#   fm.py sync FOLDER_A FOLDER_B [--to b|a|both] [--conflict newest|largest|manual]
#              [--no-recursive] [--include-hidden] [--copy] [--yes]
#   fm.py sync --profile NAME             (run a syncProfiles entry from fmConfig.json)
#   fm.py zip-subfolders  TARGET [DEST] [-r]
#   fm.py zip-view [ZIP|FOLDER]
#   fm.py find-zip ZIP|FOLDER [PATTERN]   (search filenames inside a zip, or
#                                          every zip found recursively under a
#                                          folder; wildcards ok, e.g. 'fan*.png';
#                                          filenames only — contents are not
#                                          searched; PATTERN defaults to *)
#   fm.py zip-log  TARGET  [-r]            (log .zip/.tar to CB9Inventory; -r/--recursive
#                                          with a folder TARGET also logs archives in
#                                          subfolders, not just the top level)
#   fm.py convert-video SOURCE [DEST] [-r] [--flatten] [--delete] [--pause SECONDS]
#                                          (convert .mov/.avi/.mkv/.wmv/.flv/.m4v
#                                           to .mp4 via moviepy; -r = include
#                                           subfolders; when -r is used with a
#                                           DEST different from SOURCE, subfolders
#                                           are mirrored into DEST by default —
#                                           pass --flatten to place every
#                                           converted file directly in DEST
#                                           instead; --delete = remove
#                                           originals after a successful convert;
#                                           --pause = seconds between files, e.g.
#                                           to let an external drive cool down)
#   fm.py convert-image SOURCE [DEST] [-r] [--flatten] [--format jpg|png]
#                       [--delete] [--pause SECONDS]
#                                          (convert .heic/.heif/.bmp/.tiff/.webp/.gif
#                                           to .jpg or .png via Pillow + pillow-heif;
#                                           same -r/--flatten/--delete/--pause
#                                           behavior as convert-video)
#   fm.py cleanup junk ROOT [--delete] [--yes]        (.DS_Store / desktop.ini / *.bak)
#   fm.py cleanup logs [FOLDER] [--days N] [--delete] [--yes]   (purge old log entries)
#
# NOTE: All Remove actions default to a DRY RUN (nothing is deleted). Deletion
#       only happens when explicitly opted-in (interactive confirm, or the
#       --delete flag on the CLI). Sync is the same: it previews and only
#       copies after explicit opt-in (interactive confirm, or --copy).
#
# NOTE: Every operation that creates, modifies, or deletes files (Sync,
#       Convert Video to MP4, Convert Images, Zip SubFolders, all Remove/
#       Cleanup actions, Find & Replace, Find & Rename) shows its full set of
#       chosen options, asks a final "go ahead?" confirm, then plays a
#       Backup42-style 3..2..1 countdown (ESC/Q cancels on an interactive
#       tty) right before it starts. This always plays — interactive menu
#       AND CLI runs, even with --yes/--apply/--delete/--copy — though a
#       non-tty/piped run can't read a cancel key and just counts down.
#
# NOTE: fm.log records the start and end of every run with a summary, plus
#       one timestamped [YYYY-MM-DD HH:MM:SS] line per action (throttled to
#       periodic milestones above ACTION_LOG_DETAIL_THRESHOLD items) — the
#       format LogView (~/Documents/script/LogView) expects. This is on top
#       of, not instead of, the existing full-screen mirror that's appended
#       once each run finishes.
#
# NOTE: Every batch of 2+ items in the same operations above can be paused
#       or stopped mid-run on an interactive tty: [P] pauses (press [P]
#       again to resume, or [Q]/ESC to stop from within the pause), [Q] or
#       ESC stops directly. Stopping always finishes the current item first
#       — nothing is left half-written. No effect on non-interactive
#       (piped/cron) runs.
#
# -----------------------------------------------------------------------------
# Revision History:
# -----------------------------------------------------------------------------
# v2.00 (2026-08-12)
#   • Convert Video to MP4: the "[i/N] filename" line now also shows the
#     source file's size (e.g. "(123.4 MB)") right after the filename, so
#     it's clear at a glance which file is being converted and how large it
#     is before the spinner starts ticking.
#
# v1.99 (2026-08-12)
#   • Convert Video to MP4: each file's write_videofile() call (moviepy,
#     logger=None so it prints nothing on its own) now runs in a background
#     thread while the main thread animates a spinner + elapsed time after
#     the "[i/N] filename" line — a slow/large conversion with no visible
#     change previously looked indistinguishable from a hang. New
#     _spin_while_writing() helper; shared by both Convert Video to MP4 and
#     the AWS S3 wizard's Optimize Video step via _convert_video_files_core().
#     New import: threading.
#
# v1.98 (2026-08-12)
#   • Main Menu reordered to: Display, Find, Compare, Monitor, Eject, Zip,
#     Convert, Optimize Media, Sync, Remove, Clean Up, Mount Shares,
#     Permissions, Environment Vars, Create Random UID, AWS S3, Local
#     Scripts, Admin (was Compare, Convert, Display, Find, Eject, Monitor,
#     Sync, Zip, Optimize Media, Remove, Create Random UID, Clean Up, Mount
#     Shares, Permissions, Environment Vars, ...). Purely cosmetic — no
#     functional change; each option's [N] key and dispatch target were
#     renumbered to match its new position.
#
# v1.97 (2026-08-12)
#   • Removed Server Jobs (Main Menu item + its Admin-adjacent [A]/[E]/[D]
#     add/edit/delete flow, the fmConfig.json serverJobs/serverJobLog keys,
#     and the DocInfo Manager serverJob-linking machinery — server_jobs_menu(),
#     _load_server_jobs(), _edit_server_job_flow(), _delete_server_job_flow(),
#     _SERVER_JOB_INFO_CACHE, _load_server_job_log_shared_secret(),
#     _fetch_server_job_info(), refresh_linked_server_jobs_info(),
#     _log_server_job_start(), and its post-login call site in authenticate()):
#     functionally identical to Local Scripts (same
#     _open_terminal_tab_and_run() run mechanism, same {name, command}
#     shape) minus the optional DocInfo Manager link, so it was redundant.
#     The one saved entry (fmConfig.json's "Server Health Check", which had
#     no serverJobId) was migrated into localScripts. Main Menu renumbered:
#     Environment Vars 15, AWS S3 16, Local Scripts 17, Admin 18 (was 19).
#
# v1.96 (2026-08-12)
#   • Local Scripts now always sorts alphabetically by name (case-
#     insensitive): _load_local_scripts() sorts on every load (Main Menu ->
#     Local Scripts and Admin -> Manage Local Scripts both use it), and
#     _edit_local_script_flow() re-sorts before saving an add/edit so
#     fmConfig.json's localScripts array stays alphabetical at rest too.
#
# v1.95 (2026-08-12)
#   • Added Local Scripts: a new Main Menu item (just before Admin) listing
#     saved local scripts (fmConfig.json -> localScripts, {name, command})
#     in an arrow-driven screen where [Enter] runs the highlighted script's
#     execution string in a new terminal tab via the existing
#     _open_terminal_tab_and_run() (same mechanism as Server Jobs). Adding,
#     editing, and deleting entries is done separately via a new Admin Menu
#     item, Manage Local Scripts (picker-style add/edit/delete, modeled on
#     manage_common_folders_menu()) — unlike Server Jobs, the Local Scripts
#     screen itself has no [A]/[E]/[D] and no DocInfo Manager job linking.
#
# v1.94 (2026-08-04)
#   • Dry-run removal previews (_preview_removal) now show each candidate's
#     size AND created/modified date on its numbered line (new _file_dates()
#     helper — st_birthtime with mtime fallback). All 6 duplicate-removal
#     paths (Delete Duplicates after Find Duplicates, Remove Duplicates by
#     Name/Hash/Fuzzy Name, and CLI --dupname/--duphash) now also build a
#     keep_map {removed_path: surviving_path} and pass it through
#     _finish_removal()'s new keep_map param, so each removal candidate's
#     surviving copy (folder/name + its own size/created/modified) prints
#     beneath the same numbered entry. Every other _finish_removal() caller
#     (Remove Folder/Files/Folders by Name, 0-Size, Purge/Cleanup, etc.) is
#     unaffected — keep_map defaults to None and only duplicate flows pass it.
#
# v1.93 (2026-08-02)
#   • Every dry-run-capable task (Remove Folder/Files/Folders by Name, Remove
#     Duplicates by Name/Hash/Fuzzy, Remove Files of 0 Size, Clean Up Junk
#     Files, Clean Up Purge Logs, Sync Folders) now asks "Run in dry-mode
#     first? (Y/n)" BEFORE any scanning/matching begins (previously asked
#     after the scan, only gating the preview listing), and prints an
#     unmistakable "Mode: Dry-Mode" / "Mode: Live Mode" line as the first
#     visible content on the task screen — new ask_dry_mode()/
#     print_mode_line() helpers. Live Mode also skips the detailed grouped
#     listing (Duplicates flows, Purge Logs) in favor of a one-line count, so
#     it's genuinely quieter, not just relabeled. Sync pins the Mode into its
#     stationary frame so it stays visible throughout scanning, the file
#     list, and the copy itself. Find & Replace / Find & Rename (always
#     dry-run-first by design, no skip offered) just get the Mode label
#     added — no behavior change. The existing safety invariant is
#     unchanged: nothing is ever deleted/moved/copied/overwritten without a
#     separate, explicit final confirmation, dry-run or not.
#
# v1.92 (2026-08-01)
#   • Remove Duplicates (by Hash): the dry-run listing now shows the file
#     size on the "keep" line itself (next to its path), not just once in
#     the group header — matches the "remove" lines, which now also show
#     their own size. Since this runs inside _ActivityLog, the change is
#     automatically mirrored into the fm.log activity log too.
#
# v1.91 (2026-08-01)
#   • Remove (Remove Folder / Remove Folders by Name / Remove Duplicates by
#     Name/Hash/Fuzzy — all via _finish_removal) and Sync: the dry-run
#     preview listing is no longer forced every time in interactive mode.
#     A new upfront "Dry run first?" question (default Yes) lets you skip
#     straight past the itemized preview; the final "Actually delete/copy
#     these N item(s)?" confirm still always happens either way — nothing
#     is ever deleted/copied without that explicit go-ahead, dry run or not.
#     CLI (--delete/--move-to/--copy or dry-run-only flags) is unaffected —
#     those already choose dry-run vs. live explicitly via flags.
# v1.90 (2026-08-01)
#   • Remove → Duplicates by Name/Hash/Fuzzy Name, and Find Duplicates → [D]
#     Delete Duplicates: added a "move to a folder" alternative to deleting.
#     Interactively, after the usual dry-run preview you now choose
#     [D] Delete / [M] Move to Folder / [C] Cancel instead of a plain Y/N;
#     moving asks for a destination folder (creating it if needed) and only
#     needs a normal confirm (it's reversible, unlike delete which still
#     requires typing YES).
#   • CLI: `fm.py remove dup-name|dup-hash|dup-fuzzy ... --move-to FOLDER`
#     moves instead of deleting (mutually exclusive with --delete; --yes
#     skips the confirm). Rejected for non-duplicate remove types.
#   • Moved files are never overwritten: a name collision at the destination
#     gets an auto " (1)", " (2)", ... suffix. New _unique_dest_path(),
#     _perform_move(); _finish_removal() gained allow_move/move_to params
#     shared by all four duplicate-removal call sites.
#
# v1.89 (2026-07-29)
#   • AWS S3 credentials (Admin Menu) simplified to just the auth triple —
#     Access Key ID, Secret Access Key, Region. Bucket and folder/prefix are
#     no longer stored there; they're now entered per task, right where the
#     upload happens: Zip SubFolders' "Push files to AWS S3?" prompt asks
#     for a bucket (+ optional folder) only when you say Yes, and the AWS
#     S3 wizard's step 8 (already per-task) no longer defaults from a
#     saved bucket. Lets one set of credentials push to a different bucket
#     each run. is_aws_configured() no longer requires a bucket to be
#     "configured" — just the two credential fields.
#   • _s3_push_file()/_s3_push_zip() no longer fall back to an Admin-
#     configured bucket — bucket is now a required argument from every
#     caller. zip_subfolders() gained s3_bucket/s3_folder params;
#     _prompt_zip_subfolders() now asks for them when Push to AWS S3 is Yes
#     (falls back to Push disabled for that run if no bucket is entered).
#   • Main Menu: Admin is the last option again ([18], after AWS S3 at
#     [17]) — the AWS S3 addition in v1.86 had appended it after Admin,
#     breaking the "Admin is always last" convention established in v1.65.
#   • AWS S3 is now grayed out on the Main Menu (with a reason shown
#     inline) when it can't be used yet: "boto3 not installed" or "AWS
#     credentials not configured — see Admin Menu". Recomputed fresh on
#     every Main Menu redraw, so setting credentials in Admin Menu and
#     backing out immediately un-grays it, no restart needed. Admin
#     itself is never gated — it's where those credentials get set.
# v1.87 (2026-07-29)
#   • Centralized activity logging: FM's server-side logging (via
#     api/zipFileLog.php and api/fileLog.php on doc_cloudbox9_com) now
#     writes to docInfo.activityLog instead of CB9Inventory.activityLog —
#     the two systems' activity history is no longer split across two
#     databases. fmConfig.json's logZip block gained serverId/docProjectId
#     (1/87 — "Ash Mac" / "File Management"), sent on every zipFileLog.php/
#     fileLog.php call so entries can be filtered by originating system;
#     both are optional, matching the new activityLog columns' own
#     nullability. No fm.py menu/behavior changes — this is a config +
#     wire-format addition only (_post_zip_log()/_post_file_log()).
# v1.86 (2026-07-29)
#   • New Main Menu option [18] AWS S3, with four submenu options:
#       - Display Buckets & Sizes — every bucket the saved credentials can
#         see, with object count and total size (lists every object —
#         S3 has no cheap per-bucket size API).
#       - Optimize Media, Zip, Log & Upload to AWS S3 — the wizard
#         (aws_s3_optimize_menu()): Folder/File -> Optimize Video (+ delete
#         originals) -> Resize Images (height/width, default 1200x1200,
#         fit-within-bounding-box preserve-aspect-ratio never-upscale,
#         '_HEIGHT-WIDTH' filename suffix, + delete originals) -> Create
#         Media Thumbnails (80x80, '_tmb' suffix) -> Zip to 1 File / Zip
#         Subfolders / neither -> Log Contents -> AWS S3 Bucket (+ folder/
#         prefix) -> Review screen -> 3..2..1 countdown (ESC/Q during the
#         countdown returns to Review with the same answers, not re-asked,
#         not aborted). Zipped output logs via the existing zipFileLog
#         path; unzipped files log individually via the new fileLog API.
#       - Search AWS S3 — filename PREFIX match across every object in
#         every bucket (S3 has no server-side wildcard support).
#       - Search DocInfo Manager Records — full wildcard filename search
#         merging zipFileContent (zips) and the new fileLog (individual
#         files) results into one table.
#   • New CB9Inventory.fileLog table (already existed, unused — this wires
#     it up): logging/uploading individual media files without zipping them
#     first goes through the new api/fileLog.php + api/fileLogSearch.php
#     endpoints (cb9InventoryModel::fileLog()/fileLogSearch(), doc_cloud
#     box9_com), matched by fileName + fileFolder + fileSizeBytes, same
#     dual-auth pattern as api/zipFileLog.php. Added a matching
#     fileLogBeforeUpdateArchive trigger (fileLog -> fileLogArchive).
#   • zipFileLog()/api/zipFileLog.php now accept optional s3Bucket/s3Key/
#     uploadConfirmed — the AWS S3 wizard records where a zip landed on the
#     same zipFile row (columns already existed, previously unused).
#   • Refactored convert_videos_to_mp4()/convert_images_to_format() into
#     thin UI wrappers over new no-I/O core loops
#     (_convert_video_files_core()/_convert_image_files_core()) so the
#     wizard can drive the same conversion engine directly without nested
#     screens/prompts. _convert_image_files_core() gained an optional
#     resize_to=(w,h) param (also used stand-alone by the Resize step).
#     New _create_thumbnails() and zip_folder_to_one_file() (no existing
#     feature zipped a whole folder to one archive before this).
#   • _s3_push_zip() generalized into _s3_push_file(local_path, bucket,
#     folder) so the wizard can target a different bucket/folder per run
#     than the Admin-configured default; _s3_push_zip() kept as a thin
#     backward-compatible wrapper for Zip SubFolders.
# v1.85 (2026-07-29)
#   • Main Menu: shortened the "Environment Vars — Display All" label to just
#     "Environment Vars" (the description text still explains Display All).
#   • Fixed the VERSION constant, which had been stuck at "1.62" since
#     2026-07-22 even as the Revision History (and the "# Version:" header
#     line above) kept advancing — every release from v1.63 through v1.84
#     updated this comment but never bumped the actual constant, so the
#     app's own exit/About screens were showing a version 22 releases stale.
# v1.84 (2026-07-27)
#   • Zip menu: new "Search Zip File Contents by Name" (option 6) — searches
#     filenames across every archive already LOGGED to the CB9Inventory
#     database on BPA5 (distinct from "Find Files in Zip", which searches
#     local zip files directly). Wildcards ('*', '?') are translated
#     server-side to SQL '%'/'_'; scope optionally to one zip file (exact
#     name) or a folder (+ its subfolders). New
#     search_zip_file_contents_db() and _post_zip_content_search(), calling
#     the new api/zipFileContentSearch.php endpoint (doc_cloudbox9_com) ->
#     cb9InventoryModel::zipFileContentSearch().
# v1.83 (2026-07-27)
#   • Zip -> Log Zip File Contents: new "Include subfolders?" prompt when the
#     target is a folder (default No, matches prior behavior). log_zip_files()
#     gained a `recursive` param — when set, walks every subfolder (hidden
#     folders excluded via prune_dirs(), same as Find Files in Zip) instead of
#     just the top level. New CLI flag: `fm.py zip-log TARGET -r/--recursive`.
# v1.82 (2026-07-27)
#   • ask_path() (used by every folder AND file-or-folder prompt, including
#     Zip -> Log Zip File Contents) now also resolves a saved common folder
#     by typing its name directly (case-insensitive), not just via the '+'
#     picker. New _resolve_common_folder_name(). Fixes "Path not found:
#     PendingUpload" when a common folder's name was typed as-is instead of
#     using '+'.
# v1.81 (2026-07-27)
#   • Admin Menu -> Manage Common Folders now shows "name - path" directly in
#     the option label (was: name only, path hidden behind [H] Help).
# v1.80 (2026-07-26)
#   • Display -> Drives in Use: lists every mounted external drive and
#     whether it's currently being read from or written to. Detection
#     samples each drive's underlying physical disk with iostat for ~1s and
#     checks for actual transfers (not just an open file handle), so an
#     idle-but-open file doesn't falsely read as in use. macOS only. Also
#     added as CLI command `fm.py drives-in-use`.
# v1.79 (2026-07-25)
#   • Added [A] About to the Main Menu — shows app name/version, description,
#     maintainer, Python version, and the copyright notice.
# v1.78 (2026-07-25)
#   • Removed the feature-list intro line ("Compare · Convert · Display · ...")
#     that appeared under the Main Menu title on startup.
# v1.77 (2026-07-25)
#   • New Main Menu option "Mount Shares" (13th, right after Clean Up) —
#     mounts SMB shares via `open smb://...` (macOS's own mount service;
#     /Volumes is root-owned so a plain mkdir/mount_smbfs fails for any
#     share not mounted before — same fix worked out interactively earlier
#     today for a one-off mountShares.sh script). Mount Manually (enter
#     server/username/share names on the spot) or save/run named profiles
#     (fmConfig.json -> mountProfiles: name, server, username, shares[]),
#     mirroring the existing Permissions profile Run/Set pattern. A literal
#     "@" in the username (email-style logins) is percent-encoded as %40.
#     After issuing the mount requests, briefly polls /Volumes and reports
#     which shares mounted immediately (saved Keychain credentials) versus
#     still pending (most likely a Finder password prompt — not a failure).
#     macOS only; not added to FM_Win in this pass.
# v1.76 (2026-07-25)
#   • New _check_config_file(), called right after _check_optional_dependencies()
#     at the top of main() — prints a one-line green confirmation if
#     fmConfig.json exists, or a yellow warning naming which features need it
#     (Sync/Monitor/Compare/Permission profiles, common folders, Server Jobs)
#     and pointing at fmConfig.sample.json to copy from. fmConfig.json was
#     already fully optional (every reader already handles it being absent);
#     this just surfaces that state up front instead of silently.
# v1.75 (2026-07-25)
#   • New _check_optional_dependencies(), called at the very top of main()
#     (interactive and CLI alike) — prints a one-line green confirmation if
#     boto3/moviepy/Pillow/pillow-heif are all installed, or a yellow
#     warning listing which are missing, which feature each one disables,
#     and the pip install command to fix it. These were already soft
#     dependencies (each _XXX_AVAILABLE flag gates its own feature and grays
#     it out in-menu); this just surfaces that state up front on every run
#     instead of only when the user stumbles into the specific screen that
#     needs it.
# v1.74 (2026-07-24)
#   • Server Jobs entries can now optionally link to a DocInfo Manager
#     serverJob record via a new per-job serverJobId field. Missing/blank
#     serverJobId (or a missing/blank fmConfig.json serverJobLog.
#     sharedSecret) means the job behaves exactly as before — no network
#     calls, nothing extra.
#   • On a successful DocInfo Manager login (Admin Menu), FM now calls
#     refresh_linked_server_jobs_info(), which fetches name/description/
#     last-run for every linked job via xhr/serverJobGetWithLastRun.php
#     (that endpoint requires no auth as currently implemented server-side)
#     and caches it in memory (_SERVER_JOB_INFO_CACHE) for display on the
#     Server Jobs screen — e.g. "Linked: DocInfo Manager \"Health Check\" —
#     Last run: Jul 24, 2026 10:00 am" under the job.
#   • Selecting a linked job and pressing Enter now also posts a best-effort
#     "start" event to DocInfo Manager (xhr/serverJobLogAdd.php, new
#     _log_server_job_start(), sharedSecret-authenticated) right before
#     opening its terminal tab — never blocks the launch on failure. FM only
#     ever logs "start"; the script being run is expected to log its own
#     end/error events, since FM has no visibility into a detached tab's
#     output.
#   • _edit_server_job_flow() gained a third prompt: "DocInfo Manager Server
#     Job ID (optional, blank = none)" — validates as a positive integer or
#     blank, retrying on anything else.
#   • New fmConfig.json top-level key: serverJobLog ({sharedSecret}).
# v1.73 (2026-07-24)
#   • New Main Menu option "Server Jobs" (second-to-last, before Admin): an
#     arrow-driven list of saved commands (fmConfig.json -> serverJobs, each
#     {name, command}). [Enter] opens a new tab in whichever terminal app FM
#     is running in — Terminal.app or iTerm2, detected via $TERM_PROGRAM —
#     and runs the command there via new _open_terminal_tab_and_run();
#     FM's own menu keeps running independently. Terminal.app has no native
#     "new tab" AppleScript verb, so a Cmd+T keystroke (System Events) is
#     simulated first, then `do script` targets the resulting front window.
#   • [A] Add, [E] Edit, [D] Delete are hotkeys directly on the job list
#     screen itself (not a separate menu) — new server_jobs_menu(),
#     _edit_server_job_flow(), _delete_server_job_flow(). Unlike common
#     folders, job names may contain spaces (e.g. "Server Health Check") —
#     they're just a menu label, never a picker key.
# v1.72 (2026-07-24)
#   • Every folder prompt (ask_folder()/ask_path()) now prints a static
#     "Common: name1, name2, ..." header line above the prompt, listing the
#     saved common folders (Admin Menu -> Manage Common Folders) by name —
#     visible without typing '+' to open the picker. New
#     _common_folder_names_line(). Blank/omitted when none are saved.
# v1.71 (2026-07-24)
#   • Fixed: the '+' common-folder shortcut only worked at pure-folder
#     prompts (ask_folder()) — file-or-folder prompts called plain ask()
#     instead, so typing '+' there was passed straight through as a literal
#     path and failed ("Path not found: +"), e.g. Zip -> Log Zip File
#     Contents. New ask_path() (ask() + the '+' picker loop, no
#     must-be-a-directory check) is now shared by ask_folder() and by every
#     file-or-folder / folder-destination prompt: Zip file/folder to
#     view/search/log, Zip destination, and both Optimize Media "Destination
#     folder" prompts.
# v1.70 (2026-07-24)
#   • Fixed: every prompt after the first one on a "Manage Common Folders"
#     Add/Edit screen (and any other multi-question flow) rendered
#     progressively indented, staircasing to the right. Root cause:
#     _read_line_esc() and menu_read() both call tty.setraw(fd), which
#     disables OPOST (the terminal's automatic LF -> CRLF translation);
#     their bare "\n" write on Enter/ESC therefore dropped the cursor a row
#     without returning it to column 0, leaving it wherever the previous
#     answer's text ended. Both now write "\r\n" explicitly.
# v1.69 (2026-07-24)
#   • New commonFolders shortcut: type '+' and press Enter at ANY folder
#     prompt in the app (Compare, Sync, Monitor, Find, Remove, Permissions,
#     Zip, ...) to pick from a saved list with ↑/↓ + Enter instead of typing
#     a path — Q/ESC goes back to typing one. Built into ask_folder() itself
#     (via new _pick_common_folder()) so every existing call site gets it
#     for free, no per-menu wiring.
#   • New Admin Menu entry "Manage Common Folders": add/edit/delete the
#     saved folders (name + path; name may not contain spaces, since it's
#     the picker's short label). New manage_common_folders_menu(),
#     _edit_common_folder_flow(), _delete_common_folder_flow() — reuse the
#     existing _load_config_profiles()/_save_config_profiles() helpers
#     already shared by sync/monitor/compare/permission profiles.
#   • New fmConfig.json / fmConfig.sample.json key: commonFolders (list of
#     {name, path}).
# v1.68 (2026-07-23)
#   • New Zip menu option "Find Files in Zip": search filenames INSIDE a zip
#     archive (wildcards ok, e.g. fan*.png) — matches by filename only, never
#     contents. Give a single .zip to search just that archive, or a folder to
#     recursively find every .zip under it and search each one, with results
#     grouped by which zip they were found in.
#     New functions: find_files_in_zip(), find_files_in_zips_under_folder(),
#     find_files_in_zip_target(). New CLI: `fm.py find-zip ZIP|FOLDER [PATTERN]`.
# v1.67 (2026-07-23)
#   • Zip → Log Zip File Contents now remembers the last file/folder entered
#     (LAST_PATHS["zip_log_target"], persisted to fmLastPaths.json like every
#     other keyed prompt) — Enter reuses it, typing a new path overrides it.
#     Applies to the interactive prompt and the `fm.py zip-log` CLI path
#     (both feed the same remembered value).
# v1.66 (2026-07-23)
#   • Zip menu's "Log Zip File Contents" [H] Help now spells out what
#     "matched by name + size" and "updated" actually mean for zipFileContent:
#     inserted (new file) / updated + reactivated (changed file) / left alone
#     (unchanged) / soft-deleted, deleted=1 (file no longer in the archive) —
#     and that every zipFileContent/zipFile UPDATE, including the soft-delete,
#     is archived pre-change via the BEFORE UPDATE triggers into
#     zipFileContentArchive/zipFileArchive. Same detail added to README.md's
#     "8 · Zip" section.
# v1.65 (2026-07-23)
#   • Main Menu: Admin is now the last option ([15], after Environment Vars
#     at [14] — previously Admin was [14] and Environment Vars was [15]).
# v1.64 (2026-07-23)
#   • Help screens are easier to read: _print_help_desc() now inserts a
#     blank line between paragraphs (bullet lines still stay tight against
#     each other, no blank line between one bullet and the next). Rewrote
#     the densest option descriptions across every menu — Compare, Find,
#     Remove, Zip, Optimize Media, Clean Up, Permissions, and the Main
#     Menu's help_note — as short paragraphs and bullet lists instead of
#     one run-on block of text.
# v1.63 (2026-07-23)
#   • Log Zip File Contents now computes an MD5 hash of the whole archive
#     (_file_md5(), same block-read approach as vlcmenu's duplicate finder)
#     and sends it as zipFileHash to api/zipFileLog.php, which stores it on
#     the new zipFile.zipFileHash column. Stored only for now — matching an
#     existing zip row is still by zipFileName + zipFileSizeBytes; hash is
#     there to support a future duplicate finder.
# v1.62 (2026-07-22)
#   • Added a new Main Menu option, Permissions ([13] — Admin and Environment
#     Vars each shifted down one, to [14]/[15]), with three submenu options:
#     - Set Apache Permissions: one-shot chown -R + chmod fix for /var/www
#       (directories 755, files 644, root:www-data on Linux / _www:_www on
#       macOS) via sudo — FM's first use of sudo. Grayed out with a reason
#       when /var/www doesn't exist (Apache not installed).
#     - Run a Profile: picks a saved fmConfig.json permissionProfiles entry,
#       scans its folder for files matching its pattern, previews every file
#       whose permission would change, and applies it only after you confirm
#       — a DRY RUN until then, same countdown/pause/stop/RunLog machinery as
#       Remove/Sync/Zip SubFolders.
#     - Set a Profile: creates or edits a saved permission profile (name,
#       folder, file name/type/wildcard pattern, a single octal permission
#       applied to every match, recursive) — pre-fills current values when
#       editing, same pattern as Admin Menu's Set/Update AWS Credentials.
#   • New fmConfig.json / fmConfig.sample.json key: permissionProfiles (with
#     a worked example + _help entry in the sample). Menu-only — no CLI
#     subcommand, matching Admin Menu and Environment Vars.
#   • Permission-matching scans include hidden files/dotfiles (unlike Sync/
#     Compare/Find's usual hidden-file exclusion) since permission fixes
#     commonly target .htaccess and similar; the universal junk-name
#     exclusions (.DS_Store, desktop.ini, $RECYCLE.BIN) still apply.
#   • Added _apache_available(), _apache_owner_group(), _load_permission_
#     profiles(), _valid_octal_permission(), _normalize_permission_pattern(),
#     _scan_permission_matches(), _preview_permission_changes(),
#     _apply_permission_changes(), _run_permission_profile(),
#     run_permission_profile_menu(), _edit_permission_profile_flow(),
#     set_permission_profile_menu(), set_apache_permissions(), and
#     permissions_menu(). New import: stat (for stat.S_IMODE()).
#
# v1.61 (2026-07-22)
#   • Added [14] Environment Vars to the Main Menu, with one submenu option,
#     Display All: lists every environment variable (os.environ), name and
#     value, sorted alphabetically. Read-only.
#   • Deliberately NOT wrapped in _ActivityLog()/_RunLog() — environment
#     variables commonly hold secrets (API keys, tokens, credentials), and
#     this app already takes care never to write those to fm.log (Admin
#     Menu's password/AWS key handling). Screen-only, matching that same
#     caution.
#   • Added environment_menu() and _display_environment_vars().
#
# v1.60 (2026-07-22)
#   • Compare → Compare Folder Contents → Interactive: Compare By is no
#     longer a separate full-screen menu (render_menu wiped the screen with
#     its own "Compare Folder Contents — Compare By" header right after
#     Folder A/B/Recursive/Case-sensitive were answered inline). Replaced
#     _pick_compare_by() with _ask_compare_by(), a plain inline prompt (same
#     no-screen-change style as the Recursive/Case-sensitive questions), so
#     every option is now gathered in one continuous sequence with zero
#     screen changes in between. compare_folder_contents() still does its
#     one and only redraw afterward, showing everything chosen, before
#     running the comparison.
#   • ESC during the Compare By prompt now raises EscCancelled like every
#     other question in the wizard (was a render_menu 'back'/None return
#     handled by a separate `if by is None: continue` — now redundant and
#     removed; compare_menu()'s existing try/except EscCancelled covers it).
#
# v1.59 (2026-07-22)
#   • Clean Up → Remove Junk Files now also targets *.bak files, alongside
#     .DS_Store and desktop.ini. New JUNK_FILE_SPECS list (label, description,
#     matcher) is independent of the shared EXCLUDED_FILE_NAMES/
#     is_excluded_file() used by Compare/Find — .bak files are NOT silently
#     excluded from those unrelated features, only added to this cleanup.
#   • Choosing Remove Junk Files now first shows a checklist of the 3 file
#     specs (all checked by default) — ↑/↓ move, Space toggles, Enter
#     confirms — so you can unselect any you want to keep before it asks for
#     the root folder. cleanup_junk_files() takes the picked spec_indices;
#     the CLI (fm.py cleanup junk ROOT) has no interactive checklist and
#     keeps removing all three by default.
#   • render_multiselect() gained a preselected= parameter (defaults to
#     nothing pre-checked, so the existing Find Files criteria picker is
#     unaffected) — pass a set/range of indices to start them checked; blank
#     Enter on the non-TTY fallback now keeps those defaults instead of
#     always cancelling when preselected is given.
#
# v1.58 (2026-07-22)
#   • ESC now cancels ANY question prompt instantly and returns to the menu it
#     was asked from, everywhere in the app — previously only render_menu()/
#     pause_return()/pause_rerun() reacted to ESC; free-text prompts (ask(),
#     ask_folder(), ask_file()) and Y/N prompts (safe_confirm(),
#     confirm_yes_word()) used plain input(), where ESC just typed an
#     invisible escape character into the buffer and did nothing useful.
#   • New EscCancelled exception + _read_line_esc(): a raw-mode line reader
#     (same termios/tty/select technique as the existing menu_read()/
#     read_key(), but case-preserving) that raises EscCancelled the instant a
#     bare ESC is pressed, no Enter needed. ask() and safe_confirm() (now
#     fully self-contained — no longer delegates to CB9Lib's confirm(), so
#     CB9Lib and every other script using it is unaffected) are rebuilt on
#     top of it; confirm_yes_word() too.
#   • Every top-level *_menu() (Compare, Display, Find, Remove, Monitor, Sync,
#     Zip, Optimize Media, Clean Up, Admin, Main) now wraps its option
#     dispatch in try/except EscCancelled: continue, so an ESC anywhere in a
#     multi-question wizard (e.g. Folder A entered, ESC on Folder B) abandons
#     that action and instantly redraws the menu it came from — matching the
#     request that ESC behave this way under any submenu in the script.
#     One-shot flows with no menu loop of their own (Convert, Create Random
#     UID, Interactive Sync/Monitor, the Zip/Optimize Media prompt helpers,
#     Admin's Login/Set AWS Credentials) aren't wrapped individually — ESC
#     inside them propagates up to whichever *_menu() launched them, which is
#     already exactly the "previous menu."
#   • main()'s CLI branch and the interactive main_menu() call are both also
#     wrapped in except EscCancelled as a last-resort net, so a stray
#     EscCancelled can never surface as an uncaught traceback.
#   • Known limitation: the DocInfo Manager password (Admin Menu → Login) and
#     AWS Secret Access Key still use Python's getpass.getpass(), which reads
#     in cooked/masked mode and can't detect a bare ESC mid-keystroke without
#     reimplementing masked input from scratch — left as-is rather than risk
#     that rewrite. Every other prompt in the app (including the Login
#     username and the other AWS credential fields) is ESC-aware.
#
# v1.57 (2026-07-22)
#   • Compare → Compare Folder Contents now asks Interactive (default) or Run
#     Profile first:
#     - Run Profile lists saved profiles from fmConfig.json's new
#       compareProfiles array (folderA/folderB/recursive/compareBy/
#       caseSensitive) and runs the chosen one directly.
#     - Interactive is unchanged going in, but its end-of-run footer is now
#       [R] Run Again / [S] Save as Profile / [Q/ESC] Quit/Back — [S] prompts
#       for a name and appends the folders/options just used to
#       compareProfiles (fmConfig.json), preserving every other setting in
#       the file.
#   • compare_folder_contents() refactored: the single-pass report moved into
#     _compare_folder_contents_screen(); the public function now loops
#     instead of recursing on Run Again, and takes a new offer_save flag
#     (True for Interactive, False for a profile run — already saved, no
#     need to offer again).
#   • Added _pause_compare_folders(), _save_compare_profile_flow(),
#     _run_compare_profile(), _pick_compare_mode(), _compare_profile_menu(),
#     _load_compare_profiles(), and generic _save_config_profiles() (a
#     read-modify-write helper for writing any profile list back to
#     fmConfig.json without touching auth/logZip/other profile lists).
#   • fmConfig.json / fmConfig.sample.json: added the compareProfiles key
#     (with a worked example + _help entry in the sample).
#
# v1.56 (2026-07-21)
#   • Closed the remaining fm.log coverage gaps so every menu action leaves a
#     trace, with per-file detail for anything that creates/updates/deletes/
#     uploads a file (Sync, Zip, Optimize Media, Remove, Clean Up) and a
#     start/options/errors/end summary for everything else:
#     - Admin Menu (DocInfo Manager login/logout, AWS S3 credential set/
#       update/clear) previously logged nothing at all — now each logs a
#       _RunLog start/finish, including the username or bucket/region/folder
#       chosen. Passwords and the AWS secret/access keys are never logged.
#     - Zip → Log Zip File Contents previously only wrote to the CB9Inventory
#       DB, leaving zero trace in fm.log — now wrapped in a _RunLog with one
#       action() line per archive (logged/failed/recorded-as-Failed) plus the
#       early-cancel paths (no login, no path, not a zip, etc.).
#     - Zip SubFolders' pre-zip .DS_Store/desktop.ini cleanup deleted files
#       with no record anywhere (not even a filename on screen) — now each
#       deletion is logged individually via its own _RunLog.
#     - Monitor File Activity logged nothing to fm.log at all when the user
#       chose CSV output (events went to fmMonitor.csv only) — now a
#       start/options/end summary always lands in fm.log regardless of which
#       output the events themselves go to. The "not a directory" error is
#       now logged too.
#     - Compare 2 Files, Display All Drives, Display Folder Sizes, Find
#       Folders, Find Files (combined), Find Files by Name/Extension, and
#       Find Files Over/Under N MB previously had no logging of any kind —
#       each now gets a _RunLog start (with the options entered), the "not a
#       directory" error, and a finish summary.
#     - _perform_removal() (shared by every Remove/Clean-Up action) and
#       Purge Old Log Files logged successful file removals but silently
#       dropped failures on the floor — failures now get their own
#       runlog.action() "FAILED to remove/purge ..." line too.
#
# v1.55 (2026-07-21)
#   • Added the ability to pause or stop a long-running batch (Sync, Convert
#     Video/Images, Zip SubFolders, Remove/Cleanup, Find & Replace/Rename,
#     Purge Old Log Files) while it's running. New _PauseStop class: [P]
#     pauses immediately (blocks until [P] resumes or [Q]/ESC stops from
#     within the pause), [Q] or ESC stops directly — consistent with the
#     ESC/Q convention already used everywhere else (countdown cancel,
#     Monitor). Stopping is cooperative: the current item always finishes
#     before the loop breaks, so nothing is left half-written. A hint line
#     ("Press [P] to pause, [Q] to stop early.") prints before any batch of
#     2+ items. Non-interactive runs (piped/cron, no controlling tty) are
#     unaffected — check() is a permanent no-op there. The final summary
#     (on screen, in fm.log's _RunLog finish line, and in report_result)
#     notes "Stopped early by user (N/Total processed)" when it happens.
#     Verified with a pty-based test harness: both stop-mid-batch and
#     pause-then-resume behave correctly end to end.
#
# v1.54 (2026-07-21)
#   • fm.log now logs the start and end of every run with a summary, every
#     action individually timestamped, in the [YYYY-MM-DD HH:MM:SS] format
#     LogView (~/Documents/script/LogView) already expects. New _RunLog
#     class (start()/action()/finish()) — kept ALONGSIDE the existing
#     _ActivityLog raw full-screen mirror, not replacing it, so both appear
#     per run: the structured start/action(s)/finish lines write live as the
#     run progresses, followed by the complete raw-screen block once the
#     run ends. Every function that already used _ActivityLog now also
#     creates a _RunLog with the operation name and its folders/options
#     (source/dest, root/pattern, folders, etc.) as the "started" fields.
#   • Per-action granularity is total-item-aware (new ACTION_LOG_DETAIL_
#     THRESHOLD = 100): at or under 100 items every action gets its own
#     timestamped line; above that, action() throttles to one line per
#     MILESTONE_INTERVAL_SECONDS (still always logging the very first and
#     last item so the log shows real start/end), keeping a Sync of
#     thousands of files from flooding fm.log. Sync's old separate
#     _log_activity_start()/_log_milestone() helpers are gone, folded into
#     _RunLog (Sync now passes its own runlog through _sync_screen and
#     updates its total_items/detailed flag once the file count is known).
#   • Every _finish_removal() caller (all Remove/Cleanup actions, Delete
#     Duplicates, and the 5 CLI-only remove wrappers) now passes an
#     activity name + fields — one shared code path, so all thirteen
#     covered call sites got this for free. Convert Video to MP4, Convert
#     Images, Zip SubFolders, Find & Replace, and Find & Rename each log
#     one action per file (converted/zipped/replaced-in/renamed). Compare
#     Folder Contents, Find Duplicates (by Filename/Fuzzy Name), Find
#     Missing by Filename, Convert (data file), Create Random UID, and
#     Eject All External Drives — all read-only or single-shot — get a
#     start + finish summary (Eject also logs one action per drive).
#
# v1.53 (2026-07-21)
#   • fm.log's location is now configurable: fmConfig.json's new top-level
#     "activityLogPath" key (~ expanded) overrides the default. New
#     _get_activity_log_path() reads it at import time, falling back to
#     ~/Documents/log/fm.log when the key is absent OR the config file is
#     missing/invalid JSON — a fresh install with no fmConfig.json still logs
#     to the same place as before. Moved CONFIG_FILE next to SCRIPT_DIR
#     (was defined much later, near the Admin/auth section) so it's
#     available this early. Documented in fmConfig.sample.json's _help block.
#
# v1.52 (2026-07-21)
#   • Closed a gap in ~/Documents/log/fm.log coverage: Remove Folder, Remove
#     Files/Folders by Name, Remove Duplicates (by Name/Hash/Fuzzy), Remove
#     Zero-Size Files, Clean Up (junk files + purge logs), Zip SubFolders,
#     Convert Video to MP4, and Convert Images previously ran without any
#     _ActivityLog wrapping — their file-by-file output only ever appeared on
#     screen, never persisted. All nine now wrap their full screen output in
#     `with _ActivityLog():`, same as Sync/Find & Replace/Find & Rename/Find
#     Duplicates/Find Missing/Convert(data)/UUID/Eject already did.
#   • Also fixed the CLI-only path: `fm.py remove folder|name|folder-name|
#     dup-name|dup-hash` dispatches through separate _cli_remove_folder() /
#     _cli_remove_name() / _cli_remove_folder_name() / _cli_dupname() /
#     _cli_duphash() wrappers that duplicate the interactive functions' logic
#     but hadn't been wrapped either — a CLI remove could delete files with
#     zero record in fm.log even after the fix above. Wrapped all five.
#     (`dup-fuzzy` and `zero-size` already called the real, now-wrapped
#     functions directly, so those two were fixed by the first bullet.)
#
# v1.51 (2026-07-21)
#   • New shared _run_countdown() helper: a Backup42-style 3..2..1 block-digit
#     countdown (same glyph art) shown right after the user confirms a
#     file-creating/modifying/deleting operation, right before it starts.
#     ESC/Q cancels on an interactive tty (treated the same as the confirm
#     itself answering "No"); a non-tty run (piped/cron) prints the count and
#     sleeps instead, always proceeding. Plays in BOTH the interactive menu
#     and CLI runs (including --yes/--apply/--delete/--copy), per project
#     convention that the countdown always shows.
#   • Wired the countdown into every operation that writes/deletes files:
#     Sync (after "Actually copy...?"), all Remove/Cleanup-junk actions (via
#     the shared _finish_removal), Purge Old Log Files, Find & Replace, and
#     Find & Rename. Convert Video to MP4 and Zip SubFolders previously had
#     no "go ahead?" gate at all before running — both now show a
#     "Convert/Zip these N item(s)?" confirm (default Yes) followed by the
#     countdown before any file is touched.
#   • Main Menu → Optimize Media: new "Convert Images" option alongside
#     Convert Video to MP4. Converts .heic/.heif/.bmp/.tiff/.webp/.gif files
#     to .jpg or .png (your choice each run) via Pillow + pillow-heif (HEIC/
#     HEIF support — e.g. iPhone photos). EXIF orientation is applied via
#     ImageOps.exif_transpose so rotated photos come out right-side up
#     (Pillow doesn't do this automatically on save). Same options as
#     Convert Video to MP4: source/dest folders, recursive with the
#     mirror-vs-flatten subfolder prompt, Clean Up (delete originals, gated
#     on reopening/verifying the new file first — new
#     _validate_converted_image()), and a pause-between-files setting.
#     Grayed out in the menu (like Convert Video to MP4 without moviepy)
#     when Pillow and/or pillow-heif aren't installed, with a pip install
#     hint. New CLI: fm.py convert-image SOURCE [DEST] [-r] [--flatten]
#     [--format jpg|png] [--delete] [--pause SECONDS].
#
# v1.50 (2026-07-21)
#   • Convert Video to MP4: when Subfolders is Yes AND the destination folder
#     differs from the source, a new prompt asks whether to recreate the
#     matching subfolder structure in the destination (mirror — the existing,
#     now-explicit default) or place every converted file directly in the
#     destination folder instead (flatten). Collision-safe naming already
#     handles same-named files landing in the same flattened folder. New CLI
#     flag: fm.py convert-video SOURCE DEST -r --flatten (CLI still defaults
#     to mirroring when --flatten is omitted, so existing scripted calls are
#     unaffected).
#
# v1.49 (2026-07-20)
#   • Main Menu: new "Optimize Media" option (9th, after Zip — Remove, Create
#     Random UID, Clean Up, and Admin each shifted down one). Submenu:
#     "Convert Video to MP4" — converts .mov/.avi/.mkv/.wmv/.flv/.m4v files to
#     .mp4 (H.264 + AAC) via moviepy. Asks for a source folder, an optional
#     destination (defaults to source), whether to include subfolders, and
#     a Clean Up option to delete each original after a successful convert
#     (off by default, re-confirmed right before it happens — same pattern
#     as Zip SubFolders' remove_after). Collision-safe naming (name.mp4,
#     name-2.mp4, …) never overwrites. Grayed out in the menu (like Log Zip
#     File Contents' login gate) when moviepy isn't installed, with a
#     "pip install moviepy" hint. New CLI: fm.py convert-video SOURCE [DEST]
#     [-r] [--delete].
#   • Clean Up's delete-original step now validates the new .mp4 first —
#     new _validate_converted_mp4() reopens the output with moviepy and
#     checks it's non-empty and its duration is within 2% (min 1s slack) of
#     the source's. The original is only removed when validation passes;
#     otherwise it's kept and the line is flagged "— <reason>, original
#     KEPT" instead of removed, and the run summary reports how many
#     originals were kept for that reason.
#   • Fixed: Convert Video to MP4 could crash the whole script with a raw,
#     uncaught traceback (no exit screen) if a per-file filesystem call —
#     os.makedirs(out_dir) or the output collision check — hit an error;
#     both sat outside the per-file try/except. Seen in practice when an
#     external destination drive was unmounted mid-batch: ffmpeg's own
#     write failed cleanly for that file (caught), but the NEXT file's
#     os.makedirs() raised PermissionError uncaught and killed the process.
#     Both calls now live inside the try block, so a bad file is reported
#     FAILED and the batch continues. Also added an explicit check at the
#     top of each loop iteration: if the destination folder itself has
#     stopped existing (e.g. drive disconnected), the batch stops early
#     with "Destination folder is no longer reachable" instead of hammering
#     a dead path once per remaining file.
#   • Convert Video to MP4: new "Pause between files" prompt (seconds,
#     default 5, 0 disables) to ease sustained I/O on an external drive
#     during a long batch — suspected cause of a drive dropping mid-run.
#     Sleeps between conversions (never after the last file), showing a
#     live "Pausing Ns (drive cooldown)…" line that clears itself. New CLI
#     flag: fm.py convert-video ... --pause SECONDS (default 0).
#
# v1.48 (2026-07-20)
#   • Renamed Zip menu option "Log Zip File" to "Log Zip File Contents"
#     (label, screen title, and all related help text) for clarity.
#   • An archive that can't be read (e.g. "File is not a zip file") is no
#     longer skipped — it's still logged via _post_zip_log(contents=[],
#     load_failed=True), so the zipFile row gets inserted/updated with its
#     size and zero content rows. Server-side (doc_cloudbox9_com):
#     zipFileLog.php v1.4 passes the new loadFailed flag through to
#     cb9InventoryModel::zipFileLog() v2.1, which sets zipFileStatusId = 3
#     (Failed) instead of 1 (Active) when set. Reported distinctly in the
#     summary as "recorded as Failed (unreadable)", separate from the
#     ok/fail counts for actual API failures.
#
# v1.47 (2026-07-20)
#   • Zip menu: added a 4th option "Zip SubFolders & Log" — same as Zip
#     SubFolders, then logs every zip created to CB9Inventory via
#     log_zip_files() (reused as-is, given the destination folder). Grayed
#     out under the same "login required" reason as Log Zip File.
#   • Zip SubFolders (and Zip SubFolders & Log) now also ask "Push files to
#     AWS S3?" — new _s3_push_zip() uploads each successfully created zip via
#     boto3, using credentials saved in Admin Menu. Gracefully skips (prints
#     a message, doesn't fail the zip) if boto3 is missing or AWS isn't
#     configured yet.
#   • Admin Menu: added Set/Update/Clear AWS S3 Credentials (Access Key ID,
#     Secret Access Key, region, bucket) — stored encrypted via
#     CB9Lib.secureAuth under a new authKey ("fmAwsS3"), same secure store as
#     the DocInfo Manager login but a separate namespace.
#   • FM remains fully usable with no DocInfo Manager login — only Log Zip
#     File and Zip SubFolders & Log require it; Zip SubFolders (without Log)
#     and its optional S3 push do not.
#
# v1.46 (2026-07-20)
#   • render_menu()/_render_menu_lines()/show_menu_help() now support a
#     3-tuple option (label, description, reason) — a truthy reason grays
#     the option out and blocks selection (shows the reason instead of
#     running it). Backward compatible: existing 2-tuple options unaffected.
#   • Zip menu's "Log Zip File" now uses this: shown grayed out with
#     "login required — see Admin Menu" whenever not authenticated, instead
#     of letting you select it and only then prompting for credentials.
#     (ensure_authenticated() stays in log_zip_files() itself as a fallback
#     for the `fm.py zip-log` CLI path, which has no menu to gray out.)
#   • _post_zip_log() now sends the DocInfo Manager token as an X-API-Token
#     header on every Log Zip File request (server-side: see
#     doc_cloudbox9_com's zipFileLog.php v1.3 / _auth.php requireUserAuth(),
#     which now validates it alongside the existing serverSecretKey check).
#
# v1.45 (2026-07-20)
#   • Added an Admin Menu (Show Aliases-style DocInfo Manager login) to the
#     main menu: [L] Login to DocInfo Manager / Logout, showing the current
#     auth status. Uses CB9Lib.secureAuth with its own authKey ("fm") — a
#     separate, independent session from Show Aliases' own login, even
#     though both hit the same DocInfo Manager account.
#   • Zip → Log Zip File now requires being logged in (ensure_authenticated())
#     before it will run; prompts inline for credentials if not already
#     authenticated, same pattern as Show Aliases' DB actions.
#   • New fmConfig.json "auth" block (baseUrl, loginEndpoint, secureAuthFile,
#     authKey) — all optional, defaults hardcoded in fm.py.
#
# v1.44 (2026-07-20)
#   • pause_return() (the standard end-of-task "return to menu" screen used
#     throughout the app) now reacts instantly to [Q] and [ESC], same as
#     [Enter] — previously Q/ESC had no effect there and only a literal Enter
#     press would return to the calling menu.
#   • Fixed VERSION variable drift (was still "1.42" while the header/revision
#     history had already moved to 1.43).
#
# v1.43 (2026-07-19)
#   • Compare Folder Contents now repeats the Folder A / Folder B paths right
#     before the end-of-run [R] Run Again / [Q] Quit footer, in addition to
#     the header at the top — so they're still visible after scrolling
#     through a long results list.
#   • Compare Folder Contents adds a Case-Sensitive / Case-Insensitive prompt
#     (default: Case-Insensitive, matching macOS's default filesystem). A
#     same-name entry that differs only in case (e.g. IMG_1.mov vs IMG_1.MOV
#     on two different drives) is now matched as one file instead of being
#     reported as missing on one side; when matched case-insensitively, any
#     such pair is called out separately under "Same file, different case"
#     so the casing difference is still visible. New --case-sensitive CLI
#     flag on `compare-contents`.
#   • Sync — If a File Exists on Both Sides adds a 3rd choice, "Choose per
#     file", available only for a one-way sync (A → B or B → A; not two-way).
#     For every file that exists on both sides and differs in size and/or
#     modified date, you're shown the filename with both sizes/dates (the
#     differing value(s) highlighted) and choose [I] Ignore or [S] Select
#     (copy) for that file; [IA] Ignore All or [SA] Select All applies that
#     choice to every remaining file without asking again. New
#     _sync_manual_review(), _sync_plan_manual(), _conflict_label(); new
#     --conflict manual CLI choice (rejected with a clear message for
#     two-way syncs or non-interactive/piped runs — sync profiles can't use
#     it either, since they're meant to run unattended).
# -----------------------------------------------------------------------------
# v1.42 (2026-07-17)
#   • Every folder-path prompt (Compare, Display Subfolders, Find, Monitor,
#     Sync, Zip Subfolders, Clean Up, etc.) now remembers the last path you
#     entered for that specific prompt — it becomes the new default, so
#     pressing Enter reuses it and typing a path overrides it, same as any
#     other defaulted prompt. Remembered paths are saved to fmLastPaths.json
#     (next to fmConfig.json) on exit, so they survive across runs, not just
#     within a session. New ask_folder(key=...) param, _load_last_paths(),
#     _save_last_paths(), LAST_PATHS_FILE/LAST_PATHS, saved via
#     atexit.register().
# -----------------------------------------------------------------------------
# v1.41 (2026-07-17)
#   • Find Duplicates by Filename and Find Duplicates by Fuzzy Name: once a
#     scan finds duplicates, the results footer now offers [D] Delete
#     Duplicates alongside [R] Run Again / [Q] Quit. Picking it opens a
#     keep-rule sub-menu: Keep Newest (most recent mtime wins), Keep Largest
#     (biggest file wins), or Delete from Specific Folder (you enter a
#     folder; any copy located under it — including subfolders — is deleted,
#     copies elsewhere are kept; if every copy in a group is under that
#     folder, one is kept anyway so a duplicate is never wiped out
#     entirely). Every group's KEEP/DELETE split is previewed before
#     anything happens, and deleting requires typing YES — same
#     dry-run/confirm pattern as the existing Remove → Duplicates options.
#     New _pause_find_dups(), _path_under_folder(), _resolve_dup_group(),
#     _pick_dup_delete_strategy(), _delete_duplicate_groups(); both
#     _find_duplicates_screen() and _find_fuzzy_dups_screen() now return
#     their duplicate groups instead of just printing them.
# -----------------------------------------------------------------------------
# v1.40 (2026-07-16)
#   • Sync Folders: while copying, a periodic milestone line is now written
#     straight to the activity log (~/Documents/log/fm.log) every
#     MILESTONE_INTERVAL_SECONDS (15s), e.g.
#     "2026-07-16 15:22:41 - Copied 728/2103 files, 8.3 GB of 10.4 GB".
#     New _log_milestone() helper writes immediately (unlike _ActivityLog,
#     which only flushes its buffered mirror once the whole operation
#     finishes), so an interrupted long-running sync still leaves a trail of
#     how far it got.
# v1.39 (2026-07-16)
#   • Monitor screen: stationary header (Profile/Folder/Recursive/Extensions/
#     Logging) and footer ([Q/ESC] Stop + live event counts) now stay pinned
#     at the top/bottom of the window — only the event feed between them
#     scrolls (ANSI scroll region / DECSTBM, same technique as backup42.py's
#     execution screen; CB9 Static Header/Footer permanent rule)
#   • Footer event counts update in place via save/restore-cursor without
#     disturbing the scroll region; SIGWINCH redraws the frame on resize
#   • Non-TTY runs (piped/cron) are unchanged — plain scrolling output, no
#     ANSI codes
#   • Sync screen gets the same stationary-frame treatment: header shows
#     Folder A/B, Direction, Options, and the new/updated/size summary line;
#     footer shows phase status and live "Copying... N/total" progress
#     (updated in place per file, same technique as Monitor). Only the file
#     plan/prompt/copy output between them scrolls. Frame is only used once
#     there's an actual plan to show — folder errors and "nothing to copy"
#     stay as short one-shot messages
#   • Shared the ANSI scroll-region/cursor primitives between Monitor and
#     Sync (renamed _mon_* helpers to _frm_*; added a common _frm_cleanup())
#     instead of duplicating them per screen
#   • Fixed: the header could still scroll because both frames read the
#     terminal size via shutil.get_terminal_size(), which prefers the
#     COLUMNS/LINES environment variables when present — stale values (set
#     once, never updated after a real window resize) made the frame's row
#     math target the wrong screen size. Added _frm_term_size(), which
#     queries the real size via os.get_terminal_size() (ioctl on stdout),
#     ignoring COLUMNS/LINES entirely; both _monitor_draw_frame() and
#     _sync_draw_frame() now use it. Verified via pty tests that a real
#     resize is picked up correctly even with stale COLUMNS/LINES still set
#   • Fixed a crash: Sync (and any screen run inside `with _ActivityLog():`,
#     e.g. Eject) raised AttributeError: '_ActivityLog' object has no
#     attribute 'fileno' from _frm_term_size(), because _ActivityLog
#     temporarily replaces sys.stdout with a write()/flush()-only tee proxy
#     that has no fileno(). _frm_term_size() now queries the real stdout fd
#     (1) directly instead of sys.stdout.fileno(), which is unaffected by
#     that (or any future) Python-level stdout wrapper. Verified with a pty
#     test that reproduces the exact call path from the crash (sync_folders()
#     -> with _ActivityLog(): _sync_screen()) — no crash, frame renders
#     correctly, and the activity log still receives the mirrored output
#   • Sync footer now shows "Copying N/total: filename (size)..." *before*
#     each file starts, not just the completed count afterward — a single
#     large file (e.g. video footage between two external USB drives) can
#     take a long time and the footer used to just sit at the previous
#     count the whole time, looking frozen even though the copy was
#     genuinely progressing
#   • Compare Folder Contents — Compare By submenu: relabeled the three
#     options to spell out exactly what each does ("Match Name, Display
#     Missing Files" / "Match Name, Compare Size, Display Files with Size
#     Differences" / "Display Missing Files & Existing Files with Size
#     Differences"), replacing the terse "By Name" / "By Size" / "Both"
#   • Sync footer's copy progress now shows percent complete and total
#     bytes instead of just bytes copied so far, e.g.
#     "Copying... 201/322 file(s)  (85% complete, 28.4 GB/78.2 GB complete)"
#   • Fixed: fm.log entries for Monitor/Sync (anything drawn via the
#     stationary frame) showed raw control-code junk like "^[[?25l" —
#     _ActivityLog's ANSI-stripping regex only matched plain SGR codes
#     ("\x1b[<digits><letter>"), missing private-mode codes (\x1b[?25l/h —
#     the `?` isn't in [0-9;]) and non-CSI codes (\x1b7/\x1b8, DEC save/
#     restore cursor). New _strip_screen_codes() strips every CSI sequence
#     generically, converts absolute row-jumps (used by the frame header)
#     and the save-cursor marker (used by in-place footer updates) to
#     newlines first so lines don't get mashed together, and normalizes
#     stray \r. Retroactively re-cleaned the existing fm.log with it —
#     entries logged before this fix still show one run-on line for the
#     frame header (the old regex already destroyed the row-jump codes
#     without a newline substitute, before this fix existed to convert
#     them); every entry logged from now on is fully clean.
#   • Compare Folder Contents and Sync now log a concise "started" entry to
#     fm.log the moment the operation begins (activity type + folders +
#     options — e.g. Direction/Recursive/Conflict/Hidden Files for Sync,
#     Recursive/Compare By for Compare), via the new _log_activity_start().
#     This is independent of Sync's existing full-output mirror
#     (_ActivityLog, written only when the run finishes) — a long or
#     interrupted sync now leaves a record that it was started even if it
#     never completes. Compare had no logging at all before this; it now
#     gets the same started-entry treatment. Also de-duplicated the
#     direction-description string (was hand-built in two places) into
#     _sync_direction_desc(), shared by the on-screen header, the
#     stationary frame, and this new log entry.
#

# v1.38 (2026-07-15)
#   • Remove menu: two new options. 3rd option "Duplicates by Fuzzy Name"
#     (grouped with the other Duplicates removals) — same close-name +
#     close-size grouping as Find Duplicates by Fuzzy Name (shared
#     _fuzzy_dup_groups() helper, extracted from _find_fuzzy_dups_screen);
#     keeps the shortest/cleanest name per group, removes the rest. Typed
#     YES required to delete. 6th (last) option "Files of 0 Size" — lists
#     every empty (0-byte) file under the entered folders for removal;
#     hidden files/folders skipped (protects .gitkeep-style markers);
#     standard dry-run + confirm. By File Name / By Folder Name renumbered
#     to 4/5. New remove_duplicates_by_fuzzy_name(),
#     remove_zero_size_files(). CLI: fm.py remove dup-fuzzy FOLDER... and
#     fm.py remove zero-size FOLDER... [--delete] [--yes].
# -----------------------------------------------------------------------------
# v1.37 (2026-07-15)
#   • Find menu: new 4th option "Find Duplicates by Fuzzy Name" (after Find
#     Duplicates by Filename) — finds files whose names are CLOSE (not
#     necessarily identical) AND whose sizes are close (within 1%). Example:
#     videofile1.mov and videofile.mov at the same size are duplicates.
#     Close names = same stem after stripping duplicate-style endings
#     (trailing digits, "(1)", "[2]", "copy", "copy 2") or 85%+ difflib
#     similarity; extensions must match. Matches are clustered into groups;
#     each group marks the shortest/cleanest name KEEP and the rest DELETE
#     (candidates only — read-only, nothing is deleted). Groups sort largest
#     file first; summary shows group/candidate counts + reclaimable bytes.
#     New find_duplicates_by_fuzzy_name(), _find_fuzzy_dups_screen(),
#     _fuzzy_stem(), _names_close(), _sizes_close(). Later Find options
#     renumbered (Missing 5/6, Find & Replace 7, Find & Rename 8).
#     CLI: fm.py find-fuzzy-dups FOLDER...
# -----------------------------------------------------------------------------
# v1.36 (2026-07-15)
#   • Sync: new two-way direction (A ↔ B) — both folders push to each other in
#     one run. Files only in A are copied to B, files only in B are copied to
#     A, and when a file exists on both sides the copy that wins the conflict
#     rule (newest/largest) replaces the other; ties are skipped. Still a DRY
#     RUN preview first (grouped A → B and B → A sections) and nothing is ever
#     deleted. Interactive: third Direction option "Two-way sync A ↔ B".
#     Profiles: direction accepts Both (also 2way/twoway). CLI: --to both.
#     Added _parse_sync_direction(); _sync_screen now plans/copies in passes.
# -----------------------------------------------------------------------------
# v1.35 (2026-07-15)
#   • Sync: zero-byte source files are never copied (_sync_plan skips them and
#     reports a "Skipped N zero-byte file(s)" note instead). Prevents
#     incomplete/placeholder 0-byte files from overwriting good files on the
#     destination side.
# -----------------------------------------------------------------------------
# v1.34 (2026-07-14)
#   • Main Menu labels: capitalized the first letter after the em-dash on all
#     11 options (e.g. "Compare  — compare 2 files…" → "Compare  — Compare 2
#     files…"). README overview block matched. Cosmetic only.
# -----------------------------------------------------------------------------
# v1.33 (2026-07-14)
#   • Main Menu reordered (per spec — no longer alphabetical): 1. Compare
#     2. Convert  3. Display  4. Find  5. Eject  6. Monitor  7. Sync  8. Zip
#     9. Remove  10. Create Random UID  11. Clean Up. Dispatch, [H] Help
#     order, and the menu intro line updated to match; no feature changes.
# -----------------------------------------------------------------------------
# v1.32 (2026-07-14)
#   • Find menu: new 7th (last) option "Find & Rename" — find files, then
#     rename them. Enter a folder, an optional file extension, and a mode:
#     Prepend (text.mov -> api_text.mov; files already starting with the
#     text are skipped), Append (inserted before the extension:
#     text.mov -> text_api.mov; files already ending with it are skipped),
#     or Replace (literal case-insensitive replacement inside the filename;
#     blank replacement removes the text; only matching files are listed).
#     ALWAYS dry-runs first, listing every rename old -> new; renames that
#     would overwrite an existing file (or another rename in the same run)
#     are skipped and reported. Hidden/excluded files are never touched.
#     New find_and_rename()/_find_rename_screen(); results screens log to
#     fm.log and offer [R] Run Again. CLI: fm.py find-rename ROOT
#     (--prepend TEXT | --append TEXT | --replace FIND NEW) [--ext E]
#     [--apply] [--yes].
# -----------------------------------------------------------------------------
# v1.31 (2026-07-14)
#   • Main Menu: new "Create Random UID" option (3rd, keeping the menu
#     alphabetical; Clean Up stays last) — enter a number N and get N random
#     (version 4) UUIDs displayed one per line (capped at 10,000). The list
#     logs to fm.log; [R] Run Again regenerates a fresh batch with the same
#     count. New generate_uuids()/_uuid_screen()/create_random_uid_menu().
#     CLI: fm.py uuid [N] (default 1).
# -----------------------------------------------------------------------------
# v1.30 (2026-07-14)
#   • Main Menu: new "Convert" option (2nd, keeping the menu alphabetical;
#     Clean Up stays last) — convert a data file between formats. Input:
#     .csv (delimiter sniffed, BOM tolerated), .json (array of objects or
#     array of arrays), .xlsx (first sheet; via openpyxl). Output: CSV,
#     JSON (pretty-printed array of objects), XLSX (openpyxl — only format
#     needing a package; a clear message says pip3 install openpyxl if
#     missing), or SQL (CREATE TABLE named after the file with column types
#     guessed per column — TINYINT(1)/INT/BIGINT/DECIMAL/DATE/DATETIME/
#     VARCHAR(n)/TEXT — plus multi-row INSERTs in 500-row batches; empty
#     cells become NULL, identifiers sanitized + de-duplicated). Output is
#     written beside the source with the same name and new extension,
#     collision-safe (name-2.ext, …) — never overwrites. Results screens log
#     to fm.log and offer [R] Run Again. New convert_menu()/convert_file()/
#     _read_table()/writers/_unique_path(). CLI: fm.py convert FILE
#     csv|json|xlsx|sql.
# -----------------------------------------------------------------------------
# v1.29 (2026-07-14)
#   • Find menu: new 6th (last) option "Find & Replace" — enter a folder, the
#     text to find, the replacement text, and an optional file extension.
#     Scans recursively (hidden/binary/excluded files skipped); matching is
#     literal and case-insensitive. ALWAYS dry-runs first, listing every
#     match as file + line number + line with the matched text highlighted,
#     then a [y/N] confirm performs the replacement (with a follow-up prompt
#     offering .bak backups of each modified file) — No exits unchanged.
#     .bak files are skipped by the scan (so reruns never clobber fresh
#     backups) unless --ext bak is asked for explicitly.
#     Round-trips file bytes and line endings via surrogateescape +
#     newline="". New find_and_replace()/_find_replace_screen(); results
#     screens log to fm.log and offer [R] Run Again. CLI: fm.py find-replace
#     ROOT SEARCH REPLACE [--ext E] [--apply] [--bak] [--yes].
# -----------------------------------------------------------------------------
# v1.28 (2026-07-14)
#   • Main Menu: new "Clean Up" option (9th — last, per spec). Submenu:
#     1. Remove Junk Files — find and delete every .DS_Store / desktop.ini
#     under a root folder (hidden folders included; reuses the Remove
#     preview/confirm machinery). 2. Purge Old Log Files — trim entries older
#     than N days (default 90) from every .log file in ~/Documents/log
#     (folder and days promptable), ported from purgeLog.sh but block-aware:
#     untimestamped body lines follow their [YYYY-MM-DD HH:MM:SS] header's
#     keep/purge decision, so whole entries purge together. Per-file
#     purge/keep table preview; .bak backup before each rewrite. Both are
#     DRY RUN until confirmed. CLI: fm.py cleanup junk ROOT [--delete]
#     [--yes] and fm.py cleanup logs [FOLDER] [--days N] [--delete] [--yes].
# -----------------------------------------------------------------------------
# v1.27 (2026-07-14)
#   • Display menu is now: 1. All Drives  2. Subfolders Alphabetically
#     3. Subfolders by Size (largest first). New display_all_drives() lists
#     the boot volume plus every mount under /Volumes (deduplicated by
#     device) with Size / Used / Free / Use% and the mount point — Free comes
#     from shutil.disk_usage so it matches Finder's Available; Use% colors
#     yellow at 75% and red at 90%. CLI: fm.py drives.
# -----------------------------------------------------------------------------
# v1.26 (2026-07-14)
#   • Main Menu: new "Monitor" option (5th, keeping the menu alphabetical) —
#     Monitor File Activity. Watches a folder (or a fmConfig.json
#     monitorProfiles entry: name, folder, recursive, extensions, output) and
#     reports every created/modified/deleted file in real time, on screen and
#     to the chosen log: fm.log (plain timestamped lines) or
#     ~/Documents/log/fmMonitor.csv (Timestamp,Filename,Folder,Event rows).
#     Options: recursive (default Yes) and a file-extension filter. Stdlib
#     polling (1s snapshot diffs — no watchdog dependency); events flush
#     immediately so tail -f tracks the screen. [Q/ESC] stops and returns to
#     the menu (Ctrl-C on the CLI). Ported from the standalone
#     fileActivity.py idea. CLI: fm.py monitor FOLDER [--no-recursive]
#     [--ext jpg,png] [--csv], or fm.py monitor --profile NAME.
#   • Refactored profile loading into _load_config_profiles(key), shared by
#     syncProfiles and monitorProfiles.
# -----------------------------------------------------------------------------
# v1.25 (2026-07-14)
#   • Main Menu: new "Eject" option (3rd, keeping the menu alphabetical) —
#     Eject All External Drives (macOS). Lists the mounted external drives
#     (name, size, mount point), confirms, then ejects each with a per-drive
#     Success/Failed status and offers a force eject for failures. Detection
#     and eject logic ported from ejectDrives.py (diskutil info scan of
#     /Volumes; diskutil eject with Finder/AppleScript fallback; force =
#     diskutil unmountDisk force). Results are logged to fm.log. CLI:
#     fm.py eject [--list] [--force] [--yes].
# -----------------------------------------------------------------------------
# v1.24 (2026-07-14)
#   • Main Menu: new "Sync" option (before Zip) — one-way folder sync pushing
#     new/updated files A → B or B → A. Interactive flow asks for the two
#     folders, the direction, the both-sides conflict rule (newest wins
#     default / largest wins), recursive (default Yes), and exclude hidden
#     files (default Yes). Saved profiles from fmConfig.json (syncProfiles
#     list: name, folderA, folderB, direction, recursive, conflict,
#     excludeHidden) appear as Sync menu options; the [H] Help documents the
#     profile format. Every run previews (DRY RUN) and only copies after an
#     explicit confirm; nothing is ever deleted; copies use shutil.copy2 so
#     timestamps survive for future 'newest' runs. Results screens are logged
#     to ~/Documents/log/fm.log. CLI: fm.py sync A B [--to b|a] [--conflict
#     newest|largest] [--no-recursive] [--include-hidden] [--copy] [--yes],
#     or fm.py sync --profile NAME.
# -----------------------------------------------------------------------------
# v1.23 (2026-07-14)
#   • Main Menu [H] Help: new unnumbered "Logging" section explaining that
#     commands and their results screens are appended (timestamped, colors
#     stripped) to the activity log ~/Documents/log/fm.log, which actions log
#     there, and that Log Zip File records to CB9Inventory instead.
#   • render_menu()/_render_menu_lines() gained a help_note parameter and
#     show_menu_help() a note parameter for such extra Help sections; the
#     description wrapper was factored out into _print_help_desc().
# -----------------------------------------------------------------------------
# v1.22 (2026-07-13)
#   • Find menu: new 5th option "Find Missing by Filename & Size" — same flow
#     as Find Missing by Filename, but files match only when BOTH the name AND
#     the size agree, so same-named files with different sizes are also
#     reported (size shown in both columns). CLI: find-missing --size.
# -----------------------------------------------------------------------------
# v1.21 (2026-07-13)
#   • Compare & Find results screens now end with "[R] Run Again  [Q/ESC]
#     Quit/Back" instead of the plain Enter pause — R reruns the same action
#     with the same inputs (fresh scan; Find Dup/Missing reruns re-log to
#     fm.log). R/Q/ESC react instantly; Enter also goes back. New pause_rerun();
#     CLI one-shot runs are unchanged. Remove/Zip keep the plain pause.
# -----------------------------------------------------------------------------
# v1.20 (2026-07-13)
#   • .DS_Store and desktop.ini files (case-insensitive) are now also ignored
#     by all searches and compares — every Find feature and Compare Folder
#     Contents. Added EXCLUDED_FILE_NAMES + is_excluded_file(). Remove → By
#     File Name still targets them intentionally (junk-file cleanup).
# -----------------------------------------------------------------------------
# v1.19 (2026-07-13)
#   • The Windows Recycle Bin folder ($RECYCLE.BIN, case-insensitive) is now
#     excluded from all searches and compares: every Find feature (Files,
#     Folders, Duplicates/Missing by Filename) and Compare Folder Contents
#     (recursive and top-level). Added EXCLUDED_DIR_NAMES + prune_dirs().
# -----------------------------------------------------------------------------
# v1.18 (2026-07-13)
#   • Find Duplicates/Missing by Filename: the results screen is now also
#     appended (ANSI-stripped, timestamped) to ~/Documents/log/fm.log via the
#     new _ActivityLog stdout tee.
#   • Find Missing modes "In 1st/2nd folder only": added a last 'Folder'
#     column showing the directory containing each file (_scan_filenames now
#     records (size, dir) per occurrence).
#   • Find Missing: the Folder 1/2 list is repeated after the results table.
# -----------------------------------------------------------------------------
# v1.17 (2026-07-13)
#   • Find Missing by Filename — Show menu: the two entered folders are now
#     displayed below the options as "Folder 1 - …" / "Folder 2 - …".
#   • render_menu()/_render_menu_lines() gained an `outro` parameter (context
#     line(s) below the options), shared via _print_menu_outro().
# -----------------------------------------------------------------------------
# v1.16 (2026-07-13)
#   • Find menu: new 4th option "Find Missing by Filename" — enter two folders,
#     then choose In 1st folder only / In 2nd folder only / In either folder
#     (only once). Filenames present in only one folder are shown in the same
#     size table as Find Duplicates by Filename (blank column = missing there).
#   • Refactored the folder scan + table renderer into _scan_filenames() /
#     _print_filename_size_table() shared by both features; CLI: find-missing.
# -----------------------------------------------------------------------------
# v1.15 (2026-07-13)
#   • Find menu: new 3rd option "Find Duplicates by Filename" — enter one or
#     more folders (comma-separated); files sharing the same name are shown in
#     a table with a numbered header per folder and one size column per folder
#     (multiple occurrences within one folder list each size). Read-only.
#   • Added find_duplicates_by_filename(), fmt_size_short(); CLI: find-dups.
# -----------------------------------------------------------------------------
# v1.14 (2026-07-11)
#   • Log Zip File now sends zipFileFolder (the archive's containing folder)
#     so CB9Inventory records where each zip lives.
# -----------------------------------------------------------------------------
# v1.13 (2026-07-11)
#   • Zip menu: new 2nd option "Log Zip File" — logs a .zip/.tar archive (or
#     every archive in a folder, top level only) to the CB9Inventory database
#     on BPA5 via the DocInfo Manager API (api/zipFileLog.php). zipFile rows
#     are matched by name+size (insert or update); zipFileContent is synced
#     (update by path, insert new, soft-delete missing). .gz files are ignored.
#   • New fmConfig.json (logZip: apiUrl, serverSecretKey) + CLI: fm.py zip-log.
# -----------------------------------------------------------------------------
# v1.12 (2026-07-09)
#   • Remove → Duplicates (by Name / by Hash): deleting now requires typing the
#     word YES (any other input cancels) — a stronger gate than the y/n used by
#     the other removals. Added confirm_yes_word() + _finish_removal(require_yes).
# -----------------------------------------------------------------------------
# v1.11 (2026-07-09)
#   • Fixed: after selecting a menu option, a follow-up input prompt (e.g.
#     "Folder to measure:") was printed on the same line as "Option:". The menu
#     now emits a newline on selection so each new input request starts fresh.
# -----------------------------------------------------------------------------
# v1.10 (2026-07-09)
#   • Zip menu: "View Zip" is now the first option, "Zip SubFolders" second.
# -----------------------------------------------------------------------------
# v1.9 (2026-07-09)
#   • Expanded the [H] Help descriptions for every menu (Main, Compare, Compare-
#     By, Display, Find, Remove, Zip) into full explanations with inputs, tips,
#     and examples.
#   • show_menu_help() now word-wraps descriptions to the terminal width and
#     supports multi-line / bulleted text with hanging indents.
# -----------------------------------------------------------------------------
# v1.8 (2026-07-09)
#   • Remove menu reordered/renamed: 1) Duplicates by Name, 2) Duplicates by
#     Hash, 3) By File Name, 4) By Folder Name.
#   • New "By Folder Name": deletes folders whose name matches a wildcard under a
#     search root (recursively); only top-most matches are removed. Dry-run by
#     default. Added remove_folders_by_name() + CLI 'remove folder-name'.
# -----------------------------------------------------------------------------
# v1.7 (2026-07-09)
#   • Find: reworked into "Find Files" (multi-criteria) + "Find Folders".
#     Find Files lets you multi-select any of Filename pattern / Extension /
#     Size over N / Size under N (Space toggles), then prompts for each value and
#     runs one combined AND search — e.g. .mov files under 5 MB.
#   • Added render_multiselect() (checkbox menu: ↑↓ move, Space toggle, Enter
#     confirm) and find_files_combined(); new CLI: find-files [--name/--ext/
#     --over/--under]. The old single-criterion `find` subcommand still works.
# -----------------------------------------------------------------------------
# v1.6 (2026-07-09)
#   • All menus now support Up/Down arrow navigation with a highlighted row;
#     Enter selects the highlighted option (read_key + arrow loop in render_menu).
#   • Typing a number still selects (moves the highlight); H = Help, Q/ESC = Back/
#     Exit — all instant. Non-TTY (piped/automation) keeps line-based selection.
#   • Zip-file browser reuses the same arrow-navigable menu.
# -----------------------------------------------------------------------------
# v1.5 (2026-07-09)
#   • Display (Folder Sizes): the Folder name is now the first column (was last),
#     with a computed column width so rows, header, and TOTAL stay aligned.
# -----------------------------------------------------------------------------
# v1.4 (2026-07-09)
#   • Removed the copyright line from menu/result screens; it now appears only
#     on the exit screen. Menu footers show just the separator bars + legend.
# -----------------------------------------------------------------------------
# v1.3 (2026-07-09)
#   • Replaced [B] Back with [Q/ESC] on every menu: on a submenu Q/ESC go back;
#     on the Main Menu they exit.
#   • ESC now reacts INSTANTLY (no Enter needed) via a hybrid raw-mode reader
#     (menu_read); digits/letters still buffer until Enter for multi-digit input.
#   • All submenus now default to option 1 — pressing Enter selects it.
# -----------------------------------------------------------------------------
# v1.2 (2026-07-09)
#   • Fixed: pasting a folder/file path that was wrapped in single or double
#     quotes, or drag-and-dropped (backslash-escaped spaces), no longer errors.
#   • Added clean_path() which strips surrounding quotes and undoes backslash
#     escapes; applied to every interactive prompt and CLI path argument.
# -----------------------------------------------------------------------------
# v1.1 (2026-07-09)
#   • Every menu/result screen now ends with the standard CB9 footer whose last
#     line is the copyright notice.
#   • Submenus use a custom renderer: [H] Help shows a description of each
#     option, and [Q/B] both return to the parent menu.
#   • Reworked the Compare submenu:
#       - "Compare 2 Files" — side-by-side, line-by-line file comparison.
#       - "Compare Folder Contents" — options Recursive (Y/N) and
#         Compare By Name / Size / Both (both directions always reported).
#   • CLI: replaced compare-files/compare-folders with compare-2files and
#     compare-contents [--recursive] [--by].
# -----------------------------------------------------------------------------
# v1.0 (2026-07-09)
#   • Initial version. Merges 9 legacy file-management scripts into one tool.
#   • Menu: Compare / Display / Find / Remove / Zip (with submenus).
#   • Built on CB9Lib (header, footer, menu, confirm, exit_screen, colors,
#     project sounds).
#   • Remove actions are dry-run by default; deletion is explicit opt-in.
#   • Supports both an interactive menu and direct CLI subcommands.
# -----------------------------------------------------------------------------

import sys
import os
import platform
import stat
import atexit
import argparse
import csv
import time
import threading
import zipfile
import tarfile
import hashlib
import fnmatch
import subprocess
import shutil
import difflib
import re
import textwrap
import json
import uuid
import getpass
import urllib.request
import urllib.parse
from datetime import datetime
from collections import defaultdict

try:
    import boto3
    _BOTO3_AVAILABLE = True
except ImportError:
    _BOTO3_AVAILABLE = False

try:
    from moviepy import VideoFileClip
    _MOVIEPY_AVAILABLE = True
except ImportError:
    _MOVIEPY_AVAILABLE = False

try:
    from PIL import Image, ImageOps
    _PIL_AVAILABLE = True
except ImportError:
    _PIL_AVAILABLE = False

try:
    import pillow_heif
    pillow_heif.register_heif_opener()
    _HEIF_AVAILABLE = True
except ImportError:
    _HEIF_AVAILABLE = False

_IMAGE_CONVERT_AVAILABLE = _PIL_AVAILABLE and _HEIF_AVAILABLE

# CB9Lib imports
sys.path.insert(0, os.path.expanduser("~/Documents/script/CB9Lib"))
from CB9Lib import secureAuth
from CB9Lib import (
    header, exit_screen, pause, clear_screen,
    get_width, color_text, get_project_sound, play_sound,
    RED, GREEN, YELLOW, CYAN, MAGENTA, WHITE, BOLD, DIM, RESET,
    BRIGHT_CYAN, BRIGHT_GREEN, BRIGHT_YELLOW, BRIGHT_RED,
)

# -----------------------------------------------------------------------------
# Constants
# -----------------------------------------------------------------------------
SCRIPT_NAME = "File Manager"
VERSION     = "2.00"
VER         = f"v{VERSION}"

SCRIPT_DIR    = os.path.dirname(os.path.abspath(__file__))
SOUND_SUCCESS = os.path.join(SCRIPT_DIR, "audio", "success.mp3")
SOUND_FAILURE = os.path.join(SCRIPT_DIR, "audio", "failure.wav")
CONFIG_FILE   = os.path.join(SCRIPT_DIR, "fmConfig.json")


def _check_optional_dependencies():
    """Print a one-line status for FM's optional third-party packages —
    everything the interactive/CLI entry point (main()) needs before it does
    anything else. These are all soft dependencies: each _XXX_AVAILABLE flag
    already gates its one feature (Optimize Media -> Convert Video/Images,
    Admin Menu -> S3 push) with its own in-menu grayed-out message, so a
    missing package never crashes FM. This just surfaces that state up front
    on every run instead of only when the user stumbles into the specific
    screen that needs it."""
    missing = []
    if not _BOTO3_AVAILABLE:
        missing.append(("boto3", "Admin Menu -> AWS S3 push / Zip -> Log Zip File uploads"))
    if not _MOVIEPY_AVAILABLE:
        missing.append(("moviepy", "Optimize Media -> Convert Video to MP4"))
    if not _PIL_AVAILABLE:
        missing.append(("Pillow", "Optimize Media -> Convert Images"))
    elif not _HEIF_AVAILABLE:
        missing.append(("pillow-heif", "Optimize Media -> Convert Images (HEIC/HEIF support)"))

    if not missing:
        print(color_text("  ✓ All optional packages installed (boto3, moviepy, Pillow, pillow-heif).",
                         fg=BRIGHT_GREEN))
        return

    print(color_text(f"  ⚠ {len(missing)} optional package(s) missing:", fg=BRIGHT_YELLOW, style=BOLD))
    for pkg, feature in missing:
        print(color_text(f"    • {pkg} — disables: {feature}", fg=YELLOW))
    print(color_text(f"    Install with: pip install {' '.join(p for p, _ in missing)}", fg=DIM))


def _check_config_file():
    """Print a one-line status for fmConfig.json, alongside
    _check_optional_dependencies() at startup. fmConfig.json is itself
    optional — FM runs fine without it, just with no saved sync/monitor/
    compare/permission profiles, no common-folder shortcuts, no local
    scripts, and no DocInfo Manager S3/logZip/auth settings — so a missing
    file is a heads-up, not an error."""
    if os.path.isfile(CONFIG_FILE):
        print(color_text(f"  ✓ Config file found: {CONFIG_FILE}", fg=BRIGHT_GREEN))
    else:
        sample = os.path.join(SCRIPT_DIR, "fmConfig.sample.json")
        print(color_text(f"  ⚠ No config file found at: {CONFIG_FILE}", fg=BRIGHT_YELLOW, style=BOLD))
        print(color_text(f"    Sync/Monitor/Compare/Permission profiles, common folders, and "
                         f"Local Scripts won't be available until it exists.", fg=YELLOW))
        print(color_text(f"    Copy {sample} to fmConfig.json and edit it to set one up.", fg=DIM))


# Last-used folder path per prompt (keyed by ask_folder's `key` arg), so each
# folder prompt defaults to what you entered last time — Enter reuses it,
# typing a new path overrides it. Persisted to fmLastPaths.json on exit so it
# survives across runs, not just within a session.
LAST_PATHS_FILE = os.path.join(SCRIPT_DIR, "fmLastPaths.json")


def _load_last_paths():
    try:
        with open(LAST_PATHS_FILE, "r") as fh:
            data = json.load(fh)
        return data if isinstance(data, dict) else {}
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return {}


def _save_last_paths():
    try:
        with open(LAST_PATHS_FILE, "w") as fh:
            json.dump(LAST_PATHS, fh, indent=2, sort_keys=True)
    except OSError:
        pass


LAST_PATHS = _load_last_paths()
atexit.register(_save_last_paths)

# True when driven by the interactive menu; False for one-shot CLI subcommands.
# Controls whether end-of-task screens pause for a keypress.
INTERACTIVE = True


# -----------------------------------------------------------------------------
# Small display helpers
# -----------------------------------------------------------------------------
def bar(char="="):
    return char * get_width()


def screen(subtitle=""):
    """Clear + standard CB9 header for a full-screen action view."""
    header(SCRIPT_NAME, VER, subtitle)


def ask_dry_mode(live_requested, default=True):
    """Upfront 'run in dry-mode first?' gate — asked once, before any
    scan/match work begins. Returns True (Dry-Mode) / False (Live Mode).
    Interactive (live_requested is None) -> ask. CLI (True/False) -> derived
    silently from the flag already supplied, no question shown."""
    if live_requested is None:
        return safe_confirm("  Run in dry-mode first?", default=default)
    return not live_requested


def print_mode_line(dry_mode):
    """Prints right after screen() so the run's mode is the first visible
    line on every dry-run-capable task screen."""
    label = "Dry-Mode" if dry_mode else "Live Mode"
    color = YELLOW if dry_mode else BRIGHT_RED
    print(color_text(f"  Mode: {label}", fg=color, style=BOLD))


def standard_footer(legend=""):
    """Draw the standard footer.

    Layout:
        =====================================
         <legend>                       (only if a legend was supplied)
        =====================================

    The copyright notice is intentionally NOT shown on menus/result screens —
    it appears only on the exit screen (see exit_screen()).
    """
    print(bar("="))
    if legend:
        print(color_text(" " + legend, fg=BRIGHT_YELLOW))
        print(bar("="))


def pause_return():
    """Standard end-of-task footer + pause (returns to the calling menu).
    [Enter], [Q], and [ESC] all return instantly — Q/ESC do not require an
    extra Enter press. In one-shot CLI mode there is no menu to return to,
    so draw the footer and continue without blocking on a keypress."""
    print()
    if not INTERACTIVE:
        standard_footer()
        return
    standard_footer("[Enter/Q/ESC] Return to Menu")
    if sys.stdin.isatty():
        while True:
            key = read_key()
            if key in ("ENTER", "q", "Q", "ESC"):
                print()
                return
    else:
        try:
            pause("Press Enter to return to the menu...")
        except EOFError:
            pass


def pause_rerun():
    """End-of-task footer for Compare/Find screens offering to rerun the same
    action with the same inputs.

    Returns True when the user chooses [R] Run Again, False to go back.
    [R], [Q], and ESC react instantly; Enter also goes back. In one-shot CLI
    mode there is no menu/rerun loop — draw the plain footer and return False.
    """
    print()
    if not INTERACTIVE:
        standard_footer()
        return False
    standard_footer("[R] Run Again   [Q/ESC] Quit/Back")
    sys.stdout.write(color_text(" Option: ", fg=CYAN, style=BOLD))
    sys.stdout.flush()
    if sys.stdin.isatty():
        while True:
            key = read_key()
            if key in ("r", "R"):
                print()
                return True
            if key in ("q", "Q", "ESC", "ENTER"):
                print()
                return False
            # anything else: ignore
    # Non-TTY (piped) fallback — line-based
    try:
        resp = sys.stdin.readline()
    except Exception:
        return False
    return resp.strip().lower() == "r"


def _pause_find_dups(has_dups):
    """End-of-task footer for the Find Duplicates screens (by Filename / by
    Fuzzy Name): same as pause_rerun(), plus a [D] Delete Duplicates option
    when the scan actually found duplicates.

    Returns 'delete', 'rerun', or 'back'. In one-shot CLI mode there is no
    menu/rerun loop — draw the plain footer and return 'back'.
    """
    print()
    if not INTERACTIVE:
        standard_footer()
        return "back"
    legend = "[R] Run Again   [Q/ESC] Quit/Back"
    if has_dups:
        legend = "[D] Delete Duplicates   " + legend
    standard_footer(legend)
    sys.stdout.write(color_text(" Option: ", fg=CYAN, style=BOLD))
    sys.stdout.flush()
    if sys.stdin.isatty():
        while True:
            key = read_key()
            if has_dups and key in ("d", "D"):
                print()
                return "delete"
            if key in ("r", "R"):
                print()
                return "rerun"
            if key in ("q", "Q", "ESC", "ENTER"):
                print()
                return "back"
            # anything else: ignore
    # Non-TTY (piped) fallback — line-based
    try:
        resp = sys.stdin.readline()
    except Exception:
        return "back"
    resp = resp.strip().lower()
    if has_dups and resp == "d":
        return "delete"
    return "rerun" if resp == "r" else "back"


def _pause_compare_folders():
    """End-of-task footer for Interactive Compare Folder Contents: same as
    pause_rerun(), plus a [S] Save as Profile option so the folders/options
    just used can be saved to fmConfig.json (compareProfiles) for later reuse
    via Compare Folder Contents → Run Profile.

    Returns 'save', 'rerun', or 'back'. In one-shot CLI mode there is no
    menu/rerun loop — draw the plain footer and return 'back'.
    """
    print()
    if not INTERACTIVE:
        standard_footer()
        return "back"
    standard_footer("[R] Run Again   [S] Save as Profile   [Q/ESC] Quit/Back")
    sys.stdout.write(color_text(" Option: ", fg=CYAN, style=BOLD))
    sys.stdout.flush()
    if sys.stdin.isatty():
        while True:
            key = read_key()
            if key in ("r", "R"):
                print()
                return "rerun"
            if key in ("s", "S"):
                print()
                return "save"
            if key in ("q", "Q", "ESC", "ENTER"):
                print()
                return "back"
            # anything else: ignore
    # Non-TTY (piped) fallback — line-based
    try:
        resp = sys.stdin.readline()
    except Exception:
        return "back"
    resp = resp.strip().lower()
    if resp == "s":
        return "save"
    return "rerun" if resp == "r" else "back"


def _get_activity_log_path():
    """fm.log's location, from fmConfig.json's top-level "activityLogPath"
    (e.g. "~/Documents/log/fm.log") — falling back to the default below
    when the config file is missing, invalid JSON, or doesn't set the key.
    Read once at import time, like FM's other path constants."""
    default = os.path.expanduser("~/Documents/log/fm.log")
    try:
        with open(CONFIG_FILE, "r") as fh:
            cfg = json.load(fh)
        path = cfg.get("activityLogPath")
        if path:
            return os.path.expanduser(str(path))
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        pass
    return default


ACTIVITY_LOG = _get_activity_log_path()
MILESTONE_INTERVAL_SECONDS = 15


def _strip_screen_codes(raw):
    """Turn raw terminal output (as written by print()/the stationary-frame
    helpers) into plain, readable text for the activity log.

    Beyond ordinary SGR color codes, the Monitor/Sync stationary frame (see
    _frm_mv et al.) draws with absolute cursor-position moves, erase-line,
    cursor hide/show, scroll-region set/reset, and DEC save/restore cursor —
    none of which are plain "\\x1b[<digits><letter>" SGR codes, so a naive
    strip leaves visible junk like "^[[?25l" in the log. This:
      1. Converts each absolute row move (\\x1b[R;CH) to a newline *before*
         stripping, so header/footer fields land on their own line instead
         of being silently deleted and mashed together.
      2. Strips every remaining CSI sequence (colors, erase-line, cursor
         hide/show, scroll-region, screen clear, private-mode codes — the
         `?` in \\x1b[?25l is a CSI parameter byte, covered here).
      3. Converts the save-cursor marker (ESC 7) that precedes each in-place
         footer update (see _monitor_update_footer / _sync_update_footer) to
         a newline too, so repeated live-progress updates each land on their
         own line instead of running together; the matching restore-cursor
         (ESC 8) is simply dropped.
      4. Collapses the empty/blank lines that erased-but-otherwise-unwritten
         frame rows leave behind.
    """
    text = re.sub(r"\x1b\[\d+;\d+H", "\n", raw)
    text = re.sub(r"\x1b\[[0-?]*[ -/]*[@-~]", "", text)
    text = text.replace("\x1b7", "\n").replace("\x1b8", "")
    text = text.replace("\r\n", "\n").replace("\r", "")
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text


ACTION_LOG_DETAIL_THRESHOLD = 100  # <= this many items: log every action; above: periodic


def _fmt_elapsed(seconds):
    """Human-readable elapsed time for a run summary (e.g. "3.2s", "1m 08s",
    "1h 02m")."""
    seconds = max(0, seconds)
    if seconds < 60:
        return f"{seconds:.1f}s"
    minutes, secs = divmod(int(seconds), 60)
    if minutes < 60:
        return f"{minutes}m {secs:02d}s"
    hours, minutes = divmod(minutes, 60)
    return f"{hours}h {minutes:02d}m"


def _log_line(msg):
    """Append one [YYYY-MM-DD HH:MM:SS] msg line to the FM activity log —
    the format LogView (~/Documents/script/LogView) already expects (a
    leading bracketed timestamp). Never raises — a logging failure must
    never interrupt the on-screen operation."""
    try:
        os.makedirs(os.path.dirname(ACTIVITY_LOG), exist_ok=True)
        with open(ACTIVITY_LOG, "a") as fh:
            fh.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}\n")
    except OSError:
        pass


class _RunLog:
    """Structured start/action/summary/end logging for one FM operation run,
    on top of (not instead of) _ActivityLog's raw full-screen mirror. Every
    line gets its own [YYYY-MM-DD HH:MM:SS] timestamp:

        [2026-07-21 15:20:00] Sync Folders started
        [2026-07-21 15:20:00]   Folder A: /Volumes/A
        [2026-07-21 15:20:00]   Folder B: /Volumes/B
        [2026-07-21 15:20:03] Copied IMG_0001.mov -> B (19.2 MB)
        ...
        [2026-07-21 15:22:41] Sync Folders finished — Copied 42 file(s)
        [2026-07-21 15:22:41]   (took 2m 41s)

    Written immediately as the run progresses (independent of
    _ActivityLog, which only writes once the run finishes) — a run that's
    interrupted partway (Ctrl-C, closed terminal, crash) still leaves a
    trail of how far it got.

    total_items decides per-action granularity: at or under
    ACTION_LOG_DETAIL_THRESHOLD, every action() call writes its own line;
    above it, action() throttles to one line per MILESTONE_INTERVAL_SECONDS
    (the first and last action always log, so real start/end still show)."""

    def __init__(self, activity, fields=None, total_items=None):
        self.activity = activity
        self.total_items = total_items
        self.detailed = total_items is None or total_items <= ACTION_LOG_DETAIL_THRESHOLD
        self.count = 0
        self._start = time.monotonic()
        self._last_logged = 0.0
        _log_line(f"{activity} started")
        for label, value in (fields or []):
            _log_line(f"  {label}: {value}")

    def action(self, msg):
        """Log one action (a file converted/removed/copied/renamed/etc.)."""
        self.count += 1
        if self.detailed:
            _log_line(msg)
            return
        now = time.monotonic()
        is_last = self.total_items is not None and self.count >= self.total_items
        if self.count == 1 or is_last or (now - self._last_logged) >= MILESTONE_INTERVAL_SECONDS:
            suffix = f" ({self.count}/{self.total_items})" if self.total_items else f" ({self.count})"
            _log_line(msg + suffix)
            self._last_logged = now

    def finish(self, summary):
        """Log the run's final summary line. Call exactly once, whether the
        run completed, was cancelled, or hit an error."""
        elapsed = time.monotonic() - self._start
        _log_line(f"{self.activity} finished — {summary}")
        _log_line(f"  (took {_fmt_elapsed(elapsed)})")


class _ActivityLog:
    """Context manager that mirrors everything printed to stdout into the FM
    activity log (~/Documents/log/fm.log). The screen keeps its colors; the
    log entry is written as plain readable text (see _strip_screen_codes),
    prefixed with a timestamp line. Logging failures never interrupt the
    on-screen output."""

    def __init__(self):
        self._buf = []
        self._stdout = None

    # file-like proxy -------------------------------------------------------
    def write(self, s):
        self._stdout.write(s)
        self._buf.append(s)

    def flush(self):
        self._stdout.flush()

    def isatty(self):
        return self._stdout.isatty()

    # context manager -------------------------------------------------------
    def __enter__(self):
        self._stdout = sys.stdout
        sys.stdout = self
        return self

    def __exit__(self, exc_type, exc, tb):
        sys.stdout = self._stdout
        text = _strip_screen_codes("".join(self._buf))
        try:
            os.makedirs(os.path.dirname(ACTIVITY_LOG), exist_ok=True)
            with open(ACTIVITY_LOG, "a") as fh:
                fh.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]\n")
                fh.write(text.lstrip("\n"))
                if not text.endswith("\n"):
                    fh.write("\n")
                fh.write("\n")
        except OSError:
            pass
        return False


class EscCancelled(Exception):
    """Raised by ask()/safe_confirm()/confirm_yes_word() (any question asked
    while a menu option is running) when the user presses ESC.

    Every top-level *_menu() function catches this around its option dispatch
    and treats it as an instant Back: the in-progress question/wizard is
    abandoned and that same menu is redrawn — exactly like [Q/ESC] already
    works on render_menu()/pause_return()/pause_rerun(), just extended to
    free-text and Y/N prompts, which previously used plain input() with no
    ESC awareness at all.
    """
    pass


def _read_line_esc(prompt="", default=""):
    """Read one line of text (case preserved). Enter submits — an empty
    buffer returns `default`. Backspace edits. A bare ESC raises
    EscCancelled immediately, no Enter needed.

    Falls back to a plain line read when stdin isn't an interactive TTY
    (piped input can't signal ESC instantly, so it can't cancel there).
    """
    sys.stdout.write(prompt)
    sys.stdout.flush()

    if not sys.stdin.isatty():
        try:
            line = sys.stdin.readline()
        except Exception:
            return default
        if line == "":                       # EOF
            return default
        val = line.strip()
        return val if val else default

    try:
        import termios, tty, select
    except ImportError:                      # non-POSIX fallback
        val = input().strip()
        return val if val else default

    fd = sys.stdin.fileno()
    old = termios.tcgetattr(fd)
    buf = ""
    try:
        # setraw() disables OPOST, so a bare "\n" here would just drop the
        # cursor a row without returning it to column 0 (staircasing every
        # prompt after it) — write "\r\n" explicitly instead.
        tty.setraw(fd)
        while True:
            ch = os.read(fd, 1)
            if not ch:
                raise EscCancelled
            if ch == b"\x1b":                # ESC — could be a bare ESC or an arrow seq
                r, _, _ = select.select([fd], [], [], 0.03)
                if r:
                    os.read(fd, 2)           # swallow the arrow sequence, ignore
                    continue
                sys.stdout.write("\r\n"); sys.stdout.flush()
                raise EscCancelled
            if ch in (b"\r", b"\n"):         # Enter — submit
                sys.stdout.write("\r\n"); sys.stdout.flush()
                val = buf.strip()
                return val if val else default
            if ch == b"\x03":                # Ctrl-C
                raise KeyboardInterrupt
            if ch in (b"\x7f", b"\x08"):     # Backspace
                if buf:
                    buf = buf[:-1]
                    sys.stdout.write("\b \b"); sys.stdout.flush()
                continue
            c = ch.decode("utf-8", "ignore")
            if c and c.isprintable():        # buffer printable chars (echo manually)
                buf += c
                sys.stdout.write(c); sys.stdout.flush()
    finally:
        termios.tcsetattr(fd, termios.TCSADRAIN, old)


def safe_confirm(prompt, default=False):
    """Yes/No prompt (Enter submits `default`). ESC cancels immediately —
    raises EscCancelled, which the calling menu catches and redraws the menu
    the question came from. A closed/piped stdin also returns `default`
    rather than blocking or crashing."""
    suffix = " [Y/n]: " if default else " [y/N]: "
    full_prompt = color_text(prompt + suffix, fg=YELLOW)
    while True:
        resp = _read_line_esc(full_prompt, "").lower()
        if resp == "":
            return default
        if resp in ("y", "yes"):
            return True
        if resp in ("n", "no"):
            return False


def report_result(ok, ok_msg, fail_msg=""):
    """Show a success/failure line and play the matching project sound."""
    if ok:
        print(color_text(f"  ✓ {ok_msg}", fg=BRIGHT_GREEN, style=BOLD))
        play_sound(get_project_sound(SCRIPT_NAME, "successAudio", SOUND_SUCCESS))
    else:
        print(color_text(f"  ✗ {fail_msg or ok_msg}", fg=BRIGHT_RED, style=BOLD))
        play_sound(get_project_sound(SCRIPT_NAME, "failureAudio", SOUND_FAILURE))


# 5-row block glyphs (each row 6 cols wide) for the pre-start countdown —
# same art as Backup42's _COUNTDOWN_GLYPHS.
_COUNTDOWN_GLYPHS = {
    '3': [
        "██████",
        "     █",
        " █████",
        "     █",
        "██████",
    ],
    '2': [
        "██████",
        "     █",
        "██████",
        "█     ",
        "██████",
    ],
    '1': [
        "   ██ ",
        " ████ ",
        "   ██ ",
        "   ██ ",
        " █████",
    ],
}
_COUNTDOWN_GLYPH_H = 5


def _run_countdown(seconds=3, label="Starting"):
    """Backup42-style 3..2..1 block-digit countdown, shown right after the
    user confirms a file-creating/modifying/deleting operation and right
    before it begins. Always plays — interactive menu AND CLI runs (even
    with --yes/--apply/--delete/--copy) — per project convention.

    On an interactive tty, ESC or Q cancels (returns False); the caller
    should treat that exactly like the confirm itself coming back 'No'. A
    non-tty run (piped/cron/no controlling terminal) can't read a cancel
    key, so it just prints the count and sleeps, always proceeding."""
    seconds = int(seconds)
    if seconds <= 0:
        return True
    if not sys.stdin.isatty():
        print()
        for n in range(seconds, 0, -1):
            print(color_text(f"  {label} in {n}...", fg=BRIGHT_YELLOW))
            time.sleep(1)
        return True

    import termios, tty, select
    cols = _frm_term_size().columns
    cap1 = f"{label}..."
    cap2 = "Press [ESC] or [Q] to cancel"
    block_h = _COUNTDOWN_GLYPH_H + 3   # glyph rows + blank + 2 caption rows

    fd = sys.stdin.fileno()
    old = termios.tcgetattr(fd)
    tty.setcbreak(fd)
    print()
    proceed = True
    try:
        for i, n in enumerate(range(seconds, 0, -1)):
            if i > 0:
                sys.stdout.write(f"\x1b[{block_h}A")
            glyph = _COUNTDOWN_GLYPHS.get(str(n), _COUNTDOWN_GLYPHS['1'])
            lines = []
            for row in glyph:
                pad = max(0, (cols - len(row)) // 2)
                lines.append(" " * pad + f"{BOLD}{BRIGHT_GREEN}{row}{RESET}")
            lines.append("")
            pad1 = max(0, (cols - len(cap1)) // 2)
            lines.append(" " * pad1 + f"{BOLD}{WHITE}{cap1}{RESET}")
            pad2 = max(0, (cols - len(cap2)) // 2)
            lines.append(" " * pad2 + f"{DIM}{BRIGHT_YELLOW}{cap2}{RESET}")
            for line in lines:
                sys.stdout.write("\x1b[2K" + line + "\n")
            sys.stdout.flush()

            deadline = time.monotonic() + 1.0
            while True:
                remaining = deadline - time.monotonic()
                if remaining <= 0:
                    break
                ready, _, _ = select.select([fd], [], [], remaining)
                if not ready:
                    continue
                ch = sys.stdin.read(1)
                if ch == '\x1b':
                    # Distinguish a bare ESC (cancel) from an escape sequence
                    # such as an arrow key (ignored).
                    seq, _, _ = select.select([fd], [], [], 0.05)
                    if seq:
                        sys.stdin.read(1)
                        seq2, _, _ = select.select([fd], [], [], 0.05)
                        if seq2:
                            sys.stdin.read(1)
                        continue
                    proceed = False
                    break
                if ch in ('q', 'Q'):
                    proceed = False
                    break
            if not proceed:
                break
    finally:
        termios.tcsetattr(fd, termios.TCSADRAIN, old)
        # Wipe the countdown block and leave the cursor exactly where it
        # started, so whatever prints next continues in place.
        sys.stdout.write(f"\x1b[{block_h}A")
        for _ in range(block_h):
            sys.stdout.write("\x1b[2K\n")
        sys.stdout.write(f"\x1b[{block_h}A")
        sys.stdout.flush()
    return proceed


class _PauseStop:
    """Non-blocking pause/stop control for a long-running batch loop (Sync,
    Convert Video/Images, Zip SubFolders, Remove/Cleanup, Find & Replace/
    Rename, Purge Old Log Files). Call check() once per item — cheap,
    non-blocking:

        ps = _PauseStop()
        print(color_text("  Press [P] to pause, [Q] to stop early.", fg=DIM))
        try:
            for item in items:
                if ps.check() == "stop":
                    break
                ... process item ...
        finally:
            ps.close()

    [P] pauses immediately (blocks until [P] resumes or [Q]/ESC stops from
    within the pause); [Q] or ESC stops directly. Stopping is cooperative —
    the current item always finishes; the loop breaks before starting the
    next one, so nothing is left half-written. A non-interactive run
    (piped/cron, no controlling tty) can't read a key at all: check() is
    then a permanent no-op and the batch always runs to completion.

    Always call close() in a finally block to restore the terminal — never
    skip it, even on an exception or early return."""

    def __init__(self):
        self.is_tty = sys.stdin.isatty()
        self.stopped = False
        self._fd = sys.stdin.fileno() if self.is_tty else None
        self._old = None
        if self.is_tty:
            import termios, tty
            self._old = termios.tcgetattr(self._fd)
            tty.setcbreak(self._fd)

    def _read_key(self, timeout=0):
        import select
        ready, _, _ = select.select([self._fd], [], [], timeout)
        if not ready:
            return None
        ch = sys.stdin.read(1)
        if ch == '\x1b':
            # Distinguish a bare ESC (stop) from an escape sequence such as
            # an arrow key (ignored) by peeking for more bytes right away.
            seq, _, _ = select.select([self._fd], [], [], 0.02)
            if seq:
                sys.stdin.read(1)
                seq2, _, _ = select.select([self._fd], [], [], 0.02)
                if seq2:
                    sys.stdin.read(1)
                return None
            return 'ESC'
        return ch

    def check(self):
        """Call once per item. Returns "stop" once the user has asked to
        stop (directly, or via pause -> stop); otherwise None."""
        if not self.is_tty or self.stopped:
            return "stop" if self.stopped else None
        ch = self._read_key()
        if ch is None:
            return None
        if ch in ('q', 'Q', 'ESC'):
            self.stopped = True
            print(color_text("\n  ■ Stopping — finishing the current item, "
                             "then stopping early.", fg=BRIGHT_YELLOW, style=BOLD))
            return "stop"
        if ch in ('p', 'P'):
            return self._pause()
        return None

    def _pause(self):
        print(color_text("\n  ⏸ PAUSED — press [P] to resume, [Q] to stop.",
                         fg=BRIGHT_YELLOW, style=BOLD))
        while True:
            ch = self._read_key(timeout=0.2)
            if ch is None:
                continue
            if ch in ('p', 'P'):
                print(color_text("  ▶ Resumed.", fg=BRIGHT_GREEN))
                return None
            if ch in ('q', 'Q', 'ESC'):
                self.stopped = True
                print(color_text("  ■ Stopping — finishing the current item, "
                                 "then stopping early.", fg=BRIGHT_YELLOW, style=BOLD))
                return "stop"

    def close(self):
        if self.is_tty and self._old is not None:
            import termios
            termios.tcsetattr(self._fd, termios.TCSADRAIN, self._old)


def fmt_size(n):
    """Human-readable byte size."""
    try:
        n = float(n)
    except (TypeError, ValueError):
        return "—"
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if n < 1024 or unit == "TB":
            return f"{int(n):,} B" if unit == "B" else f"{n:,.1f} {unit}"
        n /= 1024


def fmt_size_short(n):
    """Compact byte size for tight table cells (e.g. 512B, 1.1K, 2.2M)."""
    try:
        n = float(n)
    except (TypeError, ValueError):
        return "—"
    for unit in ("B", "K", "M", "G", "T"):
        if n < 1024 or unit == "T":
            return f"{int(n)}B" if unit == "B" else f"{n:.1f}{unit}"
        n /= 1024


def fmt_commas(n):
    try:
        return f"{int(n):,}"
    except (TypeError, ValueError):
        return str(n)


def fmt_date(dt):
    try:
        return dt.strftime("%b %-d, %Y %-I:%M %p").replace("AM", "am").replace("PM", "pm")
    except Exception:
        return "—"


def ask(label, default=""):
    """Prompt for a line of input with an optional default. ESC cancels
    immediately (raises EscCancelled) — the calling menu catches it and
    redraws the menu the question came from."""
    suffix = f" {DIM}[{default}]{RESET}" if default else ""
    return _read_line_esc(color_text(f"  {label}", fg=CYAN) + suffix + ": ", default)


def clean_path(raw):
    """Normalize a path the user typed or pasted, then expand ~.

    Handles the common ways a path arrives with extra characters:
      • wrapped in matching single/double quotes  →  '/My Drive/x'  "…"
      • drag-and-dropped from the terminal, which backslash-escapes spaces
        and other characters  →  /My\\ Drive/x
    A plain path (even one containing spaces) is left untouched. Backslash
    escapes are only undone if doing so points at something that exists, so a
    legitimate backslash in a name is preserved.
    """
    if raw is None:
        return ""
    s = raw.strip()
    # Strip one layer of surrounding matching quotes (copy/paste)
    if len(s) >= 2 and s[0] == s[-1] and s[0] in ("'", '"'):
        s = s[1:-1]
    expanded = os.path.expanduser(s)
    if os.path.exists(expanded):
        return expanded
    # Undo shell-style backslash escapes (terminal drag-and-drop) and retry
    unescaped = os.path.expanduser(re.sub(r"\\(.)", r"\1", s))
    if os.path.exists(unescaped):
        return unescaped
    # Nothing resolved — return the quote-stripped, expanded form for the error
    return expanded


def _load_common_folders():
    return _load_config_profiles("commonFolders")


def _pick_common_folder():
    """Arrow-driven picker for ask_folder()'s '+' shortcut.

    Returns the chosen folder path, or None if the user backs out (Q/ESC),
    no common folders are saved yet, or fmConfig.json is broken — in every
    None case the caller re-shows the plain text folder prompt.
    """
    folders, err = _load_common_folders()
    if err:
        print(color_text(f"  ⚠ {err}", fg=RED))
        return None
    if not folders:
        print(color_text(
            "  No common folders saved yet — add some via Admin Menu -> "
            "Manage Common Folders.", fg=YELLOW))
        return None
    options = [(f.get("name") or "(unnamed)", f.get("path", "")) for f in folders]
    ch = render_menu("Common Folders", options)
    if ch == "back":
        return None
    return folders[int(ch) - 1].get("path", "")


def _common_folder_names_line():
    """Return 'Common: name1, name2, ...' listing saved common folders (in
    saved order), or '' if none are saved or fmConfig.json is broken."""
    folders, err = _load_common_folders()
    if err or not folders:
        return ""
    names = ", ".join(f.get("name") or "(unnamed)" for f in folders)
    return f"Common: {names}"


def _resolve_common_folder_name(raw):
    """If raw exactly matches a saved common folder's name (case-insensitive,
    since it's typed by hand), return that folder's path. Otherwise None —
    the caller falls back to treating raw as a literal path."""
    if not raw:
        return None
    folders, err = _load_common_folders()
    if err or not folders:
        return None
    for f in folders:
        if (f.get("name") or "").strip().lower() == raw.strip().lower():
            return f.get("path", "")
    return None


def ask_path(label, default=""):
    """Like ask(), but typing '+' and pressing Enter opens an arrow-driven
    picker over the folders saved via Admin Menu -> Manage Common Folders
    (Enter picks, Q/ESC goes back to typing a path/value by hand). A static
    header line above the prompt names them too, so they're visible without
    opening the picker. Typing a saved folder's name directly (instead of
    '+') also resolves to its path — see _resolve_common_folder_name().

    Used directly by prompts that accept a file OR a folder (so they can't
    use ask_folder()'s must-be-a-directory check), and internally by
    ask_folder() itself.
    """
    common_line = _common_folder_names_line()
    if common_line:
        print(color_text(f"  {common_line}", style=DIM))
    raw = ask(label, default)
    while raw == "+":
        picked = _pick_common_folder()
        raw = picked if picked else ask(label, default)
    return _resolve_common_folder_name(raw) or raw


def ask_folder(label="Folder", default="", must_exist=True, key=None):
    """Prompt for a folder path (cleans quotes/escapes, expands ~).

    key: when given, identifies this specific prompt in LAST_PATHS. If a
    path was entered here before (this run or a prior one, via
    fmLastPaths.json), it becomes the default — Enter reuses it, typing a
    new path overrides it and becomes the new remembered value.

    Supports the '+' common-folder picker via ask_path() — see its docstring.
    """
    if key and LAST_PATHS.get(key):
        default = LAST_PATHS[key]
    raw = ask_path(label, default)
    if not raw:
        print(color_text("  No path entered.", fg=YELLOW))
        return None
    path = clean_path(raw)
    if must_exist and not os.path.isdir(path):
        print(color_text(f"  Not a directory: {path}", fg=RED))
        return None
    if key:
        LAST_PATHS[key] = path
    return path


def ask_file(label="File", default=""):
    """Prompt for a file path (cleans quotes/escapes, expands ~)."""
    raw = ask(label, default)
    if not raw:
        print(color_text("  No path entered.", fg=YELLOW))
        return None
    path = clean_path(raw)
    if not os.path.isfile(path):
        print(color_text(f"  Not a file: {path}", fg=RED))
        return None
    return path


def is_hidden(rel):
    """True if any path component starts with '.'"""
    return any(part.startswith(".") for part in rel.replace("\\", "/").split("/") if part)


# Folder/file names always excluded from searches and compares (case-insensitive).
EXCLUDED_DIR_NAMES  = {"$recycle.bin"}
EXCLUDED_FILE_NAMES = {".ds_store", "desktop.ini"}


def prune_dirs(dns):
    """In-place filter for an os.walk() dir list: drop excluded folders
    (e.g. the Windows $RECYCLE.BIN on external drives) so they are neither
    matched nor descended into."""
    dns[:] = [d for d in dns if d.lower() not in EXCLUDED_DIR_NAMES]
    return dns


def is_excluded_file(fn):
    """True for junk files ignored by searches and compares (.DS_Store,
    desktop.ini)."""
    return fn.lower() in EXCLUDED_FILE_NAMES


# -----------------------------------------------------------------------------
# Filesystem scanning helpers
# -----------------------------------------------------------------------------
def rel_files(root):
    """Set of relative file paths under root, excluding hidden files/folders
    and excluded names ($RECYCLE.BIN, .DS_Store, desktop.ini)."""
    out = set()
    for dp, dns, fns in os.walk(root):
        dns[:] = [d for d in dns if not d.startswith(".") and d.lower() not in EXCLUDED_DIR_NAMES]
        for fn in fns:
            if fn.startswith(".") or is_excluded_file(fn):
                continue
            out.add(os.path.relpath(os.path.join(dp, fn), root))
    return out


def immediate_subfolders(root):
    """Sorted list of immediate, non-hidden subfolder names."""
    try:
        return sorted(
            d for d in os.listdir(root)
            if not d.startswith(".") and os.path.isdir(os.path.join(root, d))
        )
    except OSError:
        return []


def folder_stats(path):
    """Return (total_bytes, file_count) for everything under path."""
    total = 0
    count = 0
    for dp, dns, fns in os.walk(path):
        for fn in fns:
            try:
                total += os.path.getsize(os.path.join(dp, fn))
                count += 1
            except OSError:
                pass
    return total, count


def file_size(path):
    try:
        return os.path.getsize(path)
    except OSError:
        return 0


def _file_dates(path):
    """Returns (created_str, modified_str) for a file, formatted via
    fmt_date(). Falls back to mtime for created on filesystems without
    birthtime (most Linux). Best-effort — never raises."""
    try:
        st = os.stat(path)
    except OSError:
        return "—", "—"
    created = getattr(st, "st_birthtime", st.st_mtime)
    return fmt_date(datetime.fromtimestamp(created)), fmt_date(datetime.fromtimestamp(st.st_mtime))


# =============================================================================
# COMPARE
# =============================================================================
def _read_lines(path):
    """Read a text file into a list of lines (best-effort, never raises)."""
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            return f.read().splitlines()
    except OSError as e:
        print(color_text(f"  Cannot read {path}: {e}", fg=RED))
        return None


def _sbs_row(num, left, right, mark, col, color):
    """Print one side-by-side row: line# | left | mark | right."""
    l = left[:col].ljust(col)
    r = right[:col]
    n = f"{num:>4}" if num else "    "
    print(f"{DIM}{n}{RESET} {color}{l}{RESET} {color}{mark}{RESET} {color}{r}{RESET}")


def compare_two_files(file_a, file_b):
    """Compare two files and display them side by side, line by line.

    Markers:  =  equal    ≠  changed    <  only in A    >  only in B
    """
    screen("Compare 2 Files")
    print()
    runlog = _RunLog("Compare 2 Files", [("File A", file_a), ("File B", file_b)])
    if not (os.path.isfile(file_a) and os.path.isfile(file_b)):
        print(color_text("  Both paths must be existing files.", fg=RED))
        runlog.finish("Cancelled — both paths must be existing files")
        pause_return()
        return

    a_lines = _read_lines(file_a)
    b_lines = _read_lines(file_b)
    if a_lines is None or b_lines is None:
        runlog.finish("Cancelled — could not read one or both files")
        pause_return()
        return

    w = get_width()
    col = max((w - 9) // 2, 20)   # two columns + line#(4) + marker gutters

    print(f"  {YELLOW}A{RESET}: {file_a}  {DIM}({len(a_lines)} lines, {fmt_size(file_size(file_a))}){RESET}")
    print(f"  {YELLOW}B{RESET}: {file_b}  {DIM}({len(b_lines)} lines, {fmt_size(file_size(file_b))}){RESET}")
    print()
    print(color_text(f"  {'#':>3} {os.path.basename(file_a)[:col]:<{col}}   {os.path.basename(file_b)[:col]}",
                     fg=WHITE, style=BOLD))
    print(f"  {DIM}{'-' * min(col * 2 + 8, w - 2)}{RESET}")

    sm = difflib.SequenceMatcher(None, a_lines, b_lines, autojunk=False)
    same = changed = only_a = only_b = 0
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            for k in range(i2 - i1):
                _sbs_row(i1 + k + 1, a_lines[i1 + k], b_lines[j1 + k], "=", col, WHITE)
                same += 1
        elif tag == "replace":
            la, lb = a_lines[i1:i2], b_lines[j1:j2]
            for k in range(max(len(la), len(lb))):
                left = la[k] if k < len(la) else ""
                right = lb[k] if k < len(lb) else ""
                _sbs_row((i1 + k + 1) if k < len(la) else 0, left, right, "≠", col, YELLOW)
                changed += 1
        elif tag == "delete":
            for k in range(i1, i2):
                _sbs_row(k + 1, a_lines[k], "", "<", col, RED)
                only_a += 1
        elif tag == "insert":
            for k in range(j1, j2):
                _sbs_row(0, "", b_lines[k], ">", col, GREEN)
                only_b += 1

    print(f"  {DIM}{'-' * min(col * 2 + 8, w - 2)}{RESET}")
    if changed == only_a == only_b == 0:
        report_result(True, "Files are identical.")
        runlog.finish("Files are identical")
    else:
        print(color_text(
            f"  {same} equal · {changed} changed · {only_a} only in A · {only_b} only in B",
            fg=BRIGHT_CYAN, style=BOLD))
        report_result(False, "", "Files differ.")
        runlog.finish(f"Files differ — {same} equal, {changed} changed, "
                       f"{only_a} only in A, {only_b} only in B")
    if pause_rerun():
        compare_two_files(file_a, file_b)


def _compare_folder_contents_screen(folder_a, folder_b, recursive, by, case_sensitive):
    """One pass of Compare Folder Contents: prints the report. Returns True
    if the comparison ran, False if the paths were invalid (already handled
    on screen, including pause_return())."""
    runlog = _RunLog("Compare Folder Contents", [
        ("Folder A", folder_a),
        ("Folder B", folder_b),
        ("Recursive", "Yes" if recursive else "No"),
        ("Compare By", by),
        ("Case-Sensitive", "Yes" if case_sensitive else "No"),
    ])
    screen("Compare Folder Contents")
    print()
    print(f"  {YELLOW}A{RESET}: {folder_a}")
    print(f"  {YELLOW}B{RESET}: {folder_b}")
    print(f"  {YELLOW}Recursive{RESET}: {'Yes' if recursive else 'No'}    "
          f"{YELLOW}Compare by{RESET}: {by}    "
          f"{YELLOW}Case-sensitive{RESET}: {'Yes' if case_sensitive else 'No'}")
    print()

    if not (os.path.isdir(folder_a) and os.path.isdir(folder_b)):
        print(color_text("  Both paths must be directories.", fg=RED))
        runlog.finish("Cancelled — both paths must be directories")
        pause_return()
        return False

    if recursive:
        set_a, set_b = rel_files(folder_a), rel_files(folder_b)
    else:
        def entries(root):
            try:
                return {e for e in os.listdir(root)
                        if not e.startswith(".") and e.lower() not in EXCLUDED_DIR_NAMES
                        and not is_excluded_file(e)}
            except OSError:
                return set()
        set_a, set_b = entries(folder_a), entries(folder_b)

    def size_of(root, rel):
        p = os.path.join(root, rel)
        if os.path.isdir(p):
            total, _ = folder_stats(p)
            return total
        return file_size(p)

    # Case-insensitive matching keys each set by lowercased path but keeps the
    # original (on-disk) casing for display, so a same-name/different-case
    # pair in A and B is treated as one entry instead of two "missing" ones.
    if case_sensitive:
        key_a = {rel: rel for rel in set_a}
        key_b = {rel: rel for rel in set_b}
    else:
        key_a = {rel.lower(): rel for rel in set_a}
        key_b = {rel.lower(): rel for rel in set_b}

    only_a = only_b = diffs = case_diffs = []
    if by in ("name", "both"):
        only_a = sorted(key_a[k] for k in key_a.keys() - key_b.keys())
        print(color_text(f"  In A but not in B ({len(only_a)}):", fg=BRIGHT_CYAN, style=BOLD))
        for rel in only_a:
            print(f"    {rel}")
        if not only_a:
            print(f"    {DIM}(none){RESET}")
        print()
        only_b = sorted(key_b[k] for k in key_b.keys() - key_a.keys())
        print(color_text(f"  In B but not in A ({len(only_b)}):", fg=BRIGHT_CYAN, style=BOLD))
        for rel in only_b:
            print(f"    {rel}")
        if not only_b:
            print(f"    {DIM}(none){RESET}")
        print()

    if by in ("size", "both"):
        common_keys = sorted(key_a.keys() & key_b.keys())
        diffs = []
        case_diffs = []
        for k in common_keys:
            rel_a, rel_b = key_a[k], key_b[k]
            sa = size_of(folder_a, rel_a)
            sb = size_of(folder_b, rel_b)
            label = rel_a if rel_a == rel_b else f"{rel_a}  {DIM}(B: {rel_b}){RESET}"
            if sa != sb:
                diffs.append((label, sa, sb))
            elif rel_a != rel_b:
                case_diffs.append(label)
        print(color_text(f"  In both but differing in size ({len(diffs)}):", fg=BRIGHT_CYAN, style=BOLD))
        for label, sa, sb in diffs:
            print(f"    {label}  {DIM}(A: {fmt_size(sa)}, B: {fmt_size(sb)}){RESET}")
        if not diffs:
            print(f"    {DIM}(none){RESET}")
        if not case_sensitive and case_diffs:
            print()
            print(color_text(f"  Same file, different case ({len(case_diffs)}):", fg=BRIGHT_CYAN, style=BOLD))
            for label in case_diffs:
                print(f"    {label}")

    print()
    print(f"  {YELLOW}A{RESET}: {folder_a}")
    print(f"  {YELLOW}B{RESET}: {folder_b}")

    summary = (f"{len(only_a)} only in A, {len(only_b)} only in B, "
               f"{len(diffs)} differing in size")
    runlog.finish(summary)
    return True


def compare_folder_contents(folder_a, folder_b, recursive=True, by="both",
                             case_sensitive=False, offer_save=True):
    """Compare the contents of two folders.

    recursive:      True  -> every file beneath each folder (relative paths)
                     False -> only the immediate entries (files AND subfolders)
    by:              name -> entries present in one folder but not the other
                     size -> entries in both whose size differs
                     both -> both reports
    case_sensitive:  True  -> "IMG_1.mov" and "IMG_1.MOV" are different entries
                     False -> they're matched as the same entry (default; matches
                     how macOS's default case-insensitive filesystem treats names)
    Hidden files/folders (leading '.') are ignored.
    offer_save:     True  -> end-of-run footer adds [S] Save as Profile
                             (Interactive Compare Folder Contents)
                     False -> plain [R] Run Again / [Q/ESC] Quit/Back
                             (already-saved profile run — nothing new to save)
    """
    while True:
        if not _compare_folder_contents_screen(folder_a, folder_b, recursive, by, case_sensitive):
            return
        if offer_save:
            choice = _pause_compare_folders()
            if choice == "rerun":
                continue
            if choice == "save":
                _save_compare_profile_flow(folder_a, folder_b, recursive, by, case_sensitive)
                continue
            return
        else:
            if not pause_rerun():
                return


def _save_compare_profile_flow(folder_a, folder_b, recursive, by, case_sensitive):
    """[S] Save as Profile — prompt for a name and append the folders/options
    just used to fmConfig.json's compareProfiles list."""
    print()
    name = ask("Profile name", default="").strip()
    if not name:
        print(color_text("  Cancelled — a name is required.", fg=YELLOW))
        return
    profiles, perr = _load_compare_profiles()
    if perr:
        print(color_text(f"  ⚠ {perr}", fg=RED))
        return
    profiles.append({
        "name": name,
        "folderA": folder_a,
        "folderB": folder_b,
        "recursive": recursive,
        "compareBy": by,
        "caseSensitive": case_sensitive,
    })
    err = _save_config_profiles("compareProfiles", profiles)
    if err:
        print(color_text(f"  ⚠ {err}", fg=RED))
    else:
        print(color_text(f"  ✓ Saved profile '{name}'.", fg=GREEN))


def _run_compare_profile(pr):
    """Run one compareProfiles entry from fmConfig.json."""
    folder_a = clean_path(str(pr.get("folderA", "")))
    folder_b = clean_path(str(pr.get("folderB", "")))
    recursive = bool(pr.get("recursive", True))
    by = str(pr.get("compareBy", "both")).lower()
    if by not in ("name", "size", "both"):
        by = "both"
    case_sensitive = bool(pr.get("caseSensitive", False))
    compare_folder_contents(folder_a, folder_b, recursive, by, case_sensitive,
                             offer_save=False)


# =============================================================================
# DISPLAY — Folder Sizes
# =============================================================================
def display_folder_sizes(folder, sort_mode="alpha"):
    screen("Folder Sizes")
    print()
    print(f"  {YELLOW}Folder{RESET}: {folder}")
    print(f"  {YELLOW}Sort{RESET}: {'By Size (largest first)' if sort_mode == 'size' else 'Alphabetical'}")
    print()

    runlog = _RunLog("Display Folder Sizes",
                      [("Folder", folder), ("Sort", sort_mode)])
    if not os.path.isdir(folder):
        print(color_text(f"  Not a directory: {folder}", fg=RED))
        runlog.finish(f"Cancelled — not a directory: {folder}")
        pause_return()
        return

    subs = immediate_subfolders(folder)
    rows = []
    for name in subs:
        total, count = folder_stats(os.path.join(folder, name))
        rows.append((total, count, name))

    if sort_mode == "size":
        rows.sort(key=lambda r: r[0], reverse=True)
    else:
        rows.sort(key=lambda r: r[2].lower())

    # Folder is always the first column. Size it to the longest name (capped),
    # also accounting for the header and the TOTAL label so everything aligns.
    total_label = f"TOTAL ({len(rows)} folders)"
    name_w = max([len(n) for (_, _, n) in rows] + [len("Folder"), len(total_label)])
    name_w = min(name_w, 50)

    print(color_text(f"  {'Folder':<{name_w}}  {'Size':>10}  {'Bytes':>18}  {'Files':>9}", fg=YELLOW, style=BOLD))
    print(f"  {DIM}{'-'*name_w}  {'-'*10}  {'-'*18}  {'-'*9}{RESET}")

    grand_bytes = 0
    grand_files = 0
    for total, count, name in rows:
        grand_bytes += total
        grand_files += count
        disp = name[:name_w].ljust(name_w)
        print(f"  {WHITE}{disp}{RESET}  {GREEN}{fmt_size(total):>10}{RESET}  {fmt_commas(total):>18}  {count:>9,}")

    if not rows:
        print(f"    {DIM}(no subfolders){RESET}")
    else:
        print(f"  {DIM}{'-'*name_w}  {'-'*10}  {'-'*18}  {'-'*9}{RESET}")
        print(f"  {BOLD}{total_label[:name_w].ljust(name_w)}  {fmt_size(grand_bytes):>10}  {fmt_commas(grand_bytes):>18}  {grand_files:>9,}{RESET}")

    runlog.finish(f"{len(rows)} subfolder(s), {fmt_size(grand_bytes)} total")
    pause_return()


def display_all_drives():
    """List every mounted drive — the boot volume plus each drive under
    /Volumes — with its total size, used and free space, and use%. Free is
    what shutil.disk_usage reports as available to the user, so it matches
    Finder's Available figure. Duplicate mounts of the same device are
    listed once."""
    screen("All Drives")
    print()

    runlog = _RunLog("Display All Drives")
    candidates = ["/"]
    volumes = "/Volumes"
    if os.path.isdir(volumes):
        for name in sorted(os.listdir(volumes), key=str.lower):
            path = os.path.join(volumes, name)
            if os.path.ismount(path):
                candidates.append(path)

    rows = []
    seen_devices = set()
    for path in candidates:
        try:
            dev = os.stat(path).st_dev
            usage = shutil.disk_usage(path)
        except OSError:
            continue
        if dev in seen_devices:
            continue
        seen_devices.add(dev)
        name = "/" if path == "/" else os.path.basename(path)
        pct = (usage.used / usage.total * 100) if usage.total else 0
        rows.append((name, path, usage.total, usage.used, usage.free, pct))

    if not rows:
        print(color_text("  No drives found.", fg=YELLOW))
        runlog.finish("No drives found")
        pause_return()
        return

    name_w = max([len(r[0]) for r in rows] + [len("Drive")])
    name_w = min(name_w, 40)
    print(color_text(f"  {'Drive':<{name_w}}  {'Size':>10}  {'Used':>10}  "
                     f"{'Free':>10}  {'Use%':>5}  Mount Point",
                     fg=YELLOW, style=BOLD))
    print(f"  {DIM}{'-' * name_w}  {'-' * 10}  {'-' * 10}  {'-' * 10}  {'-' * 5}  {'-' * 20}{RESET}")
    for name, path, total, used, free, pct in rows:
        pct_color = RED if pct >= 90 else (YELLOW if pct >= 75 else GREEN)
        disp = name[:name_w].ljust(name_w)
        print(f"  {WHITE}{disp}{RESET}  {fmt_size(total):>10}  {fmt_size(used):>10}  "
              f"{GREEN}{fmt_size(free):>10}{RESET}  "
              f"{color_text(f'{pct:>4.0f}%', fg=pct_color)}  {DIM}{path}{RESET}")
    print()
    print(f"  {DIM}{len(rows)} drive(s). Free = space available to you (matches "
          f"Finder's Available).{RESET}")

    runlog.finish(f"{len(rows)} drive(s) listed")
    pause_return()


IOSTAT_SAMPLE_SECONDS = 1  # how long to sample disk I/O for "Drives in Use"


def _sample_drives_in_use(drives, seconds=IOSTAT_SAMPLE_SECONDS):
    """Sample each drive's underlying physical disk for `seconds` via iostat
    and return {identifier: True/False/None}. True/False = actual read/write
    transfers did/didn't happen during the sample; None = could not be
    determined (iostat unavailable, non-macOS, or unexpected output).

    This checks real disk activity rather than open file handles (lsof), so
    a drive with an idle-but-open file (e.g. a mounted disk image someone
    forgot to close) correctly reads as not in use — only actual bytes
    moving during the sample count. One iostat call covers every drive, so
    the whole check takes one sampling window, not one per drive."""
    identifiers = []
    seen = set()
    for d in drives:
        if d["identifier"] not in seen:
            seen.add(d["identifier"])
            identifiers.append(d["identifier"])

    result = {ident: None for ident in identifiers}
    if not identifiers:
        return result

    try:
        proc = subprocess.run(
            ["iostat", "-d", "-c", "2", "-w", str(seconds)] + identifiers,
            capture_output=True, text=True, timeout=seconds + 10)
    except (OSError, subprocess.TimeoutExpired):
        return result
    if proc.returncode != 0:
        return result

    # iostat prints a since-boot average first, then the actual N-second
    # sample as the final line — that last line is the one we want. Fields
    # are 3 per disk (KB/t, tps, MB/s) in the same order as the identifiers
    # were passed on the command line.
    lines = [ln for ln in proc.stdout.splitlines() if ln.strip()]
    if len(lines) < 3:
        return result
    fields = lines[-1].split()
    if len(fields) != len(identifiers) * 3:
        return result

    for i, ident in enumerate(identifiers):
        try:
            result[ident] = float(fields[i * 3 + 1]) > 0  # tps (transfers/sec)
        except ValueError:
            result[ident] = None
    return result


def display_drives_in_use():
    """List every mounted external drive and whether it is currently being
    read from or written to, based on a short (1s) sample of its underlying
    physical disk's actual I/O activity. macOS only (iostat + diskutil)."""
    screen("Drives in Use")
    print()
    runlog = _RunLog("Display Drives in Use")

    if sys.platform != "darwin":
        print(color_text("  Drives in Use uses iostat/diskutil and is only available on macOS.", fg=RED))
        runlog.finish("Cancelled — macOS only")
        pause_return()
        return

    drives = _external_drives()
    if not drives:
        print(color_text("  No external drives found.", fg=YELLOW, style=BOLD))
        runlog.finish("No external drives found")
        pause_return()
        return

    print(color_text(f"  Sampling disk activity ({IOSTAT_SAMPLE_SECONDS}s)...", fg=DIM))
    usage = _sample_drives_in_use(drives)
    print()

    name_w = max([len(d["name"]) for d in drives] + [len("Drive")])
    name_w = min(name_w, 40)
    size_w = max([len(d["size"]) for d in drives] + [len("Size")])

    print(color_text(f"  {'Drive':<{name_w}}  {'Size':>{size_w}}  {'In Use':>7}  Mount Point",
                     fg=YELLOW, style=BOLD))
    print(f"  {DIM}{'-' * name_w}  {'-' * size_w}  {'-' * 7}  {'-' * 20}{RESET}")

    in_use_count = 0
    for d in drives:
        state = usage.get(d["identifier"])
        if state is True:
            in_use_count += 1
            label = color_text(f"{'Yes':>7}", fg=BRIGHT_GREEN, style=BOLD)
        elif state is False:
            label = color_text(f"{'No':>7}", fg=DIM)
        else:
            label = color_text(f"{'?':>7}", fg=YELLOW)
        disp = d["name"][:name_w].ljust(name_w)
        print(f"  {WHITE}{disp}{RESET}  {d['size']:>{size_w}}  {label}  {DIM}{d['mount_point']}{RESET}")

    print()
    print(f"  {DIM}{len(drives)} external drive(s), {in_use_count} in use. 'In Use' reflects "
          f"actual disk read/write activity during a {IOSTAT_SAMPLE_SECONDS}-second sample "
          f"(not just an open file).{RESET}")

    runlog.finish(f"{len(drives)} external drive(s), {in_use_count} in use")
    pause_return()


# =============================================================================
# CONVERT  (CSV / JSON / XLSX / SQL)
# =============================================================================
CONVERT_IN_FORMATS  = ("csv", "json", "xlsx")
CONVERT_OUT_FORMATS = ("csv", "json", "xlsx", "sql")


def _unique_path(path):
    """Collision-safe output path: name.ext, name-2.ext, name-3.ext, …
    (same convention as Zip SubFolders). Never overwrites."""
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    n = 2
    while os.path.exists(f"{base}-{n}{ext}"):
        n += 1
    return f"{base}-{n}{ext}"


def _require_openpyxl():
    """Import openpyxl on demand. XLSX is the only format that needs a
    third-party package; everything else is stdlib."""
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        raise ValueError("XLSX support requires the openpyxl package — "
                         "install it with: pip3 install openpyxl")


def _json_cell(value):
    """Flatten a JSON value to a table cell (nested dict/list → JSON text)."""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return "" if value is None else value


def _read_table(path):
    """Read a .csv / .json / .xlsx file into (headers, rows).

    csv  : first row = header (delimiter sniffed: , ; tab |; BOM tolerated)
    json : array of objects (headers = union of keys, first-seen order) or
           array of arrays (first array = header); a single object = one row
    xlsx : first sheet, first row = header (needs openpyxl)

    Raises ValueError with a user-friendly message on anything invalid.
    """
    ext = os.path.splitext(path)[1].lower().lstrip(".")
    if ext == "csv":
        with open(path, newline="", encoding="utf-8-sig", errors="replace") as fh:
            sample = fh.read(4096)
            fh.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except csv.Error:
                dialect = csv.excel
            data = list(csv.reader(fh, dialect))
        if not data:
            raise ValueError("The CSV file is empty.")
        return [str(h) for h in data[0]], data[1:]

    if ext == "json":
        with open(path, encoding="utf-8") as fh:
            try:
                data = json.load(fh)
            except json.JSONDecodeError as e:
                raise ValueError(f"Invalid JSON: {e}")
        if isinstance(data, dict):
            data = [data]
        if not isinstance(data, list) or not data:
            raise ValueError("JSON must be a non-empty array of objects "
                             "(or an array of arrays with a header row).")
        if isinstance(data[0], dict):
            headers = []
            for rec in data:
                if not isinstance(rec, dict):
                    raise ValueError("JSON array mixes objects with non-objects.")
                for k in rec:
                    if k not in headers:
                        headers.append(k)
            rows = [[_json_cell(rec.get(h)) for h in headers] for rec in data]
            return headers, rows
        if isinstance(data[0], list):
            return [str(h) for h in data[0]], [list(r) for r in data[1:]]
        raise ValueError("Unsupported JSON structure — expected an array of "
                         "objects or an array of arrays.")

    if ext == "xlsx":
        opx = _require_openpyxl()
        wb = opx.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        data = [["" if c is None else c for c in row]
                for row in ws.iter_rows(values_only=True)]
        wb.close()
        if not data:
            raise ValueError("The XLSX file has no rows on its first sheet.")
        return [str(h) for h in data[0]], data[1:]

    raise ValueError(f"Unsupported input type .{ext} — "
                     "supported inputs: .csv, .json, .xlsx")


def _write_csv_out(path, headers, rows):
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(headers)
        w.writerows(rows)


def _write_json_out(path, headers, rows):
    records = []
    for r in rows:
        r = list(r) + [""] * (len(headers) - len(r))
        records.append(dict(zip(headers, r)))
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(records, fh, indent=2, ensure_ascii=False, default=str)
        fh.write("\n")


def _write_xlsx_out(path, headers, rows):
    opx = _require_openpyxl()
    wb = opx.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append([str(h) for h in headers])
    for r in rows:
        ws.append([c if isinstance(c, (int, float, bool, datetime)) or c is None
                   else str(c) for c in r])
    wb.save(path)


# --- SQL output ---------------------------------------------------------------
_SQL_INT_RE  = re.compile(r"^-?\d+$")
_SQL_DEC_RE  = re.compile(r"^-?\d+\.\d+$")
_SQL_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_SQL_DT_RE   = re.compile(r"^\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}(:\d{2})?$")


def _sql_ident(name):
    """Sanitize a name into a safe SQL identifier (backtick-quoted later)."""
    ident = re.sub(r"[^A-Za-z0-9_]", "_", str(name).strip()) or "col"
    if ident[0].isdigit():
        ident = "_" + ident
    return ident


def _sql_guess_type(values):
    """Guess a MySQL column type from a column's non-empty values:
    INT/BIGINT → DECIMAL → DATE → DATETIME → VARCHAR(n) → TEXT."""
    vals = [v for v in values if v is not None and str(v).strip() != ""]
    if not vals:
        return "VARCHAR(255)"
    strs = [str(v) for v in vals]
    if all(isinstance(v, bool) for v in vals):
        return "TINYINT(1)"
    if all(_SQL_INT_RE.match(s) for s in strs):
        return "BIGINT" if any(abs(int(s)) > 2147483647 for s in strs) else "INT"
    if all(_SQL_INT_RE.match(s) or _SQL_DEC_RE.match(s) for s in strs):
        whole = max(len(s.split(".")[0].lstrip("-")) for s in strs)
        frac  = max((len(s.split(".")[1]) if "." in s else 0) for s in strs)
        return f"DECIMAL({min(whole + frac, 65)},{min(frac, 30)})"
    if all(_SQL_DATE_RE.match(s) for s in strs):
        return "DATE"
    if all(_SQL_DT_RE.match(s) or _SQL_DATE_RE.match(s) for s in strs):
        return "DATETIME"
    longest = max(len(s) for s in strs)
    if longest > 255:
        return "TEXT"
    # round up to the next 25 so similar columns get tidy consistent widths
    return f"VARCHAR({max(25, ((longest + 24) // 25) * 25)})"


def _sql_value(value, col_type):
    """Render one value for an INSERT: NULL for empty, bare numbers for
    numeric columns, otherwise a quoted + escaped string literal."""
    if value is None or str(value).strip() == "":
        return "NULL"
    s = str(value)
    if col_type.startswith(("INT", "BIGINT", "DECIMAL", "TINYINT")):
        if isinstance(value, bool):
            return "1" if value else "0"
        if _SQL_INT_RE.match(s) or _SQL_DEC_RE.match(s):
            return s
    return "'" + s.replace("\\", "\\\\").replace("'", "''") + "'"


def _write_sql_out(path, headers, rows, table):
    """CREATE TABLE (guessed column types) + multi-row INSERTs (500/batch)."""
    table = _sql_ident(table)
    cols = []
    seen = set()
    for h in headers:                       # de-duplicate sanitized names
        c = base = _sql_ident(h)
        n = 2
        while c in seen:
            c = f"{base}_{n}"; n += 1
        seen.add(c)
        cols.append(c)
    types = [_sql_guess_type([r[i] if i < len(r) else "" for r in rows])
             for i in range(len(headers))]

    with open(path, "w", encoding="utf-8") as fh:
        fh.write(f"-- Generated by File Manager v{VERSION} on "
                 f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        fh.write(f"-- Source rows: {len(rows)}\n\n")
        fh.write(f"CREATE TABLE `{table}` (\n")
        fh.write(",\n".join(f"    `{c}` {t}" for c, t in zip(cols, types)))
        fh.write("\n);\n\n")
        col_list = ", ".join(f"`{c}`" for c in cols)
        for start in range(0, len(rows), 500):
            batch = rows[start:start + 500]
            fh.write(f"INSERT INTO `{table}` ({col_list}) VALUES\n")
            lines = []
            for r in batch:
                r = list(r) + [""] * (len(cols) - len(r))
                lines.append("(" + ", ".join(
                    _sql_value(r[i], types[i]) for i in range(len(cols))) + ")")
            fh.write(",\n".join(lines) + ";\n\n")


_CONVERT_WRITERS = {
    "csv":  _write_csv_out,
    "json": _write_json_out,
    "xlsx": _write_xlsx_out,
}


def convert_file(path, out_format):
    """Convert a CSV/JSON/XLSX data file to CSV, JSON, XLSX, or SQL.

    The output file is written next to the source with the same name and the
    new extension — collision-safe (name-2.ext, …), never overwriting. SQL
    output is a CREATE TABLE (types guessed per column) plus multi-row
    INSERTs, with the table named after the file. The results screen is
    appended to ~/Documents/log/fm.log and offers [R] Run Again.
    """
    while True:
        with _ActivityLog():
            _convert_screen(path, out_format)
        if not pause_rerun():
            return


def _convert_screen(path, out_format):
    screen("Convert")
    print()
    out_format = str(out_format).lower().lstrip(".")
    in_ext = os.path.splitext(path)[1].lower().lstrip(".")
    print(f"  {YELLOW}File{RESET}  : {path}")
    print(f"  {YELLOW}From{RESET}  : .{in_ext}    {YELLOW}To{RESET}: .{out_format}\n")
    runlog = _RunLog("Convert", [("File", path), ("From", f".{in_ext}"), ("To", f".{out_format}")])
    if not os.path.isfile(path):
        print(color_text(f"  Not a file: {path}", fg=RED))
        runlog.finish(f"Cancelled — not a file: {path}")
        return
    if out_format not in CONVERT_OUT_FORMATS:
        print(color_text(f"  Unsupported output format .{out_format} — "
                         f"choose one of: {', '.join(CONVERT_OUT_FORMATS)}", fg=RED))
        runlog.finish(f"Cancelled — unsupported output format .{out_format}")
        return

    out_path = _unique_path(os.path.splitext(path)[0] + "." + out_format)
    try:
        headers, rows = _read_table(path)
        # The header row defines the columns: pad short rows, drop extras.
        width = len(headers)
        rows = [(list(r) + [""] * width)[:width] for r in rows]
        if out_format == "sql":
            table = os.path.splitext(os.path.basename(path))[0]
            _write_sql_out(out_path, headers, rows, table)
        else:
            _CONVERT_WRITERS[out_format](out_path, headers, rows)
    except ValueError as e:
        runlog.finish(f"FAILED — {e}")
        report_result(False, "", str(e))
        return
    except OSError as e:
        runlog.finish(f"FAILED — file error: {e}")
        report_result(False, "", f"File error: {e}")
        return

    print(f"  {YELLOW}Output{RESET}: {out_path}")
    print(f"  {YELLOW}Size{RESET}  : {fmt_size(file_size(out_path))}\n")
    runlog.finish(f"Converted {len(rows):,} row(s) x {len(headers)} column(s) "
                  f"to {out_path}")
    report_result(True,
                  f"Converted {len(rows):,} row(s) × {len(headers)} column(s) "
                  f"to .{out_format}.")


def convert_menu():
    """Convert flow: pick the file, pick the output format, convert."""
    path = ask_file("File to convert (.csv, .json, .xlsx)")
    if not path:
        pause_return()
        return
    in_ext = os.path.splitext(path)[1].lower().lstrip(".")
    if in_ext not in CONVERT_IN_FORMATS:
        print(color_text(f"  Unsupported input type .{in_ext} — "
                         "supported: .csv, .json, .xlsx", fg=RED))
        pause_return()
        return
    options = [
        ("CSV",
         "Comma-separated values with a header row — opens anywhere "
         "(Excel, Numbers, imports)."),
        ("JSON",
         "An array of objects, one per row, pretty-printed (2-space indent) — "
         "for APIs, scripts, and config-style data."),
        ("XLSX",
         "An Excel workbook with the data on a single sheet (header row "
         "first). Requires the openpyxl package (pip3 install openpyxl) — "
         "the other formats work without it."),
        ("SQL",
         "A MySQL script: CREATE TABLE named after the file with column types "
         "guessed from the data (INT, DECIMAL, DATE, DATETIME, VARCHAR, "
         "TEXT), then multi-row INSERT statements (500 rows per batch). "
         "Empty cells become NULL."),
    ]
    ch = render_menu("Convert — Output Format", options,
                     outro=[f"File - {path}"])
    if ch == "back":
        return
    convert_file(path, CONVERT_OUT_FORMATS[int(ch) - 1])


# =============================================================================
# CREATE RANDOM UID
# =============================================================================
UUID_MAX = 10000


def generate_uuids(count):
    """Generate `count` random (version 4) UUIDs and show them one per line.
    The list is also appended to ~/Documents/log/fm.log; [R] reruns with the
    same count (new UUIDs each time)."""
    while True:
        with _ActivityLog():
            _uuid_screen(count)
        if not pause_rerun():
            return


def _uuid_screen(count):
    screen("Create Random UID")
    print()
    runlog = _RunLog("Create Random UID", [("Requested Count", count)])
    try:
        count = int(count)
    except (TypeError, ValueError):
        print(color_text("  Invalid number.", fg=RED))
        runlog.finish("Cancelled — invalid number")
        return
    if count < 1:
        print(color_text("  Enter a number of 1 or more.", fg=RED))
        runlog.finish("Cancelled — count must be 1 or more")
        return
    if count > UUID_MAX:
        print(color_text(f"  Capped at {UUID_MAX:,}.", fg=YELLOW))
        count = UUID_MAX
    print(f"  {YELLOW}Count{RESET}: {count:,}\n")
    for _ in range(count):
        print(f"  {uuid.uuid4()}")
    print()
    print(color_text(f"  {count:,} UUID(s) generated.", fg=BRIGHT_CYAN, style=BOLD))
    runlog.finish(f"{count:,} UUID(s) generated")


def create_random_uid_menu():
    raw = ask("How many UUIDs to generate?", "1")
    try:
        count = int(raw)
    except ValueError:
        print(color_text("  Invalid number.", fg=RED))
        pause_return()
        return
    generate_uuids(count)


# =============================================================================
# FIND
# =============================================================================
def _print_find_results(results, kind, extra=None):
    if not results:
        print(color_text(f"  No matching {kind} found.", fg=YELLOW))
        return
    print(color_text(f"  {len(results)} {kind} found:", fg=BRIGHT_CYAN, style=BOLD))
    for path in results:
        if extra and path in extra:
            print(f"    {path}  {DIM}({extra[path]}){RESET}")
        else:
            print(f"    {path}")


def find_folders(root, pattern):
    screen("Find Folders")
    print()
    print(f"  {YELLOW}Root{RESET}: {root}    {YELLOW}Pattern{RESET}: {pattern}\n")
    runlog = _RunLog("Find Folders", [("Root", root), ("Pattern", pattern)])
    if not os.path.isdir(root):
        print(color_text(f"  Not a directory: {root}", fg=RED))
        runlog.finish(f"Cancelled — not a directory: {root}")
        pause_return(); return
    results = []
    for dp, dns, fns in os.walk(root):
        prune_dirs(dns)
        for d in sorted(dns):
            if fnmatch.fnmatch(d, pattern):
                results.append(os.path.join(dp, d))
    _print_find_results(sorted(results), "folders")
    runlog.finish(f"{len(results)} folder(s) found")
    if pause_rerun():
        find_folders(root, pattern)


def find_files_by_name(root, pattern):
    screen("Find Files by Name")
    print()
    print(f"  {YELLOW}Root{RESET}: {root}    {YELLOW}Pattern{RESET}: {pattern}\n")
    runlog = _RunLog("Find Files by Name", [("Root", root), ("Pattern", pattern)])
    if not os.path.isdir(root):
        print(color_text(f"  Not a directory: {root}", fg=RED))
        runlog.finish(f"Cancelled — not a directory: {root}")
        pause_return(); return
    results = []
    extra = {}
    for dp, dns, fns in os.walk(root):
        prune_dirs(dns)
        for fn in sorted(fns):
            if is_excluded_file(fn):
                continue
            if fnmatch.fnmatch(fn, pattern):
                p = os.path.join(dp, fn)
                results.append(p)
                extra[p] = fmt_size(file_size(p))
    _print_find_results(sorted(results), "files", extra)
    runlog.finish(f"{len(results)} file(s) found")
    if pause_rerun():
        find_files_by_name(root, pattern)


def find_files_by_ext(root, ext):
    ext = ext.lstrip(".").lower()
    screen("Find Files by Extension")
    print()
    print(f"  {YELLOW}Root{RESET}: {root}    {YELLOW}Extension{RESET}: .{ext}\n")
    runlog = _RunLog("Find Files by Extension", [("Root", root), ("Extension", f".{ext}")])
    if not os.path.isdir(root):
        print(color_text(f"  Not a directory: {root}", fg=RED))
        runlog.finish(f"Cancelled — not a directory: {root}")
        pause_return(); return
    results = []
    extra = {}
    for dp, dns, fns in os.walk(root):
        prune_dirs(dns)
        for fn in sorted(fns):
            if is_excluded_file(fn):
                continue
            if fn.lower().endswith("." + ext):
                p = os.path.join(dp, fn)
                results.append(p)
                extra[p] = fmt_size(file_size(p))
    _print_find_results(sorted(results), f".{ext} files", extra)
    runlog.finish(f"{len(results)} file(s) found")
    if pause_rerun():
        find_files_by_ext(root, ext)


def find_by_size(root, megabytes, over=True):
    threshold = int(megabytes * 1024 * 1024)
    label = "over" if over else "under"
    screen(f"Find Files {label.capitalize()} {megabytes} MB")
    print()
    print(f"  {YELLOW}Root{RESET}: {root}    {YELLOW}Size {label}{RESET}: {megabytes} MB\n")
    runlog = _RunLog(f"Find Files {label.capitalize()} {megabytes} MB",
                      [("Root", root), (f"Size {label}", f"{megabytes} MB")])
    if not os.path.isdir(root):
        print(color_text(f"  Not a directory: {root}", fg=RED))
        runlog.finish(f"Cancelled — not a directory: {root}")
        pause_return(); return
    results = []
    extra = {}
    for dp, dns, fns in os.walk(root):
        prune_dirs(dns)
        for fn in fns:
            if is_excluded_file(fn):
                continue
            p = os.path.join(dp, fn)
            sz = file_size(p)
            if (over and sz > threshold) or (not over and sz < threshold):
                results.append(p)
                extra[p] = fmt_size(sz)
    results.sort(key=lambda p: file_size(p), reverse=over)
    _print_find_results(results, f"files {label} {megabytes} MB", extra)
    runlog.finish(f"{len(results)} file(s) found")
    if pause_rerun():
        find_by_size(root, megabytes, over)


def find_files_combined(root, pattern=None, ext=None, over_mb=None, under_mb=None):
    """Find files matching ALL supplied criteria (logical AND).

    pattern  : filename wildcard (e.g. "IMG_*.jpg")   — None to skip
    ext      : extension without dot (e.g. "mov")      — None to skip
    over_mb  : only files strictly larger than N MB    — None to skip
    under_mb : only files strictly smaller than N MB   — None to skip

    Example: ext="mov", under_mb=5  ->  .mov files under 5 MB.
    """
    ext_l = ext.lstrip(".").lower() if ext else None
    over_b = int(over_mb * 1024 * 1024) if over_mb is not None else None
    under_b = int(under_mb * 1024 * 1024) if under_mb is not None else None

    # Human-readable summary of the active criteria
    crit = []
    if pattern:
        crit.append(f"name matches '{pattern}'")
    if ext_l:
        crit.append(f"extension .{ext_l}")
    if over_mb is not None:
        crit.append(f"over {over_mb} MB")
    if under_mb is not None:
        crit.append(f"under {under_mb} MB")

    screen("Find Files")
    print()
    print(f"  {YELLOW}Root{RESET}: {root}")
    print(f"  {YELLOW}Criteria{RESET}: {' AND '.join(crit) if crit else '(all files)'}\n")
    runlog = _RunLog("Find Files", [("Root", root),
                                     ("Criteria", ' AND '.join(crit) if crit else '(all files)')])
    if not os.path.isdir(root):
        print(color_text(f"  Not a directory: {root}", fg=RED))
        runlog.finish(f"Cancelled — not a directory: {root}")
        pause_return(); return

    results = []
    extra = {}
    for dp, dns, fns in os.walk(root):
        prune_dirs(dns)
        for fn in fns:
            if is_excluded_file(fn):
                continue
            if pattern and not fnmatch.fnmatch(fn, pattern):
                continue
            if ext_l and not fn.lower().endswith("." + ext_l):
                continue
            p = os.path.join(dp, fn)
            sz = file_size(p)
            if over_b is not None and not sz > over_b:
                continue
            if under_b is not None and not sz < under_b:
                continue
            results.append(p)
            extra[p] = fmt_size(sz)

    # Largest-first when a size criterion is in play; otherwise alphabetical.
    if over_b is not None or under_b is not None:
        results.sort(key=lambda p: file_size(p), reverse=True)
    else:
        results.sort()
    _print_find_results(results, "files", extra)
    runlog.finish(f"{len(results)} file(s) found")
    if pause_rerun():
        find_files_combined(root, pattern, ext, over_mb, under_mb)


def _scan_filenames(folders):
    """Scan each folder recursively (hidden files/folders skipped) and return
    a list parallel to `folders`: one dict per folder of
    filename -> [(size, containing_dir), ...]."""
    per_folder = []
    for folder in folders:
        occ = defaultdict(list)
        for dp, dns, fns in os.walk(folder):
            dns[:] = [d for d in dns if not d.startswith(".") and d.lower() not in EXCLUDED_DIR_NAMES]
            for fn in fns:
                if fn.startswith(".") or is_excluded_file(fn):
                    continue
                occ[fn].append((file_size(os.path.join(dp, fn)), dp))
        per_folder.append(occ)
    return per_folder


def _print_filename_size_table(names, per_folder, folder_col_idx=None):
    """Print the filename/size-per-folder table: one row per name, one size
    column per folder (blank when the name isn't in that folder; multiple
    occurrences within one folder are listed comma-separated).

    folder_col_idx: when set (0-based folder index), append a last 'Folder'
    column showing the directory containing each occurrence in that folder.
    """
    rows = []
    for name in names:
        cells = [", ".join(fmt_size_short(s) for (s, _) in occ[name]) if name in occ else ""
                 for occ in per_folder]
        if folder_col_idx is not None:
            dirs = []
            for (_, dp) in per_folder[folder_col_idx].get(name, ()):
                if dp not in dirs:
                    dirs.append(dp)
            cells.append(", ".join(dirs))
        rows.append((name, cells))

    headers = [str(i + 1) for i in range(len(per_folder))]
    if folder_col_idx is not None:
        headers.append("Folder")

    name_w = max([len(n) for (n, _) in rows] + [len("Filename")])
    name_w = min(name_w, 50)
    col_ws = [max([len(cells[i]) for (_, cells) in rows] + [len(h), 8])
              for i, h in enumerate(headers)]

    last = len(headers) - 1
    if folder_col_idx is not None:
        # Keep the header/separator inside the terminal; the path cells
        # themselves are printed unpadded (no trailing spaces).
        used = 2 + name_w + sum(cw + 2 for cw in col_ws[:-1]) + 2
        col_ws[-1] = max(min(col_ws[-1], get_width() - used), len("Folder"))

    hdr = f"  {'Filename':<{name_w}}"
    sep = f"  {'-' * name_w}"
    for i, (h, cw) in enumerate(zip(headers, col_ws)):
        # The 'Folder' path column is left-aligned; size columns right-aligned.
        hdr += f"  {h}" if (folder_col_idx is not None and i == last) else f"  {h:>{cw}}"
        sep += f"  {'-' * cw}"
    print(color_text(hdr, fg=YELLOW, style=BOLD))
    print(f"{DIM}{sep}{RESET}")

    for name, cells in rows:
        line = f"  {WHITE}{name[:name_w]:<{name_w}}{RESET}"
        for i, (cw, cell) in enumerate(zip(col_ws, cells)):
            if folder_col_idx is not None and i == last:
                line += f"  {DIM}{cell}{RESET}"
            elif cell:
                line += f"  {GREEN}{cell:>{cw}}{RESET}"
            else:
                line += f"  {'':>{cw}}"
        print(line)
    print(f"{DIM}{sep}{RESET}")


def find_duplicates_by_filename(folders):
    """Find files that share the same filename across (or within) the given
    folders — read-only, nothing is changed.

    Each folder is scanned recursively (hidden files/folders skipped). A
    filename counts as a duplicate when it appears 2+ times in total. Results
    are shown as a table: a numbered header line per folder, then one row per
    duplicated filename with a size column per folder. If a name occurs more
    than once within a single folder, each size is listed comma-separated.
    The screen output is also appended to ~/Documents/log/fm.log. After the
    results, [D] Delete Duplicates lets you clear them out — Keep Newest,
    Keep Largest, or Delete from a Specific Folder.
    """
    while True:
        with _ActivityLog():
            dup_groups = _find_duplicates_screen(folders)
        choice = _pause_find_dups(bool(dup_groups))
        if choice == "delete":
            _delete_duplicate_groups(dup_groups)
            continue
        if choice == "rerun":
            continue
        return


def _find_duplicates_screen(folders):
    """Returns dup_groups: a list of duplicate groups, each a list of
    (path, size) tuples (2+ items) — used by the [D] Delete Duplicates flow."""
    screen("Find Duplicates by Filename")
    print()

    runlog = _RunLog("Find Duplicates by Filename", [("Folders", ', '.join(folders))])
    folders = [clean_path(f) for f in folders if f]
    valid = [f for f in folders if os.path.isdir(f)]
    for f in folders:
        if f not in valid:
            print(color_text(f"  Skipping (not a directory): {f}", fg=YELLOW))
    if not valid:
        print(color_text("  No valid folders to scan.", fg=RED))
        runlog.finish("Cancelled — no valid folders to scan")
        return []

    # Numbered folder header
    for i, folder in enumerate(valid, 1):
        print(f"  {YELLOW}{i}{RESET} - {folder}")
    print()

    per_folder = _scan_filenames(valid)
    all_names = set()
    for occ in per_folder:
        all_names.update(occ)
    dups = sorted(
        (name for name in all_names
         if sum(len(occ.get(name, ())) for occ in per_folder) > 1),
        key=str.lower)

    if not dups:
        print(color_text("  No duplicate filenames found.", fg=YELLOW))
        runlog.finish("No duplicate filenames found")
        return []

    _print_filename_size_table(dups, per_folder)
    print(color_text(f"  {len(dups)} duplicated filename(s) found.", fg=BRIGHT_CYAN, style=BOLD))
    runlog.finish(f"{len(dups)} duplicated filename(s) found")

    return [[(os.path.join(dp, name), size)
              for occ in per_folder for (size, dp) in occ.get(name, ())]
             for name in dups]


FUZZY_NAME_RATIO     = 0.85   # difflib similarity threshold for "close" names
FUZZY_SIZE_TOLERANCE = 0.01   # sizes within 1% of each other count as "close"

# Duplicate-style name endings stripped before stems are compared:
# "video1", "video (1)", "video[2]", "video copy", "video copy 2", "video - Copy"
_FUZZY_SUFFIX_RE = re.compile(
    r"[ _\-.]*(?:\(\d+\)|\[\d+\]|copy(?:[ _\-]*\d+)?|\d+)$", re.IGNORECASE)


def _fuzzy_stem(stem):
    """Comparison key for a filename stem: lowercased, duplicate-style endings
    stripped, runs of separators collapsed to a single space."""
    s = _FUZZY_SUFFIX_RE.sub("", stem.lower())
    return re.sub(r"[ _\-.]+", " ", s).strip()


def _sizes_close(a, b):
    return abs(a - b) <= max(a, b) * FUZZY_SIZE_TOLERANCE


def _names_close(stem_a, stem_b):
    ka, kb = _fuzzy_stem(stem_a), _fuzzy_stem(stem_b)
    if ka and ka == kb:
        return True
    return difflib.SequenceMatcher(
        None, stem_a.lower(), stem_b.lower()).ratio() >= FUZZY_NAME_RATIO


def find_duplicates_by_fuzzy_name(folders):
    """Find files whose names are CLOSE (not necessarily identical) and whose
    sizes are close — read-only, nothing is changed.

    Each folder is scanned recursively (hidden files/folders skipped). Two
    files count as fuzzy duplicates when their extension matches, their sizes
    are within FUZZY_SIZE_TOLERANCE of each other, and their name stems either
    match after duplicate-style endings are stripped (videofile1 ~ videofile,
    photo (2) ~ photo, doc copy ~ doc) or are FUZZY_NAME_RATIO similar.
    Matches are clustered into groups; in each group the shortest/cleanest
    name is marked KEEP and the rest DELETE — candidates only, nothing is
    deleted. The screen output is also appended to ~/Documents/log/fm.log.
    After the results, [D] Delete Duplicates lets you clear them out — Keep
    Newest, Keep Largest, or Delete from a Specific Folder (this replaces the
    shortest/cleanest-name KEEP shown above with the chosen rule).
    """
    while True:
        with _ActivityLog():
            dup_groups = _find_fuzzy_dups_screen(folders)
        choice = _pause_find_dups(bool(dup_groups))
        if choice == "delete":
            path_groups = [[(os.path.join(dp, name), size)
                             for (name, _stem, _ext, size, dp) in group]
                            for group in dup_groups]
            _delete_duplicate_groups(path_groups)
            continue
        if choice == "rerun":
            continue
        return


def _fuzzy_dup_groups(folders):
    """Cluster every file under the given folders into fuzzy-duplicate groups:
    close name (_names_close) + same extension + size within
    FUZZY_SIZE_TOLERANCE. Returns a list of groups of (name, stem, ext, size,
    dir) tuples — each group sorted with the KEEP candidate (shortest/cleanest
    name) first, groups sorted largest file first. Groups have 2+ members."""
    # Flatten every occurrence across all folders: (name, stem, ext, size, dir)
    files = []
    for occ in _scan_filenames(folders):
        for name, entries in occ.items():
            stem, ext = os.path.splitext(name)
            for (size, dp) in entries:
                files.append((name, stem, ext.lower(), size, dp))

    # Sort by extension then size so only a nearby window needs pairwise name
    # comparison — once the size gap exceeds the tolerance it only grows, so
    # the inner scan can stop there.
    files.sort(key=lambda f: (f[2], f[3]))

    parent = list(range(len(files)))

    def _root(i):
        while parent[i] != i:
            parent[i] = parent[parent[i]]
            i = parent[i]
        return i

    for i in range(len(files)):
        for j in range(i + 1, len(files)):
            if files[j][2] != files[i][2] or not _sizes_close(files[i][3], files[j][3]):
                break
            if _names_close(files[i][1], files[j][1]):
                parent[_root(j)] = _root(i)

    groups = defaultdict(list)
    for i in range(len(files)):
        groups[_root(i)].append(files[i])
    dup_groups = [g for g in groups.values() if len(g) > 1]

    # KEEP the shortest (cleanest) name; largest files first overall.
    for g in dup_groups:
        g.sort(key=lambda f: (len(f[0]), f[0].lower(), f[4]))
    dup_groups.sort(key=lambda g: -max(f[3] for f in g))
    return dup_groups


def _find_fuzzy_dups_screen(folders):
    """Returns dup_groups: a list of groups of (name, stem, ext, size, dir)
    tuples (2+ items per group) — used by the [D] Delete Duplicates flow."""
    screen("Find Duplicates by Fuzzy Name")
    print()

    runlog = _RunLog("Find Duplicates by Fuzzy Name", [("Folders", ', '.join(folders))])
    folders = [clean_path(f) for f in folders if f]
    valid = [f for f in folders if os.path.isdir(f)]
    for f in folders:
        if f not in valid:
            print(color_text(f"  Skipping (not a directory): {f}", fg=YELLOW))
    if not valid:
        print(color_text("  No valid folders to scan.", fg=RED))
        runlog.finish("Cancelled — no valid folders to scan")
        return []

    # Numbered folder header
    for i, folder in enumerate(valid, 1):
        print(f"  {YELLOW}{i}{RESET} - {folder}")
    print()

    dup_groups = _fuzzy_dup_groups(valid)
    if not dup_groups:
        print(color_text("  No fuzzy-duplicate files found.", fg=YELLOW))
        runlog.finish("No fuzzy-duplicate files found")
        return []

    name_w = min(max(len(f[0]) for g in dup_groups for f in g), 50)
    del_count = del_bytes = 0
    for n, group in enumerate(dup_groups, 1):
        print(color_text(f"  {n}) {group[0][0]}", style=BOLD))
        for k, (name, _stem, _ext, size, dp) in enumerate(group):
            tag = (color_text("KEEP  ", fg=BRIGHT_CYAN, style=BOLD) if k == 0
                   else color_text("DELETE", fg=BRIGHT_YELLOW, style=BOLD))
            print(f"     {tag}  {name:<{name_w}}  {fmt_size_short(size):>8}  {dp}")
            if k:
                del_count += 1
                del_bytes += size
        print()

    print(color_text(
        f"  {len(dup_groups)} fuzzy-duplicate group(s) — {del_count} DELETE "
        f"candidate(s), {fmt_size(del_bytes)} reclaimable. Nothing was deleted.",
        fg=BRIGHT_CYAN, style=BOLD))
    runlog.finish(f"{len(dup_groups)} fuzzy-duplicate group(s), "
                  f"{del_count} DELETE candidate(s), {fmt_size(del_bytes)} reclaimable")

    return dup_groups


# -----------------------------------------------------------------------------
# Delete Duplicates (Find menu) — Keep Newest / Keep Largest / Delete from a
# Specific Folder. Shared tail for Find Duplicates by Filename and Find
# Duplicates by Fuzzy Name, invoked via the [D] Delete Duplicates option on
# the results screen.
# -----------------------------------------------------------------------------
def _path_under_folder(path, folder):
    """True if `path` is inside `folder` or any of its subfolders."""
    path = os.path.abspath(path)
    folder = os.path.abspath(folder)
    return path == folder or path.startswith(folder + os.sep)


def _resolve_dup_group(group, strategy, target_folder=None):
    """Decide which files to keep/delete in one duplicate group under the
    chosen strategy. group: list of (path, size) tuples (2+ items). Returns
    (keep, delete) — both lists of (path, size); a group's keep list is
    never empty, so a duplicate is never wiped out entirely.

    newest  -> keep the most recently modified file (mtime)
    largest -> keep the largest file (by size)
    folder  -> delete every copy located under target_folder (recursively,
               including subfolders); if every copy in the group is under
               that folder, keep the first one instead and delete the rest
    """
    if strategy == "newest":
        def mtime(ps):
            try:
                return os.path.getmtime(ps[0])
            except OSError:
                return -1
        keep = max(group, key=mtime)
        return [keep], [ps for ps in group if ps is not keep]
    if strategy == "largest":
        keep = max(group, key=lambda ps: ps[1])
        return [keep], [ps for ps in group if ps is not keep]
    # folder
    inside = [ps for ps in group if _path_under_folder(ps[0], target_folder)]
    outside = [ps for ps in group if ps not in inside]
    if outside:
        return outside, inside
    return inside[:1], inside[1:]


def _pick_dup_delete_strategy():
    """Sub-menu for choosing how to resolve duplicate groups before deleting.
    Returns 'newest', 'largest', 'folder', or None if cancelled."""
    options = [
        ("Keep Newest",
         "In each duplicate group, keep the file with the most recent "
         "modified date (mtime) and mark the rest for deletion."),
        ("Keep Largest",
         "In each duplicate group, keep the largest file by size and mark "
         "the rest for deletion."),
        ("Delete from Specific Folder",
         "Enter a folder — any copy located under that folder (including "
         "its subfolders) is marked for deletion; copies elsewhere in the "
         "group are kept. If every copy in a group happens to be under that "
         "folder, one copy is kept anyway so the file is never deleted "
         "entirely."),
    ]
    ch = render_menu("Delete Duplicates — Keep Rule", options)
    return {"1": "newest", "2": "largest", "3": "folder"}.get(ch) if ch != "back" else None


def _delete_duplicate_groups(dup_groups, live_requested=None, assume_yes=False):
    """Shared tail for the Find Duplicates [D] Delete Duplicates option: pick
    a keep rule (Newest / Largest / Specific Folder), preview the resulting
    KEEP/DELETE split per group, then either delete (typed YES) or move the
    marked copies to a folder instead — both explicit opt-in. dup_groups:
    list of groups, each a list of (path, size) tuples (2+ items per group)."""
    if not dup_groups:
        return
    strategy = _pick_dup_delete_strategy()
    if strategy is None:
        return

    target_folder = None
    if strategy == "folder":
        target_folder = ask_folder("Delete copies under this folder", key="remove_duplicates_specific_folder")
        if not target_folder:
            pause_return(); return

    screen("Delete Duplicates")
    print()
    label = {"newest": "Keep Newest", "largest": "Keep Largest",
             "folder": f"Delete from: {target_folder}"}[strategy]
    print(f"  {YELLOW}Rule{RESET}: {label}\n")

    items = []
    keep_map = {}
    any_rows = False
    for group in dup_groups:
        keep, delete = _resolve_dup_group(group, strategy, target_folder)
        if not delete:
            continue
        any_rows = True
        for (p, size) in keep:
            print(f"    {BRIGHT_CYAN}keep{RESET}   {p}  {DIM}({fmt_size(size)}){RESET}")
        for (p, size) in delete:
            print(f"    {BRIGHT_YELLOW}remove{RESET} {p}  {DIM}({fmt_size(size)}){RESET}")
            items.append((p, size, False))
            keep_map[p] = keep[0][0]
        print()

    if not any_rows:
        print(color_text("  Nothing to delete under this rule.", fg=YELLOW))
        pause_return(); return

    _finish_removal(items, live_requested, assume_yes=assume_yes, require_yes=True,
                     activity="Delete Duplicates", fields=[("Rule", label)], allow_move=True,
                     keep_map=keep_map)
    pause_return()


def find_missing_by_filename(folder_a, folder_b, mode="either", match_size=False):
    """Compare two folders by filename and list files present in only ONE of
    them (missing from the other) — read-only, nothing is changed.

    mode: first  -> filenames in folder 1 but not in folder 2
          second -> filenames in folder 2 but not in folder 1
          either -> filenames in either folder but not in both

    match_size: when True, files match only when BOTH the filename AND the
    size agree — so a same-named file whose size differs between the folders
    is also reported (its size shows in both columns, revealing the mismatch).

    Both folders are scanned recursively (hidden files/folders skipped).
    Results use the same table as Find Duplicates by Filename: a numbered
    header per folder, one row per filename, and a size column per folder —
    the blank column shows which folder the file is missing from. In the
    first/second modes a last 'Folder' column shows the directory containing
    each file. The folder list is repeated after the results, and the screen
    output is also appended to ~/Documents/log/fm.log.
    """
    while True:
        with _ActivityLog():
            _find_missing_screen(folder_a, folder_b, mode, match_size)
        if not pause_rerun():
            return


def _find_missing_screen(folder_a, folder_b, mode, match_size=False):
    screen("Find Missing by Filename & Size" if match_size else "Find Missing by Filename")
    print()

    folders = [clean_path(folder_a), clean_path(folder_b)]
    mode_labels = {
        "first":  "In 1st folder only",
        "second": "In 2nd folder only",
        "either": "In either folder (only once)",
    }
    runlog = _RunLog("Find Missing by Filename", [
        ("Folder A", folders[0]), ("Folder B", folders[1]),
        ("Show", mode_labels.get(mode, mode)), ("Match Size", "Yes" if match_size else "No"),
    ])
    ok = True
    for f in folders:
        if not os.path.isdir(f):
            print(color_text(f"  Not a directory: {f}", fg=RED))
            ok = False
    if not ok:
        runlog.finish("Cancelled — not a directory")
        return

    print(f"  {YELLOW}Show{RESET}: {mode_labels.get(mode, mode)}\n")

    def folder_list():
        for i, folder in enumerate(folders, 1):
            print(f"  {YELLOW}{i}{RESET} - {folder}")

    folder_list()
    print()

    per_folder = _scan_filenames(folders)
    if match_size:
        # Match on (filename, size): a same-named file whose size differs on
        # the other side is treated as missing there.
        set_a = {(n, s) for n, occ in per_folder[0].items() for (s, _) in occ}
        set_b = {(n, s) for n, occ in per_folder[1].items() for (s, _) in occ}
    else:
        set_a, set_b = set(per_folder[0]), set(per_folder[1])
    if mode == "first":
        found = set_a - set_b
    elif mode == "second":
        found = set_b - set_a
    else:
        found = set_a ^ set_b
    names = {n for (n, _) in found} if match_size else found
    missing = sorted(names, key=str.lower)

    if not missing:
        print(color_text("  No missing files found for the selected option.", fg=YELLOW))
        runlog.finish("No missing files found")
        return

    # In the single-folder modes every file lives in one known folder, so a
    # last column shows the directory actually containing each file.
    folder_col_idx = {"first": 0, "second": 1}.get(mode)
    _print_filename_size_table(missing, per_folder, folder_col_idx)
    what = "missing or differing in size" if match_size else "found in only one folder"
    print(color_text(f"  {len(missing)} filename(s) {what}.", fg=BRIGHT_CYAN, style=BOLD))
    print()
    folder_list()
    runlog.finish(f"{len(missing)} filename(s) {what}")


def find_and_replace(root, search, replace, ext=None, live_requested=None,
                     assume_yes=False, backup=False):
    """Find a text string in files under a folder and replace it.

    ALWAYS a dry run first: every match is listed (file, line number, line
    with the match highlighted) before anything is touched. Matching is
    literal and case-insensitive; the replacement is inserted exactly as
    typed. Hidden files/folders, excluded names, and binary files are never
    scanned or modified. `ext` limits the scan to one file extension.

    live_requested: None  -> interactive (confirm Y/N, then ask about .bak)
                    True  -> CLI --apply (confirm unless assume_yes)
                    False -> CLI dry run (report only)
    The screen output is also appended to ~/Documents/log/fm.log.
    """
    while True:
        with _ActivityLog():
            _find_replace_screen(root, search, replace, ext,
                                 live_requested, assume_yes, backup)
        if not pause_rerun():
            return


def _find_replace_screen(root, search, replace, ext, live_requested,
                         assume_yes, backup):
    screen("Find & Replace")
    print()
    print_mode_line(True)   # always dry-run-first by design; no upfront choice
    ext_l = ext.lstrip(".").lower() if ext else None
    print(f"  {YELLOW}Folder{RESET} : {root}")
    print(f"  {YELLOW}Find{RESET}   : {search}  {DIM}(case-insensitive){RESET}")
    print(f"  {YELLOW}Replace{RESET}: "
          f"{replace if replace else DIM + '(remove the text)' + RESET}")
    if ext_l:
        print(f"  {YELLOW}Ext{RESET}    : .{ext_l}")
    print()
    runlog = _RunLog("Find & Replace", [
        ("Folder", root), ("Find", search), ("Replace", replace),
    ] + ([("Ext", f".{ext_l}")] if ext_l else []))
    if not os.path.isdir(root):
        print(color_text(f"  Not a directory: {root}", fg=RED))
        runlog.finish(f"Cancelled — not a directory: {root}")
        return
    if not search:
        print(color_text("  Search text required.", fg=RED))
        runlog.finish("Cancelled — search text required")
        return

    pattern = re.compile(re.escape(search), re.IGNORECASE)

    plans = []                    # (path, text, match_count, [(line_no, line)])
    scanned = 0
    for dp, dns, fns in os.walk(root):
        dns[:] = [d for d in dns if not d.startswith(".")
                  and d.lower() not in EXCLUDED_DIR_NAMES]
        for fn in sorted(fns):
            if fn.startswith(".") or is_excluded_file(fn):
                continue
            # Never touch .bak backups (this feature creates them) unless
            # .bak files were explicitly requested via the extension filter.
            if ext_l != "bak" and fn.lower().endswith(".bak"):
                continue
            if ext_l and not fn.lower().endswith("." + ext_l):
                continue
            path = os.path.join(dp, fn)
            try:
                with open(path, "rb") as fh:
                    if b"\x00" in fh.read(8192):
                        continue              # binary — never touched
                # surrogateescape + newline="" round-trips any byte sequence
                # and line-ending style unchanged apart from the replacement.
                with open(path, "r", encoding="utf-8",
                          errors="surrogateescape", newline="") as fh:
                    text = fh.read()
            except OSError:
                continue
            scanned += 1
            if not pattern.search(text):
                continue
            hits = [(no, line) for no, line in enumerate(text.splitlines(), 1)
                    if pattern.search(line)]
            plans.append((path, text, len(pattern.findall(text)), hits))

    if not plans:
        print(color_text(f"  No matches found ({scanned:,} file(s) scanned).",
                         fg=YELLOW))
        runlog.finish(f"No matches found ({scanned:,} file(s) scanned)")
        return

    # ---- Dry-run report: file, then line number + line per match ----------
    width = get_width()
    total = line_total = 0
    for path, _text, count, hits in plans:
        total += count
        line_total += len(hits)
        print(color_text(
            f"  {path}  ({count} match{'es' if count != 1 else ''})",
            fg=BRIGHT_CYAN, style=BOLD))
        for no, line in hits:
            plain = line.rstrip()
            prefix = f"    {no:>6}: "
            avail = max(20, width - len(prefix) - 1)
            if len(plain) > avail:
                plain = plain[:avail - 1] + "…"
            shown = pattern.sub(
                lambda m: f"{BRIGHT_RED}{BOLD}{m.group(0)}{RESET}", plain)
            print(f"{DIM}{prefix}{RESET}{shown}")
        print()

    print(color_text(
        f"  DRY RUN — {total:,} match(es) on {line_total:,} line(s) in "
        f"{len(plans)} file(s). Nothing has been changed.",
        fg=YELLOW, style=BOLD))

    do_replace = False
    make_bak = backup
    if live_requested is None:                # interactive
        do_replace = safe_confirm(
            f"  Replace all {total:,} occurrence(s) in {len(plans)} file(s)?",
            default=False)
        if do_replace:
            make_bak = safe_confirm(
                "  Create a .bak backup of each file before replacing?",
                default=True)
    elif live_requested is True:              # CLI --apply
        do_replace = True if assume_yes else safe_confirm(
            f"  Replace all {total:,} occurrence(s) in {len(plans)} file(s)?",
            default=False)
    else:                                     # CLI dry run
        print(color_text("  Re-run with --apply to perform the replacement.",
                         fg=YELLOW))
        runlog.finish("DRY RUN — re-run with --apply to perform the replacement")
        return
    if not do_replace:
        print(color_text("  Cancelled — nothing changed.", fg=YELLOW))
        runlog.finish("Cancelled — nothing changed")
        return
    if not _run_countdown(label="Starting replace"):
        print(color_text("  Cancelled — nothing changed.", fg=YELLOW))
        runlog.finish("Cancelled — nothing changed")
        return

    runlog.total_items = len(plans)
    runlog.detailed = len(plans) <= ACTION_LOG_DETAIL_THRESHOLD
    ok = fail = 0
    stopped_early = False
    ps = _PauseStop()
    if ps.is_tty and len(plans) > 1:
        print(color_text("  Press [P] to pause, [Q] to stop early.", fg=DIM))
    try:
        for path, text, _count, _hits in plans:
            if ps.check() == "stop":
                stopped_early = True
                break
            try:
                if make_bak:
                    shutil.copy2(path, path + ".bak")
                with open(path, "w", encoding="utf-8",
                          errors="surrogateescape", newline="") as fh:
                    fh.write(pattern.sub(lambda m: replace, text))
                ok += 1
                runlog.action(f"Replaced {_count} occurrence(s) in {path}")
            except OSError as e:
                fail += 1
                print(color_text(f"  ✗ {path}: {e}", fg=RED))
                runlog.action(f"FAILED to replace in {path}: {e}")
    finally:
        ps.close()
    print()
    summary = (f"Replaced {total:,} occurrence(s) in {ok} file(s)."
               + (" Backups saved with a .bak extension." if make_bak else "")
               + (f" {fail} failed." if fail else ""))
    if stopped_early:
        summary += f" Stopped early by user ({ok + fail}/{len(plans)} processed)."
    runlog.finish(summary)
    report_result(fail == 0, summary, f"Replaced in {ok} file(s), {fail} failed.")


def find_and_rename(root, mode, text, replace_with=None, ext=None,
                    live_requested=None, assume_yes=False):
    """Find files under a folder and rename them.

    mode: 'prepend' — new name = text + filename (files already starting
                      with the text are skipped)
          'append'  — text is inserted at the end of the name, BEFORE the
                      extension (files already ending with it are skipped)
          'replace' — occurrences of `text` inside the filename are replaced
                      with `replace_with` (matching is literal and
                      case-insensitive; only matching files are listed)

    ALWAYS a dry run first: every planned rename is listed old → new before
    anything is touched. Renames that would collide with an existing file
    (or with another rename in the same run) are skipped and reported.
    `ext` limits the scan to one file extension.

    live_requested: None  -> interactive (confirm Y/N after the preview)
                    True  -> CLI --apply (confirm unless assume_yes)
                    False -> CLI dry run (report only)
    The screen output is also appended to ~/Documents/log/fm.log.
    """
    while True:
        with _ActivityLog():
            _find_rename_screen(root, mode, text, replace_with, ext,
                                live_requested, assume_yes)
        if not pause_rerun():
            return


def _find_rename_screen(root, mode, text, replace_with, ext,
                        live_requested, assume_yes):
    screen("Find & Rename")
    print()
    print_mode_line(True)   # always dry-run-first by design; no upfront choice
    ext_l = ext.lstrip(".").lower() if ext else None
    mode_disp = {"prepend": "Prepend (text + filename)",
                 "append":  "Append (before the extension)",
                 "replace": "Replace (inside the filename)"}[mode]
    print(f"  {YELLOW}Folder{RESET} : {root}")
    print(f"  {YELLOW}Mode{RESET}   : {mode_disp}")
    if mode == "replace":
        print(f"  {YELLOW}Find{RESET}   : {text}  {DIM}(case-insensitive){RESET}")
        print(f"  {YELLOW}Replace{RESET}: "
              f"{replace_with if replace_with else DIM + '(remove the text)' + RESET}")
    else:
        print(f"  {YELLOW}Text{RESET}   : {text}")
    if ext_l:
        print(f"  {YELLOW}Ext{RESET}    : .{ext_l}")
    print()
    runlog = _RunLog("Find & Rename", [
        ("Folder", root), ("Mode", mode_disp),
    ] + ([("Ext", f".{ext_l}")] if ext_l else []))
    if not os.path.isdir(root):
        print(color_text(f"  Not a directory: {root}", fg=RED))
        runlog.finish(f"Cancelled — not a directory: {root}")
        return
    if not text:
        print(color_text("  Text required.", fg=RED))
        runlog.finish("Cancelled — text required")
        return

    pattern = (re.compile(re.escape(text), re.IGNORECASE)
               if mode == "replace" else None)

    plans = []                    # (dirpath, old_name, new_name)
    collisions = []               # (dirpath, old_name, new_name)
    taken = set()                 # target paths already claimed this run
    scanned = already = 0
    for dp, dns, fns in os.walk(root):
        dns[:] = [d for d in dns if not d.startswith(".")
                  and d.lower() not in EXCLUDED_DIR_NAMES]
        for fn in sorted(fns):
            if fn.startswith(".") or is_excluded_file(fn):
                continue
            if ext_l and not fn.lower().endswith("." + ext_l):
                continue
            scanned += 1
            stem, dot_ext = os.path.splitext(fn)
            if mode == "prepend":
                if fn.lower().startswith(text.lower()):
                    already += 1              # already has the prefix
                    continue
                new_name = text + fn
            elif mode == "append":
                if stem.lower().endswith(text.lower()):
                    already += 1              # already has the suffix
                    continue
                new_name = stem + text + dot_ext
            else:                             # replace
                if not pattern.search(fn):
                    continue                  # filename doesn't contain it
                new_name = pattern.sub(lambda m: replace_with, fn)
            # A rename must produce a real, different, visible filename.
            if new_name == fn or not new_name or new_name.startswith("."):
                already += 1
                continue
            target = os.path.join(dp, new_name)
            if os.path.exists(target) or target in taken:
                collisions.append((dp, fn, new_name))
                continue
            taken.add(target)
            plans.append((dp, fn, new_name))

    if not plans and not collisions:
        print(color_text(
            f"  Nothing to rename ({scanned:,} file(s) scanned"
            + (f", {already:,} already named that way" if already else "")
            + ").", fg=YELLOW))
        runlog.finish(f"Nothing to rename ({scanned:,} file(s) scanned)")
        return

    # ---- Dry-run report: old → new, path shown relative to the root -------
    for dp, fn, new_name in plans:
        rel = os.path.relpath(os.path.join(dp, fn), root)
        print(f"  {WHITE}{rel}{RESET}  {DIM}->{RESET}  "
              f"{BRIGHT_CYAN}{new_name}{RESET}")
    for dp, fn, new_name in collisions:
        rel = os.path.relpath(os.path.join(dp, fn), root)
        print(f"  {YELLOW}SKIP{RESET} {WHITE}{rel}{RESET}  {DIM}->{RESET}  "
              f"{new_name}  {YELLOW}(target already exists){RESET}")
    print()

    summary = f"  DRY RUN — {len(plans):,} file(s) would be renamed"
    extras = []
    if already:
        extras.append(f"{already:,} already named that way")
    if collisions:
        extras.append(f"{len(collisions):,} skipped — target exists")
    if extras:
        summary += f" ({', '.join(extras)})"
    summary += ". Nothing has been changed."
    print(color_text(summary, fg=YELLOW, style=BOLD))

    if not plans:
        runlog.finish("Nothing to rename (only skipped collisions)" if collisions
                       else "Nothing to rename")
        return

    do_rename = False
    if live_requested is None:                # interactive
        do_rename = safe_confirm(
            f"  Rename these {len(plans):,} file(s)?", default=False)
    elif live_requested is True:              # CLI --apply
        do_rename = True if assume_yes else safe_confirm(
            f"  Rename these {len(plans):,} file(s)?", default=False)
    else:                                     # CLI dry run
        print(color_text("  Re-run with --apply to perform the rename.",
                         fg=YELLOW))
        runlog.finish("DRY RUN — re-run with --apply to perform the rename")
        return
    if not do_rename:
        print(color_text("  Cancelled — nothing renamed.", fg=YELLOW))
        runlog.finish("Cancelled — nothing renamed")
        return
    if not _run_countdown(label="Starting rename"):
        print(color_text("  Cancelled — nothing renamed.", fg=YELLOW))
        runlog.finish("Cancelled — nothing renamed")
        return

    runlog.total_items = len(plans)
    runlog.detailed = len(plans) <= ACTION_LOG_DETAIL_THRESHOLD
    ok = fail = 0
    stopped_early = False
    ps = _PauseStop()
    if ps.is_tty and len(plans) > 1:
        print(color_text("  Press [P] to pause, [Q] to stop early.", fg=DIM))
    try:
        for dp, fn, new_name in plans:
            if ps.check() == "stop":
                stopped_early = True
                break
            try:
                os.rename(os.path.join(dp, fn), os.path.join(dp, new_name))
                ok += 1
                runlog.action(f"Renamed {os.path.join(dp, fn)} -> {new_name}")
            except OSError as e:
                fail += 1
                print(color_text(f"  ✗ {os.path.join(dp, fn)}: {e}", fg=RED))
                runlog.action(f"FAILED to rename {os.path.join(dp, fn)}: {e}")
    finally:
        ps.close()
    print()
    summary = f"Renamed {ok} file(s)." + (f" {fail} failed." if fail else "")
    if stopped_early:
        summary += f" Stopped early by user ({ok + fail}/{len(plans)} processed)."
    runlog.finish(summary)
    report_result(fail == 0, summary, f"Renamed {ok} file(s), {fail} failed.")


# =============================================================================
# REMOVE  (dry-run by default)
# =============================================================================
def _preview_removal(items, keep_map=None):
    """items: list of (path, size, is_dir). Prints a numbered preview, each
    item's line including size + created/modified date. keep_map (optional):
    {path: keep_path} — for duplicate-removal flows, shows the surviving
    copy (folder/name + its own size/created/modified) beneath the same
    number. Returns total size in bytes."""
    total = 0
    print(color_text(f"  {len(items)} item(s) matched:", fg=BRIGHT_CYAN, style=BOLD))
    for i, (path, size, is_dir) in enumerate(items, 1):
        total += size
        tag = f"{YELLOW}[DIR]{RESET} " if is_dir else ""
        created, modified = _file_dates(path)
        print(f"    {i:>3}. {tag}{path}  "
              f"{DIM}({fmt_size(size)} / Created: {created} / Modified: {modified}){RESET}")
        keepPath = keep_map.get(path) if keep_map else None
        if keepPath:
            keepCreated, keepModified = _file_dates(keepPath)
            keepFolder = os.path.basename(os.path.dirname(keepPath))
            keepName = os.path.basename(keepPath)
            print(f"         {GREEN}Surviving copy:{RESET} {keepFolder}/{keepName}")
            print(f"         {DIM}({fmt_size(file_size(keepPath))} / Created: {keepCreated} / "
                  f"Modified: {keepModified}){RESET}")
    if not items:
        print(f"    {DIM}(nothing to remove){RESET}")
    else:
        print(f"  {DIM}Total: {fmt_size(total)}{RESET}")
    return total


def _perform_removal(items, runlog=None):
    """Actually delete. Returns (ok_count, fail_count, stopped_early).
    Logs each removal (success or failure) via runlog.action() when a
    _RunLog is given. [P] pauses / [Q] stops early (interactive tty only)."""
    ok = 0
    fail = 0
    stopped_early = False
    ps = _PauseStop()
    if ps.is_tty and len(items) > 1:
        print(color_text("  Press [P] to pause, [Q] to stop early.", fg=DIM))
    try:
        for path, size, is_dir in items:
            if ps.check() == "stop":
                stopped_early = True
                break
            try:
                if is_dir:
                    shutil.rmtree(path)
                else:
                    os.remove(path)
                ok += 1
                if runlog:
                    runlog.action(f"Removed {path} ({fmt_size(size)})")
            except OSError as e:
                print(color_text(f"    Failed: {path} ({e})", fg=RED))
                fail += 1
                if runlog:
                    runlog.action(f"FAILED to remove {path}: {e}")
    finally:
        ps.close()
    return ok, fail, stopped_early


def _unique_dest_path(dest_folder, filename):
    """Return a destination path for `filename` inside dest_folder, appending
    a " (1)", " (2)", ... suffix before the extension if a file with that
    name is already there — so a move never overwrites an existing file."""
    candidate = os.path.join(dest_folder, filename)
    if not os.path.exists(candidate):
        return candidate
    base, ext = os.path.splitext(filename)
    n = 1
    while True:
        candidate = os.path.join(dest_folder, f"{base} ({n}){ext}")
        if not os.path.exists(candidate):
            return candidate
        n += 1


def _perform_move(items, dest_folder, runlog=None):
    """Move items into dest_folder instead of deleting them (auto-renaming on
    a name collision so nothing already there is ever overwritten). Returns
    (ok_count, fail_count, stopped_early); mirrors _perform_removal's
    pause/stop/runlog behavior."""
    ok = 0
    fail = 0
    stopped_early = False

    try:
        os.makedirs(dest_folder, exist_ok=True)
    except OSError as e:
        print(color_text(f"  Could not create destination folder: {e}", fg=RED))
        return 0, len(items), False

    ps = _PauseStop()
    if ps.is_tty and len(items) > 1:
        print(color_text("  Press [P] to pause, [Q] to stop early.", fg=DIM))
    try:
        for path, size, is_dir in items:
            if ps.check() == "stop":
                stopped_early = True
                break
            try:
                dest = _unique_dest_path(dest_folder, os.path.basename(path))
                shutil.move(path, dest)
                ok += 1
                if runlog:
                    runlog.action(f"Moved {path} -> {dest} ({fmt_size(size)})")
            except OSError as e:
                print(color_text(f"    Failed: {path} ({e})", fg=RED))
                fail += 1
                if runlog:
                    runlog.action(f"FAILED to move {path}: {e}")
    finally:
        ps.close()
    return ok, fail, stopped_early


def confirm_yes_word(prompt):
    """Require the user to type the word YES (case-insensitive) to proceed.
    ANY other input — including a bare 'y' — is treated as no. ESC raises
    EscCancelled, same as every other question prompt."""
    resp = _read_line_esc(color_text(prompt, fg=YELLOW), "")
    return resp.upper() == "YES"


def _finish_removal(items, live_requested, assume_yes=False, require_yes=False,
                     activity="Remove", fields=None, allow_move=False, move_to=None,
                     dry_mode=True, keep_map=None):
    """Common tail: preview, then delete (or move) only on explicit opt-in.

    live_requested:
        None  -> interactive mode: ask after preview whether to delete/move
        True  -> CLI --delete or --move-to: act after confirm (or assume_yes)
        False -> CLI default: dry run only
    require_yes:
        True  -> deleting requires typing the word YES (any other input
                 cancels). Used for the higher-stakes Duplicates removals.
                 Moving is reversible, so it only ever needs a plain Y/N.
    allow_move:
        True  -> offers "move to a folder" as an alternative to deleting
                 (interactive: a [D]/[M]/[C] choice; CLI: via --move-to).
                 Used by the Duplicates removal flows only.
    move_to:
        CLI — a destination folder supplied via --move-to, skipping the
        interactive folder prompt. Its presence implies the action is "move".
    dry_mode:
        The caller's already-answered upfront "Run in dry-mode first?" choice
        (see ask_dry_mode()) — controls whether the preview listing prints.
        The final delete/move confirm below always still happens regardless.
    keep_map:
        Optional {path: keep_path} — for duplicate-removal flows, passed
        straight through to _preview_removal() so each removal candidate's
        surviving copy is shown beneath it.
    activity/fields: passed straight to _RunLog for the fm.log "started"
        entry — activity is the operation name (e.g. "Remove Folder"),
        fields a list of (label, value) tuples (source folder(s)/pattern/etc).
    """
    do_preview = dry_mode

    if do_preview:
        _preview_removal(items, keep_map=keep_map)
    elif items:
        total = sum(size for _, size, _ in items)
        print(color_text(
            f"  Skipping preview — {len(items)} item(s) matched ({fmt_size(total)}).", fg=DIM))

    if not items:
        print()
        print(color_text("  DRY RUN — nothing to delete.", fg=YELLOW, style=BOLD))
        return

    action = None          # "delete" | "move" | None (cancelled)
    dest_folder = None

    if live_requested is None:            # interactive
        print()
        if do_preview:
            print(color_text("  This was a DRY RUN — nothing has been deleted yet.", fg=YELLOW, style=BOLD))
        if allow_move:
            ch = menu_read(color_text(
                "  [D] Delete  [M] Move to Folder  [C] Cancel   Option: ", fg=YELLOW, style=BOLD))
            if ch == "ESC":
                ch = "c"
            if ch in ("d", "delete"):
                if require_yes:
                    if confirm_yes_word(f"  Type YES to delete these {len(items)} item(s) (anything else cancels): "):
                        action = "delete"
                elif safe_confirm(f"  Actually delete these {len(items)} item(s)?", default=False):
                    action = "delete"
            elif ch in ("m", "move"):
                dest_folder = ask_folder("Move to folder", must_exist=False, key="remove_move_to_folder")
                if dest_folder and safe_confirm(f"  Move these {len(items)} item(s) to {dest_folder}?", default=False):
                    action = "move"
        else:
            if require_yes:
                if confirm_yes_word(f"  Type YES to delete these {len(items)} item(s) (anything else cancels): "):
                    action = "delete"
            elif safe_confirm(f"  Actually delete these {len(items)} item(s)?", default=False):
                action = "delete"

    elif live_requested is True:          # CLI --delete or --move-to
        if move_to:
            dest_folder = move_to
            if assume_yes or safe_confirm(f"  Move these {len(items)} item(s) to {dest_folder}?", default=False):
                action = "move"
        elif assume_yes:
            action = "delete"
        elif require_yes:
            if confirm_yes_word(f"  Type YES to delete these {len(items)} item(s) (anything else cancels): "):
                action = "delete"
        elif safe_confirm(f"  Delete these {len(items)} item(s)?", default=False):
            action = "delete"

    else:                                 # CLI dry run
        print()
        tip = " (or --move-to FOLDER)" if allow_move else ""
        print(color_text(f"  DRY RUN — nothing deleted. Re-run with --delete{tip} to act.", fg=YELLOW, style=BOLD))
        return

    if action is None:
        print(color_text("  Cancelled — nothing deleted.", fg=YELLOW))
        return

    verb = "deletion" if action == "delete" else "move"
    if not _run_countdown(label=f"Starting {verb}"):
        print(color_text(f"  Cancelled — nothing {'deleted' if action == 'delete' else 'moved'}.", fg=YELLOW))
        return

    runlog_fields = list(fields or [])
    if action == "move":
        runlog_fields.append(("Moved To", dest_folder))
    runlog = _RunLog(activity, runlog_fields, total_items=len(items))

    if action == "delete":
        ok, fail, stopped_early = _perform_removal(items, runlog=runlog)
        verb_past = "Removed"
    else:
        ok, fail, stopped_early = _perform_move(items, dest_folder, runlog=runlog)
        verb_past = "Moved"

    print()
    summary = f"{verb_past} {ok} item(s)." + (f" {fail} failed." if fail else "")
    if stopped_early:
        summary += f" Stopped early by user ({ok + fail}/{len(items)} processed)."
    runlog.finish(summary)
    report_result(fail == 0, summary, f"{verb_past} {ok} item(s), {fail} failed.")


def remove_folder(path, live_requested=None):
    with _ActivityLog():
        screen("Remove Folder")
        dry_mode = ask_dry_mode(live_requested)
        print()
        print_mode_line(dry_mode)
        path = clean_path(path)
        print(f"  {YELLOW}Target{RESET}: {path}\n")
        if not os.path.isdir(path):
            print(color_text(f"  Not a directory: {path}", fg=RED)); pause_return(); return
        total, count = folder_stats(path)
        items = [(path, total, True)]
        print(color_text(f"  Contains {count:,} file(s), {fmt_size(total)}.", fg=DIM))
        _finish_removal(items, live_requested, activity="Remove Folder",
                         fields=[("Target", path)], dry_mode=dry_mode)
        pause_return()


def remove_by_name(root, pattern, live_requested=None):
    with _ActivityLog():
        screen("Remove Files by Name")
        dry_mode = ask_dry_mode(live_requested)
        print()
        print_mode_line(dry_mode)
        root = clean_path(root)
        print(f"  {YELLOW}Root{RESET}: {root}    {YELLOW}Pattern{RESET}: {pattern}\n")
        if not os.path.isdir(root):
            print(color_text(f"  Not a directory: {root}", fg=RED)); pause_return(); return
        items = []
        for dp, dns, fns in os.walk(root):
            for fn in fns:
                if fnmatch.fnmatch(fn, pattern):
                    p = os.path.join(dp, fn)
                    items.append((p, file_size(p), False))
        items.sort(key=lambda t: t[0])
        _finish_removal(items, live_requested, activity="Remove Files by Name",
                         fields=[("Root", root), ("Pattern", pattern)], dry_mode=dry_mode)
        pause_return()


def _topmost(paths):
    """Drop any path that is nested inside another path in the list, so a parent
    and its descendant aren't both scheduled for deletion (which would error)."""
    paths = sorted(set(paths))
    return [p for p in paths if not any(o != p and p.startswith(o + os.sep) for o in paths)]


def remove_folders_by_name(root, pattern, live_requested=None):
    with _ActivityLog():
        screen("Remove Folders by Name")
        dry_mode = ask_dry_mode(live_requested)
        print()
        print_mode_line(dry_mode)
        root = clean_path(root)
        print(f"  {YELLOW}Root{RESET}: {root}    {YELLOW}Pattern{RESET}: {pattern}\n")
        if not os.path.isdir(root):
            print(color_text(f"  Not a directory: {root}", fg=RED)); pause_return(); return
        matches = []
        for dp, dns, fns in os.walk(root):
            for d in dns:
                if fnmatch.fnmatch(d, pattern):
                    matches.append(os.path.join(dp, d))
        # Only delete top-most matches (a matched parent already covers matched children).
        items = [(m, folder_stats(m)[0], True) for m in _topmost(matches)]
        _finish_removal(items, live_requested, activity="Remove Folders by Name",
                         fields=[("Root", root), ("Pattern", pattern)], dry_mode=dry_mode)
        pause_return()


def _gather_files(folders, ext=None):
    """All files under the given folders, optionally filtered by extension."""
    files = []
    for folder in folders:
        folder = clean_path(folder)
        if not os.path.isdir(folder):
            continue
        for dp, dns, fns in os.walk(folder):
            for fn in fns:
                if ext and not fn.lower().endswith("." + ext.lstrip(".").lower()):
                    continue
                files.append(os.path.join(dp, fn))
    return files


def remove_duplicates_by_name(folders, live_requested=None, ext=None, move_to=None):
    with _ActivityLog():
        screen("Remove Duplicates (by Name)")
        dry_mode = ask_dry_mode(live_requested)
        print()
        print_mode_line(dry_mode)
        print(f"  {YELLOW}Folders{RESET}: {', '.join(folders)}")
        if ext:
            print(f"  {YELLOW}Extension{RESET}: .{ext.lstrip('.')}")
        print(f"  {DIM}Keeps the first occurrence of each name; removes the rest.{RESET}\n")

        files = _gather_files(folders, ext)
        groups = defaultdict(list)
        for p in files:
            groups[os.path.basename(p)].append(p)

        dups = {name: sorted(paths) for name, paths in groups.items() if len(paths) > 1}
        if not dups:
            print(color_text("  No duplicate filenames found.", fg=YELLOW))
            pause_return(); return

        items = []
        keep_map = {}
        for name in sorted(dups):
            keepPath = dups[name][0]
            for p in dups[name][1:]:
                items.append((p, file_size(p), False))
                keep_map[p] = keepPath

        if dry_mode:
            print(color_text(f"  {len(dups)} duplicated name(s):", fg=BRIGHT_CYAN, style=BOLD))
            for name in sorted(dups):
                paths = dups[name]
                print(f"    {WHITE}{name}{RESET}")
                print(f"      {GREEN}keep{RESET}   {paths[0]}")
                for p in paths[1:]:
                    print(f"      {RED}remove{RESET} {p}")
        else:
            print(color_text(f"  {len(dups)} duplicated name(s) found — skipping detailed "
                              f"listing (Live Mode).", fg=DIM))
        print()
        _finish_removal(items, live_requested, require_yes=True,
                         activity="Remove Duplicates (by Name)",
                         fields=[("Folders", ', '.join(folders))] +
                                 ([("Extension", f".{ext.lstrip('.')}")] if ext else []),
                         allow_move=True, move_to=move_to, dry_mode=dry_mode,
                         keep_map=keep_map)
        pause_return()


def remove_duplicates_by_hash(folders, live_requested=None, move_to=None):
    with _ActivityLog():
        screen("Remove Duplicates (by Hash)")
        dry_mode = ask_dry_mode(live_requested)
        print()
        print_mode_line(dry_mode)
        print(f"  {YELLOW}Folders{RESET}: {', '.join(folders)}")
        print(f"  {DIM}Groups by SHA-256 content hash; keeps the first, removes the rest.{RESET}\n")

        files = _gather_files(folders)
        # Pre-group by size to avoid hashing unique-size files
        by_size = defaultdict(list)
        for p in files:
            by_size[file_size(p)].append(p)

        groups = defaultdict(list)
        for size, paths in by_size.items():
            if len(paths) < 2:
                continue
            for p in paths:
                h = _hash_file(p)
                if h:
                    groups[h].append(p)

        dups = {h: sorted(paths) for h, paths in groups.items() if len(paths) > 1}
        if not dups:
            print(color_text("  No duplicate content found.", fg=YELLOW))
            pause_return(); return

        items = []
        keep_map = {}
        for h in sorted(dups):
            keepPath = dups[h][0]
            for p in dups[h][1:]:
                items.append((p, file_size(p), False))
                keep_map[p] = keepPath

        if dry_mode:
            print(color_text(f"  {len(dups)} set(s) of identical files:", fg=BRIGHT_CYAN, style=BOLD))
            for h in sorted(dups):
                paths = dups[h]
                print(f"    {DIM}{h[:16]}…{RESET}")
                print(f"      {GREEN}keep{RESET}   {paths[0]}  {DIM}({fmt_size(file_size(paths[0]))}){RESET}")
                for p in paths[1:]:
                    print(f"      {RED}remove{RESET} {p}  {DIM}({fmt_size(file_size(p))}){RESET}")
        else:
            print(color_text(f"  {len(dups)} set(s) of identical files found — skipping "
                              f"detailed listing (Live Mode).", fg=DIM))
        print()
        _finish_removal(items, live_requested, require_yes=True,
                         activity="Remove Duplicates (by Hash)",
                         fields=[("Folders", ', '.join(folders))],
                         allow_move=True, move_to=move_to, dry_mode=dry_mode,
                         keep_map=keep_map)
        pause_return()


def remove_duplicates_by_fuzzy_name(folders, live_requested=None, assume_yes=False, move_to=None):
    """Remove fuzzy duplicates: same grouping as Find Duplicates by Fuzzy Name
    (close name — dup-style endings stripped or 85%+ similar, same extension —
    plus sizes within 1%). Keeps the shortest/cleanest name in each group;
    the rest are listed for removal. Dry run + typed YES to delete, or move
    them to a folder instead (interactively, or via --move-to on the CLI)."""
    with _ActivityLog():
        screen("Remove Duplicates (by Fuzzy Name)")
        dry_mode = ask_dry_mode(live_requested)
        print()
        print_mode_line(dry_mode)
        folders = [clean_path(f) for f in folders if f]
        valid = [f for f in folders if os.path.isdir(f)]
        for f in folders:
            if f not in valid:
                print(color_text(f"  Skipping (not a directory): {f}", fg=YELLOW))
        if not valid:
            print(color_text("  No valid folders to scan.", fg=RED)); pause_return(); return
        print(f"  {YELLOW}Folders{RESET}: {', '.join(valid)}")
        print(f"  {DIM}Close names (duplicate-style endings stripped or 85%+ similar, same "
              f"extension) with sizes within 1%. Keeps the shortest/cleanest name; "
              f"removes the rest. Hidden files/folders are skipped.{RESET}\n")

        dup_groups = _fuzzy_dup_groups(valid)
        if not dup_groups:
            print(color_text("  No fuzzy-duplicate files found.", fg=YELLOW))
            pause_return(); return

        items = []
        keep_map = {}
        for group in dup_groups:
            (kname, _kstem, _kext, _ksize, kdp) = group[0]
            keepPath = os.path.join(kdp, kname)
            for (name, _stem, _ext, size, dp) in group[1:]:
                p = os.path.join(dp, name)
                items.append((p, size, False))
                keep_map[p] = keepPath

        if dry_mode:
            print(color_text(f"  {len(dup_groups)} fuzzy-duplicate group(s):", fg=BRIGHT_CYAN, style=BOLD))
            for group in dup_groups:
                print(f"    {WHITE}{group[0][0]}{RESET}")
                (name, _stem, _ext, size, dp) = group[0]
                print(f"      {BRIGHT_CYAN}keep{RESET}   {os.path.join(dp, name)}  {DIM}({fmt_size(size)}){RESET}")
                for (name, _stem, _ext, size, dp) in group[1:]:
                    p = os.path.join(dp, name)
                    print(f"      {BRIGHT_YELLOW}remove{RESET} {p}  {DIM}({fmt_size(size)}){RESET}")
        else:
            print(color_text(f"  {len(dup_groups)} fuzzy-duplicate group(s) found — skipping "
                              f"detailed listing (Live Mode).", fg=DIM))
        print()
        _finish_removal(items, live_requested, assume_yes=assume_yes, require_yes=True,
                         activity="Remove Duplicates (by Fuzzy Name)",
                         fields=[("Folders", ', '.join(valid))],
                         allow_move=True, move_to=move_to, dry_mode=dry_mode,
                         keep_map=keep_map)
        pause_return()


def remove_zero_size_files(folders, live_requested=None, assume_yes=False):
    """Remove empty (0-byte) files under the given folders. Hidden
    files/folders are skipped — so intentional placeholders like .gitkeep are
    never touched. Dry run + confirm before anything is deleted."""
    with _ActivityLog():
        screen("Remove Files of 0 Size")
        dry_mode = ask_dry_mode(live_requested)
        print()
        print_mode_line(dry_mode)
        folders = [clean_path(f) for f in folders if f]
        valid = [f for f in folders if os.path.isdir(f)]
        for f in folders:
            if f not in valid:
                print(color_text(f"  Skipping (not a directory): {f}", fg=YELLOW))
        if not valid:
            print(color_text("  No valid folders to scan.", fg=RED)); pause_return(); return
        print(f"  {YELLOW}Folders{RESET}: {', '.join(valid)}")
        print(f"  {DIM}Removes empty (0-byte) files. Hidden files/folders are skipped.{RESET}\n")

        items = []
        for folder in valid:
            for dp, dns, fns in os.walk(folder):
                dns[:] = [d for d in dns if not d.startswith(".") and d.lower() not in EXCLUDED_DIR_NAMES]
                for fn in fns:
                    if fn.startswith(".") or is_excluded_file(fn):
                        continue
                    p = os.path.join(dp, fn)
                    try:
                        if os.path.getsize(p) == 0 and os.path.isfile(p):
                            items.append((p, 0, False))
                    except OSError:
                        continue
        items.sort(key=lambda t: t[0])

        if not items:
            print(color_text("  No 0-byte files found.", fg=YELLOW))
            pause_return(); return

        _finish_removal(items, live_requested, assume_yes=assume_yes,
                         activity="Remove Files of 0 Size",
                         fields=[("Folders", ', '.join(valid))], dry_mode=dry_mode)
        pause_return()


def _hash_file(path, chunk=1 << 20):
    try:
        h = hashlib.sha256()
        with open(path, "rb") as f:
            while True:
                b = f.read(chunk)
                if not b:
                    break
                h.update(b)
        return h.hexdigest()
    except OSError:
        return None


# =============================================================================
# CLEAN UP  (junk files, old log entries; dry-run by default)
# =============================================================================
LOG_PURGE_DIR   = os.path.expanduser("~/Documents/log")
LOG_PURGE_DAYS  = 90
_LOG_TS_RE      = re.compile(r"\[(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\]")


JUNK_FILE_SPECS = [
    (".DS_Store",   "macOS Finder metadata file",
     lambda fn: fn.lower() == ".ds_store"),
    ("desktop.ini", "Windows folder metadata file",
     lambda fn: fn.lower() == "desktop.ini"),
    ("*.bak",       "Backup files (e.g. saved by Purge Old Log Files)",
     lambda fn: fn.lower().endswith(".bak")),
]


def cleanup_junk_files(root, live_requested=None, assume_yes=False, spec_indices=None):
    """Find every selected junk-file spec under root (recursively, hidden
    folders included) and remove them — DRY RUN preview + confirm, exactly
    like the Remove actions.

    spec_indices: which JUNK_FILE_SPECS entries to target (default: all —
    .DS_Store, desktop.ini, *.bak)."""
    indices = list(range(len(JUNK_FILE_SPECS))) if spec_indices is None else list(spec_indices)
    matchers = [JUNK_FILE_SPECS[i][2] for i in indices]
    labels = ", ".join(JUNK_FILE_SPECS[i][0] for i in indices) or "(none selected)"
    with _ActivityLog():
        screen("Clean Up — Remove Junk Files")
        dry_mode = ask_dry_mode(live_requested)
        print()
        print_mode_line(dry_mode)
        root = clean_path(root)
        print(f"  {YELLOW}Root{RESET}: {root}    {YELLOW}Targets{RESET}: {labels}\n")
        if not os.path.isdir(root):
            print(color_text(f"  Not a directory: {root}", fg=RED))
            return
        items = []
        for dp, dns, fns in os.walk(root):
            prune_dirs(dns)
            for fn in fns:
                if any(m(fn) for m in matchers):
                    full = os.path.join(dp, fn)
                    try:
                        size = os.path.getsize(full)
                    except OSError:
                        size = 0
                    items.append((full, size, False))
        items.sort(key=lambda i: i[0])
        _finish_removal(items, live_requested, assume_yes=assume_yes,
                         activity="Clean Up — Remove Junk Files",
                         fields=[("Root", root), ("Targets", labels)], dry_mode=dry_mode)


def _purge_scan_file(path, cutoff):
    """Read one log file and split it at cutoff. Returns (kept_lines,
    purge_count).

    Block-aware: a line with a [YYYY-MM-DD HH:MM:SS] timestamp decides
    keep/purge for itself AND every untimestamped line that follows it (an
    entry's body and trailing blank line belong to its header). Lines before
    any timestamp are kept."""
    kept = []
    purged = 0
    keep_current = True
    with open(path, "r", errors="replace") as fh:
        for line in fh:
            m = _LOG_TS_RE.search(line)
            if m:
                try:
                    ts = datetime.strptime(m.group(1), "%Y-%m-%d %H:%M:%S")
                    keep_current = ts >= cutoff
                except ValueError:
                    keep_current = True
            if keep_current:
                kept.append(line)
            else:
                purged += 1
    return kept, purged


def cleanup_purge_logs(folder=LOG_PURGE_DIR, days=LOG_PURGE_DAYS,
                       live_requested=None, assume_yes=False):
    """Purge entries older than N days from every .log file in folder.
    DRY RUN preview + confirm; a .bak backup of each changed file is written
    before it is rewritten. Empty files after purging are preserved."""
    with _ActivityLog():
        screen("Clean Up — Purge Old Log Files")
        dry_mode = ask_dry_mode(live_requested)
        print()
        print_mode_line(dry_mode)
        folder = clean_path(folder)
        try:
            days = max(0, int(days))
        except (TypeError, ValueError):
            days = LOG_PURGE_DAYS
        cutoff = datetime.fromtimestamp(time.time() - days * 86400)
        print(f"  {YELLOW}Log folder{RESET}  : {folder}")
        print(f"  {YELLOW}Days to keep{RESET}: {days}")
        print(f"  {YELLOW}Cutoff{RESET}      : entries before "
              f"{cutoff.strftime('%-m/%-d/%y %-I:%M ') + cutoff.strftime('%p').lower()} are purged")
        print()
        if not os.path.isdir(folder):
            print(color_text(f"  Not a directory: {folder}", fg=RED))
            return

        log_files = sorted(f for f in os.listdir(folder) if f.lower().endswith(".log"))
        if not log_files:
            print(color_text(f"  No .log files found in {folder}.", fg=YELLOW))
            return

        plans = []                       # (path, kept_lines, purge_count)
        for fn in log_files:
            path = os.path.join(folder, fn)
            try:
                kept, purge_count = _purge_scan_file(path, cutoff)
            except OSError as e:
                print(color_text(f"  ⚠ {fn}: {e}", fg=YELLOW))
                continue
            plans.append((path, kept, purge_count))

        total_purge = sum(purge_count for _, _, purge_count in plans)

        if dry_mode:
            name_w = max([len(os.path.basename(p)) for p, _, _ in plans] + [len("Log File")])
            name_w = min(name_w, 40)
            print(color_text(f"  {'Log File':<{name_w}}  {'Purge':>8}  {'Keep':>8}  {'Size':>10}",
                             fg=YELLOW, style=BOLD))
            print(f"  {DIM}{'-' * name_w}  {'-' * 8}  {'-' * 8}  {'-' * 10}{RESET}")
            for path, kept, purge_count in plans:
                try:
                    size = os.path.getsize(path)
                except OSError:
                    size = 0
                purge_disp = (color_text(f"{purge_count:>8,}", fg=RED, style=BOLD)
                              if purge_count else f"{purge_count:>8,}")
                print(f"  {WHITE}{os.path.basename(path)[:name_w].ljust(name_w)}{RESET}  "
                      f"{purge_disp}  {len(kept):>8,}  {fmt_size(size):>10}")
        else:
            print(color_text(f"  {len(plans)} log file(s) scanned — skipping detailed "
                              f"listing (Live Mode).", fg=DIM))

        print()
        if total_purge == 0:
            print(color_text(f"  Nothing to purge — no entries older than {days} days.",
                             fg=GREEN, style=BOLD))
            return

        do_purge = False
        if live_requested is None:                # interactive
            print(color_text("  This was a DRY RUN — nothing has been purged yet.",
                             fg=YELLOW, style=BOLD))
            do_purge = safe_confirm(f"  Actually purge {total_purge:,} old line(s) "
                                    f"from {len(plans)} file(s)?", default=False)
        elif live_requested is True:              # CLI --delete
            do_purge = True if assume_yes else safe_confirm(
                f"  Purge {total_purge:,} old line(s) from {len(plans)} file(s)?", default=False)
        else:                                     # CLI dry run
            print(color_text("  DRY RUN — nothing purged. Re-run with --delete to purge.",
                             fg=YELLOW, style=BOLD))
            return
        if not do_purge:
            print(color_text("  Cancelled — nothing purged.", fg=YELLOW))
            return
        if not _run_countdown(label="Starting purge"):
            print(color_text("  Cancelled — nothing purged.", fg=YELLOW))
            return

        purge_plans = [p for p in plans if p[2] > 0]
        runlog = _RunLog("Clean Up — Purge Old Log Files",
                          [("Log folder", folder), ("Days to keep", days)],
                          total_items=len(purge_plans))
        ok = fail = 0
        stopped_early = False
        ps = _PauseStop()
        if ps.is_tty and len(purge_plans) > 1:
            print(color_text("  Press [P] to pause, [Q] to stop early.", fg=DIM))
        try:
            for path, kept, purge_count in plans:
                if purge_count == 0:
                    continue
                if ps.check() == "stop":
                    stopped_early = True
                    break
                try:
                    shutil.copy2(path, path + ".bak")
                    with open(path, "w") as fh:
                        fh.writelines(kept)
                    ok += 1
                    runlog.action(f"Purged {purge_count:,} line(s) from {os.path.basename(path)}")
                except OSError as e:
                    fail += 1
                    print(color_text(f"  ✗ {os.path.basename(path)}: {e}", fg=RED))
                    runlog.action(f"FAILED to purge {os.path.basename(path)}: {e}")
        finally:
            ps.close()
        print()
        summary = (f"Purged {total_purge:,} line(s) from {ok} file(s)."
                   + (f" {fail} failed." if fail else ""))
        if stopped_early:
            summary += f" Stopped early by user ({ok + fail}/{len(purge_plans)} processed)."
        runlog.finish(summary)
        report_result(fail == 0,
                      f"Purged {total_purge:,} line(s) from {ok} file(s). "
                      "Backups saved with a .bak extension.",
                      f"Purged {ok} file(s), {fail} failed.")


# =============================================================================
# EJECT EXTERNAL DRIVES  (macOS)
# =============================================================================
def _external_drives():
    """Scan /Volumes and return the mounted external drives as a list of
    {identifier, name, size, mount_point} dicts (deduplicated by disk).

    A volume counts as external when `diskutil info` reports Internal: No,
    a Protocol of USB/Thunderbolt/SATA/FireWire, or Removable Media. Same
    detection logic as the standalone ejectDrives.py script."""
    drives = []
    volumes_path = "/Volumes"
    if not os.path.isdir(volumes_path):
        return drives
    for volume_name in sorted(os.listdir(volumes_path)):
        volume_path = os.path.join(volumes_path, volume_name)
        if not os.path.ismount(volume_path):
            continue
        try:
            info_result = subprocess.run(["diskutil", "info", volume_path],
                                         capture_output=True, text=True)
        except OSError:
            return drives                    # diskutil unavailable (not macOS)
        if info_result.returncode != 0:
            continue

        is_external = False
        disk_identifier = None
        size = "—"
        for line in info_result.stdout.split("\n"):
            if "Protocol:" in line:
                if line.split(":")[1].strip().lower() in ("usb", "thunderbolt",
                                                          "sata", "firewire"):
                    is_external = True
            elif "Removable Media:" in line:
                if "Removable" in line or "Yes" in line:
                    is_external = True
            elif "Internal:" in line:
                if "No" in line:
                    is_external = True
            elif "Part of Whole:" in line:
                disk_identifier = line.split(":")[1].strip()
            elif "Device Identifier:" in line and not disk_identifier:
                m = re.match(r"(disk\d+)", line.split(":")[1].strip())
                if m:
                    disk_identifier = m.group(1)
            elif ("Disk Size:" in line or "Container Total Space:" in line
                  or "Volume Total Space:" in line):
                m = re.search(r"\((\d[\d,]*)\s*Bytes\)", line)
                if m:
                    size = fmt_size(int(m.group(1).replace(",", "")))

        if is_external and disk_identifier:
            drives.append({"identifier": disk_identifier, "name": volume_name,
                           "size": size, "mount_point": volume_path})

    # One entry per physical disk (a disk may mount several volumes)
    seen = set()
    unique = []
    for d in drives:
        if d["identifier"] not in seen:
            seen.add(d["identifier"])
            unique.append(d)
    return unique


def _eject_via_finder(volume_name):
    """AppleScript fallback — same as clicking the eject button in Finder."""
    script = f'tell application "Finder"\n    eject disk "{volume_name}"\nend tell'
    try:
        result = subprocess.run(["osascript", "-e", script],
                                capture_output=True, text=True)
    except OSError as e:
        return False, str(e)
    if result.returncode == 0:
        return True, "Ejected via Finder"
    return False, result.stderr.strip()


def _eject_drive(drive, force=False):
    """Eject one drive dict. force=True force-unmounts the whole disk first.
    Returns (success, message)."""
    identifier = drive["identifier"]
    try:
        if force:
            unmount = subprocess.run(["diskutil", "unmountDisk", "force", identifier],
                                     capture_output=True, text=True)
            if unmount.returncode != 0:
                return False, f"Force unmount failed: {unmount.stderr.strip()}"
        result = subprocess.run(["diskutil", "eject", identifier],
                                capture_output=True, text=True)
        if result.returncode == 0:
            return True, result.stdout.strip()
        if not force:                         # Finder fallback (non-force only)
            ok, msg = _eject_via_finder(drive["name"])
            if ok:
                return True, msg
        return False, result.stderr.strip()
    except OSError as e:
        return False, str(e)


def _print_drive_list(drives):
    name_w = max([len(d["name"]) for d in drives] + [4])
    size_w = max([len(d["size"]) for d in drives] + [4])
    print(color_text(f"  {'#':>2}  {'Name':<{name_w}}  {'Size':>{size_w}}  Mount Point",
                     fg=WHITE, style=BOLD))
    print(f"  {'-' * 2}  {'-' * name_w}  {'-' * size_w}  {'-' * 20}")
    for i, d in enumerate(drives, 1):
        print(f"  {i:>2}  {d['name']:<{name_w}}  {d['size']:>{size_w}}  "
              f"{DIM}{d['mount_point']}{RESET}")


def _eject_loop(drives, force=False, runlog=None):
    """Eject each drive with a per-drive status line.
    Returns (ok_count, failed_drives)."""
    ok = 0
    failed = []
    for d in drives:
        verb = "Force ejecting" if force else "Ejecting"
        sys.stdout.write(color_text(f"  {verb} {d['name']}... ", fg=CYAN))
        sys.stdout.flush()
        success, msg = _eject_drive(d, force=force)
        if success:
            print(color_text("Success", fg=BRIGHT_GREEN, style=BOLD))
            ok += 1
            if runlog:
                runlog.action(f"Ejected {d['name']}")
        else:
            print(color_text("Failed", fg=BRIGHT_RED, style=BOLD))
            if msg:
                print(color_text(f"    {msg}", fg=RED))
            failed.append(d)
            if runlog:
                runlog.action(f"FAILED to eject {d['name']}" + (f": {msg}" if msg else ""))
    return ok, failed


def _eject_screen(live_requested, assume_yes, force):
    screen("Eject All External Drives")
    print()
    runlog = _RunLog("Eject All External Drives")
    if sys.platform != "darwin":
        print(color_text("  Eject uses diskutil and is only available on macOS.", fg=RED))
        runlog.finish("Cancelled — macOS only")
        return

    drives = _external_drives()
    if not drives:
        print(color_text("  No external drives found.", fg=YELLOW, style=BOLD))
        runlog.finish("No external drives found")
        return

    _print_drive_list(drives)
    print()

    do_eject = False
    if live_requested is None or (live_requested is True and not assume_yes):
        do_eject = safe_confirm(f"  Eject all {len(drives)} drive(s)?", default=False)
    else:
        do_eject = True                       # CLI --yes
    if not do_eject:
        print(color_text("  Cancelled — nothing ejected.", fg=YELLOW))
        runlog.finish("Cancelled — nothing ejected")
        return

    print()
    runlog.total_items = len(drives)
    runlog.detailed = len(drives) <= ACTION_LOG_DETAIL_THRESHOLD
    ok, failed = _eject_loop(drives, force=force, runlog=runlog)

    # Offer to force-eject anything that would not let go (Spotlight, etc.)
    if failed and not force:
        print()
        force_it = (True if assume_yes and live_requested is True
                    else safe_confirm(f"  Force eject the {len(failed)} failed drive(s)?",
                                      default=False))
        if force_it:
            print()
            ok2, failed = _eject_loop(failed, force=True, runlog=runlog)
            ok += ok2

    print()
    runlog.finish(f"Ejected {ok} drive(s)." + (f" {len(failed)} failed." if failed else ""))
    report_result(not failed,
                  f"Ejected {ok} drive(s).",
                  f"Ejected {ok} drive(s), {len(failed)} failed — "
                  "close any apps using the drive(s) and try again.")


def eject_external_drives(live_requested=None, assume_yes=False, force=False):
    """List the mounted external drives and eject them all after a confirm
    (macOS). Failures offer a force eject. The screen output is also appended
    to ~/Documents/log/fm.log."""
    with _ActivityLog():
        _eject_screen(live_requested, assume_yes, force)


# =============================================================================
# MONITOR FILE ACTIVITY  (real-time, stdlib polling)
# =============================================================================
MONITOR_CSV          = os.path.expanduser("~/Documents/log/fmMonitor.csv")
MONITOR_POLL_SECONDS = 1.0


def _parse_ext_filter(raw):
    """Normalize an extension filter ('jpg, .PNG' or ['jpg','png']) to a
    lowercase set without dots. Empty/blank input -> empty set (= all files)."""
    if not raw:
        return set()
    parts = raw if isinstance(raw, (list, tuple)) else str(raw).split(",")
    return {str(p).strip().lstrip(".").lower() for p in parts if str(p).strip()}


def _monitor_snapshot(root, recursive, exts):
    """One poll of the watched folder: {relative path: (mtime, size)}.

    Hidden files/folders ARE included (their activity is often the point);
    junk files (.DS_Store, desktop.ini) and $RECYCLE.BIN are skipped.
    exts: set of extensions to watch (empty = every file)."""
    snap = {}

    def want(fn):
        if is_excluded_file(fn):
            return False
        return not exts or os.path.splitext(fn)[1][1:].lower() in exts

    if recursive:
        for dp, dns, fns in os.walk(root):
            prune_dirs(dns)
            for fn in fns:
                if not want(fn):
                    continue
                full = os.path.join(dp, fn)
                try:
                    st = os.stat(full)
                except OSError:
                    continue
                snap[os.path.relpath(full, root)] = (st.st_mtime, st.st_size)
    else:
        try:
            names = os.listdir(root)
        except OSError:
            return snap
        for fn in names:
            full = os.path.join(root, fn)
            if not want(fn) or not os.path.isfile(full):
                continue
            try:
                st = os.stat(full)
            except OSError:
                continue
            snap[fn] = (st.st_mtime, st.st_size)
    return snap


def _diff_snapshots(old, new):
    """Compare two snapshots. Returns [(event, relpath, size)] sorted by
    path; event is CREATED / MODIFIED / DELETED."""
    events = []
    for rel, meta in new.items():
        if rel not in old:
            events.append(("CREATED", rel, meta[1]))
        elif meta != old[rel]:
            events.append(("MODIFIED", rel, meta[1]))
    for rel in old:
        if rel not in new:
            events.append(("DELETED", rel, None))
    return sorted(events, key=lambda e: e[1])


def _monitor_ts():
    """Event timestamp in the display format m/d/yy h:mm:ss am."""
    now = datetime.now()
    return now.strftime("%-m/%-d/%y %-I:%M:%S ") + now.strftime("%p").lower()


# =============================================================================
# Stationary header/footer primitives (ANSI scroll region / DECSTBM)
# =============================================================================
# Shared by any FM screen that shows a long/scrolling body (Monitor's live
# event feed, Sync's file plan + copy progress, ...): the header stays pinned
# to the top of the window with the task's own context (folder(s), options,
# a running summary), the footer stays pinned to the bottom, and only the
# body between them scrolls — same technique as backup42.py's execution
# screen (CB9 Static Header/Footer permanent rule). Each screen defines its
# own HEADER_LINES/FOOTER_LINES and draws its own frame content; only the
# low-level cursor/scroll-region escape codes are shared here.


def _frm_mv(row, col=1):
    """Move cursor to absolute row/col (1-indexed)."""
    return f"\033[{row};{col}H"


def _frm_el():
    """Erase entire current line."""
    return "\033[2K"


def _frm_set_sr(top, bottom):
    """Set ANSI scroll region (DECSTBM)."""
    return f"\033[{top};{bottom}r"


def _frm_rst_sr():
    """Reset scroll region to full screen."""
    return "\033[r"


def _frm_hide_cur():
    return "\033[?25l"


def _frm_show_cur():
    return "\033[?25h"


def _frm_term_size():
    """Real current terminal size via ioctl on the process's real stdout fd.

    shutil.get_terminal_size() prefers the COLUMNS/LINES environment
    variables when present, which can be stale (set once at shell start,
    inherited by this process, and never updated on a later window resize).
    Absolute-row frame math needs the terminal's *actual* current size, so
    query it directly.

    Uses the raw fd 1 rather than sys.stdout.fileno(): several FM screens
    (Sync, Eject, ...) run inside `with _ActivityLog():`, which temporarily
    replaces sys.stdout with a write()/flush()-only tee proxy that has no
    fileno() at all — fd 1 is unaffected by that (or any future) Python-level
    stdout wrapper. Falls back to shutil's version (env vars or its 80x24
    default) if the ioctl isn't available (e.g. no real tty)."""
    try:
        return os.get_terminal_size(1)
    except OSError:
        return shutil.get_terminal_size()


def _frm_save_cur():
    """Save cursor position (DEC DECSC — reliable across scroll-region changes)."""
    return "\0337"


def _frm_rest_cur():
    """Restore saved cursor position (DEC DECRC)."""
    return "\0338"


def _frm_cleanup(rows):
    """Reset the scroll region, show the cursor, and move to the bottom of
    the screen so subsequent print() calls appear after the frozen area."""
    sys.stdout.write(_frm_rst_sr() + _frm_show_cur() + _frm_mv(rows) + "\r\n")
    sys.stdout.flush()


# -----------------------------------------------------------------------------
# Monitor screen frame
# -----------------------------------------------------------------------------
MONITOR_HEADER_LINES = 8   # rows 1-8: ===, title, ===, Profile, Folder, Recursive/Ext, Logging, ===
MONITOR_FOOTER_LINES = 2   # bottom 2: ===, legend (stop hint + live counts)
MONITOR_MIN_HEIGHT   = MONITOR_HEADER_LINES + MONITOR_FOOTER_LINES + 3


def _monitor_legend(counts):
    total = sum(counts.values())
    return (f" [Q/ESC] Stop monitoring   {BRIGHT_YELLOW}Events{RESET}: {total}"
            f"  ({GREEN}{counts['CREATED']} created{RESET}, "
            f"{YELLOW}{counts['MODIFIED']} modified{RESET}, "
            f"{RED}{counts['DELETED']} deleted{RESET})")


def _monitor_draw_frame(folder, recursive, exts, log_disp, profile_name, counts):
    """Draw the stationary header (rows 1-8) and footer (last 2 rows) at the
    current terminal size, set the scroll region to the rows between them,
    and park the cursor at its top. Called on first draw and on SIGWINCH."""
    cols, rows = _frm_term_size()

    if rows < MONITOR_MIN_HEIGHT:
        sys.stdout.write("\033[2J" + _frm_mv(1))
        print(color_text(f" Terminal too small — need at least {MONITOR_MIN_HEIGHT} rows "
                         f"(current: {rows}).", fg=YELLOW))
        sys.stdout.flush()
        return None

    w = cols
    ext_disp = ", ".join(sorted(exts)) if exts else "all files"
    out = [_frm_rst_sr(), "\033[2J", _frm_hide_cur()]

    out.append(_frm_mv(1) + _frm_el() + "=" * w)
    out.append(_frm_mv(2) + _frm_el() +
               f" {BOLD}{CYAN}{SCRIPT_NAME}{RESET} {MAGENTA}{VER}{RESET}  "
               f"{DIM}{WHITE}[Monitor File Activity]{RESET}")
    out.append(_frm_mv(3) + _frm_el() + "=" * w)
    out.append(_frm_mv(4) + _frm_el() + f" {YELLOW}Profile{RESET}   : {profile_name or '—'}")
    out.append(_frm_mv(5) + _frm_el() + f" {YELLOW}Folder{RESET}    : {folder}")
    out.append(_frm_mv(6) + _frm_el() +
               f" {YELLOW}Recursive{RESET} : {'Yes' if recursive else 'No'}     "
               f"{YELLOW}Extensions{RESET}: {ext_disp}")
    out.append(_frm_mv(7) + _frm_el() + f" {YELLOW}Logging to{RESET}: {log_disp}")
    out.append(_frm_mv(8) + _frm_el() + "=" * w)

    footer_top = rows - MONITOR_FOOTER_LINES + 1
    out.append(_frm_mv(footer_top) + _frm_el() + "=" * w)
    out.append(_frm_mv(footer_top + 1) + _frm_el() + _monitor_legend(counts))

    scroll_top    = MONITOR_HEADER_LINES + 1
    scroll_bottom = rows - MONITOR_FOOTER_LINES
    out.append(_frm_set_sr(scroll_top, scroll_bottom))
    out.append(_frm_mv(scroll_top))

    sys.stdout.write("".join(out))
    sys.stdout.flush()
    return rows


def _monitor_update_footer(counts, rows):
    """In-place update of the footer legend (live event counts) without
    disturbing the scroll region or the cursor position within it."""
    footer_top = rows - MONITOR_FOOTER_LINES + 1
    out = (_frm_save_cur()
           + _frm_mv(footer_top + 1) + _frm_el() + _monitor_legend(counts)
           + _frm_rest_cur())
    sys.stdout.write(out)
    sys.stdout.flush()


# -----------------------------------------------------------------------------
# Sync screen frame
# -----------------------------------------------------------------------------
# Unlike Monitor (fixed header every run), Sync's header height depends on
# whether a profile name is shown, so it's computed per-run rather than a
# fixed module constant.
SYNC_FOOTER_LINES = 2   # bottom 2: ===, legend (phase / live copy progress)


def _sync_header_lines(folder_a, folder_b, direction, recursive, conflict,
                       exclude_hidden, profile_name, total_new, total_upd,
                       total_bytes, copy_desc, w, dry_mode=True):
    """Build the Sync frame's fixed header content, one already-colored
    string per row (row position is assigned by the caller)."""
    direction_disp = _sync_direction_desc(direction)
    options_disp = (f"Recursive: {'Yes' if recursive else 'No'}   "
                    f"Conflict: {_conflict_label(conflict)}   "
                    f"Hidden files: {'Excluded' if exclude_hidden else 'Included'}")
    mode_label = "Dry-Mode" if dry_mode else "Live Mode"
    mode_color = YELLOW if dry_mode else BRIGHT_RED

    lines = [
        "=" * w,
        f" {BOLD}{CYAN}{SCRIPT_NAME}{RESET} {MAGENTA}{VER}{RESET}  {DIM}{WHITE}[Sync Folders]{RESET}",
        "=" * w,
        "",
    ]
    lines.append(color_text(f" Mode: {mode_label}", fg=mode_color, style=BOLD))
    if profile_name:
        lines.append(f" {YELLOW}Profile{RESET}  : {profile_name}")
    lines.append(f" {YELLOW}Folder A{RESET} : {folder_a}")
    lines.append(f" {YELLOW}Folder B{RESET} : {folder_b}")
    lines.append(f" {YELLOW}Direction{RESET}: {direction_disp}")
    lines.append(f" {YELLOW}Options{RESET}  : {options_disp}")
    lines.append("")
    lines.append(color_text(f" {total_new} new, {total_upd} updated — "
                            f"{fmt_size(total_bytes)} to copy {copy_desc}:",
                            fg=WHITE, style=BOLD))
    lines.append("=" * w)
    return lines


def _sync_draw_frame(folder_a, folder_b, direction, recursive, conflict,
                     exclude_hidden, profile_name, total_new, total_upd,
                     total_bytes, copy_desc, status, dry_mode=True):
    """Draw the stationary Sync header (folders/direction/options + the new/
    updated/size summary) and footer (phase/progress status), set the
    scroll region between them, and park the cursor at its top. Called on
    first draw and again on SIGWINCH. Returns (rows, header_line_count) or
    (None, None) if the terminal is too short."""
    cols, rows = _frm_term_size()
    header_lines      = _sync_header_lines(folder_a, folder_b, direction, recursive,
                                           conflict, exclude_hidden, profile_name,
                                           total_new, total_upd, total_bytes,
                                           copy_desc, cols, dry_mode=dry_mode)
    header_line_count = len(header_lines)
    min_height        = header_line_count + SYNC_FOOTER_LINES + 3

    if rows < min_height:
        sys.stdout.write("\033[2J" + _frm_mv(1))
        print(color_text(f" Terminal too small — need at least {min_height} rows "
                         f"(current: {rows}).", fg=YELLOW))
        sys.stdout.flush()
        return None, None

    out = [_frm_rst_sr(), "\033[2J", _frm_hide_cur()]
    for i, text in enumerate(header_lines, 1):
        out.append(_frm_mv(i) + _frm_el() + text)

    footer_top = rows - SYNC_FOOTER_LINES + 1
    out.append(_frm_mv(footer_top) + _frm_el() + "=" * cols)
    out.append(_frm_mv(footer_top + 1) + _frm_el() + f" {status}")

    scroll_top    = header_line_count + 1
    scroll_bottom = rows - SYNC_FOOTER_LINES
    out.append(_frm_set_sr(scroll_top, scroll_bottom))
    out.append(_frm_mv(scroll_top))

    sys.stdout.write("".join(out))
    sys.stdout.flush()
    return rows, header_line_count


def _sync_update_footer(rows, status):
    """In-place update of the footer legend (phase / live copy progress)
    without disturbing the scroll region or the cursor position within it."""
    footer_top = rows - SYNC_FOOTER_LINES + 1
    out = (_frm_save_cur()
           + _frm_mv(footer_top + 1) + _frm_el() + f" {status}"
           + _frm_rest_cur())
    sys.stdout.write(out)
    sys.stdout.flush()


class _MonitorLog:
    """Real-time event writer. output='log' appends lines to the FM activity
    log (~/Documents/log/fm.log); output='csv' appends rows to fmMonitor.csv
    in the same folder (header written when the file is new). Every event is
    flushed immediately so `tail -f` tracks the screen live. Write failures
    never interrupt monitoring."""

    def __init__(self, output, folder, recursive, exts):
        self.output = output
        self.path = MONITOR_CSV if output == "csv" else ACTIVITY_LOG
        self._fh = None
        try:
            os.makedirs(os.path.dirname(self.path), exist_ok=True)
            is_new = not os.path.exists(self.path) or os.path.getsize(self.path) == 0
            self._fh = open(self.path, "a", newline="" if output == "csv" else None)
            if output == "csv":
                self._csv = csv.writer(self._fh)
                if is_new:
                    self._csv.writerow(["Timestamp", "Filename", "Folder", "Event"])
            else:
                ext_disp = ", ".join(sorted(exts)) if exts else "all"
                self._fh.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]\n")
                self._fh.write(f"Monitor File Activity — {folder}\n")
                self._fh.write(f"Recursive: {'Yes' if recursive else 'No'}   "
                               f"Extensions: {ext_disp}\n")
            self._fh.flush()
        except OSError:
            self._fh = None
        self.opened = self._fh is not None

    def event(self, ts, event, rel, folder):
        if not self._fh:
            return
        try:
            if self.output == "csv":
                self._csv.writerow([ts, os.path.basename(rel),
                                    os.path.join(folder, os.path.dirname(rel)).rstrip("/"),
                                    event.lower()])
            else:
                self._fh.write(f"{ts}  {event:<9} {rel}\n")
            self._fh.flush()
        except OSError:
            pass

    def close(self, summary=""):
        if not self._fh:
            return
        try:
            if self.output == "log":
                if summary:
                    self._fh.write(summary + "\n")
                self._fh.write("\n")
            self._fh.flush()
            self._fh.close()
        except OSError:
            pass
        self._fh = None


_MONITOR_EVENT_COLORS = {"CREATED": GREEN, "MODIFIED": YELLOW, "DELETED": RED}


def monitor_activity(folder, recursive=True, exts=None, output="log",
                     profile_name=None):
    """Watch a folder for file activity (created / modified / deleted) and
    report each event in real time — on screen and to fm.log (or
    fmMonitor.csv). Polls once a second using snapshot diffs (no third-party
    packages). [Q/ESC] stops and returns to the menu (Ctrl-C non-TTY).

    In an interactive terminal the header (folder/recursive/extensions/
    logging info) and footer (stop hint + live event counts) are stationary
    at the top/bottom of the window; only the event feed between them
    scrolls (ANSI scroll region — see _monitor_draw_frame)."""
    exts = _parse_ext_filter(exts)

    if not os.path.isdir(folder):
        screen("Monitor File Activity")
        print()
        print(color_text(f"  Not a directory: {folder}", fg=RED))
        _log_line(f"Monitor File Activity cancelled — not a directory: {folder}")
        return

    log_disp = MONITOR_CSV.replace(os.path.expanduser('~'), '~') if output == "csv" \
        else ACTIVITY_LOG.replace(os.path.expanduser('~'), '~')
    is_tty = sys.stdin.isatty()
    counts = {"CREATED": 0, "MODIFIED": 0, "DELETED": 0}
    frame_rows = None

    if is_tty:
        frame_rows = _monitor_draw_frame(folder, recursive, exts, log_disp, profile_name, counts)
        if frame_rows is None:
            pause_return()
            return
    else:
        screen("Monitor File Activity")
        print()
        if profile_name:
            print(f"  {YELLOW}Profile{RESET}   : {profile_name}")
        print(f"  {YELLOW}Folder{RESET}    : {folder}")
        print(f"  {YELLOW}Recursive{RESET} : {'Yes' if recursive else 'No'}")
        print(f"  {YELLOW}Extensions{RESET}: {', '.join(sorted(exts)) if exts else 'all files'}")
        print(f"  {YELLOW}Logging to{RESET}: {log_disp}")
        print()
        print(color_text("  Monitoring... Ctrl-C to stop. Events appear below as they happen.",
                         fg=CYAN, style=BOLD))
        print()

    ext_disp = ", ".join(sorted(exts)) if exts else "all"
    if output == "csv":
        # Events go to fmMonitor.csv, but fm.log still records that this run
        # happened (start/options/end) so no Monitor run is ever invisible
        # to fm.log, per the framework's "every run" guarantee.
        _log_line(f"Monitor File Activity started (logging events to {MONITOR_CSV})")
        _log_line(f"  Folder: {folder}")
        _log_line(f"  Recursive: {'Yes' if recursive else 'No'}   Extensions: {ext_disp}")

    mlog = _MonitorLog(output, folder, recursive, exts)
    if mlog._fh is None:
        print(color_text(f"  ⚠ Could not open {log_disp} — events will show on screen only.",
                         fg=YELLOW))
    baseline = _monitor_snapshot(folder, recursive, exts)

    def poll_once():
        nonlocal baseline
        snap = _monitor_snapshot(folder, recursive, exts)
        events = _diff_snapshots(baseline, snap)
        for event, rel, size in events:
            ts = _monitor_ts()
            counts[event] += 1
            colored = color_text(f"{event:<9}", fg=_MONITOR_EVENT_COLORS[event], style=BOLD)
            size_disp = f"  {DIM}({fmt_size(size)}){RESET}" if size is not None else ""
            print(f"  {DIM}{ts}{RESET}  {colored} {rel}{size_disp}")
            mlog.event(ts, event, rel, folder)
        baseline = snap
        if events and is_tty and frame_rows:
            _monitor_update_footer(counts, frame_rows)

    sigwinch_prev = None
    if is_tty:
        import signal

        def _on_resize(signum, frame):
            nonlocal frame_rows
            new_rows = _monitor_draw_frame(folder, recursive, exts, log_disp, profile_name, counts)
            if new_rows:
                frame_rows = new_rows

        sigwinch_prev = signal.signal(signal.SIGWINCH, _on_resize)

    try:
        if is_tty:
            import termios, tty, select
            fd = sys.stdin.fileno()
            old_attr = termios.tcgetattr(fd)
            try:
                tty.setcbreak(fd)
                while True:
                    r, _, _ = select.select([fd], [], [], MONITOR_POLL_SECONDS)
                    if r:
                        ch = os.read(fd, 1)
                        if ch in (b"q", b"Q", b"\x1b"):
                            break
                    poll_once()
            finally:
                termios.tcsetattr(fd, termios.TCSADRAIN, old_attr)
        else:
            while True:
                time.sleep(MONITOR_POLL_SECONDS)
                poll_once()
    except KeyboardInterrupt:
        pass
    finally:
        if is_tty:
            import signal
            signal.signal(signal.SIGWINCH, sigwinch_prev if sigwinch_prev else signal.SIG_DFL)
            if frame_rows:
                _frm_cleanup(frame_rows)
        total = sum(counts.values())
        summary = (f"Stopped — {total} event(s): {counts['CREATED']} created, "
                   f"{counts['MODIFIED']} modified, {counts['DELETED']} deleted.")
        mlog.close(summary)
        if output == "csv":
            _log_line(f"Monitor File Activity finished — {summary}")

    print()
    print(color_text(f"  {summary}", fg=WHITE, style=BOLD))
    if mlog.opened:
        print(f"  {DIM}Activity logged to {log_disp}{RESET}")


# =============================================================================
# SYNC  (one-way push; dry-run by default)
# =============================================================================
def _load_config_profiles(key):
    """Return (profiles, error) — a profile list (e.g. syncProfiles,
    monitorProfiles) from fmConfig.json. A missing config file or a missing
    key simply means no profiles (empty list, no error); a broken file
    returns an error string."""
    try:
        with open(CONFIG_FILE, "r") as fh:
            cfg = json.load(fh)
    except OSError:
        return [], ""
    except ValueError as e:
        return [], f"fmConfig.json is not valid JSON: {e}"
    profs = cfg.get(key, [])
    if not isinstance(profs, list):
        return [], f"fmConfig.json: {key} must be a list"
    return [p for p in profs if isinstance(p, dict)], ""


def _load_sync_profiles():
    return _load_config_profiles("syncProfiles")


def _load_monitor_profiles():
    return _load_config_profiles("monitorProfiles")


def _load_compare_profiles():
    return _load_config_profiles("compareProfiles")


def _save_config_profiles(key, profiles):
    """Write `profiles` back to fmConfig.json under `key`, preserving every
    other setting in the file (auth, logZip, other profile lists, ...).
    Returns "" on success, or an error string. Refuses to touch the file if
    it exists but isn't valid JSON, so a save can't silently clobber it."""
    try:
        with open(CONFIG_FILE, "r") as fh:
            cfg = json.load(fh)
    except FileNotFoundError:
        cfg = {}
    except ValueError as e:
        return f"fmConfig.json is not valid JSON — not saved: {e}"
    cfg[key] = profiles
    try:
        with open(CONFIG_FILE, "w") as fh:
            json.dump(cfg, fh, indent=2)
            fh.write("\n")
    except OSError as e:
        return f"Could not write fmConfig.json: {e}"
    return ""


def _scan_sync_side(root, recursive, exclude_hidden):
    """Scan one side of a sync: {relative path: (size, mtime)}.

    Junk files (.DS_Store, desktop.ini) and excluded folders ($RECYCLE.BIN)
    are always skipped; hidden files/folders only when exclude_hidden."""
    out = {}
    if recursive:
        for dp, dns, fns in os.walk(root):
            prune_dirs(dns)
            if exclude_hidden:
                dns[:] = [d for d in dns if not d.startswith(".")]
            for fn in fns:
                if is_excluded_file(fn):
                    continue
                if exclude_hidden and fn.startswith("."):
                    continue
                full = os.path.join(dp, fn)
                try:
                    st = os.stat(full)
                except OSError:
                    continue
                out[os.path.relpath(full, root)] = (st.st_size, st.st_mtime)
    else:
        try:
            names = os.listdir(root)
        except OSError:
            return out
        for fn in names:
            full = os.path.join(root, fn)
            if not os.path.isfile(full) or is_excluded_file(fn):
                continue
            if exclude_hidden and fn.startswith("."):
                continue
            try:
                st = os.stat(full)
            except OSError:
                continue
            out[fn] = (st.st_size, st.st_mtime)
    return out


def _sync_direction_desc(direction):
    """Human-readable description of a sync direction, e.g.
    'A → B  (push new/updated files from A to B)' or the two-way
    description for 'both'. Shared by the on-screen header, the stationary
    frame, and the activity-log start entry so the wording never drifts."""
    if direction == "both":
        return "A ↔ B  (two-way: each side receives the other's new/updated files)"
    src_lbl, dst_lbl = ("B", "A") if direction == "b2a" else ("A", "B")
    return f"{src_lbl} → {dst_lbl}  (push new/updated files from {src_lbl} to {dst_lbl})"


def _sync_plan(src_map, dst_map, conflict):
    """Decide what to push src → dst. Returns (new_rels, upd_rels, skipped_zero).

    A file that exists on both sides is only copied when the source wins the
    conflict rule: 'newest' = source modified time is newer (1s tolerance for
    coarse filesystem timestamps), 'largest' = source is bigger.

    Zero-byte source files are never copied (usually incomplete/placeholder
    files) and are counted in skipped_zero instead."""
    new_rels, upd_rels = [], []
    skipped_zero = 0
    for rel in sorted(src_map):
        s_size, s_mtime = src_map[rel]
        if s_size == 0:
            skipped_zero += 1
            continue
        if rel not in dst_map:
            new_rels.append(rel)
            continue
        d_size, d_mtime = dst_map[rel]
        if conflict == "largest":
            if s_size > d_size:
                upd_rels.append(rel)
        else:                                   # newest (default)
            if s_mtime > d_mtime + 1:
                upd_rels.append(rel)
    return new_rels, upd_rels, skipped_zero


def _conflict_label(conflict):
    """Short display label for the Sync 'if a file exists on both sides' rule."""
    if conflict == "largest":
        return "Largest wins"
    if conflict == "manual":
        return "Manual (per file)"
    return "Newest wins"


def _fmt_mtime(ts):
    """Modified time in the display format m/d/yy h:mm am."""
    try:
        dt = datetime.fromtimestamp(ts)
        return dt.strftime("%-m/%-d/%y %-I:%M ") + dt.strftime("%p").lower()
    except (OSError, OverflowError, ValueError):
        return "—"


def _sync_manual_review(rels, src_map, dst_map, src_lbl, dst_lbl):
    """Walk the caller through each file that exists on both sides and
    differs (in size and/or modified time), asking Ignore or Select (copy)
    for each one. [IA]/[SA] apply that same choice to every remaining file
    without asking again. The differing Size/Modified values are colored so
    the mismatch is obvious at a glance.

    Returns the sorted list of selected (to-copy) relative paths, or None if
    the user cancelled the whole sync ([Q] or ESC)."""
    selected = []
    bulk = None   # once set to 'i' or 's', silently applies to the rest
    total = len(rels)
    w = get_width()
    rule_w = min(max(w - 2, 20), 78)

    print()
    print(color_text(f"  {total} file(s) exist on both sides and differ — "
                     f"decide each one:", fg=BRIGHT_CYAN, style=BOLD))

    for i, rel in enumerate(rels, 1):
        s_size, s_mtime = src_map[rel]
        d_size, d_mtime = dst_map[rel]

        if bulk is not None:
            if bulk == "s":
                selected.append(rel)
            continue

        size_diff = s_size != d_size
        date_diff = abs(s_mtime - d_mtime) > 1

        size_a_txt = f"{fmt_size(s_size):>12}"
        size_b_txt = f"{fmt_size(d_size):>12}"
        date_a_txt = f"{_fmt_mtime(s_mtime):<17}"
        date_b_txt = f"{_fmt_mtime(d_mtime):<17}"
        if size_diff:
            size_a_txt = color_text(size_a_txt, fg=YELLOW)
            size_b_txt = color_text(size_b_txt, fg=YELLOW)
        if date_diff:
            date_a_txt = color_text(date_a_txt, fg=YELLOW)
            date_b_txt = color_text(date_b_txt, fg=YELLOW)

        print()
        print(f"  {DIM}{'-' * rule_w}{RESET}")
        print(color_text(f"  Reviewing conflicts: {i} of {total}", fg=BRIGHT_CYAN, style=BOLD))
        print(f"  {WHITE}{rel}{RESET}")
        print()
        print(f"  {'':<16}{'Size':>12}   {'Modified':<17}")
        print(f"  {src_lbl + ' (source)':<16}{size_a_txt}   {date_a_txt}")
        print(f"  {dst_lbl + ' (dest)':<16}{size_b_txt}   {date_b_txt}")
        print()

        while True:
            ch = menu_read(color_text(
                "  [I] Ignore  [S] Select  [IA] Ignore All  [SA] Select All   Option: ",
                fg=CYAN, style=BOLD))
            if ch == "ESC":
                ch = "q"
            if ch in ("i", "ignore"):
                break
            if ch in ("s", "select"):
                selected.append(rel)
                break
            if ch == "ia":
                bulk = "i"
                break
            if ch == "sa":
                bulk = "s"
                selected.append(rel)
                break
            if ch in ("q", "quit", "cancel"):
                print(color_text("  Cancelled — nothing copied.", fg=YELLOW))
                return None
            print(color_text("  Please enter I, S, IA, or SA (or Q to cancel the sync).", fg=RED))

    print()
    print(color_text(f"  Reviewed {total} file(s): {len(selected)} selected, "
                     f"{total - len(selected)} ignored.", fg=BRIGHT_CYAN, style=BOLD))
    return sorted(selected)


def _sync_plan_manual(src_map, dst_map, src_lbl, dst_lbl):
    """Manual conflict mode: files that exist only in src are queued as NEW
    automatically (nothing to decide). Files that exist on both sides and
    differ in size or modified time are reviewed one at a time via
    _sync_manual_review() so the caller picks Ignore or Select for each.

    Returns (new_rels, upd_rels, skipped_zero) — same shape as _sync_plan()
    — or (None, None, skipped_zero) if the user cancelled the review."""
    new_rels = []
    conflicts = []
    skipped_zero = 0
    for rel in sorted(src_map):
        s_size, s_mtime = src_map[rel]
        if s_size == 0:
            skipped_zero += 1
            continue
        if rel not in dst_map:
            new_rels.append(rel)
            continue
        d_size, d_mtime = dst_map[rel]
        if s_size != d_size or abs(s_mtime - d_mtime) > 1:
            conflicts.append(rel)

    if not conflicts:
        return new_rels, [], skipped_zero

    upd_rels = _sync_manual_review(conflicts, src_map, dst_map, src_lbl, dst_lbl)
    if upd_rels is None:
        return None, None, skipped_zero
    return new_rels, upd_rels, skipped_zero


def _sync_screen(folder_a, folder_b, direction, recursive, conflict,
                 exclude_hidden, live_requested, assume_yes, profile_name, runlog):
    # Folder validation and "nothing to copy" are short, one-shot messages —
    # no stationary frame needed for those, just the standard screen header.
    screen("Sync Folders")
    dry_mode = ask_dry_mode(live_requested)
    print()
    print_mode_line(dry_mode)
    if profile_name:
        print(f"  {YELLOW}Profile{RESET}  : {profile_name}")
    print(f"  {YELLOW}Folder A{RESET} : {folder_a}")
    print(f"  {YELLOW}Folder B{RESET} : {folder_b}")
    print(f"  {YELLOW}Direction{RESET}: {_sync_direction_desc(direction)}")
    print(f"  {YELLOW}Options{RESET}  : Recursive: {'Yes' if recursive else 'No'}   "
          f"Conflict: {_conflict_label(conflict)}   "
          f"Hidden files: {'Excluded' if exclude_hidden else 'Included'}")
    print()

    if conflict == "manual" and direction == "both":
        print(color_text("  Manual conflict review is only available for one-way "
                         "syncs (A → B or B → A), not two-way.", fg=RED))
        runlog.finish("Cancelled — manual conflict review isn't available for two-way sync")
        return
    if conflict == "manual" and not sys.stdin.isatty():
        print(color_text("  Manual conflict review needs an interactive terminal "
                         "to ask about each file — not available here (piped "
                         "input or a non-interactive run). Use --conflict newest "
                         "or largest instead.", fg=RED))
        runlog.finish("Cancelled — manual conflict review needs an interactive terminal")
        return

    for path, lbl in ((folder_a, "A"), (folder_b, "B")):
        if not os.path.isdir(path):
            print(color_text(f"  Folder {lbl} is not a directory: {path}", fg=RED))
            runlog.finish(f"Cancelled — Folder {lbl} is not a directory: {path}")
            return

    a_map = _scan_sync_side(folder_a, recursive, exclude_hidden)
    b_map = _scan_sync_side(folder_b, recursive, exclude_hidden)

    # Each pass is one push: (src, dst, src_lbl, dst_lbl, src_map, dst_map,
    # new_rels, upd_rels, skipped_zero). Two-way = both passes; a file on both
    # sides can only win the conflict rule in one direction, so the two passes
    # never fight over the same file. Manual conflict mode is one-way only
    # (enforced above), so only one of these two branches ever runs when
    # conflict == "manual".
    passes = []
    if direction in ("a2b", "both"):
        if conflict == "manual":
            new_r, upd_r, sk = _sync_plan_manual(a_map, b_map, "A", "B")
            if new_r is None:
                runlog.finish("Cancelled — manual conflict review aborted")
                return
        else:
            new_r, upd_r, sk = _sync_plan(a_map, b_map, conflict)
        passes.append((folder_a, folder_b, "A", "B", a_map, b_map, new_r, upd_r, sk))
    if direction in ("b2a", "both"):
        if conflict == "manual":
            new_r, upd_r, sk = _sync_plan_manual(b_map, a_map, "B", "A")
            if new_r is None:
                runlog.finish("Cancelled — manual conflict review aborted")
                return
        else:
            new_r, upd_r, sk = _sync_plan(b_map, a_map, conflict)
        passes.append((folder_b, folder_a, "B", "A", b_map, a_map, new_r, upd_r, sk))

    total = sum(len(p[6]) + len(p[7]) for p in passes)
    copy_desc = ("between A and B" if direction == "both"
                 else f"from {passes[0][2]} to {passes[0][3]}")
    runlog.total_items = total
    runlog.detailed = total <= ACTION_LOG_DETAIL_THRESHOLD

    if total == 0:
        if direction == "both":
            print(color_text("  Nothing to copy — A and B already have each "
                             "other's new/updated files.", fg=GREEN, style=BOLD))
        else:
            print(color_text(f"  Nothing to copy — {passes[0][3]} already has every "
                             f"new/updated file from {passes[0][2]}.", fg=GREEN, style=BOLD))
        print(f"  {DIM}Scanned {len(a_map):,} file(s) in A, "
              f"{len(b_map):,} in B.{RESET}")
        for p in passes:
            if p[8]:
                print(f"  {DIM}Skipped {p[8]:,} zero-byte file(s) in {p[2]} "
                      f"(never copied).{RESET}")
        runlog.finish("Nothing to copy — already in sync")
        return

    # There's a real plan to show — from here on the folders/direction/
    # options and the new/updated/size summary stay pinned in a stationary
    # header, with the file plan, prompt, and copy progress scrolling below.
    total_new   = sum(len(p[6]) for p in passes)
    total_upd   = sum(len(p[7]) for p in passes)
    total_bytes = sum(p[4][r][0] for p in passes for r in p[6] + p[7])

    is_tty = sys.stdin.isatty()
    frame_rows = header_line_count = None
    status = "Reviewing sync plan below..."

    if is_tty:
        frame_rows, header_line_count = _sync_draw_frame(
            folder_a, folder_b, direction, recursive, conflict, exclude_hidden,
            profile_name, total_new, total_upd, total_bytes, copy_desc, status,
            dry_mode=dry_mode)
        if frame_rows is None:
            runlog.finish("Cancelled — terminal window too small")
            pause_return()
            return
    else:
        print(color_text(f"  {total_new} new, {total_upd} updated — "
                         f"{fmt_size(total_bytes)} to copy {copy_desc}:",
                         fg=WHITE, style=BOLD))

    sigwinch_prev = None
    if is_tty:
        import signal

        def _on_resize(signum, frame):
            nonlocal frame_rows, header_line_count
            new_rows, new_hdr = _sync_draw_frame(
                folder_a, folder_b, direction, recursive, conflict, exclude_hidden,
                profile_name, total_new, total_upd, total_bytes, copy_desc, status,
                dry_mode=dry_mode)
            if new_rows:
                frame_rows, header_line_count = new_rows, new_hdr

        sigwinch_prev = signal.signal(signal.SIGWINCH, _on_resize)

    ps = None
    try:
        for p in passes:
            if p[8]:
                print(f"  {DIM}Skipped {p[8]:,} zero-byte file(s) in {p[2]} "
                      f"(never copied).{RESET}")

        # dry_mode (asked upfront, before scanning) gates the detailed per-file
        # NEW/UPDATE table below. The stationary frame above already shows the
        # new/updated/size totals either way. The final "Actually copy...?"
        # confirm always still happens regardless.
        show_file_list = dry_mode
        if not show_file_list:
            print()
            print(color_text(
                f"  Skipping file list — {total_new} new, {total_upd} updated, "
                f"{fmt_size(total_bytes)} to copy {copy_desc}.", fg=DIM))

        if show_file_list:
            for src, dst, src_lbl, dst_lbl, src_map, dst_map, new_rels, upd_rels, _sk in passes:
                if not new_rels and not upd_rels:
                    if direction == "both":
                        print()
                        print(color_text(f"  {src_lbl} → {dst_lbl}: nothing to copy.", fg=GREEN))
                    continue
                print()
                if direction == "both":
                    print(color_text(f"  {src_lbl} → {dst_lbl}:", fg=CYAN, style=BOLD))
                print(color_text(f"  {'Action':<7} {'Size':>10}  {'Modified':<17} File",
                                 fg=WHITE, style=BOLD))
                print(f"  {'-' * 7} {'-' * 10}  {'-' * 17} {'-' * 30}")
                upd_set = set(upd_rels)
                for rel in sorted(new_rels + upd_rels):
                    s_size, s_mtime = src_map[rel]
                    if rel in upd_set:
                        act = color_text(f"{'UPDATE':<7}", fg=YELLOW)
                        d_size, d_mtime = dst_map[rel]
                        tail = f"  {DIM}(replaces {fmt_size(d_size)}, {_fmt_mtime(d_mtime)}){RESET}"
                    else:
                        act = color_text(f"{'NEW':<7}", fg=GREEN)
                        tail = ""
                    print(f"  {act} {fmt_size(s_size):>10}  {_fmt_mtime(s_mtime):<17} {rel}{tail}")

        # Copy only on explicit opt-in — same safety model as Remove.
        print()
        do_copy = False
        if live_requested is None:                # interactive
            print(color_text("  This was a DRY RUN — nothing has been copied yet.",
                             fg=YELLOW, style=BOLD))
            do_copy = safe_confirm(f"  Actually copy these {total} file(s) "
                                   f"{copy_desc}?", default=False)
        elif live_requested is True:              # CLI --copy
            do_copy = True if assume_yes else safe_confirm(
                f"  Copy these {total} file(s) {copy_desc}?", default=False)
        else:                                     # CLI dry run
            print(color_text("  DRY RUN — nothing copied. Re-run with --copy to sync.",
                             fg=YELLOW, style=BOLD))
            return

        if not do_copy:
            print(color_text("  Cancelled — nothing copied.", fg=YELLOW))
            runlog.finish("Cancelled — nothing copied")
            return
        if not _run_countdown(label="Starting copy"):
            print(color_text("  Cancelled — nothing copied.", fg=YELLOW))
            runlog.finish("Cancelled — nothing copied")
            return

        if is_tty and frame_rows:
            status = f"Copying... 0/{total} file(s)"
            _sync_update_footer(frame_rows, status)

        ok = fail = 0
        copied_bytes = 0
        copied_so_far = 0
        stopped_early = False
        ps = _PauseStop()
        if ps.is_tty and total > 1:
            print(color_text("  Press [P] to pause, [Q] to stop early.", fg=DIM))
        for src, dst, src_lbl, dst_lbl, src_map, dst_map, new_rels, upd_rels, _sk in passes:
            if stopped_early:
                break
            for rel in new_rels + upd_rels:
                if ps.check() == "stop":
                    stopped_early = True
                    break
                s_path = os.path.join(src, rel)
                d_path = os.path.join(dst, rel)
                s_size = src_map[rel][0]
                # Show which file is copying *before* it starts — a single
                # large file (over USB to a slow external drive especially)
                # can take a while, and without this the footer just sits
                # at the previous count the whole time, looking frozen.
                if is_tty and frame_rows:
                    status = (f"Copying {copied_so_far + 1}/{total}: {rel} "
                              f"({fmt_size(s_size)})...")
                    _sync_update_footer(frame_rows, status)
                try:
                    os.makedirs(os.path.dirname(d_path), exist_ok=True)
                    shutil.copy2(s_path, d_path)  # copy2 preserves mtime for future 'newest' runs
                    ok += 1
                    copied_bytes += s_size
                    runlog.action(f"Copied {rel} ({fmt_size(s_size)}) {src_lbl} -> {dst_lbl}")
                except OSError as e:
                    fail += 1
                    print(color_text(f"  ✗ {src_lbl} → {dst_lbl} {rel}: {e}", fg=RED))
                    runlog.action(f"FAILED to copy {rel} {src_lbl} -> {dst_lbl}: {e}")
                copied_so_far += 1
                if is_tty and frame_rows:
                    pct = int(copied_bytes / total_bytes * 100) if total_bytes else 100
                    status = (f"Copying... {copied_so_far}/{total} file(s)"
                              f"  ({pct}% complete, {fmt_size(copied_bytes)}/{fmt_size(total_bytes)} complete"
                              + (f", {fail} failed" if fail else "") + ")")
                    _sync_update_footer(frame_rows, status)
        print()
        summary = f"Copied {ok} file(s) ({fmt_size(copied_bytes)}) {copy_desc}."
        if fail:
            summary += f" {fail} failed."
        if stopped_early:
            summary += f" Stopped early by user ({copied_so_far}/{total} processed)."
        runlog.finish(summary)
        report_result(fail == 0,
                      f"Copied {ok} file(s) ({fmt_size(copied_bytes)}) {copy_desc}.",
                      f"Copied {ok} file(s) ({fmt_size(copied_bytes)}) {copy_desc}, {fail} failed.")
    finally:
        if ps is not None:
            ps.close()
        if is_tty:
            import signal
            signal.signal(signal.SIGWINCH, sigwinch_prev if sigwinch_prev else signal.SIG_DFL)
            if frame_rows:
                _frm_cleanup(frame_rows)


def sync_folders(folder_a, folder_b, direction, recursive=True,
                 conflict="newest", exclude_hidden=True, live_requested=None,
                 assume_yes=False, profile_name=None):
    """Folder sync: push new/updated files from one folder into the other
    (A → B or B → A), or both ways at once (direction 'both' — each side
    receives the other's new/updated files; the conflict rule decides which
    copy wins when a file exists on both sides). Nothing is ever deleted,
    and nothing is copied until explicitly confirmed (interactive y, or
    --copy on the CLI) — the preview is a DRY RUN. The screen output is
    also appended to ~/Documents/log/fm.log."""
    runlog = _RunLog("Sync Folders", [
        ("Folder A", folder_a),
        ("Folder B", folder_b),
        ("Direction", _sync_direction_desc(direction)),
        ("Recursive", "Yes" if recursive else "No"),
        ("Conflict", _conflict_label(conflict)),
        ("Hidden Files", "Excluded" if exclude_hidden else "Included"),
    ])
    with _ActivityLog():
        _sync_screen(folder_a, folder_b, direction, recursive, conflict,
                     exclude_hidden, live_requested, assume_yes, profile_name, runlog)


def _parse_sync_direction(raw):
    """Normalize a profile/CLI direction value to a2b, b2a, or both."""
    d = str(raw).lower()
    if d in ("both", "2way", "twoway", "two-way", "ab", "a2b+b2a"):
        return "both"
    if d in ("btoa", "b2a"):
        return "b2a"
    return "a2b"


def _run_sync_profile(pr, live_requested=None, assume_yes=False):
    """Run one syncProfiles entry from fmConfig.json (defaults: AtoB,
    recursive, newest wins, hidden excluded). Still previews + confirms."""
    folder_a = clean_path(str(pr.get("folderA", "")))
    folder_b = clean_path(str(pr.get("folderB", "")))
    direction = _parse_sync_direction(pr.get("direction", "AtoB"))
    conflict = "largest" if str(pr.get("conflict", "newest")).lower() == "largest" else "newest"
    sync_folders(folder_a, folder_b, direction,
                 recursive=bool(pr.get("recursive", True)),
                 conflict=conflict,
                 exclude_hidden=bool(pr.get("excludeHidden", True)),
                 live_requested=live_requested, assume_yes=assume_yes,
                 profile_name=pr.get("name") or "(unnamed profile)")
    pause_return()


# =============================================================================
# ZIP
# =============================================================================
def _s3_client():
    """A boto3 S3 client using the AWS credentials saved in Admin Menu.
    Raises if boto3 is missing or credentials aren't configured — callers
    that want the print-and-return-None behavior should check
    _BOTO3_AVAILABLE / is_aws_configured() first (see _s3_push_file)."""
    aws_cfg = get_aws_config()
    return boto3.client(
        "s3",
        region_name=aws_cfg.get("region") or "us-east-1",
        aws_access_key_id=aws_cfg["accessKeyId"],
        aws_secret_access_key=aws_cfg["secretAccessKey"],
    )


def _s3_push_file(local_path, bucket, folder=""):
    """Upload one file to S3 using the AWS credentials saved in Admin Menu.
    bucket/folder are task-specific and NOT read from Admin config — every
    caller (Zip SubFolders' push prompt, the AWS S3 wizard) asks for them
    itself. Prints its own status line. Returns (True, object_key) on
    success, (False, None) on any error (missing boto3, missing
    credentials, or an upload failure)."""
    if not _BOTO3_AVAILABLE:
        print(color_text("        S3: boto3 not installed — skipping push. Run: pip install boto3", fg=RED))
        return False, None

    if not is_aws_configured():
        print(color_text("        S3: AWS credentials not configured — see Admin Menu. Skipping push.", fg=YELLOW))
        return False, None

    folder = (folder or "").strip("/")
    filename = os.path.basename(local_path)
    object_key = f"{folder}/{filename}" if folder else filename

    try:
        client = _s3_client()
        client.upload_file(local_path, bucket, object_key)
        print(color_text(f"        S3: Uploaded → s3://{bucket}/{object_key}", fg=GREEN))
        return True, object_key
    except Exception as e:
        print(color_text(f"        S3: Upload failed for {filename}: {e}", fg=RED))
        return False, None


def _s3_push_zip(zip_path, bucket, folder=""):
    """Convenience wrapper around _s3_push_file() for callers (Zip
    SubFolders) that only need a True/False result."""
    ok, _key = _s3_push_file(zip_path, bucket, folder)
    return ok


def zip_subfolders(target, dest=None, remove_after=False, push_to_s3=False, s3_bucket="", s3_folder=""):
    with _ActivityLog():
        screen("Zip SubFolders")
        print()
        target = clean_path(target)
        dest = clean_path(dest) if dest else os.getcwd()
        print(f"  {YELLOW}Source{RESET}: {target}")
        print(f"  {YELLOW}Dest{RESET}  : {dest}")
        print(f"  {YELLOW}Remove source after zip{RESET}: {'Yes' if remove_after else 'No'}")
        s3_note = f" — s3://{s3_bucket}/{s3_folder}".rstrip("/") if push_to_s3 else ""
        print(f"  {YELLOW}Push to AWS S3{RESET}: {'Yes' if push_to_s3 else 'No'}{s3_note}\n")

        if not os.path.isdir(target):
            print(color_text(f"  Not a directory: {target}", fg=RED)); pause_return(); return

        os.makedirs(dest, exist_ok=True)

        # Cleanup .DS_Store / desktop.ini
        junk_files = []
        for dp, dns, fns in os.walk(target):
            for fn in fns:
                if fn == ".DS_Store" or fn.lower() == "desktop.ini":
                    junk_files.append(os.path.join(dp, fn))

        junk = 0
        if junk_files:
            junk_runlog = _RunLog("Zip SubFolders — Junk Cleanup", [("Root", target)],
                                   total_items=len(junk_files))
            for jf in junk_files:
                rel = os.path.relpath(jf, target)
                try:
                    os.remove(jf)
                    junk += 1
                    junk_runlog.action(f"Removed {rel}")
                except OSError as e:
                    junk_runlog.action(f"FAILED to remove {rel}: {e}")
            junk_runlog.finish(f"Removed {junk} of {len(junk_files)} junk file(s)")
        print(color_text(f"  Cleaned {junk} .DS_Store/desktop.ini file(s).", fg=DIM))

        subs = immediate_subfolders(target)
        if not subs:
            print(color_text("  No subfolders to zip.", fg=YELLOW)); pause_return(); return

        if remove_after:
            remove_after = safe_confirm(color_text("  Remove each source folder after a successful zip?", fg=YELLOW), default=False)

        print()
        if not safe_confirm(f"  Zip these {len(subs)} folder(s)?", default=True):
            print(color_text("  Cancelled — nothing zipped.", fg=YELLOW))
            pause_return(); return
        if not _run_countdown(label="Starting zip"):
            print(color_text("  Cancelled — nothing zipped.", fg=YELLOW))
            pause_return(); return

        print(color_text(f"\n  Zipping {len(subs)} folder(s)…", fg=BRIGHT_CYAN, style=BOLD))
        runlog = _RunLog("Zip SubFolders", [("Source", target), ("Dest", dest)],
                          total_items=len(subs))
        ok = 0
        fail = 0
        removed = 0
        s3_ok = 0
        s3_fail = 0
        stopped_early = False
        ps = _PauseStop()
        if ps.is_tty and len(subs) > 1:
            print(color_text("  Press [P] to pause, [Q] to stop early.", fg=DIM))
        try:
            for i, name in enumerate(subs, 1):
                if ps.check() == "stop":
                    stopped_early = True
                    break
                zip_path = os.path.join(dest, name + ".zip")
                suffix = 2
                while os.path.exists(zip_path):
                    zip_path = os.path.join(dest, f"{name}-{suffix}.zip")
                    suffix += 1
                base = os.path.basename(zip_path)
                print(f"    {CYAN}[{i}/{len(subs)}]{RESET} {WHITE}{name:<40}{RESET}", end="", flush=True)
                rc = subprocess.run(
                    ["zip", "-r", zip_path, name],
                    cwd=target, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                ).returncode
                if rc == 0:
                    note = f" [{base}]" if base != name + ".zip" else ""
                    size_note = f"{fmt_size(file_size(zip_path))}{note}"
                    print(f" {BRIGHT_GREEN}OK{RESET} {DIM}({size_note}){RESET}")
                    ok += 1
                    action_msg = f"Zipped {name} -> {base} ({size_note})"
                    if push_to_s3:
                        if _s3_push_zip(zip_path, s3_bucket, s3_folder):
                            s3_ok += 1
                            action_msg += ", uploaded to S3"
                        else:
                            s3_fail += 1
                            action_msg += ", S3 upload FAILED"
                    if remove_after:
                        try:
                            shutil.rmtree(os.path.join(target, name)); removed += 1
                            action_msg += ", source removed"
                        except OSError:
                            pass
                    runlog.action(action_msg)
                else:
                    print(f" {BRIGHT_RED}FAILED{RESET}")
                    fail += 1
                    runlog.action(f"FAILED to zip {name}")
        finally:
            ps.close()

        print()
        summary = f"Zipped {ok} folder(s)."
        if remove_after:
            summary += f" Removed {removed} source(s)."
        if push_to_s3:
            summary += f" S3: {s3_ok} uploaded" + (f", {s3_fail} failed" if s3_fail else "") + "."
        if fail:
            summary += f" {fail} failed."
        if stopped_early:
            summary += f" Stopped early by user ({ok + fail}/{len(subs)} processed)."
        runlog.finish(summary)
        report_result(fail == 0 and s3_fail == 0,
                      summary,
                      f"Zipped {ok}, {fail} failed." + (f" S3: {s3_fail} failed." if push_to_s3 and s3_fail else ""))
        pause_return()


def zip_folder_to_one_file(folder, dest_zip_path, runlog=None):
    """Zip an entire folder (recursively) into ONE archive at dest_zip_path —
    the AWS S3 wizard's 'Zip to 1 File' step. Distinct from zip_subfolders(),
    which zips each immediate subfolder into its own archive; no existing FM
    feature zips a whole folder to a single file before this.

    Cleans .DS_Store/desktop.ini first (same as zip_subfolders()) and picks
    a collision-safe dest_zip_path if the target already exists. No
    screen()/prompt/pause_return() I/O — called directly from the wizard,
    which owns its own review/confirm/countdown flow.

    Returns (ok: bool, final_zip_path or None, file_count).
    """
    junk = 0
    for dp, dns, fns in os.walk(folder):
        for fn in fns:
            if fn == ".DS_Store" or fn.lower() == "desktop.ini":
                try:
                    os.remove(os.path.join(dp, fn))
                    junk += 1
                except OSError:
                    pass
    if runlog and junk:
        runlog.action(f"Cleaned {junk} .DS_Store/desktop.ini file(s) before zipping")

    dest_dir = os.path.dirname(dest_zip_path) or "."
    os.makedirs(dest_dir, exist_ok=True)
    base, ext = os.path.splitext(dest_zip_path)
    final_path = dest_zip_path
    suffix = 2
    while os.path.exists(final_path):
        final_path = f"{base}-{suffix}{ext}"
        suffix += 1

    file_count = 0
    try:
        with zipfile.ZipFile(final_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for dp, dns, fns in os.walk(folder):
                prune_dirs(dns)
                for fn in fns:
                    if fn.startswith(".") or is_excluded_file(fn):
                        continue
                    full = os.path.join(dp, fn)
                    arcname = os.path.relpath(full, folder)
                    zf.write(full, arcname)
                    file_count += 1
        if runlog:
            runlog.action(f"Zipped {folder} -> {os.path.basename(final_path)} ({file_count} file(s))")
        return True, final_path, file_count
    except Exception as e:
        if runlog:
            runlog.action(f"FAILED to zip {folder}: {e}")
        return False, None, 0


# ---- Zip viewing (ported from zipView.py) -----------------------------------
def zip_view(target):
    target = clean_path(target) if target else None
    if not target:
        screen("View Zip")
        print()
        target = ask_path("Zip file or folder path")
        if target:
            target = clean_path(target)
    if not target:
        return
    if os.path.isdir(target):
        _browse_zip_folder(target)
    else:
        screen("View Zip")
        print()
        _view_zip_file(target)
        pause_return()


def _browse_zip_folder(folder):
    while True:
        try:
            zips = sorted(
                f for f in os.listdir(folder)
                if not f.startswith(".") and f.lower().endswith(".zip")
                and os.path.isfile(os.path.join(folder, f))
            )
        except OSError as e:
            screen("View Zip"); print(color_text(f"\n  Cannot read folder: {e}", fg=RED)); pause_return(); return

        if not zips:
            screen(f"Select Zip — {os.path.basename(os.path.abspath(folder))}")
            print(color_text("\n  No zip files found here.", fg=YELLOW)); pause_return(); return

        # Reuse the standard arrow-navigable menu to pick a zip.
        options = [(f"{z}  ({fmt_size(file_size(os.path.join(folder, z)))})", "") for z in zips]
        choice = render_menu(
            f"Select Zip — {os.path.basename(os.path.abspath(folder))}", options,
            intro=f"Folder: {folder}   ({len(zips)} zip file(s))")
        if choice == "back":
            return
        screen("View Zip"); print()
        _view_zip_file(os.path.join(folder, zips[int(choice) - 1]))
        pause_return()


def _view_zip_file(path):
    if not os.path.isfile(path):
        print(color_text(f"  File not found: {path}", fg=RED)); return
    if not zipfile.is_zipfile(path):
        print(color_text(f"  Not a valid zip file: {path}", fg=RED)); return

    w = get_width()
    print(f"  {BOLD}{WHITE}{os.path.basename(path)}{RESET}   {DIM}{path}{RESET}\n")
    with zipfile.ZipFile(path, "r") as zf:
        entries = [e for e in zf.infolist() if not is_hidden(e.filename)]
        files = [e for e in entries if not e.is_dir()]
        dirs = [e for e in entries if e.is_dir()]

        name_w = max((len(e.filename) for e in entries), default=20)
        name_w = min(max(name_w, 20), max(w - 55, 20))
        hdr = f"  {'Name':<{name_w}}  {'Uncompressed':>13}  {'Compressed':>11}  {'Ratio':>6}  Modified"
        print(color_text(hdr, fg=WHITE, style=BOLD))
        print(f"  {DIM}{'-' * min(len(hdr), w)}{RESET}")

        tu = tc = 0
        for e in sorted(entries, key=lambda x: x.filename):
            name = e.filename[:name_w]
            try:
                mod = fmt_date(datetime(*e.date_time)) if e.date_time and e.date_time[0] >= 1980 else "—"
            except Exception:
                mod = "—"
            if e.is_dir():
                print(f"  {DIM}{YELLOW}{name:<{name_w}}{RESET}  {DIM}{'<DIR>':>13}  {'':>11}  {'':>6}  {mod}{RESET}")
                continue
            uncomp, comp = e.file_size, e.compress_size
            ratio = (1 - comp / uncomp) * 100 if uncomp > 0 else 0.0
            tu += uncomp; tc += comp
            col = GREEN if ratio > 20 else (CYAN if ratio > 0 else DIM)
            print(f"  {WHITE}{name:<{name_w}}{RESET}  {fmt_size(uncomp):>13}  "
                  f"{col}{fmt_size(comp):>11}{RESET}  {col}{ratio:5.1f}%{RESET}  {DIM}{mod}{RESET}")

        print(f"  {DIM}{'-' * min(len(hdr), w)}{RESET}")
        tr = (1 - tc / tu) * 100 if tu > 0 else 0.0
        print(f"  {BOLD}{'TOTAL (' + str(len(files)) + ' files, ' + str(len(dirs)) + ' dirs)':<{name_w}}"
              f"  {fmt_size(tu):>13}  {fmt_size(tc):>11}  {tr:5.1f}%{RESET}")
        print(f"\n  {DIM}Zip size on disk: {RESET}{BOLD}{fmt_size(file_size(path))}{RESET}")


def find_files_in_zip(zip_path, pattern):
    """Search filenames inside a single zip archive for a wildcard pattern
    (e.g. 'fan*.png'). Matches by filename only — contents are not searched."""
    screen("Find Files in Zip")
    print()
    print(f"  {YELLOW}Zip{RESET}: {zip_path}    {YELLOW}Pattern{RESET}: {pattern}\n")
    runlog = _RunLog("Find Files in Zip", [("Zip", zip_path), ("Pattern", pattern)])
    if not os.path.isfile(zip_path):
        print(color_text(f"  Not a file: {zip_path}", fg=RED))
        runlog.finish(f"Cancelled — not a file: {zip_path}")
        pause_return(); return
    if not zipfile.is_zipfile(zip_path):
        print(color_text(f"  Not a valid zip file: {zip_path}", fg=RED))
        runlog.finish(f"Cancelled — not a valid zip file: {zip_path}")
        pause_return(); return

    results = []
    extra = {}
    with zipfile.ZipFile(zip_path, "r") as zf:
        for e in zf.infolist():
            if e.is_dir() or is_hidden(e.filename):
                continue
            if fnmatch.fnmatch(os.path.basename(e.filename), pattern):
                results.append(e.filename)
                extra[e.filename] = fmt_size(e.file_size)
    _print_find_results(sorted(results), "files", extra)
    runlog.finish(f"{len(results)} file(s) found")
    if pause_rerun():
        find_files_in_zip(zip_path, pattern)


def find_files_in_zips_under_folder(root, pattern):
    """Recursively find every .zip file under root, then search each one's
    filenames for a wildcard pattern. Matches are grouped by the zip they were
    found in. Matches by filename only — contents are not searched."""
    screen("Find Files in Zip")
    print()
    print(f"  {YELLOW}Root{RESET}: {root}    {YELLOW}Pattern{RESET}: {pattern}\n")
    runlog = _RunLog("Find Files in Zip", [("Root", root), ("Pattern", pattern)])
    if not os.path.isdir(root):
        print(color_text(f"  Not a directory: {root}", fg=RED))
        runlog.finish(f"Cancelled — not a directory: {root}")
        pause_return(); return

    zip_paths = []
    for dp, dns, fns in os.walk(root):
        prune_dirs(dns)
        for fn in sorted(fns):
            if fn.startswith(".") or not fn.lower().endswith(".zip"):
                continue
            zip_paths.append(os.path.join(dp, fn))
    zip_paths.sort()

    total_matches = 0
    bad_zips = []
    for zp in zip_paths:
        if not zipfile.is_zipfile(zp):
            bad_zips.append(zp)
            continue
        matches = []
        extra = {}
        with zipfile.ZipFile(zp, "r") as zf:
            for e in zf.infolist():
                if e.is_dir() or is_hidden(e.filename):
                    continue
                if fnmatch.fnmatch(os.path.basename(e.filename), pattern):
                    matches.append(e.filename)
                    extra[e.filename] = fmt_size(e.file_size)
        if matches:
            print(f"  {BOLD}{WHITE}{zp}{RESET}")
            for m in sorted(matches):
                print(f"    {m}  {DIM}({extra[m]}){RESET}")
            print()
            total_matches += len(matches)

    if total_matches == 0:
        print(color_text(f"  No matching files found in {len(zip_paths)} zip file(s).", fg=YELLOW))
    else:
        print(color_text(f"  {total_matches} file(s) found across {len(zip_paths)} zip file(s).",
                          fg=BRIGHT_CYAN, style=BOLD))
    if bad_zips:
        print(color_text(f"  Skipped {len(bad_zips)} unreadable zip file(s).", fg=YELLOW))
    runlog.finish(f"{total_matches} file(s) found across {len(zip_paths)} zip(s)")
    if pause_rerun():
        find_files_in_zips_under_folder(root, pattern)


def find_files_in_zip_target(target, pattern):
    """Search filenames matching `pattern`. If target is a .zip file, search
    inside it directly; if it's a folder, recursively search every .zip found
    under it."""
    target = clean_path(target) if target else None
    if not target:
        return
    if os.path.isdir(target):
        find_files_in_zips_under_folder(target, pattern)
    else:
        find_files_in_zip(target, pattern)


# ---- DocInfo Manager Authentication (Admin Menu: Login / Logout) -----------
# Same secureAuth-backed pattern as Show Aliases, but with its own authKey so
# an FM login is independent of a Show Aliases login (separate sessions, same
# encrypted file).
DEFAULT_AUTH_SETTINGS = {
    "baseUrl":         "https://docinfo.cloudbox9.com/xhr",
    "loginEndpoint":   "login.php",
    "secureAuthFile":  "~/.cb9Auth.enc",
    "authKey":         "fm",
    "timeout":         30,
}


def _load_auth_settings():
    """Load the 'auth' block from fmConfig.json, falling back to
    DEFAULT_AUTH_SETTINGS for any missing keys (the block is optional)."""
    settings = dict(DEFAULT_AUTH_SETTINGS)
    try:
        with open(CONFIG_FILE, "r") as fh:
            cfg = json.load(fh)
        settings.update(cfg.get("auth", {}))
    except (OSError, ValueError):
        pass
    return settings


_AUTH_SETTINGS   = _load_auth_settings()
SECURE_AUTH_FILE = _AUTH_SETTINGS["secureAuthFile"]
AUTH_KEY         = _AUTH_SETTINGS["authKey"]
API_LOGIN_URL    = _AUTH_SETTINGS["baseUrl"].rstrip("/") + "/" + _AUTH_SETTINGS["loginEndpoint"]
API_TIMEOUT      = _AUTH_SETTINGS["timeout"]


def get_auth_config():
    """Get authentication config from the encrypted auth file."""
    try:
        return secureAuth.getAuthConfig(AUTH_KEY, SECURE_AUTH_FILE)
    except secureAuth.SecureAuthError as e:
        print(color_text(f"[WARN] {e}", fg=YELLOW))
        return {}


def save_auth_config(auth_data):
    """Save authentication config to the encrypted auth file."""
    try:
        if secureAuth.saveAuthConfig(AUTH_KEY, auth_data, SECURE_AUTH_FILE):
            return True
        print(color_text(f"[ERROR] Failed to save {SECURE_AUTH_FILE}", fg=RED))
    except secureAuth.SecureAuthError as e:
        print(color_text(f"[ERROR] {e}", fg=RED))
    return False


def get_api_token():
    """Get the current DocInfo Manager API token, if any."""
    return get_auth_config().get("token")


def is_token_valid():
    """Check whether the saved token exists and is not expired."""
    token = get_api_token()
    if not token:
        return False
    expires_str = get_auth_config().get("tokenExpires")
    if expires_str:
        try:
            expires = datetime.strptime(expires_str, "%Y-%m-%d %H:%M:%S")
            if datetime.now() > expires:
                return False
        except ValueError:
            pass
    return True


def authenticate():
    """Prompt for DocInfo Manager credentials and obtain an API token."""
    print(f"\n{BOLD}{CYAN}DocInfo Manager Login{RESET}")
    print("-" * 40)
    print(f"A token will be saved encrypted to {SECURE_AUTH_FILE} (valid for 90 days).\n")

    username = _read_line_esc(f"{YELLOW}Email/Username: {RESET}", "")
    if not username:
        _log_line("DocInfo Manager Login cancelled — no username entered")
        return False
    password = getpass.getpass(f"{YELLOW}Password: {RESET}")
    if not password:
        _log_line(f"DocInfo Manager Login cancelled — no password entered (user: {username})")
        return False

    runlog = _RunLog("DocInfo Manager Login", [("Username", username)])
    print(f"\n{CYAN}Authenticating...{RESET}")
    try:
        post_data = urllib.parse.urlencode({
            "usrname":          username,
            "pwd":              password,
            "returnToken":      "1",
            "tokenDescription": "FM CLI",
        }).encode("utf-8")
        req = urllib.request.Request(API_LOGIN_URL, data=post_data)
        req.add_header("User-Agent", f"FM/{VERSION}")
        req.add_header("Content-Type", "application/x-www-form-urlencoded")
        with urllib.request.urlopen(req, timeout=API_TIMEOUT) as response:
            data = json.loads(response.read().decode("utf-8"))

        if str(data.get("success")) == "1" and data.get("apiToken"):
            auth_config = get_auth_config()
            auth_config["token"]             = data["apiToken"]
            auth_config["tokenExpires"]      = data.get("tokenExpires")
            auth_config["authenticatedUser"] = username
            auth_config["userId"]            = data.get("userId")
            auth_config["userAltId"]         = data.get("userAltId")
            if save_auth_config(auth_config):
                print(color_text("\nAuthentication successful!", fg=GREEN))
                print(f"User: {username} (userId: {data.get('userId', '?')})")
                print(f"Expires: {data.get('tokenExpires', 'Unknown')}")
                runlog.finish(f"Authentication successful (user: {username}, "
                               f"userId: {data.get('userId', '?')}, "
                               f"expires: {data.get('tokenExpires', 'Unknown')})")
                return True
            runlog.finish("Authentication succeeded but failed to save token locally")
            return False
        print(color_text(f"[ERROR] Authentication failed: {data.get('msg', 'Unknown error')}", fg=RED))
        runlog.finish(f"Authentication failed: {data.get('msg', 'Unknown error')}")
        return False
    except Exception as e:
        print(color_text(f"[ERROR] Authentication failed: {e}", fg=RED))
        runlog.finish(f"Authentication failed: {e}")
        return False


def clear_authentication():
    """Log out — clear the saved DocInfo Manager token."""
    runlog = _RunLog("DocInfo Manager Logout")
    if secureAuth.clearAuthConfig(AUTH_KEY, SECURE_AUTH_FILE):
        print(color_text(f"Logged out (cleared {SECURE_AUTH_FILE}).", fg=GREEN))
        runlog.finish("Logged out successfully")
    else:
        print(color_text("Failed to clear authentication.", fg=RED))
        runlog.finish("Logout FAILED")


def ensure_authenticated():
    """Ensure we have a valid DocInfo Manager token, prompting for login if
    needed. Returns True once authenticated (or already was), False if the
    user declined/failed to log in."""
    if is_token_valid():
        return True
    print(color_text("\nNo valid DocInfo Manager token found.", fg=YELLOW))
    return authenticate()


# ---- AWS S3 Push Credentials (Admin Menu) -----------------------------------
# Same secureAuth-backed encrypted store as the DocInfo Manager login, under
# its own authKey — used by Zip SubFolders' optional "Push files to AWS S3".
AWS_AUTH_KEY = "fmAwsS3"


def get_aws_config():
    """Get the saved AWS S3 push credentials from the encrypted auth file."""
    try:
        return secureAuth.getAuthConfig(AWS_AUTH_KEY, SECURE_AUTH_FILE)
    except secureAuth.SecureAuthError as e:
        print(color_text(f"[WARN] {e}", fg=YELLOW))
        return {}


def save_aws_config(aws_data):
    """Save AWS S3 push credentials to the encrypted auth file."""
    try:
        if secureAuth.saveAuthConfig(AWS_AUTH_KEY, aws_data, SECURE_AUTH_FILE):
            return True
        print(color_text(f"[ERROR] Failed to save {SECURE_AUTH_FILE}", fg=RED))
    except secureAuth.SecureAuthError as e:
        print(color_text(f"[ERROR] {e}", fg=RED))
    return False


def clear_aws_config():
    """Remove the saved AWS S3 push credentials."""
    runlog = _RunLog("AWS S3 Credentials Clear")
    if secureAuth.clearAuthConfig(AWS_AUTH_KEY, SECURE_AUTH_FILE):
        print(color_text(f"AWS S3 credentials cleared (removed from {SECURE_AUTH_FILE}).", fg=GREEN))
        runlog.finish("AWS S3 credentials cleared")
    else:
        print(color_text("Failed to clear AWS S3 credentials.", fg=RED))
        runlog.finish("Clear FAILED")


def is_aws_configured():
    """Whether enough AWS S3 credentials are saved to attempt an upload.
    Bucket isn't part of this check — it's a per-task value, not stored
    here (see set_aws_credentials())."""
    cfg = get_aws_config()
    return bool(cfg.get("accessKeyId") and cfg.get("secretAccessKey"))


def set_aws_credentials():
    """Prompt for and save AWS S3 credentials (Admin Menu). Just the auth
    triple — Access Key ID, Secret Access Key, Region. The bucket and an
    optional folder/prefix are NOT stored here; they're task-specific and
    entered where the upload actually happens (Zip SubFolders' 'Push files
    to AWS S3?' prompt, the AWS S3 wizard's step 8)."""
    print(f"\n{BOLD}{CYAN}AWS S3 Credentials{RESET}")
    print("-" * 40)
    print(f"Used by Zip SubFolders' optional 'Push files to AWS S3' and the AWS S3 menu. Saved encrypted to {SECURE_AUTH_FILE}.")
    print(color_text("Press Enter on any field to keep its current saved value.\n", style=DIM))

    existing = get_aws_config()

    access_key = _read_line_esc(color_text("AWS Access Key ID: ", fg=YELLOW), "") or existing.get("accessKeyId", "")
    secret_key = getpass.getpass(color_text("AWS Secret Access Key: ", fg=YELLOW)) or existing.get("secretAccessKey", "")

    region_default = existing.get("region") or "us-east-1"
    region = _read_line_esc(color_text(f"AWS Region [{region_default}]: ", fg=YELLOW), "") or region_default

    if not access_key or not secret_key:
        print(color_text("\nCancelled — Access Key ID and Secret Access Key are both required.", fg=YELLOW))
        _log_line("AWS S3 Credentials update cancelled — missing required field(s)")
        return

    runlog = _RunLog("AWS S3 Credentials Update", [("Region", region)])
    if save_aws_config({
        "accessKeyId":     access_key,
        "secretAccessKey": secret_key,
        "region":          region,
    }):
        print(color_text("\nAWS S3 credentials saved.", fg=GREEN))
        runlog.finish("AWS S3 credentials saved")
    else:
        runlog.finish("AWS S3 credentials save FAILED")


# ---- Common Folders (Admin Menu: Add/Edit/Delete) --------------------------

def _edit_common_folder_flow(existing, folders, idx):
    """Prompt for a common folder's name + path (pre-filled with `existing`'s
    values when editing) and save it into `folders` at `idx` (None = append
    a new one). The name can't contain spaces — it's the short label shown
    in ask_folder()'s '+' picker."""
    print()
    default_name = existing.get("name", "") if existing else ""
    name = default_name
    while True:
        typed = ask("Folder name (no spaces)", default=name).strip()
        if not typed:
            if not name:
                print(color_text("  Cancelled — a name is required.", fg=YELLOW))
                return
            break
        if " " in typed:
            print(color_text("  Name can't contain spaces — try again (e.g. use - or _).", fg=YELLOW))
            continue
        name = typed
        break

    default_path = existing.get("path", "") if existing else ""
    path = ask_folder("Folder path", default=default_path, must_exist=True)
    if not path:
        if not default_path:
            print(color_text("  Cancelled — a folder is required.", fg=YELLOW))
            return
        path = default_path

    folder = {"name": name, "path": path}
    if idx is None:
        folders.append(folder)
    else:
        folders[idx] = folder
    err = _save_config_profiles("commonFolders", folders)
    if err:
        print(color_text(f"  ⚠ {err}", fg=RED))
    else:
        verb = "Updated" if idx is not None else "Saved"
        print(color_text(f"  ✓ {verb} common folder '{name}'.", fg=GREEN))


def _delete_common_folder_flow(folder, folders, idx):
    """Confirm, then remove `folders[idx]` and save."""
    name = folder.get("name") or "(unnamed)"
    if not safe_confirm(f"  Delete common folder '{name}'?", default=False):
        print(color_text("  Cancelled — nothing deleted.", fg=YELLOW))
        return
    del folders[idx]
    err = _save_config_profiles("commonFolders", folders)
    if err:
        print(color_text(f"  ⚠ {err}", fg=RED))
    else:
        print(color_text(f"  ✓ Deleted common folder '{name}'.", fg=GREEN))


def manage_common_folders_menu():
    """Admin Menu -> Manage Common Folders: add/edit/delete the folders
    offered by the '+' shortcut at every folder prompt in the app."""
    while True:
        folders, err = _load_common_folders()
        if err:
            screen("Manage Common Folders")
            print()
            print(color_text(f"  ⚠ {err}", fg=RED))
            pause_return()
            return

        options = [("Add Common Folder", "Save a new folder to the '+' picker.")]
        for f in folders:
            name = f.get("name") or "(unnamed)"
            path = f.get("path", "?")
            options.append((f"{name} - {path}", f"Edit or delete '{name}'."))
        ch = render_menu(
            "Manage Common Folders", options,
            intro="These folders appear when you type '+' and press Enter "
                  "at any folder prompt.")
        if ch == "back":
            return

        picked = int(ch) - 1
        screen("Manage Common Folders")
        if picked == 0:
            _edit_common_folder_flow(None, folders, None)
            pause_return()
            continue

        idx = picked - 1
        folder = folders[idx]
        name = folder.get("name") or "(unnamed)"
        action = render_menu(
            f"Manage Common Folders — {name}",
            [("Edit", f"Change the name/path of '{name}'."),
             ("Delete", f"Remove '{name}' from the common folders list.")])
        if action == "back":
            continue
        screen("Manage Common Folders")
        if action == "1":
            _edit_common_folder_flow(folder, folders, idx)
        elif action == "2":
            _delete_common_folder_flow(folder, folders, idx)
        pause_return()


def admin_menu():
    """Admin Menu — DocInfo Manager login/logout, AWS S3 push credentials,
    and Manage Common Folders (same header as every other FM menu). [Q/ESC]
    Back is handled by render_menu()."""
    while True:
        authenticated = is_token_valid()
        if authenticated:
            auth_config = get_auth_config()
            user    = auth_config.get("authenticatedUser", "Unknown")
            expires = auth_config.get("tokenExpires", "Unknown")
            docinfo_status = f"authenticated as {user} — token expires {expires}"
        else:
            docinfo_status = "not authenticated — login required for Log Zip File Contents / Zip SubFolders & Log"

        aws_ready = is_aws_configured()
        aws_cfg   = get_aws_config()
        if aws_ready:
            aws_status = f"configured — region {aws_cfg.get('region') or 'us-east-1'} (bucket entered per task)"
        else:
            aws_status = "not configured — required for Zip SubFolders' optional Push to S3, and the AWS S3 menu"

        intro = f"DocInfo Manager: {docinfo_status}\n  AWS S3: {aws_status}"

        entries = []
        if authenticated:
            entries.append(("Logout of DocInfo Manager",
                             "Clear the saved DocInfo Manager login token.",
                             clear_authentication))
        else:
            entries.append(("Login to DocInfo Manager",
                             "Enter your DocInfo Manager email and password to "
                             "obtain an API token, saved encrypted for 90 days.",
                             authenticate))

        if aws_ready:
            entries.append(("Update AWS S3 Credentials",
                             "Replace the saved AWS Access Key ID / Secret Access "
                             "Key / Region, encrypted in the same secure store as "
                             "the DocInfo Manager login. Bucket and folder/prefix "
                             "aren't stored here — they're entered per task.",
                             set_aws_credentials))
            entries.append(("Clear AWS S3 Credentials",
                             "Remove the saved AWS credentials. Zip SubFolders' "
                             "optional Push to S3, and the AWS S3 menu, become "
                             "unavailable until re-entered.",
                             clear_aws_config))
        else:
            entries.append(("Set AWS S3 Credentials",
                             "Enter an AWS Access Key ID, Secret Access Key, and "
                             "Region, saved encrypted, so Zip SubFolders can "
                             "optionally push the resulting zips to S3 and the "
                             "AWS S3 menu becomes usable. Bucket and folder/"
                             "prefix are entered per task, not stored here.",
                             set_aws_credentials))

        entries.append(("Manage Common Folders",
                         "Add, edit, or delete the folders offered by the "
                         "'+' shortcut at any folder prompt.",
                         manage_common_folders_menu))

        entries.append(("Manage Local Scripts",
                         "Add, edit, or delete the scripts offered on the "
                         "Main Menu -> Local Scripts screen (name + "
                         "execution string).",
                         manage_local_scripts_menu))

        ch = render_menu("Admin Menu", [(label, desc) for label, desc, _ in entries],
                          intro=intro)
        if ch == "back":
            return
        action = entries[int(ch) - 1][2]
        try:
            action()
        except EscCancelled:
            continue
        # manage_common_folders_menu() and manage_local_scripts_menu() are
        # full submenus that already pause after each add/edit/delete — an
        # extra pause here would just be a redundant keypress when the user
        # backs straight out of one of them.
        if action not in (manage_common_folders_menu, manage_local_scripts_menu):
            pause_return()


# ---- Log Zip File (CB9Inventory via DocInfo Manager API) --------------------
LOGZIP_EXTS = (".zip", ".tar")   # .gz is intentionally not supported


def _load_logzip_config():
    """Load fmConfig.json and return the logZip settings, or (None, error)."""
    try:
        with open(CONFIG_FILE, "r") as fh:
            cfg = json.load(fh)
    except OSError:
        return None, f"Config file not found: {CONFIG_FILE}"
    except ValueError as e:
        return None, f"Config file is not valid JSON: {e}"
    logzip = cfg.get("logZip", {})
    if not logzip.get("apiUrl") or not logzip.get("serverSecretKey"):
        return None, "fmConfig.json is missing logZip.apiUrl or logZip.serverSecretKey"
    return logzip, ""


def _sql_datetime(ts):
    try:
        return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return ""


def _file_md5(path, block_size=65536):
    """MD5 hash of a file's full contents, for duplicate detection (same
    approach as vlcmenu's get_file_hash/calculateFileHash)."""
    hasher = hashlib.md5()
    with open(path, "rb") as fh:
        for block in iter(lambda: fh.read(block_size), b""):
            hasher.update(block)
    return hasher.hexdigest()


def _archive_contents(path):
    """Return the file listing of a .zip or .tar archive as a list of dicts
    matching the zipFileContent columns. Directories are skipped."""
    contents = []
    if path.lower().endswith(".zip"):
        with zipfile.ZipFile(path, "r") as zf:
            for e in zf.infolist():
                if e.is_dir():
                    continue
                try:
                    mod = (datetime(*e.date_time).strftime("%Y-%m-%d %H:%M:%S")
                           if e.date_time and e.date_time[0] >= 1980 else "")
                except Exception:
                    mod = ""
                name = os.path.basename(e.filename.rstrip("/"))
                contents.append({
                    "fileName":         name[:255],
                    "filePath":         e.filename[:500],
                    "fileSizeBytes":    e.file_size,
                    "fileModifiedDate": mod,
                    "fileExtension":    os.path.splitext(name)[1].lstrip(".").lower()[:20],
                })
    else:  # .tar
        with tarfile.open(path, "r") as tf:
            for m in tf.getmembers():
                if not m.isfile():
                    continue
                name = os.path.basename(m.name.rstrip("/"))
                contents.append({
                    "fileName":         name[:255],
                    "filePath":         m.name[:500],
                    "fileSizeBytes":    m.size,
                    "fileModifiedDate": _sql_datetime(m.mtime) if m.mtime else "",
                    "fileExtension":    os.path.splitext(name)[1].lstrip(".").lower()[:20],
                })
    return contents


def _post_zip_log(logzip, path, contents, load_failed=False, s3_bucket="", s3_key="", upload_confirmed=False):
    """POST one archive's info + contents to the DocInfo Manager API.

    load_failed: set when the archive's contents couldn't be read (corrupt/
    not a zip) — the server still records the zipFile row (size, zero
    content rows) but with zipFileStatusId = 3 (Failed) instead of 1
    (Active), rather than the row being skipped entirely.

    s3_bucket/s3_key/upload_confirmed: set together when the AWS S3 wizard
    also uploaded this zip — records where it landed on the same zipFile
    row. Omitted (blank) on a routine log-only call, which leaves any
    previously-recorded upload untouched server-side (COALESCE).

    Sends logzip's serverId/docProjectId (fmConfig.json) on every call —
    tags the resulting docInfo.activityLog entry with FM as the source
    system (File Management / Ash Mac), not what got changed."""
    st = os.stat(path)
    created = getattr(st, "st_birthtime", st.st_mtime)
    try:
        zip_hash = _file_md5(path)
    except OSError:
        zip_hash = ""
    fields = {
        "serverSecretKey":         logzip["serverSecretKey"],
        "zipFileName":             os.path.basename(path),
        "zipFileFolder":           os.path.dirname(os.path.abspath(path))[:300],
        "zipFileSizeBytes":        st.st_size,
        "zipFileHash":             zip_hash,
        "zipFileCount":            len(contents),
        "zipFileCreatedDate":      _sql_datetime(created),
        "zipFileLastModifiedDate": _sql_datetime(st.st_mtime),
        "zipDate":                 _sql_datetime(st.st_mtime),
        "contents":                json.dumps(contents),
        "loadFailed":              "1" if load_failed else "0",
        "s3Bucket":                s3_bucket,
        "s3Key":                   s3_key,
        "uploadConfirmed":         "1" if upload_confirmed else "0",
        "serverId":                logzip.get("serverId") or "",
        "docProjectId":            logzip.get("docProjectId") or "",
    }
    req = urllib.request.Request(
        logzip["apiUrl"],
        data=urllib.parse.urlencode(fields).encode("utf-8"),
        method="POST")
    req.add_header("X-API-Token", get_api_token() or "")
    with urllib.request.urlopen(req, timeout=120) as resp:
        return json.loads(resp.read().decode("utf-8"))


def _post_zip_content_search(logzip, file_name_pattern, zip_file_name="", zip_file_folder="",
                              show_hidden=False, page_no=1, page_per_count=1000):
    """POST a filename search to the DocInfo Manager zipFileContentSearch API
    (searches zipFileContent across every archive already logged to
    CB9Inventory). Reuses the same host/serverSecretKey as logZip in
    fmConfig.json — the search endpoint lives alongside zipFileLog.php."""
    api_url = logzip["apiUrl"].replace("zipFileLog.php", "zipFileContentSearch.php")
    fields = {
        "serverSecretKey": logzip["serverSecretKey"],
        "fileNamePattern": file_name_pattern,
        "zipFileName":     zip_file_name,
        "zipFileFolder":   zip_file_folder,
        "showHidden":      "1" if show_hidden else "0",
        "pageNo":          page_no,
        "pagePerCount":    page_per_count,
    }
    req = urllib.request.Request(
        api_url,
        data=urllib.parse.urlencode(fields).encode("utf-8"),
        method="POST")
    req.add_header("X-API-Token", get_api_token() or "")
    with urllib.request.urlopen(req, timeout=60) as resp:
        return json.loads(resp.read().decode("utf-8"))


def _post_file_log(logzip, path, s3_bucket="", s3_key="", upload_confirmed=False,
                    load_failed=False, error_msg=""):
    """POST one individual file's info to the DocInfo Manager fileLog API —
    the fileLog counterpart to _post_zip_log(), for files the AWS S3 wizard
    uploads without zipping them first. Reuses the same host/serverSecretKey
    as logZip in fmConfig.json (api/fileLog.php lives alongside zipFileLog.php).
    Also sends logzip's serverId/docProjectId — see _post_zip_log()."""
    api_url = logzip["apiUrl"].replace("zipFileLog.php", "fileLog.php")
    st = os.stat(path)
    fields = {
        "serverSecretKey":      logzip["serverSecretKey"],
        "fileName":             os.path.basename(path),
        "fileFolder":           os.path.dirname(os.path.abspath(path))[:300],
        "fileSizeBytes":        st.st_size,
        "fileExtension":        os.path.splitext(path)[1].lstrip("."),
        "fileCreatedDate":      _sql_datetime(getattr(st, "st_birthtime", st.st_mtime)),
        "fileLastModifiedDate": _sql_datetime(st.st_mtime),
        "s3Bucket":             s3_bucket,
        "s3Key":                s3_key,
        "uploadConfirmed":      "1" if upload_confirmed else "0",
        "loadFailed":           "1" if load_failed else "0",
        "errorMsg":             error_msg,
        "serverId":             logzip.get("serverId") or "",
        "docProjectId":         logzip.get("docProjectId") or "",
    }
    req = urllib.request.Request(
        api_url,
        data=urllib.parse.urlencode(fields).encode("utf-8"),
        method="POST")
    req.add_header("X-API-Token", get_api_token() or "")
    with urllib.request.urlopen(req, timeout=60) as resp:
        return json.loads(resp.read().decode("utf-8"))


def _post_file_log_search(logzip, file_name_pattern, file_folder="", page_no=1, page_per_count=1000):
    """POST a filename search to the DocInfo Manager fileLogSearch API
    (searches fileLog — individual files logged without zipping). The
    fileLog counterpart to _post_zip_content_search()."""
    api_url = logzip["apiUrl"].replace("zipFileLog.php", "fileLogSearch.php")
    fields = {
        "serverSecretKey": logzip["serverSecretKey"],
        "fileNamePattern": file_name_pattern,
        "fileFolder":      file_folder,
        "pageNo":          page_no,
        "pagePerCount":    page_per_count,
    }
    req = urllib.request.Request(
        api_url,
        data=urllib.parse.urlencode(fields).encode("utf-8"),
        method="POST")
    req.add_header("X-API-Token", get_api_token() or "")
    with urllib.request.urlopen(req, timeout=60) as resp:
        return json.loads(resp.read().decode("utf-8"))


def log_zip_files(target, recursive=False):
    """Log a .zip/.tar file — or every one in a folder — to the CB9Inventory
    database via the DocInfo Manager API. A folder is scanned top-level only
    unless recursive=True, which walks every subfolder too (hidden folders
    excluded via prune_dirs(), same as Find Files in Zip)."""
    screen("Log Zip File Contents")
    print()

    if not ensure_authenticated():
        print(color_text("  DocInfo Manager login is required to log zip files.", fg=RED))
        _log_line("Log Zip File Contents cancelled — DocInfo Manager login required")
        pause_return(); return

    logzip, err = _load_logzip_config()
    if not logzip:
        print(color_text(f"  {err}", fg=RED))
        _log_line(f"Log Zip File Contents cancelled — {err}")
        pause_return(); return

    target = clean_path(target) if target else ""
    if not target:
        print(color_text("  No path entered.", fg=YELLOW))
        _log_line("Log Zip File Contents cancelled — no path entered")
        pause_return(); return

    if os.path.isfile(target):
        if not target.lower().endswith(LOGZIP_EXTS):
            print(color_text(f"  Not a .zip or .tar file: {target}", fg=RED))
            _log_line(f"Log Zip File Contents cancelled — not a .zip or .tar file: {target}")
            pause_return(); return
        archives = [target]
    elif os.path.isdir(target):
        if recursive:
            archives = []
            for dp, dns, fns in os.walk(target):
                prune_dirs(dns)
                for fn in fns:
                    if not fn.startswith(".") and fn.lower().endswith(LOGZIP_EXTS):
                        archives.append(os.path.join(dp, fn))
            archives.sort()
        else:
            archives = sorted(
                os.path.join(target, f) for f in os.listdir(target)
                if not f.startswith(".") and f.lower().endswith(LOGZIP_EXTS)
                and os.path.isfile(os.path.join(target, f))
            )
        if not archives:
            scope = "under" if recursive else "(top level) in"
            print(color_text(f"  No .zip or .tar files found {scope} {target}.", fg=YELLOW))
            _log_line(f"Log Zip File Contents cancelled — no .zip/.tar files found {scope} {target}")
            pause_return(); return
    else:
        print(color_text(f"  Path not found: {target}", fg=RED))
        _log_line(f"Log Zip File Contents cancelled — path not found: {target}")
        pause_return(); return

    LAST_PATHS["zip_log_target"] = target

    print(f"  {YELLOW}Target{RESET}: {target}")
    if os.path.isdir(target):
        print(f"  {YELLOW}Subfolders{RESET}: {'Included' if recursive else 'Not included'}")
    print(f"  {YELLOW}Archives to log{RESET}: {len(archives)}\n")

    runlog_fields = [("Target", target), ("Archives", len(archives))]
    if os.path.isdir(target):
        runlog_fields.insert(1, ("Subfolders", "Included" if recursive else "Not included"))
    runlog = _RunLog("Log Zip File Contents", runlog_fields, total_items=len(archives))
    ok = fail = load_failed_count = 0
    for path in archives:
        base = os.path.basename(path)
        print(f"  {WHITE}{base}{RESET} {DIM}({fmt_size(file_size(path))}){RESET} ... ", end="", flush=True)

        # If the archive can't be read, still log the zipFile row (size,
        # zero content rows) marked Failed instead of skipping it entirely.
        load_error = None
        try:
            contents = _archive_contents(path)
        except (OSError, ValueError, zipfile.BadZipFile, tarfile.TarError) as e:
            contents = []
            load_error = str(e)

        try:
            result = _post_zip_log(logzip, path, contents, load_failed=bool(load_error))
        except Exception as e:
            print(f"{BRIGHT_RED}FAILED{RESET} {DIM}{e}{RESET}")
            runlog.action(f"FAILED to log {base}: {e}")
            fail += 1
            continue

        if str(result.get("success", "0")) == "1":
            if load_error:
                load_failed_count += 1
                print(f"{BRIGHT_YELLOW}FAILED TO LOAD{RESET} "
                      f"{DIM}[zipFileId {result.get('zipFileId', 0)}] recorded as Failed — {load_error}{RESET}")
                runlog.action(f"{base} recorded as Failed (unreadable) [zipFileId {result.get('zipFileId', 0)}] — {load_error}")
            else:
                counts = (f"{len(contents)} files: "
                          f"+{result.get('contentInserted', 0)} "
                          f"~{result.get('contentUpdated', 0)} "
                          f"-{result.get('contentDeleted', 0)} "
                          f"={result.get('contentUnchanged', 0)}")
                print(f"{BRIGHT_GREEN}{result.get('action', 'logged').upper()}{RESET} "
                      f"{DIM}[zipFileId {result.get('zipFileId', 0)}] ({counts}){RESET}")
                runlog.action(f"{result.get('action', 'logged').capitalize()} {base} "
                               f"[zipFileId {result.get('zipFileId', 0)}] ({counts})")
                ok += 1
        else:
            print(f"{BRIGHT_RED}FAILED{RESET} {DIM}{result.get('msg', 'Unknown error')}{RESET}")
            runlog.action(f"FAILED to log {base}: {result.get('msg', 'Unknown error')}")
            fail += 1

    print()
    summary_ok = f"Logged {ok} archive(s)"
    if load_failed_count:
        summary_ok += f", {load_failed_count} recorded as Failed (unreadable)"
    summary_ok += " to CB9Inventory."
    summary_fail = f"Logged {ok}, {fail} failed."
    if load_failed_count:
        summary_fail += f" {load_failed_count} recorded as Failed (unreadable)."
    report_result(fail == 0, summary_ok, summary_fail)
    runlog.finish(summary_ok if fail == 0 else summary_fail)
    pause_return()


# =============================================================================
# INTERACTIVE MENUS
# =============================================================================
def menu_read(prompt="", default=""):
    """Read a menu selection with a hybrid input model.

    • ESC acts IMMEDIATELY — it does not require the user to press Enter.
    • Enter submits the current buffer; on an empty buffer it returns `default`.
    • All other keys (digits, letters) buffer until Enter is pressed, so
      multi-digit choices still work.

    Returns 'ESC', or the typed text (lower-cased, stripped / or `default`).
    Falls back to a normal line read when stdin is not an interactive TTY
    (e.g. piped input), where ESC cannot be detected instantly.
    """
    sys.stdout.write(prompt)
    sys.stdout.flush()

    # Non-TTY (piped) — behave like a normal line read.
    if not sys.stdin.isatty():
        try:
            line = sys.stdin.readline()
        except Exception:
            return "ESC"
        if line == "":                       # EOF
            return "ESC"
        val = line.strip().lower()
        return val if val else default

    try:
        import termios, tty, select
    except ImportError:                      # non-POSIX fallback
        val = input().strip().lower()
        return val if val else default

    fd = sys.stdin.fileno()
    old = termios.tcgetattr(fd)
    buf = ""
    try:
        tty.setraw(fd)
        while True:
            ch = os.read(fd, 1)
            if not ch:
                return "ESC"
            if ch == b"\x1b":                # ESC — could be a bare ESC or an arrow seq
                r, _, _ = select.select([fd], [], [], 0.03)
                if r:
                    os.read(fd, 2)           # swallow the arrow sequence, ignore
                    continue
                sys.stdout.write("\r\n"); sys.stdout.flush()
                return "ESC"
            if ch in (b"\r", b"\n"):         # Enter — submit
                sys.stdout.write("\r\n"); sys.stdout.flush()
                val = buf.strip().lower()
                return val if val else default
            if ch == b"\x03":                # Ctrl-C
                raise KeyboardInterrupt
            if ch in (b"\x7f", b"\x08"):     # Backspace
                if buf:
                    buf = buf[:-1]
                    sys.stdout.write("\b \b"); sys.stdout.flush()
                continue
            c = ch.decode("utf-8", "ignore")
            if c and c.isprintable():        # buffer printable chars (echo manually)
                buf += c
                sys.stdout.write(c); sys.stdout.flush()
    finally:
        termios.tcsetattr(fd, termios.TCSADRAIN, old)


def read_key():
    """Read ONE keypress in raw mode (TTY only), reacting instantly.

    Returns one of: 'UP', 'DOWN', 'LEFT', 'RIGHT', 'ENTER', 'ESC',
    'BACKSPACE', or the typed character.
    """
    import termios, tty, select
    fd = sys.stdin.fileno()
    old = termios.tcgetattr(fd)
    try:
        tty.setraw(fd)
        ch = os.read(fd, 1)
        if not ch:
            return "ESC"
        if ch == b"\x1b":                    # ESC alone, or an arrow-key sequence
            r, _, _ = select.select([fd], [], [], 0.03)
            if not r:
                return "ESC"
            seq = os.read(fd, 2)
            return {b"[A": "UP", b"[B": "DOWN", b"[C": "RIGHT", b"[D": "LEFT"}.get(seq, "ESC")
        if ch in (b"\r", b"\n"):
            return "ENTER"
        if ch == b"\x03":                    # Ctrl-C
            raise KeyboardInterrupt
        if ch in (b"\x7f", b"\x08"):
            return "BACKSPACE"
        return ch.decode("utf-8", "ignore")
    finally:
        termios.tcsetattr(fd, termios.TCSADRAIN, old)


def _opt_disabled_reason(opt):
    """Return the disabled-reason string for a menu option tuple, or None if
    it's enabled. Options are (label, description) or, when conditionally
    unavailable, (label, description, reason) — e.g. ("Log Zip File Contents", "...",
    "login required")."""
    return opt[2] if len(opt) > 2 and opt[2] else None


def render_menu(subtitle, options, is_main=False, intro=None, outro=None,
                help_note=None):
    """Render a CB9 menu and return the user's choice.

    Navigation (interactive TTY):
      • ↑/↓ arrows move the highlighted option; Enter selects it.
      • Typing a number also selects that option (multi-digit buffered).
      • H shows Help; Q or ESC go Back (submenu) / Exit (main) — instant.

    options   : list of (label, description) tuples. Description is shown by
                [H]. An option may instead be a (label, description, reason)
                triple — a truthy reason (e.g. "login required") grays out
                the option and blocks selection, showing that reason.
    intro     : optional context line shown above the options.
    outro     : optional context line(s) — str or list of str — shown below the
                options (e.g. the folders already entered).
    help_note : optional (label, description) tuple shown at the bottom of the
                [H] Help screen after the option descriptions (e.g. logging info).
    Returns str(number) for a selection, or 'back' (submenu) / 'quit' (main).
    Falls back to line input when stdin is not an interactive TTY (piped/CLI).
    """
    n = len(options)
    if n == 0:
        return "back"
    if not sys.stdin.isatty():
        return _render_menu_lines(subtitle, options, is_main, intro, outro,
                                  help_note)

    back_word = "Exit" if is_main else "Back"
    selected = 0        # highlighted row (starts on option 1, the default)
    buf = ""            # numeric typing buffer (for multi-digit menus)
    blocked_msg = ""    # set when a disabled option was picked; shown once
    while True:
        screen(subtitle)
        if intro:
            print(f"  {DIM}{intro}{RESET}")
        print()
        for i, opt in enumerate(options):
            label  = opt[0]
            reason = _opt_disabled_reason(opt)
            if reason:
                print(f"    {color_text(f'{i + 1}. {label}  ({reason})', style=DIM)}")
            elif i == selected:
                print(color_text(f"  ▶ {i + 1}. {label}", fg=BRIGHT_CYAN, style=BOLD))
            else:
                print(f"    {color_text(str(i + 1), fg=YELLOW)}. {label}")
        print()
        _print_menu_outro(outro)
        about_legend = "  [A] About" if is_main else ""
        standard_footer(f"[↑↓] Move   [Enter] Select   [H] Help{about_legend}   [Q/ESC] {back_word}")
        if blocked_msg:
            print(color_text(f" {blocked_msg}", fg=YELLOW))
            blocked_msg = ""
        sys.stdout.write(color_text(" Option: ", fg=CYAN, style=BOLD) + buf)
        sys.stdout.flush()

        key = read_key()
        if key == "UP":
            selected = (selected - 1) % n; buf = ""
        elif key == "DOWN":
            selected = (selected + 1) % n; buf = ""
        elif key == "ENTER":
            picked = int(buf) - 1 if buf.isdigit() and 1 <= int(buf) <= n else selected
            reason = _opt_disabled_reason(options[picked])
            if reason:
                blocked_msg = f"{options[picked][0]} is unavailable — {reason}."
                buf = ""
                continue
            print()   # leave the prompt line so any follow-up prompt starts fresh
            return str(picked + 1)
        elif key == "ESC" or key in ("q", "Q"):
            return "quit" if is_main else "back"
        elif key in ("h", "H"):
            show_menu_help(subtitle, options, help_note)
        elif is_main and key in ("a", "A"):
            return "about"
        elif key == "BACKSPACE":
            buf = buf[:-1]
        elif key and key.isdigit():
            buf += key
            if 1 <= int(buf) <= n:            # move highlight to the typed number
                selected = int(buf) - 1
        # anything else: ignore and redraw


def _print_menu_outro(outro):
    """Print the optional context line(s) shown below a menu's options."""
    if not outro:
        return
    lines = [outro] if isinstance(outro, str) else outro
    for line in lines:
        print(f"  {DIM}{line}{RESET}")
    print()


def _render_menu_lines(subtitle, options, is_main, intro, outro=None,
                       help_note=None):
    """Line-based menu fallback for non-TTY stdin (piped input / automation).
    Preserves numbered selection, Enter=default option 1, H help, Q/ESC back."""
    default = "" if is_main else "1"
    while True:
        screen(subtitle)
        if intro:
            print(f"  {DIM}{intro}{RESET}")
        print()
        for i, opt in enumerate(options, 1):
            label  = opt[0]
            reason = _opt_disabled_reason(opt)
            if reason:
                print(f"  {color_text(f'{i}. {label}  ({reason})', style=DIM)}")
            else:
                tag = f" {DIM}(default){RESET}" if (i == 1 and not is_main) else ""
                print(f"  {color_text(str(i), fg=YELLOW, style=BOLD)}. {label}{tag}")
        print()
        _print_menu_outro(outro)
        back_word = "Exit" if is_main else "Back"
        about_legend = "  [A] About" if is_main else ""
        standard_footer(f"[↑↓] Move   [Enter] Select   [H] Help{about_legend}   [Q/ESC] {back_word}")
        prompt = " Option: " if is_main else " Option [1]: "
        choice = menu_read(color_text(prompt, fg=CYAN, style=BOLD), default=default)
        if choice == "ESC":
            return "quit" if is_main else "back"
        if choice.isdigit() and 1 <= int(choice) <= len(options):
            reason = _opt_disabled_reason(options[int(choice) - 1])
            if reason:
                print(color_text(f"  {options[int(choice) - 1][0]} is unavailable — {reason}.", fg=YELLOW))
                continue
            return choice
        if choice == "q":
            return "quit" if is_main else "back"
        if choice == "h":
            show_menu_help(subtitle, options, help_note)
        if is_main and choice == "a":
            return "about"


def render_multiselect(subtitle, options, intro=None, preselected=None):
    """Multi-select menu — the user can toggle several options on.

    Interactive TTY: ↑/↓ move, Space toggles the highlighted item, Enter
    confirms, Q/ESC cancels.
    Non-TTY: enter comma-separated numbers (or 'all'); blank keeps
    `preselected` (or cancels if none was given).

    preselected: iterable of indices checked by default — e.g. range(n) so
                 everything starts checked and the user unselects what they
                 don't want. Default: nothing pre-checked (unchanged
                 behavior for callers that don't pass it).
    Returns a list of selected indices (possibly empty), or None if cancelled.
    """
    n = len(options)
    if n == 0:
        return None
    default_set = {i for i in (preselected or []) if 0 <= i < n}

    if not sys.stdin.isatty():
        screen(subtitle)
        if intro:
            print(f"  {DIM}{intro}{RESET}")
        print()
        for i, (label, _d) in enumerate(options, 1):
            mark = color_text(" (default: checked)", style=DIM) if (i - 1) in default_set else ""
            print(f"  {color_text(str(i), fg=YELLOW, style=BOLD)}. {label}{mark}")
        print()
        blank_hint = "keeps the defaults shown above" if default_set else "cancels"
        standard_footer(f"Enter numbers (comma-separated) or 'all'; blank {blank_hint}")
        raw = menu_read(color_text(" Select: ", fg=CYAN, style=BOLD))
        if raw in ("ESC", "q"):
            return None
        if raw == "":
            return sorted(default_set) if default_set else None
        if raw == "all":
            return list(range(n))
        picks = [int(t) - 1 for t in raw.split(",") if t.strip().isdigit() and 1 <= int(t) <= n]
        return sorted(set(picks))

    selected = set(default_set)
    cursor = 0
    while True:
        screen(subtitle)
        if intro:
            print(f"  {DIM}{intro}{RESET}")
        print()
        for i, (label, _d) in enumerate(options):
            box = "☑" if i in selected else "☐"
            if i == cursor:
                print(color_text(f"  ▶ {box} {label}", fg=BRIGHT_CYAN, style=BOLD))
            else:
                mark = color_text(box, fg=GREEN) if i in selected else box
                print(f"    {mark} {label}")
        print()
        standard_footer("[↑↓] Move   [Space] Toggle   [Enter] Confirm   [Q/ESC] Cancel")
        sys.stdout.write(color_text(f" Selected: {len(selected)} ", fg=CYAN, style=BOLD))
        sys.stdout.flush()

        key = read_key()
        if key == "UP":
            cursor = (cursor - 1) % n
        elif key == "DOWN":
            cursor = (cursor + 1) % n
        elif key == " ":
            selected.discard(cursor) if cursor in selected else selected.add(cursor)
        elif key == "ENTER":
            print()   # leave the prompt line so any follow-up prompt starts fresh
            return sorted(selected)
        elif key == "ESC" or key in ("q", "Q"):
            return None
        # anything else: ignore and redraw


def _print_help_desc(desc, width):
    """Word-wrap and print one Help description block (dim, indented).

    A blank line separates consecutive paragraphs so long descriptions don't
    read as one dense wall of text. Bullet lines (starting with "•" or "-")
    stay tight against each other — no blank line between one bullet and the
    next — but still get a blank line before/after the list as a whole."""
    paras = desc.split("\n")
    for i, para in enumerate(paras):
        is_bullet = para.lstrip().startswith(("•", "-"))
        hang = "       " if is_bullet else "     "
        wrapped = textwrap.fill(para.strip(), width=width,
                                initial_indent="     ", subsequent_indent=hang)
        print(f"{DIM}{wrapped}{RESET}")
        if i < len(paras) - 1:
            next_is_bullet = paras[i + 1].lstrip().startswith(("•", "-"))
            if not (is_bullet and next_is_bullet):
                print()


def show_menu_help(subtitle, options, note=None):
    """Show a Help screen describing every option in the current menu.

    Descriptions are word-wrapped to the terminal width. A description may
    contain explicit newlines to force line breaks (e.g. for bullet lists);
    a line starting with a bullet keeps its hanging indent when it wraps.
    note : optional (label, description) tuple printed after the options —
           an unnumbered extra section (e.g. where activity is logged).
    """
    screen(f"{subtitle} — Help")
    print()
    width = max(get_width() - 6, 40)
    for i, opt in enumerate(options, 1):
        label, desc = opt[0], opt[1]
        reason = _opt_disabled_reason(opt)
        suffix = color_text(f"  ({reason})", style=DIM) if reason else ""
        print(f"  {color_text(f'{i}. {label}', fg=WHITE, style=BOLD)}{suffix}")
        if desc:
            _print_help_desc(desc, width)
        print()
    if note:
        note_label, note_desc = note
        print(f"  {color_text(note_label, fg=WHITE, style=BOLD)}")
        if note_desc:
            _print_help_desc(note_desc, width)
        print()
    standard_footer("[Any key] Back to menu")
    sys.stdout.write(color_text(" Option: ", fg=CYAN, style=BOLD)); sys.stdout.flush()
    if sys.stdin.isatty():
        read_key()          # any key dismisses
    else:
        menu_read("")


def _ask_compare_by():
    """Inline prompt (no screen redraw) for the comparison basis — asked
    right after Recursive/Case-sensitive so every Compare Folder Contents
    option is gathered in one continuous flow, with no full-screen menu
    interrupting midway. compare_folder_contents() then does the one and
    only redraw, showing everything chosen, before it runs. Returns
    name|size|both. ESC raises EscCancelled like every other question here,
    caught by compare_menu()'s own try/except."""
    print()
    print(color_text("  Compare by:", fg=CYAN))
    print(f"    {color_text('1', fg=YELLOW)}. Name — entries missing from either folder (default)")
    print(f"    {color_text('2', fg=YELLOW)}. Size — same-named entries whose size differs")
    print(f"    {color_text('3', fg=YELLOW)}. Both — name and size together")
    choice = ask("  Compare by (1-3)", default="1")
    return {"1": "name", "2": "size", "3": "both"}.get(choice, "name")


def _pick_compare_mode():
    """Interactive vs Run Profile for Compare Folder Contents. Interactive is
    the default (highlighted first). Returns 'interactive', 'profile', or
    None if cancelled."""
    options = [
        ("Interactive",
         "Enter Folder A and Folder B yourself and choose Recursive, "
         "Case-Sensitive, and Compare By each time."),
        ("Run Profile",
         "Pick a saved Compare profile (fmConfig.json → compareProfiles) and "
         "run it with its saved folders and options."),
    ]
    ch = render_menu("Compare Folder Contents", options)
    return {"1": "interactive", "2": "profile"}.get(ch) if ch != "back" else None


def _compare_profile_menu():
    """[Run Profile] picker for Compare Folder Contents."""
    profiles, perr = _load_compare_profiles()
    if perr:
        screen("Compare Folder Contents — Run Profile")
        print()
        print(color_text(f"  ⚠ {perr}", fg=RED))
        pause_return()
        return
    if not profiles:
        screen("Compare Folder Contents — Run Profile")
        print()
        print(color_text("  No saved Compare profiles yet.", fg=YELLOW))
        print(f"  Add them to fmConfig.json as a compareProfiles list, or run "
              f"Interactive and press [S] Save as Profile afterward.")
        pause_return()
        return
    options = []
    for pr in profiles:
        name = pr.get("name") or "(unnamed profile)"
        by = pr.get("compareBy", "both")
        options.append((
            f"Profile: {name}",
            f"Compare what is inside two folders using this saved profile:\n"
            f"• Folder A — {pr.get('folderA', '?')}\n"
            f"• Folder B — {pr.get('folderB', '?')}\n"
            f"• Recursive: {'Yes' if pr.get('recursive', True) else 'No'}   "
            f"Compare By: {by}   "
            f"Case-sensitive: {'Yes' if pr.get('caseSensitive', False) else 'No'}"))
    ch = render_menu("Compare Folder Contents — Run Profile", options)
    if ch == "back":
        return
    _run_compare_profile(profiles[int(ch) - 1])


def compare_menu():
    options = [
        ("Compare 2 Files",
         "Pick two text files; they are shown side by side, line by line, "
         "aligned by a diff. Each row is marked:\n"
         "• = — the lines are identical\n"
         "• ≠ — the line changed\n"
         "• < — the line exists only in file A\n"
         "• > — the line exists only in file B\n"
         "A summary counts equal / changed / only-in-A / only-in-B lines, or "
         "reports that the files are identical."),
        ("Compare Folder Contents",
         "First choose Interactive (default) or Run Profile.\n"
         "Interactive — Folder A, Folder B, Recursive?, Case-sensitive?, and "
         "Compare By are all asked one after another with no screen changes "
         "in between:\n"
         "• Recursive? — Yes compares every file beneath each folder; No compares "
         "only the immediate top-level files and subfolders.\n"
         "• Case-sensitive? — No (default) treats IMG_1.mov and IMG_1.MOV as the "
         "same file, matching how macOS's default filesystem sees them; Yes treats "
         "them as different files. When matched case-insensitively, a same-name "
         "pair that differs only in case is called out separately so nothing is "
         "silently hidden.\n"
         "• Compare By — Name (entries in one folder only, both directions), Size "
         "(entries in both whose size differs), or Both.\n"
         "Once everything is answered, the results screen redraws once, showing "
         "every choice you made, and runs the comparison. When it finishes you "
         "can [R] Run Again, [S] Save as Profile (fmConfig.json → "
         "compareProfiles), or Quit/Back.\n"
         "Run Profile — pick a saved profile from the list and run it directly.\n"
         "Hidden files and folders (names starting with a dot) are ignored."),
    ]
    while True:
        ch = render_menu("Compare", options)
        if ch == "back":
            return
        try:
            if ch == "1":
                fa = ask_file("File A")
                if not fa:
                    pause_return(); continue
                fb = ask_file("File B")
                if not fb:
                    pause_return(); continue
                compare_two_files(fa, fb)
            elif ch == "2":
                mode = _pick_compare_mode()
                if mode is None:
                    continue
                if mode == "profile":
                    _compare_profile_menu()
                    continue
                a = ask_folder("Folder A", key="compare_folder_a")
                if not a:
                    pause_return(); continue
                b = ask_folder("Folder B", key="compare_folder_b")
                if not b:
                    pause_return(); continue
                recursive = safe_confirm("  Recursive (descend into subfolders)?", default=True)
                case_sensitive = safe_confirm(
                    "  Case-sensitive filename matching? (No = IMG_1.mov and IMG_1.MOV are treated as the same file)",
                    default=False)
                by = _ask_compare_by()
                compare_folder_contents(a, b, recursive, by, case_sensitive)
        except EscCancelled:
            continue


def display_menu():
    options = [
        ("All Drives",
         "Size and free space of every mounted drive — the boot volume plus "
         "each drive under /Volumes. Shows Size, Used, Free, and Use% per "
         "drive (Use% turns yellow at 75% and red at 90%). Free is the space "
         "actually available to you, matching Finder's Available figure."),
        ("Drives in Use",
         "Lists every mounted external drive and whether it is currently "
         "being read from or written to. Detection samples each drive's "
         "actual disk I/O for a second (not just whether a file is open), so "
         "an idle-but-open file doesn't read as in use. macOS only."),
        ("Subfolders Alphabetically",
         "Enter a folder; its subfolders are listed sorted A→Z by name. Each "
         "row shows the folder, its size (human-readable), the exact byte "
         "count, and how many files it contains; a TOTAL row sums them. Best "
         "when you know the folder name you're looking for."),
        ("Subfolders by Size (largest first)",
         "Same columns, but sorted by total size with the biggest folder first — "
         "the quickest way to see what is using the most space."),
    ]
    while True:
        ch = render_menu("Display", options,
                         intro="Drive space · subfolder sizes, byte counts, and file counts.")
        if ch == "back":
            return
        try:
            if ch == "1":
                display_all_drives()
                continue
            if ch == "2":
                display_drives_in_use()
                continue
            folder = ask_folder("Folder to measure", default=os.getcwd(), key="display_subfolders")
            if not folder:
                pause_return(); continue
            display_folder_sizes(folder, "size" if ch == "4" else "alpha")
        except EscCancelled:
            continue


def _pick_missing_mode(folder_a, folder_b, subtitle="Find Missing by Filename — Show"):
    """Sub-selection for the Find Missing features — shows the two entered
    folders below the options. Returns first|second|either, or None if
    cancelled."""
    options = [
        ("In 1st folder only",
         "List filenames that exist in the 1st folder but NOT in the 2nd — the "
         "files the 2nd folder is missing."),
        ("In 2nd folder only",
         "List filenames that exist in the 2nd folder but NOT in the 1st — the "
         "files the 1st folder is missing."),
        ("In either folder (only once)",
         "List filenames that exist in one folder but not the other, in both "
         "directions — the full picture of what doesn't match."),
    ]
    ch = render_menu(subtitle, options,
                     outro=[f"Folder 1 - {folder_a}", f"Folder 2 - {folder_b}"])
    return {"1": "first", "2": "second", "3": "either"}.get(ch) if ch != "back" else None


def find_menu():
    options = [
        ("Find Files   (one or more criteria)",
         "Search for files that match one or more criteria combined with AND. On "
         "the next screen, use Space to toggle any of:\n"
         "• Filename pattern — a wildcard like IMG_*.jpg\n"
         "• File extension — e.g. mov, jpg\n"
         "• Size over N MB — larger than N\n"
         "• Size under N MB — smaller than N\n"
         "You then enter a value for each toggled criterion and a search root. "
         "Example: extension = mov and size under 5 finds .mov files under 5 MB. "
         "Each result shows its size (largest first when a size filter is on)."),
        ("Find Folders (by name/pattern)",
         "Find directories whose name matches a wildcard pattern, searched "
         "recursively under a chosen root. Example pattern: *_backup. Results are "
         "just listed — nothing is changed."),
        ("Find Duplicates by Filename",
         "Enter one or more folders (comma-separated); each is scanned "
         "recursively.\n"
         "Files sharing the same filename are reported in a table: a "
         "numbered header lists the folders, then one row per duplicated "
         "filename with a size column per folder — so you can see at a "
         "glance where each copy lives and whether the sizes match. If a "
         "name occurs more than once inside a single folder, every size is "
         "listed in that column. Matching is by filename only — contents "
         "are not compared.\n"
         "Results are just listed; nothing is changed unless you press "
         "[D] Delete Duplicates afterward to pick a keep rule:\n"
         "• Keep Newest\n"
         "• Keep Largest\n"
         "• Delete from Specific Folder — enter a folder; any copy under "
         "it (including subfolders) is deleted, and copies elsewhere are "
         "kept. If every copy in a group is under that folder, one is "
         "kept anyway.\n"
         "Every group's KEEP/DELETE split is shown before anything "
         "happens; deleting requires typing YES."),
        ("Find Duplicates by Fuzzy Name",
         "Enter one or more folders (comma-separated); each is scanned "
         "recursively. Files are grouped as duplicates when their names "
         "are CLOSE — not necessarily identical — AND their sizes are "
         "close (within 1%). Example: videofile1.mov and videofile.mov at "
         "the same size are duplicates.\n"
         "Names are close when the stems match after duplicate-style "
         "endings are stripped (trailing digits, \"(1)\", \"[2]\", "
         "\"copy\", \"copy 2\") or are 85%+ similar; extensions must "
         "match. Each group marks the shortest/cleanest name KEEP and the "
         "rest DELETE — so videofile1.mov is the one flagged to delete.\n"
         "Results are just listed, largest files first, with a "
         "reclaimable-space total; nothing is deleted unless you press "
         "[D] Delete Duplicates afterward — same Keep Newest / Keep "
         "Largest / Delete from Specific Folder rule as Find Duplicates "
         "by Filename, overriding the shortest-name KEEP shown above "
         "with your chosen rule. Deleting requires typing YES."),
        ("Find Missing by Filename",
         "Enter two folders, then choose what to show:\n"
         "• In 1st folder only — files the 2nd folder is missing.\n"
         "• In 2nd folder only — files the 1st folder is missing.\n"
         "• In either folder (only once) — both directions together.\n"
         "Each folder is scanned recursively and files are matched by filename "
         "only — contents are not compared. Results use the same table as Find "
         "Duplicates by Filename: a numbered header lists the folders, then one "
         "row per filename with a size column per folder — the blank column "
         "shows where the file is missing. Nothing is changed; results are "
         "just listed."),
        ("Find Missing by Filename & Size",
         "Same as Find Missing by Filename, but files only count as a match "
         "when BOTH the filename AND the file size agree. That means a "
         "same-named file whose size differs between the two folders is also "
         "reported — with its size shown in both columns so the mismatch is "
         "obvious. Enter two folders, then choose In 1st folder only / In 2nd "
         "folder only / In either folder. Nothing is changed; results are "
         "just listed."),
        ("Find & Replace",
         "Enter a folder, the text to find, the replacement text, and an "
         "optional file extension. Every text file under the folder is "
         "scanned recursively (hidden files/folders and binary files are "
         "skipped); matching is literal and case-insensitive.\n"
         "ALWAYS a dry run first — every match is listed with its file, "
         "line number, and the line with the matched text highlighted, "
         "and nothing is touched. Then confirm [y/N] to replace every "
         "occurrence (you are also asked whether to save a .bak backup "
         "of each file first), or answer No to exit with nothing "
         "changed."),
        ("Find & Rename",
         "Find files, then rename them. Enter a folder, an optional file "
         "extension, and choose how the text is applied:\n"
         "• Prepend — add the text to the START of the filename "
         "(text.mov -> api_text.mov). Files already starting with it are "
         "skipped.\n"
         "• Append — add the text to the END of the name, before the "
         "extension (text.mov -> text_api.mov). Files already ending with "
         "it are skipped.\n"
         "• Replace — replace text inside the filename (draft_a.mov -> "
         "final_a.mov); matching is literal and case-insensitive, and a "
         "blank replacement removes the text.\n"
         "ALWAYS a dry run first — every rename is listed old -> new, and "
         "nothing is touched until you confirm [y/N]. Renames that would "
         "overwrite an existing file are skipped and reported."),
    ]
    while True:
        ch = render_menu("Find", options)
        if ch == "back":
            return
        try:
            if ch == "2":
                root = ask_folder("Search root", default=os.getcwd(), key="find_folders_root")
                if not root:
                    pause_return(); continue
                pat = ask("Folder name/pattern (wildcards ok)", "*")
                find_folders(root, pat or "*")
                continue
            if ch in ("3", "4"):
                folders = _ask_folders_multi()
                if folders:
                    if ch == "3":
                        find_duplicates_by_filename(folders)
                    else:
                        find_duplicates_by_fuzzy_name(folders)
                else:
                    pause_return()
                continue
            if ch in ("5", "6"):
                match_size = (ch == "6")
                a = ask_folder("Folder 1", key="find_missing_folder1")
                if not a:
                    pause_return(); continue
                b = ask_folder("Folder 2", key="find_missing_folder2")
                if not b:
                    pause_return(); continue
                subtitle = ("Find Missing by Filename & Size — Show" if match_size
                            else "Find Missing by Filename — Show")
                mode = _pick_missing_mode(a, b, subtitle)
                if mode is None:
                    continue
                find_missing_by_filename(a, b, mode, match_size)
                continue
            if ch == "7":
                root = ask_folder("Folder to search", key="find_replace_root")
                if not root:
                    pause_return(); continue
                search = ask("Text to find")
                if not search:
                    print(color_text("  Search text required — cancelled.", fg=YELLOW))
                    pause_return(); continue
                replace = ask("Replace with (blank = remove the text)")
                ext = ask("File extension (optional, e.g. php)")
                find_and_replace(root, search, replace, ext or None)
                continue
            if ch == "8":
                root = ask_folder("Folder to search", key="find_rename_root")
                if not root:
                    pause_return(); continue
                mode_opts = [
                    ("Prepend — add text to the start of the filename",
                     "The text is added to the front of every found filename: "
                     "text.mov -> api_text.mov. Files whose name already starts "
                     "with the text are skipped, so reruns are safe."),
                    ("Append  — add text to the end, before the extension",
                     "The text is inserted between the name and its extension: "
                     "text.mov -> text_api.mov. Files whose name already ends "
                     "with the text are skipped, so reruns are safe."),
                    ("Replace — replace text within the filename",
                     "Enter the text to find in the filename and its replacement "
                     "(blank = remove the text): draft_a.mov -> final_a.mov. "
                     "Matching is literal and case-insensitive; only files whose "
                     "name contains the text are listed."),
                ]
                mch = render_menu("Find & Rename — Mode", mode_opts,
                                  outro=[f"Folder - {root}"])
                if mch == "back":
                    continue
                replace_with = None
                if mch == "1":
                    mode = "prepend"
                    text = ask("Text to prepend")
                elif mch == "2":
                    mode = "append"
                    text = ask("Text to append (before the extension)")
                else:
                    mode = "replace"
                    text = ask("Text to find in the filename")
                    if text:
                        replace_with = ask("Replacement text (blank = remove the text)")
                if not text:
                    print(color_text("  Text required — cancelled.", fg=YELLOW))
                    pause_return(); continue
                ext = ask("File extension (optional, e.g. mov)")
                find_and_rename(root, mode, text, replace_with, ext or None)
                continue

            # Find Files — pick one or more criteria, then prompt for each value.
            crit_opts = [
                ("Filename pattern", "Match file names against a wildcard (e.g. IMG_*.jpg)."),
                ("File extension",   "Match a file extension (e.g. mov, jpg)."),
                ("Size over N MB",   "Only files larger than N megabytes."),
                ("Size under N MB",  "Only files smaller than N megabytes."),
            ]
            picks = render_multiselect(
                "Find Files — Select Criteria", crit_opts,
                intro="Toggle any criteria to combine (e.g. extension + size under). They are AND-ed together.")
            if picks is None:
                continue                       # cancelled — back to Find menu
            if not picks:
                print(color_text("  No criteria selected.", fg=YELLOW)); pause_return(); continue

            pattern = ext = over_mb = under_mb = None
            if 0 in picks:
                pattern = ask("Filename pattern (wildcards ok)", "*") or "*"
            if 1 in picks:
                ext = ask("Extension (e.g. mov)")
                if not ext:
                    print(color_text("  Extension required — cancelled.", fg=YELLOW)); pause_return(); continue
            if 2 in picks:
                try:
                    over_mb = float(ask("Size OVER how many MB?", "5"))
                except ValueError:
                    print(color_text("  Invalid number.", fg=RED)); pause_return(); continue
            if 3 in picks:
                try:
                    under_mb = float(ask("Size UNDER how many MB?", "5"))
                except ValueError:
                    print(color_text("  Invalid number.", fg=RED)); pause_return(); continue

            root = ask_folder("Search root", default=os.getcwd(), key="find_files_root")
            if not root:
                pause_return(); continue
            find_files_combined(root, pattern, ext, over_mb, under_mb)
        except EscCancelled:
            continue


def _ask_folders_multi():
    raw = ask("Folder(s) — comma-separated for multiple")
    if not raw:
        return None
    folders = [clean_path(f) for f in raw.split(",") if f.strip()]
    valid = [f for f in folders if os.path.isdir(f)]
    for f in folders:
        if f not in valid:
            print(color_text(f"  Skipping (not a directory): {f}", fg=YELLOW))
    return valid or None


def remove_menu():
    options = [
        ("Duplicates by Name",
         "Scan one or more folders (enter several comma-separated) and group "
         "files that share the same file name. The first occurrence of each "
         "name is kept; the rest are listed for removal. You can optionally "
         "limit the scan to a single extension. Matches by name only — the "
         "contents are not checked.\n"
         "Previewed and dry-run; you can then delete (type the word YES; "
         "anything else cancels) or move the duplicates to a folder instead."),
        ("Duplicates by Hash",
         "Scan one or more folders and group files with identical CONTENT "
         "using a SHA-256 hash. Files are pre-grouped by size first, so "
         "files of different sizes are never hashed (fast). The first of "
         "each identical set is kept; the rest are listed for removal. "
         "This catches duplicates even when they have different names.\n"
         "Previewed and dry-run; you can then delete (type the word YES; "
         "anything else cancels) or move the duplicates to a folder instead."),
        ("Duplicates by Fuzzy Name",
         "Scan one or more folders and group files whose names are CLOSE — "
         "not necessarily identical — AND whose sizes are close (within "
         "1%). Example: videofile1.mov and videofile.mov at the same size "
         "are duplicates, and videofile1.mov is the one removed.\n"
         "Names are close when the stems match after duplicate-style "
         "endings are stripped (trailing digits, \"(1)\", \"[2]\", "
         "\"copy\") or are 85%+ similar; extensions must match. The "
         "shortest/cleanest name in each group is kept; the rest are "
         "listed for removal. Hidden files are skipped.\n"
         "Previewed and dry-run; you can then delete (type the word YES; "
         "anything else cancels) or move the duplicates to a folder instead."),
        ("By File Name",
         "Delete files whose name matches a wildcard pattern, searched recursively "
         "under a chosen root. Every match is previewed with its size before "
         "anything is deleted.\n"
         "• Tip: enter '.DS_Store' or 'desktop.ini' to clean macOS/Windows junk "
         "files. Wildcards work too, e.g. *.tmp."),
        ("By Folder Name",
         "Delete folders whose name matches a wildcard pattern, searched "
         "recursively under a chosen root; each matching folder tree is removed. "
         "When a matched folder sits inside another matched folder, only the "
         "top-most one is removed (it already contains the child). Previewed and "
         "dry-run until you confirm. Example pattern: *_tmp."),
        ("Files of 0 Size",
         "Scan one or more folders (enter several comma-separated) and list "
         "every empty (0-byte) file for removal — incomplete downloads, "
         "placeholder files, failed copies. Hidden files/folders are "
         "skipped, so intentional markers like .gitkeep are never "
         "touched.\n"
         "Previewed and dry-run until you confirm."),
    ]
    while True:
        ch = render_menu("Remove", options,
                         intro="DRY RUN by default — you always confirm before anything is deleted.")
        if ch == "back":
            return
        try:
            if ch == "1":
                folders = _ask_folders_multi()
                if folders:
                    ext = ask("Limit to extension (blank = all)")
                    remove_duplicates_by_name(folders, live_requested=None, ext=ext or None)
            elif ch == "2":
                folders = _ask_folders_multi()
                if folders:
                    remove_duplicates_by_hash(folders, live_requested=None)
            elif ch == "3":
                folders = _ask_folders_multi()
                if folders:
                    remove_duplicates_by_fuzzy_name(folders, live_requested=None)
            elif ch == "4":
                root = ask_folder("Search root", default=os.getcwd(), key="remove_by_name_root")
                if not root:
                    pause_return(); continue
                print(f"  {DIM}Tip: use '.DS_Store' or 'desktop.ini' to clean junk files.{RESET}")
                pat = ask("File name/pattern to remove (wildcards ok)")
                if pat:
                    remove_by_name(root, pat, live_requested=None)
            elif ch == "5":
                root = ask_folder("Search root", default=os.getcwd(), key="remove_folders_root")
                if not root:
                    pause_return(); continue
                pat = ask("Folder name/pattern to remove (wildcards ok)")
                if pat:
                    remove_folders_by_name(root, pat, live_requested=None)
            elif ch == "6":
                folders = _ask_folders_multi()
                if folders:
                    remove_zero_size_files(folders, live_requested=None)
        except EscCancelled:
            continue


def _monitor_output_choice():
    """Ask where monitor events should be logged. Returns 'log' | 'csv' | None."""
    log_disp = ACTIVITY_LOG.replace(os.path.expanduser("~"), "~")
    csv_disp = MONITOR_CSV.replace(os.path.expanduser("~"), "~")
    options = [
        (f"FM activity log — {log_disp}",
         "Each event is appended to the shared FM activity log as a plain "
         "timestamped line, right below the session header. Everything FM "
         "does stays in one log."),
        (f"CSV file — {csv_disp}",
         "Each event is appended as a CSV row (Timestamp, Filename, Folder, "
         "Event) — easy to open in a spreadsheet or parse. The header row is "
         "written when the file is first created; sessions append."),
    ]
    ch = render_menu("Monitor — Log Destination", options)
    if ch == "back":
        return None
    return "log" if ch == "1" else "csv"


def _monitor_interactive():
    """Interactive Monitor — ask for the folder and options, then watch it."""
    screen("Monitor File Activity")
    print()
    folder = ask_folder("Folder to monitor", key="monitor_folder")
    if not folder:
        pause_return(); return
    recursive = safe_confirm("  Include subfolders (recursive)?", default=True)
    ext_raw = ask("File extensions to watch (comma-separated, blank = all)")
    output = _monitor_output_choice()
    if output is None:
        return
    monitor_activity(folder, recursive, ext_raw, output)
    pause_return()


def _run_monitor_profile(pr):
    """Run one monitorProfiles entry from fmConfig.json (defaults: recursive,
    all extensions, log output)."""
    folder = clean_path(str(pr.get("folder", "")))
    output = "csv" if str(pr.get("output", "log")).lower() == "csv" else "log"
    monitor_activity(folder,
                     recursive=bool(pr.get("recursive", True)),
                     exts=pr.get("extensions", ""),
                     output=output,
                     profile_name=pr.get("name") or "(unnamed profile)")
    pause_return()


def monitor_menu():
    profile_note = (
        "Monitor Profiles (fmConfig.json)",
        "Saved monitor setups appear as options above Interactive Monitor. Add "
        "them to fmConfig.json as a monitorProfiles list; each profile supports:\n"
        "• name — the label shown in this menu\n"
        "• folder — the folder to watch (~ is expanded)\n"
        "• recursive — true/false (default true)\n"
        "• extensions — extensions to watch, e.g. \"jpg,png\" or a list "
        "(default: all files)\n"
        "• output — log (fm.log) or csv (fmMonitor.csv); default log\n"
        "Monitoring runs until you press [Q/ESC].")
    while True:
        profiles, perr = _load_monitor_profiles()
        options = []
        for pr in profiles:
            name = pr.get("name") or "(unnamed profile)"
            exts = _parse_ext_filter(pr.get("extensions", ""))
            options.append((
                f"Profile: {name}",
                f"Watch this saved profile's folder for created/modified/deleted "
                f"files:\n"
                f"• Folder — {pr.get('folder', '?')}\n"
                f"• Recursive: {'Yes' if pr.get('recursive', True) else 'No'}   "
                f"Extensions: {', '.join(sorted(exts)) if exts else 'all'}   "
                f"Output: {'CSV' if str(pr.get('output', 'log')).lower() == 'csv' else 'fm.log'}\n"
                "Runs until you press [Q/ESC]."))
        options.append((
            "Interactive Monitor — enter a folder & options",
            "Enter the folder to watch, whether to include subfolders (default "
            "Yes), an optional extension filter (e.g. jpg, png — blank watches "
            "everything), and where to log (fm.log or fmMonitor.csv). Every "
            "created, modified, or deleted file then appears on screen and in "
            "the log in real time until you press [Q/ESC]."))
        ch = render_menu("Monitor File Activity", options,
                         outro=(f"⚠ {perr}" if perr else None),
                         help_note=profile_note)
        if ch == "back":
            return
        try:
            if int(ch) <= len(profiles):
                _run_monitor_profile(profiles[int(ch) - 1])
            else:
                _monitor_interactive()
        except EscCancelled:
            continue


def _sync_interactive():
    """Interactive Sync — ask for the two folders, the direction, and the
    options, then run the preview/confirm sync."""
    screen("Sync Folders")
    print()
    folder_a = ask_folder("Folder A", key="sync_folder_a")
    if not folder_a:
        pause_return(); return
    folder_b = ask_folder("Folder B", key="sync_folder_b")
    if not folder_b:
        pause_return(); return

    dir_options = [
        ("Push new/updated files from A → B",
         "Folder A is the source. Files that exist only in A are copied into B, "
         "and files existing on both sides are copied when A's copy wins the "
         "conflict rule you pick next. Folder B is never read from — nothing in "
         "A changes, and nothing is ever deleted."),
        ("Push new/updated files from B → A",
         "Folder B is the source. Files that exist only in B are copied into A, "
         "and files existing on both sides are copied when B's copy wins the "
         "conflict rule you pick next. Folder A is never read from — nothing in "
         "B changes, and nothing is ever deleted."),
        ("Two-way sync A ↔ B",
         "Both folders push to each other in one run. Files that exist only in "
         "A are copied into B, files that exist only in B are copied into A, "
         "and when a file exists on both sides the copy that wins the conflict "
         "rule you pick next replaces the other. Nothing is ever deleted."),
    ]
    ch = render_menu("Sync — Direction", dir_options,
                     outro=[f"Folder A - {folder_a}", f"Folder B - {folder_b}"])
    if ch == "back":
        return
    direction = {"1": "a2b", "2": "b2a", "3": "both"}[ch]

    conflict_options = [
        ("Choose the newest  — copy only when the source file is newer",
         "When a file exists in both folders, its modified times are compared "
         "and the source copy is pushed only if it is newer than the "
         "destination copy. The usual choice for keeping a backup current."),
        ("Choose the largest — copy only when the source file is larger",
         "When a file exists in both folders, their sizes are compared and the "
         "source copy is pushed only if it is larger than the destination "
         "copy. Useful when bigger means better (e.g. re-exported media)."),
    ]
    if direction != "both":
        conflict_options.append((
            "Choose per file — you decide, one file at a time",
            "For every file that exists on both sides and differs, you're shown "
            "the filename with both sizes and both modified dates — whichever "
            "of the two differs is highlighted — then choose [I] Ignore or "
            "[S] Select (copy) for that file. [IA] Ignore All or [SA] Select "
            "All applies your choice to every remaining file without asking "
            "again. Only available for a one-way sync (A → B or B → A), not "
            "two-way."))
    ch = render_menu("Sync — If a File Exists on Both Sides", conflict_options,
                     outro=[f"Folder A - {folder_a}", f"Folder B - {folder_b}"])
    if ch == "back":
        return
    conflict = {"1": "newest", "2": "largest"}.get(ch, "manual")

    recursive = safe_confirm("  Include subfolders (recursive)?", default=True)
    exclude_hidden = safe_confirm("  Exclude hidden files?", default=True)

    sync_folders(folder_a, folder_b, direction, recursive, conflict, exclude_hidden)
    pause_return()


def sync_menu():
    profile_note = (
        "Sync Profiles (fmConfig.json)",
        "Saved sync setups appear as options above Interactive Sync. Add them to "
        "fmConfig.json as a syncProfiles list; each profile supports:\n"
        "• name — the label shown in this menu\n"
        "• folderA / folderB — the two folders (~ is expanded)\n"
        "• direction — AtoB, BtoA, or Both (which side pushes its new/updated "
        "files; Both = two-way, each side receives the other's)\n"
        "• recursive — true/false (default true)\n"
        "• conflict — newest or largest: when a file exists on both sides, copy "
        "only when the source is newer / larger (default newest). Manual per-file "
        "review isn't available for profiles — use Interactive Sync for that.\n"
        "• excludeHidden — true/false (default true)\n"
        "Profiles run exactly like Interactive Sync: a preview first, and "
        "nothing is copied until you confirm.")
    while True:
        profiles, perr = _load_sync_profiles()
        options = []
        for pr in profiles:
            name = pr.get("name") or "(unnamed profile)"
            pdir = _parse_sync_direction(pr.get("direction", "AtoB"))
            arrow = {"b2a": "B → A", "both": "A ↔ B"}.get(pdir, "A → B")
            conflict = ("Largest"
                        if str(pr.get("conflict", "newest")).lower() == "largest"
                        else "Newest")
            options.append((
                f"Profile: {name}  ({arrow})",
                f"Push new/updated files {arrow} using this saved profile:\n"
                f"• Folder A — {pr.get('folderA', '?')}\n"
                f"• Folder B — {pr.get('folderB', '?')}\n"
                f"• Recursive: {'Yes' if pr.get('recursive', True) else 'No'}   "
                f"Conflict: {conflict} wins   Hidden files: "
                f"{'Excluded' if pr.get('excludeHidden', True) else 'Included'}\n"
                "Previews first — nothing is copied until you confirm."))
        options.append((
            "Interactive Sync — enter folders, direction & options",
            "Enter Folder A and Folder B, choose the push direction (A → B, "
            "B → A, or two-way A ↔ B), what happens when a file exists on both "
            "sides (newest wins, largest wins, or — for a one-way sync only — "
            "choose per file yourself), whether to include subfolders (default "
            "Yes), and whether to exclude hidden files (default Yes). A preview "
            "lists every file that would be copied — nothing is copied until "
            "you confirm."))
        ch = render_menu("Sync", options,
                         outro=(f"⚠ {perr}" if perr else None),
                         help_note=profile_note)
        if ch == "back":
            return
        try:
            if int(ch) <= len(profiles):
                _run_sync_profile(profiles[int(ch) - 1])
            else:
                _sync_interactive()
        except EscCancelled:
            continue


def _prompt_zip_subfolders():
    """Shared prompt sequence for Zip SubFolders / Zip SubFolders & Log.
    Returns (target, dest, remove_after, push_to_s3, s3_bucket, s3_folder),
    or None if cancelled."""
    target = ask_folder("Target folder (contains subfolders to zip)", key="zip_subfolders_target")
    if not target:
        return None
    dest = ask_path("Zip destination", default=os.getcwd())
    remove_after = safe_confirm("  Remove source folders after zipping?", default=False)
    push_to_s3 = safe_confirm("  Push files to AWS S3?", default=False)
    s3_bucket, s3_folder = "", ""
    if push_to_s3:
        s3_bucket = ask("    AWS S3 Bucket")
        if not s3_bucket:
            print(color_text("    No bucket entered — Push to AWS S3 disabled for this run.", fg=YELLOW))
            push_to_s3 = False
        elif safe_confirm("    Folder? (S3 prefix inside the bucket)", default=False):
            s3_folder = ask("    S3 Folder/prefix", default="").strip("/")
    return target, dest, remove_after, push_to_s3, s3_bucket, s3_folder


def search_zip_file_contents_db(pattern=None, scope=None):
    """Search zipFileContent by file NAME (never contents) across every
    archive already logged to the CB9Inventory database on BPA5, via the
    DocInfo Manager API — distinct from Find Files in Zip, which searches
    local zip files directly instead of the database. Wildcards are
    translated server-side ('*' -> '%', '?' -> '_'); a pattern with no
    wildcard is an exact match.

    scope may be blank (search every logged zip), a .zip/.tar file name
    (scope to that one archive), or a folder (scope to zips logged from
    that folder or any subfolder beneath it)."""
    screen("Search Zip File Contents by Name")
    print()

    if not ensure_authenticated():
        print(color_text("  DocInfo Manager login is required to search zip file contents.", fg=RED))
        _log_line("Search Zip File Contents by Name cancelled — DocInfo Manager login required")
        pause_return(); return

    logzip, err = _load_logzip_config()
    if err:
        print(color_text(f"  {err}", fg=RED))
        pause_return(); return

    if pattern is None:
        pattern = ask("Filename pattern (wildcards ok)", "*")
        if not pattern:
            pause_return(); return
    if scope is None:
        scope = ask_path("Zip file name or folder to scope the search (blank = search all logged zips)")

    zip_file_name = ""
    zip_file_folder = ""
    if scope:
        cleaned = clean_path(scope)
        if cleaned.lower().endswith((".zip", ".tar")):
            zip_file_name = os.path.basename(cleaned)
        else:
            zip_file_folder = os.path.abspath(cleaned)

    print(f"  {YELLOW}Pattern{RESET}: {pattern}")
    if zip_file_name:
        print(f"  {YELLOW}Scope{RESET}: zip file '{zip_file_name}'")
    elif zip_file_folder:
        print(f"  {YELLOW}Scope{RESET}: folder '{zip_file_folder}' (+ subfolders)")
    else:
        print(f"  {YELLOW}Scope{RESET}: all logged zip files")
    print()

    runlog = _RunLog("Search Zip File Contents by Name",
                      [("Pattern", pattern), ("Zip File", zip_file_name), ("Folder", zip_file_folder)])
    try:
        data = _post_zip_content_search(logzip, pattern, zip_file_name, zip_file_folder)
    except Exception as e:
        print(color_text(f"  [ERROR] Search failed: {e}", fg=RED))
        runlog.finish(f"Search FAILED: {e}")
        pause_return(); return

    if str(data.get("success")) != "1":
        print(color_text(f"  [ERROR] {data.get('msg', 'Search failed')}", fg=RED))
        runlog.finish(f"Search FAILED: {data.get('msg', 'Unknown error')}")
        pause_return(); return

    rows = data.get("recordList", [])
    total = data.get("recordCount", 0)
    if not rows:
        print(color_text("  No matching files found.", fg=YELLOW))
    else:
        current_zip = None
        for row in rows:
            if row.get("zipFileName") != current_zip:
                current_zip = row.get("zipFileName")
                print(f"  {BOLD}{WHITE}{current_zip}{RESET}  {DIM}({row.get('zipFileFolder', '')}){RESET}")
            print(f"    {row.get('filePath')}  {DIM}({row.get('fileSizeFormatted', '')}){RESET}")
        print()
        shown = len(rows)
        if total > shown:
            print(color_text(f"  {shown} of {total} file(s) shown — narrow your pattern/scope to see the rest.",
                              fg=BRIGHT_CYAN, style=BOLD))
        else:
            print(color_text(f"  {total} file(s) found.", fg=BRIGHT_CYAN, style=BOLD))

    runlog.finish(f"{total} file(s) found")
    if pause_rerun():
        search_zip_file_contents_db()


def search_docinfo_records_menu(pattern=None, scope=None):
    """AWS S3 -> Search DocInfo Manager Records: search filenames across
    BOTH tables FM logs to — zipFileContent (files inside logged zips) and
    fileLog (individual files uploaded without zipping) — and show one
    merged table. Unlike Search AWS S3 (prefix-only, against live S3
    listings), this is a cheap server-side DB LIKE query, so full wildcard
    matching ('*', '?') stays supported here."""
    screen("Search DocInfo Manager Records")
    print()

    if not ensure_authenticated():
        print(color_text("  DocInfo Manager login is required to search DocInfo Manager records.", fg=RED))
        _log_line("Search DocInfo Manager Records cancelled — DocInfo Manager login required")
        pause_return(); return

    logzip, err = _load_logzip_config()
    if err:
        print(color_text(f"  {err}", fg=RED))
        pause_return(); return

    if pattern is None:
        pattern = ask("Filename pattern (wildcards ok)", "*")
        if not pattern:
            pause_return(); return
    if scope is None:
        scope = ask_path("Folder to scope the search (blank = search everything)")
    folder = os.path.abspath(clean_path(scope)) if scope else ""

    print(f"  {YELLOW}Pattern{RESET}: {pattern}")
    print(f"  {YELLOW}Scope{RESET}: {folder + ' (+ subfolders)' if folder else 'everything logged'}")
    print()

    runlog = _RunLog("Search DocInfo Manager Records", [("Pattern", pattern), ("Folder", folder)])

    rows = []
    error = None
    try:
        zip_data = _post_zip_content_search(logzip, pattern, "", folder)
        if str(zip_data.get("success")) == "1":
            for row in zip_data.get("recordList", []):
                rows.append({
                    "source": "Zip", "name": row.get("fileName"),
                    "location": row.get("zipFileName"),
                    "size": row.get("fileSizeFormatted", ""),
                    "s3": "",
                })
        else:
            error = zip_data.get("msg", "zip content search failed")
    except Exception as e:
        error = f"zip content search failed: {e}"

    try:
        file_data = _post_file_log_search(logzip, pattern, folder)
        if str(file_data.get("success")) == "1":
            for row in file_data.get("recordList", []):
                s3 = f"s3://{row['s3Bucket']}/{row['s3Key']}" if row.get("s3Bucket") else ""
                rows.append({
                    "source": "File", "name": row.get("fileName"),
                    "location": row.get("fileFolder", ""),
                    "size": row.get("fileSizeFormatted", ""),
                    "s3": s3,
                })
        elif not error:
            error = file_data.get("msg", "file log search failed")
    except Exception as e:
        error = f"{error}; " if error else ""
        error += f"file log search failed: {e}"

    if error:
        print(color_text(f"  [ERROR] {error}", fg=RED))

    if not rows:
        print(color_text("  No matching records found.", fg=YELLOW))
    else:
        rows.sort(key=lambda r: (r["source"], r["name"] or ""))
        name_w = max(len(r["name"] or "") for r in rows)
        for row in rows:
            tag = color_text(f"[{row['source']}]", fg=CYAN)
            s3_note = f"  {DIM}{row['s3']}{RESET}" if row["s3"] else ""
            print(f"  {tag} {row['name']:<{name_w}}  {DIM}({row['size']}) {row['location']}{RESET}{s3_note}")
        print()
        print(color_text(f"  {len(rows)} record(s) found.", fg=BRIGHT_CYAN, style=BOLD))

    runlog.finish(f"{len(rows)} record(s) found" + (f" — {error}" if error else ""))
    if pause_rerun():
        search_docinfo_records_menu()


def zip_menu():
    while True:
        log_zip_reason = None if is_token_valid() else "login required — see Admin Menu"
        options = [
            ("View Zip",
             "Inspect a zip WITHOUT extracting it: each entry's uncompressed and "
             "compressed size, compression ratio, and modified date, plus totals. "
             "Give a single .zip file to view it directly, or give a folder to pick "
             "from the zips inside it."),
            ("Find Files in Zip",
             "Search for filenames INSIDE zip archives — matching is by "
             "filename only, contents are never searched. Wildcards are "
             "supported (e.g. fan*.png).\n"
             "Give a single .zip file to search just that archive, or give a "
             "folder to recursively find every .zip under it and search each "
             "one — results are grouped by which zip they were found in."),
            ("Log Zip File Contents",
             "Log a .zip or .tar archive — or every archive in a folder — to "
             "the CB9Inventory database on BPA5 via the DocInfo Manager API. "
             "A folder is scanned top level only unless you answer Yes to "
             "'Include subfolders?'. .gz files are ignored.\n"
             "• Each archive is recorded in zipFile (matched by name + size: "
             "insert when new, update when seen before). An MD5 hash of the "
             "whole archive is stored too (zipFileHash), for a future "
             "duplicate finder — it isn't part of the match yet.\n"
             "• Its file listing is synced to zipFileContent by filePath: "
             "files new to the archive are inserted; files whose name, "
             "size, modified date, or extension changed are updated "
             "(reactivated if previously deleted); unchanged files are "
             "left alone; files no longer in the archive are soft-deleted "
             "(deleted = 1) — never removed outright.\n"
             "• Every zipFileContent change is preserved first: a database "
             "trigger copies the row's prior state into "
             "zipFileContentArchive before ANY update applies, including "
             "the soft-delete. Re-logging an existing zip archives the "
             "zipFile row's prior state into zipFileArchive the same way.\n"
             "• If an archive can't be read (corrupt/not a zip), it's still "
             "recorded — with its size and zero content rows — with a "
             "Failed status instead of being skipped.\n"
             "Settings come from fmConfig.json (logZip: apiUrl, "
             "serverSecretKey). Requires a DocInfo Manager login (Admin "
             "Menu).",
             log_zip_reason),
            ("Zip SubFolders",
             "Zip each immediate subfolder of a target folder into its own "
             ".zip file in a destination folder (defaults to the current "
             "directory). Before zipping it removes .DS_Store and "
             "desktop.ini files, and it resolves name collisions "
             "automatically (name.zip, name-2.zip, …).\n"
             "You'll be asked whether to delete each source folder after it "
             "zips successfully — that removal only happens on a "
             "successful zip. You'll also be asked whether to push each "
             "resulting zip to AWS S3 (credentials saved via Admin Menu)."),
            ("Zip SubFolders & Log",
             "Same as Zip SubFolders (including the optional Push to AWS S3), and "
             "afterward also logs every zip created in the destination folder to "
             "the CB9Inventory database — same as Log Zip File Contents. Requires "
             "a DocInfo Manager login (Admin Menu).",
             log_zip_reason),
            ("Search Zip File Contents by Name",
             "Search for filenames across every archive already LOGGED to the "
             "CB9Inventory database on BPA5 — distinct from 'Find Files in "
             "Zip', which searches local zip files directly instead of the "
             "database. Matching is by filename only, contents are never "
             "searched. Wildcards are supported (e.g. fan*.png) — '*' and "
             "'?' are translated to SQL '%'/'_'; a pattern with no wildcard "
             "is an exact match (e.g. johncabibbo*.doc -> fileName LIKE "
             "'johncabibbo%.doc').\n"
             "Optionally scope the search to a single zip file (exact name) "
             "or a folder (matches archives logged from that folder or any "
             "subfolder beneath it) — leave blank to search every logged "
             "zip. Requires a DocInfo Manager login (Admin Menu).",
             log_zip_reason),
        ]
        ch = render_menu("Zip", options)
        if ch == "back":
            return
        try:
            if ch == "1":
                target = ask_path("Zip file or folder to view")
                if target:
                    zip_view(target)
            elif ch == "2":
                target = ask_path("Zip file or folder to search")
                if not target:
                    pause_return(); continue
                pattern = ask("Filename pattern (wildcards ok)", "*")
                find_files_in_zip_target(target, pattern or "*")
            elif ch == "3":
                target = ask_path("Zip/tar file or folder to log",
                             default=LAST_PATHS.get("zip_log_target", ""))
                if target:
                    recursive = False
                    if os.path.isdir(clean_path(target)):
                        recursive = safe_confirm("  Include subfolders?", default=False)
                    log_zip_files(target, recursive=recursive)
            elif ch == "4":
                prompted = _prompt_zip_subfolders()
                if prompted is None:
                    pause_return(); continue
                zip_subfolders(*prompted)
            elif ch == "5":
                prompted = _prompt_zip_subfolders()
                if prompted is None:
                    pause_return(); continue
                zip_subfolders(*prompted)
                log_zip_files(prompted[1])
            elif ch == "6":
                search_zip_file_contents_db()
        except EscCancelled:
            continue


# ---- Optimize Media ---------------------------------------------------------
VIDEO_CONVERT_EXTS = {".mov", ".avi", ".mkv", ".wmv", ".flv", ".m4v"}
IMAGE_CONVERT_EXTS = {".heic", ".heif", ".bmp", ".tiff", ".tif", ".webp", ".gif"}
IMAGE_CONVERT_OUT_FORMATS = ("jpg", "png")


def _find_convertible_videos(source, recursive):
    """Relative paths (from source) of video files with a convertible
    extension (VIDEO_CONVERT_EXTS), sorted. .mp4 files are left alone —
    they're already the target format."""
    out = []
    if recursive:
        for dp, dns, fns in os.walk(source):
            prune_dirs(dns)
            for fn in fns:
                if fn.startswith(".") or is_excluded_file(fn):
                    continue
                if os.path.splitext(fn)[1].lower() in VIDEO_CONVERT_EXTS:
                    out.append(os.path.relpath(os.path.join(dp, fn), source))
    else:
        try:
            for fn in sorted(os.listdir(source)):
                full = os.path.join(source, fn)
                if fn.startswith(".") or not os.path.isfile(full):
                    continue
                if os.path.splitext(fn)[1].lower() in VIDEO_CONVERT_EXTS:
                    out.append(fn)
        except OSError:
            pass
    return sorted(out)


def _validate_converted_mp4(out_path, source_duration):
    """Sanity-check a freshly converted .mp4 before its original is deleted:
    the file must exist, be non-empty, reopen cleanly, and have a duration
    close to the source's (within 2%, minimum 1 second of slack — re-encoding
    can shift frame-accurate duration slightly). Returns (ok, reason);
    reason is only set when ok is False."""
    if not os.path.isfile(out_path) or os.path.getsize(out_path) == 0:
        return False, "output file missing or empty"
    check_clip = None
    try:
        check_clip = VideoFileClip(out_path)
        duration = check_clip.duration
    except Exception as e:
        return False, f"could not reopen output ({e})"
    finally:
        if check_clip is not None:
            check_clip.close()
    if not duration:
        return False, "output has no readable duration"
    if source_duration:
        tolerance = max(1.0, source_duration * 0.02)
        if abs(duration - source_duration) > tolerance:
            return False, f"duration mismatch (source {source_duration:.1f}s vs output {duration:.1f}s)"
    return True, ""


def _prompt_optimize_media():
    """Shared prompt sequence for Optimize Media → Convert Video to MP4.
    Returns (source, dest, recursive, remove_after, pause_seconds, flatten), or
    None if cancelled."""
    source = ask_folder("Source folder", key="optimize_media_source")
    if not source:
        return None
    dest = ask_path("Destination folder (blank = source folder)", default=source)
    dest = clean_path(dest) if dest else source
    recursive = safe_confirm("  Include subfolders?", default=False)
    flatten = False
    if recursive and os.path.abspath(dest) != os.path.abspath(source):
        mirror = safe_confirm(
            "  Recreate matching subfolders in the destination? "
            "(No = place all converted files in one folder)", default=True)
        flatten = not mirror
    remove_after = safe_confirm(
        "  Clean Up — delete original files after a successful conversion?", default=False)
    pause_raw = ask("  Pause between files, in seconds (0 = no pause)", default="5")
    try:
        pause_seconds = max(0.0, float(pause_raw))
    except (TypeError, ValueError):
        pause_seconds = 5.0
    return source, dest, recursive, remove_after, pause_seconds, flatten


def _spin_while_writing(clip, out_path, prefix):
    """Run clip.write_videofile() in a background thread while animating a
    spinner + elapsed time right after `prefix` (already printed with
    end=""), so a single large/slow conversion doesn't look frozen — moviepy
    itself prints nothing here (logger=None), and a multi-minute encode with
    no visible change is indistinguishable from a hang otherwise. Leaves the
    cursor positioned right after `prefix` when done, ready for the caller's
    own trailing OK/FAILED print. Any exception from the write is re-raised
    in the calling thread."""
    err = []

    def _target():
        try:
            clip.write_videofile(out_path, codec="libx264", audio_codec="aac", logger=None)
        except Exception as e:
            err.append(e)

    t = threading.Thread(target=_target, daemon=True)
    t.start()
    frames = "|/-\\"
    start = time.time()
    i = 0
    while t.is_alive():
        elapsed = time.time() - start
        spin = f" {frames[i % len(frames)]} {elapsed:>4.0f}s working…"
        sys.stdout.write("\r" + prefix + color_text(spin, fg=DIM))
        sys.stdout.flush()
        i += 1
        t.join(0.2)
    sys.stdout.write("\r" + prefix + " " * 24 + "\r" + prefix)
    sys.stdout.flush()
    if err:
        raise err[0]


def _convert_video_files_core(files, source, dest, flatten, remove_after, pause_seconds, runlog):
    """Core per-file MP4 conversion loop — no screen()/prompt/pause_return()
    I/O, so it's callable directly from the AWS S3 wizard as well as from
    convert_videos_to_mp4()'s interactive/CLI wrapper below. Prints its own
    progress lines (same as before the refactor) and drives `runlog`, but
    leaves confirmation, the countdown, and the final report to the caller.

    Returns (ok, fail, removed, kept, stopped_early)."""
    ok = 0
    fail = 0
    removed = 0
    kept = 0
    stopped_early = False
    ps = _PauseStop()
    if ps.is_tty and len(files) > 1:
        print(color_text("  Press [P] to pause, [Q] to stop early.", fg=DIM))
    try:
        for i, rel in enumerate(files, 1):
            if ps.check() == "stop":
                stopped_early = True
                break
            if not os.path.isdir(dest):
                print(color_text(f"  Destination folder is no longer reachable: {dest}", fg=RED))
                print(color_text("  Stopping — check that the drive is still connected/mounted.", fg=RED))
                runlog.action(f"Destination folder no longer reachable: {dest} — stopping")
                fail += len(files) - i + 1
                break

            src_path = os.path.join(source, rel)
            rel_dir = "" if flatten else os.path.dirname(rel)
            out_dir = os.path.join(dest, rel_dir) if rel_dir else dest
            name = os.path.splitext(os.path.basename(rel))[0]
            label = rel if len(rel) <= 40 else "…" + rel[-39:]
            src_size = fmt_size(file_size(src_path))
            prefix = (f"    {CYAN}[{i}/{len(files)}]{RESET} {WHITE}{label:<40}{RESET} "
                      f"{DIM}({src_size:>9}){RESET}")
            print(prefix, end="", flush=True)
            clip = None
            try:
                os.makedirs(out_dir, exist_ok=True)
                out_path = os.path.join(out_dir, name + ".mp4")
                suffix = 2
                while os.path.exists(out_path):
                    out_path = os.path.join(out_dir, f"{name}-{suffix}.mp4")
                    suffix += 1
                base = os.path.basename(out_path)

                clip = VideoFileClip(src_path)
                source_duration = clip.duration
                _spin_while_writing(clip, out_path, prefix)
                clip.close()
                clip = None
                ok += 1
                note = f" [{base}]" if base != name + ".mp4" else ""
                size_note = f"{fmt_size(file_size(out_path))}{note}"

                if remove_after:
                    valid, reason = _validate_converted_mp4(out_path, source_duration)
                    if valid:
                        try:
                            os.remove(src_path); removed += 1
                        except OSError:
                            pass
                        print(f" {BRIGHT_GREEN}OK{RESET} {DIM}({size_note}){RESET} {GREEN}— validated, original removed{RESET}")
                        runlog.action(f"Converted {rel} -> {base} ({size_note}) — original removed")
                    else:
                        kept += 1
                        print(f" {BRIGHT_GREEN}OK{RESET} {DIM}({size_note}){RESET} {BRIGHT_YELLOW}— {reason}, original KEPT{RESET}")
                        runlog.action(f"Converted {rel} -> {base} ({size_note}) — original KEPT ({reason})")
                else:
                    print(f" {BRIGHT_GREEN}OK{RESET} {DIM}({size_note}){RESET}")
                    runlog.action(f"Converted {rel} -> {base} ({size_note})")
            except Exception as e:
                print(f" {BRIGHT_RED}FAILED{RESET} {DIM}({e}){RESET}")
                fail += 1
                runlog.action(f"FAILED to convert {rel} ({e})")
            finally:
                if clip is not None:
                    clip.close()

            if pause_seconds and i < len(files):
                print(color_text(f"        Pausing {pause_seconds:g}s (drive cooldown)…", fg=DIM), end="\r")
                time.sleep(pause_seconds)
                print(" " * 60, end="\r")
    finally:
        ps.close()

    return ok, fail, removed, kept, stopped_early


def convert_videos_to_mp4(source, dest=None, recursive=False, remove_after=False, pause_seconds=0,
                           flatten=False):
    with _ActivityLog():
        screen("Convert Video to MP4")
        print()
        source = clean_path(source)
        dest = clean_path(dest) if dest else source
        dest_differs = os.path.abspath(dest) != os.path.abspath(source)
        flatten = bool(flatten) and recursive and dest_differs
        try:
            pause_seconds = max(0.0, float(pause_seconds))
        except (TypeError, ValueError):
            pause_seconds = 0.0
        print(f"  {YELLOW}Source{RESET}    : {source}")
        print(f"  {YELLOW}Dest{RESET}      : {dest}")
        print(f"  {YELLOW}Subfolders{RESET}: {'Yes' if recursive else 'No'}")
        if recursive and dest_differs:
            print(f"  {YELLOW}Subfolder Mode{RESET}: {'Flatten — all files into Dest' if flatten else 'Mirror — recreate subfolders in Dest'}")
        print(f"  {YELLOW}Clean Up (delete originals){RESET}: {'Yes' if remove_after else 'No'}")
        print(f"  {YELLOW}Pause between files{RESET}: {f'{pause_seconds:g}s' if pause_seconds else 'None'}\n")

        if not os.path.isdir(source):
            print(color_text(f"  Not a directory: {source}", fg=RED)); pause_return(); return

        if not _MOVIEPY_AVAILABLE:
            print(color_text("  moviepy not installed — skipping. Run: pip install moviepy", fg=RED))
            pause_return(); return

        files = _find_convertible_videos(source, recursive)
        if not files:
            exts = ", ".join(sorted(VIDEO_CONVERT_EXTS))
            print(color_text(f"  No convertible video files found ({exts}).", fg=YELLOW))
            pause_return(); return

        if remove_after:
            remove_after = safe_confirm(
                color_text("  Delete each original file after a successful conversion?", fg=YELLOW),
                default=False)

        print()
        if not safe_confirm(f"  Convert these {len(files)} file(s) to MP4?", default=True):
            print(color_text("  Cancelled — nothing converted.", fg=YELLOW))
            pause_return(); return
        if not _run_countdown(label="Starting conversion"):
            print(color_text("  Cancelled — nothing converted.", fg=YELLOW))
            pause_return(); return

        os.makedirs(dest, exist_ok=True)
        print(color_text(f"\n  Converting {len(files)} file(s)…", fg=BRIGHT_CYAN, style=BOLD))
        runlog = _RunLog("Convert Video to MP4", [("Source", source), ("Dest", dest)],
                          total_items=len(files))

        ok, fail, removed, kept, stopped_early = _convert_video_files_core(
            files, source, dest, flatten, remove_after, pause_seconds, runlog)

        print()
        summary = f"Converted {ok} file(s) to MP4."
        if remove_after:
            summary += f" Removed {removed} original(s)."
            if kept:
                summary += f" Kept {kept} original(s) (validation failed)."
        if fail:
            summary += f" {fail} failed."
        if stopped_early:
            summary += f" Stopped early by user ({ok + fail}/{len(files)} processed)."
        runlog.finish(summary)
        report_result(fail == 0, summary, f"Converted {ok}, {fail} failed.")
        pause_return()


def _find_convertible_images(source, recursive):
    """Relative paths (from source) of image files with a convertible
    extension (IMAGE_CONVERT_EXTS), sorted. .jpg/.jpeg/.png files are left
    alone — they're already common target formats."""
    out = []
    if recursive:
        for dp, dns, fns in os.walk(source):
            prune_dirs(dns)
            for fn in fns:
                if fn.startswith(".") or is_excluded_file(fn):
                    continue
                if os.path.splitext(fn)[1].lower() in IMAGE_CONVERT_EXTS:
                    out.append(os.path.relpath(os.path.join(dp, fn), source))
    else:
        try:
            for fn in sorted(os.listdir(source)):
                full = os.path.join(source, fn)
                if fn.startswith(".") or not os.path.isfile(full):
                    continue
                if os.path.splitext(fn)[1].lower() in IMAGE_CONVERT_EXTS:
                    out.append(fn)
        except OSError:
            pass
    return sorted(out)


def _convert_image_file(src_path, out_path, out_format):
    """Open src_path with Pillow (HEIC/HEIF via the pillow_heif opener
    registered at import time) and save it as out_path in out_format
    (jpg or png). EXIF orientation is baked into the pixels via
    ImageOps.exif_transpose so rotated iPhone photos come out right-side
    up — Pillow doesn't apply that automatically on save."""
    with Image.open(src_path) as img:
        img = ImageOps.exif_transpose(img) or img
        if out_format == "jpg":
            if img.mode in ("RGBA", "LA", "P"):
                img = img.convert("RGB")
            img.save(out_path, "JPEG", quality=92)
        else:
            img.save(out_path, "PNG")


def _validate_converted_image(out_path):
    """Sanity-check a freshly converted image before its original is
    deleted: the file must exist, be non-empty, and reopen/verify cleanly.
    Returns (ok, reason); reason is only set when ok is False."""
    if not os.path.isfile(out_path) or os.path.getsize(out_path) == 0:
        return False, "output file missing or empty"
    try:
        with Image.open(out_path) as chk:
            chk.verify()
    except Exception as e:
        return False, f"could not reopen output ({e})"
    return True, ""


def _prompt_optimize_images():
    """Shared prompt sequence for Optimize Media → Convert Images.
    Returns (source, dest, recursive, out_format, remove_after,
    pause_seconds, flatten), or None if cancelled."""
    source = ask_folder("Source folder", key="optimize_images_source")
    if not source:
        return None
    dest = ask_path("Destination folder (blank = source folder)", default=source)
    dest = clean_path(dest) if dest else source
    recursive = safe_confirm("  Include subfolders?", default=False)
    flatten = False
    if recursive and os.path.abspath(dest) != os.path.abspath(source):
        mirror = safe_confirm(
            "  Recreate matching subfolders in the destination? "
            "(No = place all converted files in one folder)", default=True)
        flatten = not mirror
    format_options = [
        ("JPG", "Smaller files, no transparency — best for photos."),
        ("PNG", "Lossless, keeps transparency — best for graphics/screenshots."),
    ]
    ch = render_menu("Convert Images — Output Format", format_options)
    if ch == "back":
        return None
    out_format = IMAGE_CONVERT_OUT_FORMATS[int(ch) - 1]
    remove_after = safe_confirm(
        "  Clean Up — delete original files after a successful conversion?", default=False)
    pause_raw = ask("  Pause between files, in seconds (0 = no pause)", default="5")
    try:
        pause_seconds = max(0.0, float(pause_raw))
    except (TypeError, ValueError):
        pause_seconds = 5.0
    return source, dest, recursive, out_format, remove_after, pause_seconds, flatten


def _convert_image_files_core(files, source, dest, out_format, remove_after, pause_seconds,
                               runlog, flatten, resize_to=None):
    """Core per-file image conversion loop — no screen()/prompt/pause_return()
    I/O, so it's callable directly from the AWS S3 wizard as well as from
    convert_images_to_format()'s interactive/CLI wrapper below.

    resize_to=(max_width, max_height), when given, fits the image within that
    bounding box (preserve aspect ratio, never upscale — Image.thumbnail())
    BEFORE the collision-safe output filename is picked, so the filename can
    carry the image's ACTUAL resulting height/width as a '_HEIGHT-WIDTH'
    suffix (e.g. photo_900-1200.jpg) — the box you asked for, not what you
    typed. resize_to=None (the default, used by plain Convert Images)
    reproduces the exact pre-refactor behavior/filenames.

    Returns (ok, fail, removed, kept, stopped_early)."""
    ok = 0
    fail = 0
    removed = 0
    kept = 0
    stopped_early = False
    ps = _PauseStop()
    if ps.is_tty and len(files) > 1:
        print(color_text("  Press [P] to pause, [Q] to stop early.", fg=DIM))
    try:
        for i, rel in enumerate(files, 1):
            if ps.check() == "stop":
                stopped_early = True
                break
            if not os.path.isdir(dest):
                print(color_text(f"  Destination folder is no longer reachable: {dest}", fg=RED))
                print(color_text("  Stopping — check that the drive is still connected/mounted.", fg=RED))
                runlog.action(f"Destination folder no longer reachable: {dest} — stopping")
                fail += len(files) - i + 1
                break

            src_path = os.path.join(source, rel)
            rel_dir = "" if flatten else os.path.dirname(rel)
            out_dir = os.path.join(dest, rel_dir) if rel_dir else dest
            name = os.path.splitext(os.path.basename(rel))[0]
            label = rel if len(rel) <= 40 else "…" + rel[-39:]
            print(f"    {CYAN}[{i}/{len(files)}]{RESET} {WHITE}{label:<40}{RESET}", end="", flush=True)
            try:
                os.makedirs(out_dir, exist_ok=True)

                if resize_to:
                    with Image.open(src_path) as img:
                        img = ImageOps.exif_transpose(img) or img
                        img.thumbnail(resize_to, Image.LANCZOS)
                        w, h = img.size
                        # Auto-detect alpha rather than trusting a fixed
                        # out_format here — a transparent source (webp/gif
                        # with alpha) resized straight to JPG would silently
                        # lose its transparency. PNG when the source has
                        # alpha, out_format otherwise.
                        has_alpha = img.mode in ("RGBA", "LA") or (img.mode == "P" and "transparency" in img.info)
                        resize_format = "png" if has_alpha else out_format
                        base_name = f"{name}_{h}-{w}"
                        out_path = os.path.join(out_dir, f"{base_name}.{resize_format}")
                        suffix = 2
                        while os.path.exists(out_path):
                            out_path = os.path.join(out_dir, f"{base_name}-{suffix}.{resize_format}")
                            suffix += 1
                        if resize_format == "jpg":
                            if img.mode in ("RGBA", "LA", "P"):
                                img = img.convert("RGB")
                            img.save(out_path, "JPEG", quality=92)
                        else:
                            img.save(out_path, "PNG")
                else:
                    out_path = os.path.join(out_dir, f"{name}.{out_format}")
                    suffix = 2
                    while os.path.exists(out_path):
                        out_path = os.path.join(out_dir, f"{name}-{suffix}.{out_format}")
                        suffix += 1
                    _convert_image_file(src_path, out_path, out_format)

                base = os.path.basename(out_path)
                ok += 1
                default_base = f"{name}.{out_format}" if not resize_to else None
                note = f" [{base}]" if (default_base and base != default_base) else ""
                size_note = f"{fmt_size(file_size(out_path))}{note}"

                if remove_after:
                    valid, reason = _validate_converted_image(out_path)
                    if valid:
                        try:
                            os.remove(src_path); removed += 1
                        except OSError:
                            pass
                        print(f" {BRIGHT_GREEN}OK{RESET} {DIM}({size_note}){RESET} {GREEN}— validated, original removed{RESET}")
                        runlog.action(f"Converted {rel} -> {base} ({size_note}) — original removed")
                    else:
                        kept += 1
                        print(f" {BRIGHT_GREEN}OK{RESET} {DIM}({size_note}){RESET} {BRIGHT_YELLOW}— {reason}, original KEPT{RESET}")
                        runlog.action(f"Converted {rel} -> {base} ({size_note}) — original KEPT ({reason})")
                else:
                    print(f" {BRIGHT_GREEN}OK{RESET} {DIM}({size_note}){RESET}")
                    runlog.action(f"Converted {rel} -> {base} ({size_note})")
            except Exception as e:
                print(f" {BRIGHT_RED}FAILED{RESET} {DIM}({e}){RESET}")
                fail += 1
                runlog.action(f"FAILED to convert {rel} ({e})")

            if pause_seconds and i < len(files):
                print(color_text(f"        Pausing {pause_seconds:g}s (drive cooldown)…", fg=DIM), end="\r")
                time.sleep(pause_seconds)
                print(" " * 60, end="\r")
    finally:
        ps.close()

    return ok, fail, removed, kept, stopped_early


def convert_images_to_format(source, dest=None, recursive=False, out_format="jpg",
                              remove_after=False, pause_seconds=0, flatten=False):
    with _ActivityLog():
        screen("Convert Images")
        print()
        source = clean_path(source)
        dest = clean_path(dest) if dest else source
        dest_differs = os.path.abspath(dest) != os.path.abspath(source)
        flatten = bool(flatten) and recursive and dest_differs
        out_format = out_format.lower() if out_format else "jpg"
        if out_format not in IMAGE_CONVERT_OUT_FORMATS:
            out_format = "jpg"
        try:
            pause_seconds = max(0.0, float(pause_seconds))
        except (TypeError, ValueError):
            pause_seconds = 0.0
        print(f"  {YELLOW}Source{RESET}       : {source}")
        print(f"  {YELLOW}Dest{RESET}         : {dest}")
        print(f"  {YELLOW}Subfolders{RESET}   : {'Yes' if recursive else 'No'}")
        if recursive and dest_differs:
            print(f"  {YELLOW}Subfolder Mode{RESET}: {'Flatten — all files into Dest' if flatten else 'Mirror — recreate subfolders in Dest'}")
        print(f"  {YELLOW}Output Format{RESET}: {out_format.upper()}")
        print(f"  {YELLOW}Clean Up (delete originals){RESET}: {'Yes' if remove_after else 'No'}")
        print(f"  {YELLOW}Pause between files{RESET}: {f'{pause_seconds:g}s' if pause_seconds else 'None'}\n")

        if not os.path.isdir(source):
            print(color_text(f"  Not a directory: {source}", fg=RED)); pause_return(); return

        if not _IMAGE_CONVERT_AVAILABLE:
            missing = []
            if not _PIL_AVAILABLE:
                missing.append("Pillow")
            if not _HEIF_AVAILABLE:
                missing.append("pillow-heif")
            print(color_text(f"  {' and '.join(missing)} not installed — skipping. "
                             f"Run: pip install {' '.join(m.lower() for m in missing)}", fg=RED))
            pause_return(); return

        files = _find_convertible_images(source, recursive)
        if not files:
            exts = ", ".join(sorted(IMAGE_CONVERT_EXTS))
            print(color_text(f"  No convertible image files found ({exts}).", fg=YELLOW))
            pause_return(); return

        if remove_after:
            remove_after = safe_confirm(
                color_text("  Delete each original file after a successful conversion?", fg=YELLOW),
                default=False)

        print()
        if not safe_confirm(f"  Convert these {len(files)} file(s) to {out_format.upper()}?", default=True):
            print(color_text("  Cancelled — nothing converted.", fg=YELLOW))
            pause_return(); return
        if not _run_countdown(label="Starting conversion"):
            print(color_text("  Cancelled — nothing converted.", fg=YELLOW))
            pause_return(); return

        os.makedirs(dest, exist_ok=True)
        print(color_text(f"\n  Converting {len(files)} file(s)…", fg=BRIGHT_CYAN, style=BOLD))
        runlog = _RunLog("Convert Images", [("Source", source), ("Dest", dest)],
                          total_items=len(files))

        ok, fail, removed, kept, stopped_early = _convert_image_files_core(
            files, source, dest, out_format, remove_after, pause_seconds, runlog, flatten)

        print()
        summary = f"Converted {ok} file(s) to {out_format.upper()}."
        if remove_after:
            summary += f" Removed {removed} original(s)."
            if kept:
                summary += f" Kept {kept} original(s) (validation failed)."
        if fail:
            summary += f" {fail} failed."
        if stopped_early:
            summary += f" Stopped early by user ({ok + fail}/{len(files)} processed)."
        runlog.finish(summary)
        report_result(fail == 0, summary, f"Converted {ok}, {fail} failed.")
        pause_return()


THUMBNAIL_MAX_PX = 80


def _create_thumbnails(files, source, dest, runlog, flatten, max_px=THUMBNAIL_MAX_PX):
    """Core thumbnail-generation loop for the AWS S3 wizard's 'Create Media
    Thumbnails' step. Same discovery scope as Convert Images/Resize
    (IMAGE_CONVERT_EXTS), same no-screen/no-pause_return() shape as
    _convert_video_files_core()/_convert_image_files_core() — menu-only
    prompts live in the wizard, not here.

    Each thumbnail fits within max_px x max_px (preserve aspect ratio, never
    upscale — Image.thumbnail()) and is saved as '<name>_tmb.<ext>' beside
    the original, collision-safe. Saved as PNG when the source has an alpha
    channel (keeps transparency), JPG otherwise. No delete-originals option
    — thumbnails are a new derived file, never a replacement for the source.

    Returns (ok, fail, stopped_early)."""
    ok = 0
    fail = 0
    stopped_early = False
    ps = _PauseStop()
    if ps.is_tty and len(files) > 1:
        print(color_text("  Press [P] to pause, [Q] to stop early.", fg=DIM))
    try:
        for i, rel in enumerate(files, 1):
            if ps.check() == "stop":
                stopped_early = True
                break
            if not os.path.isdir(dest):
                print(color_text(f"  Destination folder is no longer reachable: {dest}", fg=RED))
                runlog.action(f"Destination folder no longer reachable: {dest} — stopping")
                fail += len(files) - i + 1
                break

            src_path = os.path.join(source, rel)
            rel_dir = "" if flatten else os.path.dirname(rel)
            out_dir = os.path.join(dest, rel_dir) if rel_dir else dest
            name = os.path.splitext(os.path.basename(rel))[0]
            label = rel if len(rel) <= 40 else "…" + rel[-39:]
            print(f"    {CYAN}[{i}/{len(files)}]{RESET} {WHITE}{label:<40}{RESET}", end="", flush=True)
            try:
                os.makedirs(out_dir, exist_ok=True)
                with Image.open(src_path) as img:
                    img = ImageOps.exif_transpose(img) or img
                    img.thumbnail((max_px, max_px), Image.LANCZOS)
                    has_alpha = img.mode in ("RGBA", "LA") or (img.mode == "P" and "transparency" in img.info)
                    out_format = "png" if has_alpha else "jpg"

                    base_name = f"{name}_tmb"
                    out_path = os.path.join(out_dir, f"{base_name}.{out_format}")
                    suffix = 2
                    while os.path.exists(out_path):
                        out_path = os.path.join(out_dir, f"{base_name}-{suffix}.{out_format}")
                        suffix += 1

                    if out_format == "jpg":
                        if img.mode in ("RGBA", "LA", "P"):
                            img = img.convert("RGB")
                        img.save(out_path, "JPEG", quality=85)
                    else:
                        img.save(out_path, "PNG")

                ok += 1
                base = os.path.basename(out_path)
                size_note = fmt_size(file_size(out_path))
                print(f" {BRIGHT_GREEN}OK{RESET} {DIM}({size_note}, {base}){RESET}")
                runlog.action(f"Thumbnail {rel} -> {base} ({size_note})")
            except Exception as e:
                print(f" {BRIGHT_RED}FAILED{RESET} {DIM}({e}){RESET}")
                fail += 1
                runlog.action(f"FAILED to thumbnail {rel} ({e})")
    finally:
        ps.close()

    return ok, fail, stopped_early


def optimize_media_menu():
    while True:
        moviepy_reason = None if _MOVIEPY_AVAILABLE else "moviepy not installed — run: pip install moviepy"
        image_reason = (None if _IMAGE_CONVERT_AVAILABLE
                         else "Pillow/pillow-heif not installed — run: pip install Pillow pillow-heif")
        options = [
            ("Convert Video to MP4",
             "Convert video files of other formats (.mov, .avi, .mkv, .wmv, "
             ".flv, .m4v) to .mp4 (H.264 video + AAC audio), using moviepy.\n"
             "• Source / destination — enter a source folder and, "
             "optionally, a destination folder (defaults to the source "
             "folder if left blank), and whether to include subfolders. If "
             "subfolders are included AND the destination differs from the "
             "source, you're asked whether to recreate the matching "
             "subfolders in the destination (default) or flatten every "
             "converted file into that one destination folder instead.\n"
             "• Naming — collision-safe (name.mp4, name-2.mp4, …); never "
             "overwrites.\n"
             "• Clean Up — optionally deletes each original file after it "
             "converts successfully. Off by default, confirmed again right "
             "before it happens, and only after the new .mp4 is validated "
             "(reopened, checked for a matching duration). If validation "
             "fails the original is kept and flagged instead of deleted.\n"
             "• Pause — a pause (default 5s) between files, to let an "
             "external drive cool down during a long batch; 0 disables it.",
             moviepy_reason),
            ("Convert Images",
             "Convert image files of other formats (.heic, .heif, .bmp, "
             ".tiff, .webp, .gif) to JPG or PNG (your choice each run), "
             "using Pillow + pillow-heif (for HEIC/HEIF, e.g. iPhone "
             "photos). EXIF orientation is applied so rotated photos come "
             "out right-side up.\n"
             "• Source / destination — enter a source folder and, "
             "optionally, a destination folder (defaults to the source "
             "folder if left blank), and whether to include subfolders. If "
             "subfolders are included AND the destination differs from the "
             "source, you're asked whether to recreate the matching "
             "subfolders in the destination (default) or flatten every "
             "converted file into that one destination folder instead.\n"
             "• Naming — collision-safe (name.jpg, name-2.jpg, …); never "
             "overwrites.\n"
             "• Clean Up — optionally deletes each original file after it "
             "converts successfully. Off by default, confirmed again right "
             "before it happens, and only after the new image is validated "
             "(reopened and verified). If validation fails the original is "
             "kept and flagged instead of deleted.\n"
             "• Pause — a pause (default 5s) between files, to let an "
             "external drive cool down during a long batch; 0 disables it.",
             image_reason),
        ]
        ch = render_menu("Optimize Media", options)
        if ch == "back":
            return
        try:
            if ch == "1":
                prompted = _prompt_optimize_media()
                if prompted is None:
                    pause_return(); continue
                convert_videos_to_mp4(*prompted)
            elif ch == "2":
                prompted = _prompt_optimize_images()
                if prompted is None:
                    pause_return(); continue
                convert_images_to_format(*prompted)
        except EscCancelled:
            continue


# =============================================================================
# PERMISSIONS
#   - Set Apache Permissions: one-shot chown/chmod fix for /var/www (sudo),
#     grayed out when /var/www doesn't exist (Apache not installed).
#   - Run a Profile / Set a Profile: fmConfig.json's permissionProfiles list
#     (folder + file/file-type pattern + a single octal permission +
#     recursive), same shape as syncProfiles/monitorProfiles/compareProfiles.
# =============================================================================
APACHE_DOCROOT = "/var/www"


def _apache_available():
    return os.path.isdir(APACHE_DOCROOT)


def _apache_owner_group():
    """(owner, group) used to chown Apache's docroot, by platform."""
    if sys.platform == "darwin":
        return "_www", "_www"
    return "root", "www-data"


def _load_permission_profiles():
    return _load_config_profiles("permissionProfiles")


def _valid_octal_permission(text):
    """True for a 3- or 4-digit chmod octal string (each digit 0-7),
    e.g. '644', '755', '0755'."""
    text = text.strip()
    return len(text) in (3, 4) and text.isdigit() and all(c in "01234567" for c in text)


def _normalize_permission_pattern(pattern):
    """A bare extension with no dot ('php', 'log') becomes a '*.ext' wildcard
    so it matches every file with that extension. Anything else — an existing
    wildcard pattern ('*.php', 'IMG_*.jpg'), a dotted name, or an exact
    filename ('.htaccess', 'config.json') — is matched exactly as typed."""
    p = pattern.strip()
    if not p:
        return "*"
    if "." in p or any(c in p for c in "*?["):
        return p
    return f"*.{p}"


def _scan_permission_matches(folder, pattern, recursive):
    """Files under `folder` matching `pattern` (fnmatch). Hidden files/dotfiles
    ARE included (permission fixes commonly target .htaccess and similar), but
    the usual junk names (.DS_Store, desktop.ini, $RECYCLE.BIN) are skipped.
    Returns a sorted list of full paths."""
    out = []
    if recursive:
        for dp, dns, fns in os.walk(folder):
            dns[:] = [d for d in dns if d.lower() not in EXCLUDED_DIR_NAMES]
            for fn in fns:
                if is_excluded_file(fn):
                    continue
                if fnmatch.fnmatch(fn, pattern):
                    out.append(os.path.join(dp, fn))
    else:
        try:
            for fn in os.listdir(folder):
                full = os.path.join(folder, fn)
                if not os.path.isfile(full) or is_excluded_file(fn):
                    continue
                if fnmatch.fnmatch(fn, pattern):
                    out.append(full)
        except OSError:
            pass
    return sorted(out)


def _preview_permission_changes(changes, already_ok):
    """changes: list of (path, current_octal, target_octal). Prints a
    numbered preview table."""
    print(color_text(f"  {len(changes)} file(s) need a permission change:", fg=BRIGHT_CYAN, style=BOLD))
    for i, (path, cur, new) in enumerate(changes, 1):
        print(f"    {i:>3}. {path}  {DIM}({cur} -> {RESET}{YELLOW}{new}{RESET}{DIM}){RESET}")
    if not changes:
        print(f"    {DIM}(nothing to change){RESET}")
    if already_ok:
        print(f"  {DIM}{already_ok} file(s) already at the target permission.{RESET}")


def _apply_permission_changes(changes, runlog=None):
    """Actually chmod. Returns (ok_count, fail_count, stopped_early).
    [P] pauses / [Q] stops early (interactive tty only)."""
    ok = 0
    fail = 0
    stopped_early = False
    ps = _PauseStop()
    if ps.is_tty and len(changes) > 1:
        print(color_text("  Press [P] to pause, [Q] to stop early.", fg=DIM))
    try:
        for path, cur, new in changes:
            if ps.check() == "stop":
                stopped_early = True
                break
            try:
                os.chmod(path, int(new, 8))
                ok += 1
                if runlog:
                    runlog.action(f"Changed {path} ({cur} -> {new})")
            except OSError as e:
                fail += 1
                if runlog:
                    runlog.action(f"FAILED to change {path}: {e}")
    finally:
        ps.close()
    return ok, fail, stopped_early


def _run_permission_profile(pr):
    """Run one permissionProfiles entry from fmConfig.json: scan, preview,
    confirm, apply (dry-run until confirmed, same as every other FM
    permission/removal/copy flow)."""
    with _ActivityLog():
        name      = pr.get("name") or "(unnamed profile)"
        folder    = clean_path(str(pr.get("folder", "")))
        pattern   = _normalize_permission_pattern(str(pr.get("pattern", "*")))
        target    = str(pr.get("permission", "644")).strip()
        recursive = bool(pr.get("recursive", True))

        screen(f"Permissions — Run Profile: {name}")
        print()
        print(f"  {YELLOW}Folder{RESET}    : {folder}")
        print(f"  {YELLOW}Pattern{RESET}   : {pattern}")
        print(f"  {YELLOW}Permission{RESET}: {target}")
        print(f"  {YELLOW}Recursive{RESET} : {'Yes' if recursive else 'No'}\n")

        if not os.path.isdir(folder):
            print(color_text(f"  Not a directory: {folder}", fg=RED)); pause_return(); return
        if not _valid_octal_permission(target):
            print(color_text(f"  Invalid permission in profile: '{target}' (must be 3-4 octal digits, e.g. 644)", fg=RED))
            pause_return(); return

        matches = _scan_permission_matches(folder, pattern, recursive)
        changes = []
        already_ok = 0
        for path in matches:
            try:
                cur = oct(stat.S_IMODE(os.stat(path).st_mode))[-3:]
            except OSError:
                continue
            new = target[-3:]
            if cur == new:
                already_ok += 1
            else:
                changes.append((path, cur, new))

        _preview_permission_changes(changes, already_ok)
        if not changes:
            print()
            print(color_text("  DRY RUN — nothing to change.", fg=YELLOW, style=BOLD))
            pause_return(); return

        print()
        print(color_text("  This was a DRY RUN — nothing has been changed yet.", fg=YELLOW, style=BOLD))
        if not safe_confirm(f"  Actually change permission on these {len(changes)} file(s)?", default=False):
            print(color_text("  Cancelled — nothing changed.", fg=YELLOW))
            pause_return(); return
        if not _run_countdown(label="Starting permission change"):
            print(color_text("  Cancelled — nothing changed.", fg=YELLOW))
            pause_return(); return

        runlog = _RunLog("Permissions — Run Profile", [("Profile", name), ("Folder", folder),
                          ("Pattern", pattern), ("Permission", target)], total_items=len(changes))
        ok, fail, stopped_early = _apply_permission_changes(changes, runlog=runlog)
        print()
        summary = f"Changed {ok} file(s)." + (f" {fail} failed." if fail else "")
        if stopped_early:
            summary += f" Stopped early by user ({ok + fail}/{len(changes)} processed)."
        runlog.finish(summary)
        report_result(fail == 0, summary, f"Changed {ok}, {fail} failed.")
        pause_return()


def run_permission_profile_menu():
    """[Run a Profile] picker."""
    profiles, perr = _load_permission_profiles()
    if perr:
        screen("Permissions — Run a Profile")
        print()
        print(color_text(f"  ⚠ {perr}", fg=RED))
        pause_return()
        return
    if not profiles:
        screen("Permissions — Run a Profile")
        print()
        print(color_text("  No saved Permission profiles yet.", fg=YELLOW))
        print("  Create one first via Permissions -> Set a Profile.")
        pause_return()
        return
    options = []
    for pr in profiles:
        name = pr.get("name") or "(unnamed profile)"
        options.append((
            f"Profile: {name}",
            f"Change permission on files matching this saved profile:\n"
            f"• Folder     — {pr.get('folder', '?')}\n"
            f"• Pattern    — {pr.get('pattern', '*')}\n"
            f"• Permission — {pr.get('permission', '644')}\n"
            f"• Recursive: {'Yes' if pr.get('recursive', True) else 'No'}"))
    ch = render_menu("Permissions — Run a Profile", options)
    if ch == "back":
        return
    _run_permission_profile(profiles[int(ch) - 1])


def _edit_permission_profile_flow(existing, profiles, idx):
    """Prompt for a permission profile's fields (pre-filled with `existing`'s
    values when editing) and save it into `profiles` at `idx` (None = append
    a new profile). Same shape as Admin Menu's Set/Update AWS Credentials:
    Enter keeps the current value when editing."""
    print()
    default_name = existing.get("name", "") if existing else ""
    name = ask("Profile name", default=default_name).strip() or default_name
    if not name:
        print(color_text("  Cancelled — a name is required.", fg=YELLOW))
        return

    default_folder = existing.get("folder", "") if existing else ""
    folder = ask_folder("Folder path", default=default_folder, must_exist=True)
    if not folder:
        if not default_folder:
            print(color_text("  Cancelled — a folder is required.", fg=YELLOW))
            return
        folder = default_folder

    default_pattern = existing.get("pattern", "*") if existing else "*"
    print(color_text("  Enter an exact filename (.htaccess, config.json), a bare "
                      "extension with no dot to match every file of that type "
                      "(php, log), or a full wildcard pattern (*.php, IMG_*.jpg).",
                      style=DIM))
    pattern = ask("File name or file type", default=default_pattern).strip() or default_pattern

    default_perm = existing.get("permission", "644") if existing else "644"
    perm = ask("Permission (octal, e.g. 644 or 755)", default=default_perm).strip() or default_perm
    if not _valid_octal_permission(perm):
        print(color_text(f"  Cancelled — '{perm}' isn't a valid permission (3-4 octal digits, e.g. 644).", fg=YELLOW))
        return

    default_recursive = bool(existing.get("recursive", True)) if existing else True
    recursive = safe_confirm("  Include subfolders (recursive)?", default=default_recursive)

    profile = {
        "name": name,
        "folder": folder,
        "pattern": pattern,
        "permission": perm,
        "recursive": recursive,
    }
    if idx is None:
        profiles.append(profile)
    else:
        profiles[idx] = profile
    err = _save_config_profiles("permissionProfiles", profiles)
    if err:
        print(color_text(f"  ⚠ {err}", fg=RED))
    else:
        verb = "Updated" if idx is not None else "Saved"
        print(color_text(f"  ✓ {verb} profile '{name}'.", fg=GREEN))


def set_permission_profile_menu():
    """[Set a Profile] — pick an existing profile to edit, or create a new one."""
    profiles, perr = _load_permission_profiles()
    if perr:
        screen("Permissions — Set a Profile")
        print()
        print(color_text(f"  ⚠ {perr}", fg=RED))
        pause_return()
        return
    options = [(f"Edit: {p.get('name') or '(unnamed profile)'}",
                f"Folder: {p.get('folder', '?')}   Pattern: {p.get('pattern', '*')}   "
                f"Permission: {p.get('permission', '644')}") for p in profiles]
    options.append(("New Profile", "Create a new saved permission profile."))
    ch = render_menu("Permissions — Set a Profile", options)
    if ch == "back":
        return
    picked = int(ch) - 1
    screen("Permissions — Set a Profile")
    if picked == len(profiles):
        _edit_permission_profile_flow(None, profiles, None)
    else:
        _edit_permission_profile_flow(profiles[picked], profiles, picked)
    pause_return()


def set_apache_permissions():
    """One-shot fix for Apache's docroot: chown -R (root:www-data on Linux,
    _www:_www on macOS) and chmod (directories 755, files 644), via sudo.
    Only reachable when /var/www exists (see permissions_menu())."""
    with _ActivityLog():
        screen("Set Apache Permissions")
        print()
        owner, group = _apache_owner_group()
        print(f"  {YELLOW}Target{RESET}: {APACHE_DOCROOT}")
        print(f"  {YELLOW}Owner{RESET} : {owner}:{group}")
        print(f"  {YELLOW}Mode{RESET}  : directories 755, files 644\n")
        print(color_text("  This runs chown/chmod via sudo and may prompt for your password.", fg=DIM))

        if not safe_confirm(f"  Reset ownership and permissions under {APACHE_DOCROOT}?", default=False):
            print(color_text("  Cancelled — nothing changed.", fg=YELLOW))
            pause_return(); return
        if not _run_countdown(label="Starting permission reset"):
            print(color_text("  Cancelled — nothing changed.", fg=YELLOW))
            pause_return(); return

        runlog = _RunLog("Set Apache Permissions", [("Target", APACHE_DOCROOT), ("Owner", f"{owner}:{group}")])
        steps = [
            ("Ownership", ["sudo", "chown", "-R", f"{owner}:{group}", APACHE_DOCROOT]),
            ("Directory permissions (755)", ["sudo", "find", APACHE_DOCROOT, "-type", "d", "-exec", "chmod", "755", "{}", "+"]),
            ("File permissions (644)", ["sudo", "find", APACHE_DOCROOT, "-type", "f", "-exec", "chmod", "644", "{}", "+"]),
        ]
        all_ok = True
        for label, cmd in steps:
            print(f"\n  {CYAN}{label}...{RESET}", flush=True)
            rc = subprocess.run(cmd).returncode
            if rc == 0:
                print(color_text(f"  {label}: OK", fg=BRIGHT_GREEN))
                runlog.action(f"{label}: OK")
            else:
                print(color_text(f"  {label}: FAILED (exit {rc})", fg=BRIGHT_RED))
                runlog.action(f"{label}: FAILED (exit {rc})")
                all_ok = False

        print()
        summary = f"Ownership {owner}:{group}, directories 755, files 644 under {APACHE_DOCROOT}"
        if not all_ok:
            summary += " — one or more steps failed"
        runlog.finish(summary)
        report_result(all_ok, summary, "One or more steps failed — see above.")
        pause_return()


def permissions_menu():
    apache_reason = None if _apache_available() else f"{APACHE_DOCROOT} not found — Apache not installed"
    options = [
        ("Set Apache Permissions",
         f"Reset ownership and permissions under {APACHE_DOCROOT} so Apache can "
         "serve it: directories to 755, files to 644, owned by the Apache "
         "user/group (root:www-data on Linux, _www:_www on macOS) via sudo "
         "(may prompt for your password). Grayed out when Apache isn't "
         f"installed ({APACHE_DOCROOT} doesn't exist).",
         apache_reason),
        ("Run a Profile",
         "Pick a saved permission profile (fmConfig.json -> permissionProfiles) "
         "and apply it: scans the profile's folder for files matching its "
         "pattern, previews every file whose permission would change, and "
         "changes it only after you confirm — a DRY RUN until then."),
        ("Set a Profile",
         "Create a new saved permission profile, or edit an existing one: a "
         "folder path, a file name or file type/wildcard pattern to match, a "
         "single octal permission (e.g. 644) to set on every match, and "
         "whether to include subfolders."),
    ]
    while True:
        ch = render_menu("Permissions", options)
        if ch == "back":
            return
        try:
            if ch == "1":
                set_apache_permissions()
            elif ch == "2":
                run_permission_profile_menu()
            elif ch == "3":
                set_permission_profile_menu()
        except EscCancelled:
            continue


def cleanup_menu():
    options = [
        ("Remove Junk Files (.DS_Store / desktop.ini / *.bak)",
         "First choose which file types to include — all three are checked "
         "by default; use ↑/↓ and Space to unselect any you want to keep, "
         "then Enter. Then enter a root folder; every matching file beneath "
         "it (hidden folders included) is found and listed. A DRY RUN "
         "preview shows each file and its size — nothing is deleted until "
         "you confirm."),
        ("Purge Old Log Files",
         "Trim old entries out of every .log file in ~/Documents/log (or a "
         "folder you choose). You pick how many days to keep (default 90); "
         "entries with a [YYYY-MM-DD HH:MM:SS] timestamp older than that are "
         "removed — an entry's untimestamped body lines go with their header. "
         "A DRY RUN table shows purge/keep line counts per file first, and a "
         ".bak backup of each changed file is saved before it is rewritten."),
    ]
    while True:
        ch = render_menu("Clean Up", options,
                         intro="Junk-file cleanup · log purging — dry run first, always.")
        if ch == "back":
            return
        try:
            if ch == "1":
                spec_options = [(label, desc) for label, desc, _m in JUNK_FILE_SPECS]
                picks = render_multiselect(
                    "Clean Up — Remove Junk Files", spec_options,
                    intro="File types to remove — unselect any you want to keep.",
                    preselected=range(len(JUNK_FILE_SPECS)))
                if picks is None:
                    continue
                if not picks:
                    print(color_text("  No file types selected.", fg=YELLOW)); pause_return(); continue
                root = ask_folder("Root folder to clean", default=os.path.expanduser("~"), key="cleanup_junk_root")
                if not root:
                    pause_return(); continue
                cleanup_junk_files(root, spec_indices=picks)
                pause_return()
            elif ch == "2":
                folder = ask_folder("Log folder", default=LOG_PURGE_DIR, key="cleanup_purge_log_folder")
                if not folder:
                    pause_return(); continue
                days = ask("Days to keep", default=str(LOG_PURGE_DAYS))
                cleanup_purge_logs(folder, days)
                pause_return()
        except EscCancelled:
            continue


# =============================================================================
# MOUNT SHARES  (SMB, via `open smb://...` — macOS only)
# =============================================================================
# /Volumes is root-owned (drwxr-xr-x root:wheel) — a plain user can't mkdir a
# new entry there directly, so mount_smbfs run straight from a shell fails
# for any share that hasn't been mounted before. `open smb://user@host/share`
# hands off to macOS's own mount service (the same mechanism Finder's
# Cmd+K/"Connect to Server" uses), which is allowed to create the mount
# point. A literal "@" in the username (e.g. an email-style login) is
# percent-encoded as %40 since it would otherwise collide with the
# user@host separator in the smb:// URL.
def _load_mount_profiles():
    return _load_config_profiles("mountProfiles")


def _mount_one_share(server, username, share):
    """Fire off `open smb://...` for one share. Non-blocking — Finder
    handles the actual mount (and any password prompt) independently."""
    encoded_user = username.replace("@", "%40")
    url = f"smb://{encoded_user}@{server}/{share}"
    subprocess.Popen(["open", url], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)


def _run_mount_shares(server, username, shares, runlog):
    """Issue `open smb://` for every share, then briefly poll /Volumes to
    report which ones mounted right away (a saved Keychain credential —
    common on a repeat connection) versus which are still pending (most
    likely a Finder password prompt waiting for input — NOT a failure, just
    not yet confirmable from here)."""
    for share in shares:
        print(f"  Requesting smb://{username}@{server}/{share} ...")
        _mount_one_share(server, username, share)
        runlog.action(f"Requested mount: {share}")
        time.sleep(1)

    print()
    print(color_text("  Checking /Volumes for shares that mounted immediately "
                     "(saved credentials)...", fg=DIM))
    time.sleep(3)
    mounted, pending = [], []
    for share in shares:
        if os.path.ismount(os.path.join("/Volumes", share)):
            mounted.append(share)
        else:
            pending.append(share)

    print()
    if mounted:
        print(color_text(f"  ✓ Mounted: {', '.join(mounted)}", fg=BRIGHT_GREEN))
    if pending:
        print(color_text(f"  ⏳ Still pending — check Finder for a password prompt, or "
                         f"double-check the server/username/share name: {', '.join(pending)}",
                         fg=BRIGHT_YELLOW))
    summary = f"{len(mounted)}/{len(shares)} mounted immediately, {len(pending)} pending."
    runlog.finish(summary)
    report_result(True, summary)
    pause_return()


def mount_shares_manual():
    with _ActivityLog():
        screen("Mount Shares — Manual")
        print()
        server = ask("Server address or hostname").strip()
        if not server:
            print(color_text("  Cancelled — a server is required.", fg=YELLOW)); pause_return(); return
        username = ask("Username").strip()
        if not username:
            print(color_text("  Cancelled — a username is required.", fg=YELLOW)); pause_return(); return
        shares_raw = ask("Share name(s) — comma-separated for more than one").strip()
        shares = [s.strip() for s in shares_raw.split(",") if s.strip()]
        if not shares:
            print(color_text("  Cancelled — at least one share name is required.", fg=YELLOW))
            pause_return(); return

        print()
        print(f"  {YELLOW}Server{RESET}  : {server}")
        print(f"  {YELLOW}Username{RESET}: {username}")
        print(f"  {YELLOW}Shares{RESET}  : {', '.join(shares)}\n")

        if not safe_confirm(f"  Mount these {len(shares)} share(s)?", default=True):
            print(color_text("  Cancelled — nothing mounted.", fg=YELLOW)); pause_return(); return
        if not _run_countdown(label="Starting mount"):
            print(color_text("  Cancelled — nothing mounted.", fg=YELLOW)); pause_return(); return

        runlog = _RunLog("Mount Shares — Manual", [("Server", server), ("Username", username),
                          ("Shares", ", ".join(shares))], total_items=len(shares))
        _run_mount_shares(server, username, shares, runlog)


def _run_mount_profile(pr):
    """Run one mountProfiles entry from fmConfig.json."""
    with _ActivityLog():
        name     = pr.get("name") or "(unnamed profile)"
        server   = str(pr.get("server", "")).strip()
        username = str(pr.get("username", "")).strip()
        shares   = [s for s in pr.get("shares", []) if isinstance(s, str) and s.strip()]

        screen(f"Mount Shares — Run Profile: {name}")
        print()
        print(f"  {YELLOW}Server{RESET}  : {server}")
        print(f"  {YELLOW}Username{RESET}: {username}")
        print(f"  {YELLOW}Shares{RESET}  : {', '.join(shares)}\n")

        if not server or not username or not shares:
            print(color_text("  Profile is missing a server, username, or share list.", fg=RED))
            pause_return(); return

        if not safe_confirm(f"  Mount these {len(shares)} share(s)?", default=True):
            print(color_text("  Cancelled — nothing mounted.", fg=YELLOW)); pause_return(); return
        if not _run_countdown(label="Starting mount"):
            print(color_text("  Cancelled — nothing mounted.", fg=YELLOW)); pause_return(); return

        runlog = _RunLog("Mount Shares — Run Profile", [("Profile", name), ("Server", server),
                          ("Username", username), ("Shares", ", ".join(shares))], total_items=len(shares))
        _run_mount_shares(server, username, shares, runlog)


def run_mount_profile_menu():
    """[Run a Profile] picker."""
    profiles, perr = _load_mount_profiles()
    if perr:
        screen("Mount Shares — Run a Profile")
        print()
        print(color_text(f"  ⚠ {perr}", fg=RED))
        pause_return()
        return
    if not profiles:
        screen("Mount Shares — Run a Profile")
        print()
        print(color_text("  No saved Mount profiles yet.", fg=YELLOW))
        print("  Create one first via Mount Shares -> Set a Profile.")
        pause_return()
        return
    options = []
    for pr in profiles:
        name = pr.get("name") or "(unnamed profile)"
        shares = [s for s in pr.get("shares", []) if isinstance(s, str) and s.strip()]
        options.append((
            f"Profile: {name}",
            f"Mount the shares saved in this profile:\n"
            f"• Server   — {pr.get('server', '?')}\n"
            f"• Username — {pr.get('username', '?')}\n"
            f"• Shares   — {', '.join(shares) or '(none)'}"))
    ch = render_menu("Mount Shares — Run a Profile", options)
    if ch == "back":
        return
    _run_mount_profile(profiles[int(ch) - 1])


def _edit_mount_profile_flow(existing, profiles, idx):
    """Prompt for a mount profile's fields (pre-filled with `existing`'s
    values when editing) and save it into `profiles` at `idx` (None = append
    a new profile). Enter keeps the current value when editing."""
    print()
    default_name = existing.get("name", "") if existing else ""
    name = ask("Profile name", default=default_name).strip() or default_name
    if not name:
        print(color_text("  Cancelled — a name is required.", fg=YELLOW))
        return

    default_server = existing.get("server", "") if existing else ""
    server = ask("Server address or hostname", default=default_server).strip() or default_server
    if not server:
        print(color_text("  Cancelled — a server is required.", fg=YELLOW))
        return

    default_username = existing.get("username", "") if existing else ""
    username = ask("Username", default=default_username).strip() or default_username
    if not username:
        print(color_text("  Cancelled — a username is required.", fg=YELLOW))
        return

    default_shares = ", ".join(existing.get("shares", [])) if existing else ""
    shares_raw = ask("Share name(s) — comma-separated for more than one",
                     default=default_shares).strip() or default_shares
    shares = [s.strip() for s in shares_raw.split(",") if s.strip()]
    if not shares:
        print(color_text("  Cancelled — at least one share name is required.", fg=YELLOW))
        return

    profile = {"name": name, "server": server, "username": username, "shares": shares}
    if idx is None:
        profiles.append(profile)
    else:
        profiles[idx] = profile
    err = _save_config_profiles("mountProfiles", profiles)
    if err:
        print(color_text(f"  ⚠ {err}", fg=RED))
    else:
        verb = "Updated" if idx is not None else "Saved"
        print(color_text(f"  ✓ {verb} profile '{name}'.", fg=GREEN))


def set_mount_profile_menu():
    """[Set a Profile] — pick an existing profile to edit, or create a new one."""
    profiles, perr = _load_mount_profiles()
    if perr:
        screen("Mount Shares — Set a Profile")
        print()
        print(color_text(f"  ⚠ {perr}", fg=RED))
        pause_return()
        return
    options = [(f"Edit: {p.get('name') or '(unnamed profile)'}",
                f"Server: {p.get('server', '?')}   Username: {p.get('username', '?')}   "
                f"Shares: {', '.join(p.get('shares', [])) or '(none)'}") for p in profiles]
    options.append(("New Profile", "Create a new saved mount profile."))
    ch = render_menu("Mount Shares — Set a Profile", options)
    if ch == "back":
        return
    picked = int(ch) - 1
    screen("Mount Shares — Set a Profile")
    if picked == len(profiles):
        _edit_mount_profile_flow(None, profiles, None)
    else:
        _edit_mount_profile_flow(profiles[picked], profiles, picked)
    pause_return()


def mount_shares_menu():
    options = [
        ("Mount Manually",
         "Enter a server, username, and one or more share names (comma-"
         "separated) and mount them right away — nothing is saved."),
        ("Run a Profile",
         "Pick a saved mount profile (fmConfig.json -> mountProfiles) and "
         "mount all of its shares."),
        ("Set a Profile",
         "Create a new saved mount profile, or edit an existing one: a "
         "name, server address, username, and share name list."),
    ]
    while True:
        ch = render_menu("Mount Shares", options,
                         intro="Mount SMB shares via Finder's mount service — macOS only.")
        if ch == "back":
            return
        try:
            if ch == "1":
                mount_shares_manual()
            elif ch == "2":
                run_mount_profile_menu()
            elif ch == "3":
                set_mount_profile_menu()
        except EscCancelled:
            continue


# =============================================================================
# ENVIRONMENT VARS
# =============================================================================
def _display_environment_vars():
    """List every environment variable in this process, name and value,
    sorted alphabetically (case-insensitive). Read-only.

    Deliberately NOT wrapped in _ActivityLog()/_RunLog() — environment
    variables commonly hold secrets (API keys, tokens, credentials), and
    this app is careful never to write those to fm.log (see Admin Menu:
    passwords and AWS keys are never logged either). This stays screen-only.
    """
    screen("Environment Vars — Display All")
    print()
    env = dict(os.environ)
    if not env:
        print(color_text("  No environment variables found.", fg=YELLOW))
        pause_return()
        return
    names = sorted(env.keys(), key=str.lower)
    name_w = max(len(n) for n in names)
    print(color_text(f"  {len(names)} environment variable(s):", fg=CYAN, style=BOLD))
    print()
    for name in names:
        print(f"  {color_text(f'{name:<{name_w}}', fg=YELLOW)}  {env[name]}")
    pause_return()


def environment_menu():
    options = [
        ("Display All",
         "List every environment variable currently set for this session — "
         "name and value, sorted alphabetically. Read-only; nothing is "
         "changed. Not written to the activity log, since environment "
         "variables commonly hold secrets (API keys, tokens, credentials)."),
    ]
    while True:
        ch = render_menu("Environment Vars", options)
        if ch == "back":
            return
        if ch == "1":
            _display_environment_vars()


def _open_terminal_tab_and_run(command):
    """Open a new tab in the terminal app FM is currently running in and run
    `command` there. Detected via $TERM_PROGRAM (iTerm.app -> iTerm2;
    anything else, including Apple_Terminal, falls back to Terminal.app —
    the common case and the only other terminal FM's osascript integrations
    target elsewhere). Returns (success: bool, message: str).

    Terminal.app has no native "new tab" scripting verb, so the standard
    workaround is used: simulate Cmd+T via System Events (requires
    Accessibility permission for Terminal/System Events, a one-time macOS
    prompt), then target the resulting front window with do script.
    """
    escaped = command.replace('\\', '\\\\').replace('"', '\\"')
    if os.environ.get('TERM_PROGRAM') == 'iTerm.app':
        app_label = "iTerm"
        script = (
            'tell application "iTerm"\n'
            '  tell current window\n'
            '    create tab with default profile\n'
            '    tell current session to write text "' + escaped + '"\n'
            '  end tell\n'
            'end tell'
        )
    else:
        app_label = "Terminal"
        script = (
            'tell application "Terminal"\n'
            '  activate\n'
            '  tell application "System Events" to keystroke "t" using command down\n'
            '  delay 0.3\n'
            '  do script "' + escaped + '" in front window\n'
            'end tell'
        )
    try:
        result = subprocess.run(["osascript", "-e", script],
                                capture_output=True, text=True, timeout=10)
    except Exception as e:
        return False, str(e)
    if result.returncode == 0:
        return True, f"Launched in a new {app_label} tab"
    return False, result.stderr.strip() or f"Failed to open a {app_label} tab"


# =============================================================================
# LOCAL SCRIPTS
# =============================================================================
def _load_local_scripts():
    """Load localScripts from fmConfig.json, sorted alphabetically by name
    (case-insensitive) — the list is a plain menu, so alphabetical order
    (the CB9 dropdown fallback sort) is the most useful default."""
    scripts, err = _load_config_profiles("localScripts")
    if not err:
        scripts.sort(key=lambda s: (s.get("name") or "").lower())
    return scripts, err


def _edit_local_script_flow(existing, scripts, idx):
    """Prompt for a local script's name + execution string (pre-filled with
    `existing`'s values when editing) and save it into `scripts` at `idx`
    (None = append a new one)."""
    print()
    default_name = existing.get("name", "") if existing else ""
    name = ask("Script name", default=default_name).strip() or default_name
    if not name:
        print(color_text("  Cancelled — a name is required.", fg=YELLOW))
        return

    default_command = existing.get("command", "") if existing else ""
    command = ask("Execution string", default=default_command).strip() or default_command
    if not command:
        print(color_text("  Cancelled — an execution string is required.", fg=YELLOW))
        return

    script = {"name": name, "command": command}
    if idx is None:
        scripts.append(script)
    else:
        scripts[idx] = script
    scripts.sort(key=lambda s: (s.get("name") or "").lower())
    err = _save_config_profiles("localScripts", scripts)
    if err:
        print(color_text(f"  ⚠ {err}", fg=RED))
    else:
        verb = "Updated" if idx is not None else "Saved"
        print(color_text(f"  ✓ {verb} script '{name}'.", fg=GREEN))


def _delete_local_script_flow(script, scripts, idx):
    """Confirm, then remove `scripts[idx]` and save."""
    name = script.get("name") or "(unnamed)"
    if not safe_confirm(f"  Delete local script '{name}'?", default=False):
        print(color_text("  Cancelled — nothing deleted.", fg=YELLOW))
        return
    del scripts[idx]
    err = _save_config_profiles("localScripts", scripts)
    if err:
        print(color_text(f"  ⚠ {err}", fg=RED))
    else:
        print(color_text(f"  ✓ Deleted local script '{name}'.", fg=GREEN))


def local_scripts_menu():
    """Main Menu -> Local Scripts: arrow-driven list of saved local scripts
    (fmConfig.json -> localScripts). [Enter] opens a new tab in whichever
    terminal app FM is running in (Terminal.app or iTerm2) and runs the
    highlighted script's execution string there — FM's own menu keeps
    running independently. Adding, editing, and deleting scripts is done
    from Admin Menu -> Manage Local Scripts, not from this screen."""
    selected = 0
    status_msg, status_color = "", GREEN
    while True:
        scripts, err = _load_local_scripts()
        if err:
            screen("Local Scripts")
            print()
            print(color_text(f"  ⚠ {err}", fg=RED))
            pause_return()
            return
        if scripts:
            selected = max(0, min(selected, len(scripts) - 1))

        screen("Local Scripts")
        print()
        if not scripts:
            print(color_text("  No local scripts saved yet — add one from "
                              "Admin Menu -> Manage Local Scripts.", fg=YELLOW))
        else:
            for i, s in enumerate(scripts):
                name = s.get("name") or "(unnamed)"
                command = s.get("command", "")
                if i == selected:
                    arrow = color_text("▶", fg=BRIGHT_CYAN, style=BOLD)
                    label = color_text(f"{i + 1}. {name}", fg=BRIGHT_CYAN, style=BOLD)
                else:
                    arrow = " "
                    label = f"{color_text(str(i + 1), fg=YELLOW)}. {name}"
                print(f"  {arrow} {label}")
                print(color_text(f"       {command}", style=DIM))
        print()
        if status_msg:
            print(color_text(f"  {status_msg}", fg=status_color))
            print()
            status_msg = ""
        standard_footer("[↑↓] Move   [Enter] Run   [Q/ESC] Back")
        sys.stdout.write(color_text(" Option: ", fg=CYAN, style=BOLD))
        sys.stdout.flush()

        key = read_key()
        if key == "UP" and scripts:
            selected = (selected - 1) % len(scripts)
        elif key == "DOWN" and scripts:
            selected = (selected + 1) % len(scripts)
        elif key in ("q", "Q", "ESC"):
            return
        elif key == "ENTER" and scripts:
            s = scripts[selected]
            name = s.get("name") or "(unnamed)"
            command = s.get("command", "")
            success, message = _open_terminal_tab_and_run(command)
            status_msg, status_color = message, (GREEN if success else RED)
            if success:
                _log_line(f"Local Script launched: {name} — {command}")
            else:
                _log_line(f"Local Script launch FAILED: {name} — {command} — {message}")


def manage_local_scripts_menu():
    """Admin Menu -> Manage Local Scripts: add/edit/delete the scripts
    offered on the Main Menu -> Local Scripts screen."""
    while True:
        scripts, err = _load_local_scripts()
        if err:
            screen("Manage Local Scripts")
            print()
            print(color_text(f"  ⚠ {err}", fg=RED))
            pause_return()
            return

        options = [("Add Local Script", "Save a new script name + execution string.")]
        for s in scripts:
            name = s.get("name") or "(unnamed)"
            command = s.get("command", "?")
            options.append((f"{name} - {command}", f"Edit or delete '{name}'."))
        ch = render_menu(
            "Manage Local Scripts", options,
            intro="These scripts appear on the Main Menu -> Local Scripts screen.")
        if ch == "back":
            return

        picked = int(ch) - 1
        screen("Manage Local Scripts")
        if picked == 0:
            _edit_local_script_flow(None, scripts, None)
            pause_return()
            continue

        idx = picked - 1
        script = scripts[idx]
        name = script.get("name") or "(unnamed)"
        action = render_menu(
            f"Manage Local Scripts — {name}",
            [("Edit", f"Change the name/execution string of '{name}'."),
             ("Delete", f"Remove '{name}' from the local scripts list.")])
        if action == "back":
            continue
        screen("Manage Local Scripts")
        if action == "1":
            _edit_local_script_flow(script, scripts, idx)
        elif action == "2":
            _delete_local_script_flow(script, scripts, idx)
        pause_return()


def _s3_iter_all_objects(client, bucket, status_label=None):
    """Yield every object dict ({Key, Size, LastModified, ...}) in `bucket`,
    paginating with boto3's built-in paginator. If status_label is given,
    prints a single overwriting progress line (object count so far) — S3 has
    no cheap way to know a bucket's size/object count up front, so this can
    run for a while on a large bucket."""
    paginator = client.get_paginator("list_objects_v2")
    seen = 0
    for page in paginator.paginate(Bucket=bucket):
        for obj in page.get("Contents", []):
            seen += 1
            if status_label and seen % 200 == 0:
                print(color_text(f"        {status_label}: {seen:,} object(s) scanned…", fg=DIM), end="\r")
            yield obj
    if status_label and seen:
        print(" " * 80, end="\r")


def display_buckets_menu():
    """AWS S3 -> Display Buckets & Sizes: every bucket the saved credentials
    can see, with total size and object count. S3 has no cheap per-bucket
    size API without extra CloudWatch permissions, so this lists every
    object in every bucket directly — can be slow on large buckets."""
    screen("AWS S3 — Display Buckets & Sizes")
    print()

    if not _BOTO3_AVAILABLE:
        print(color_text("  boto3 not installed. Run: pip install boto3", fg=RED))
        pause_return(); return
    if not is_aws_configured():
        print(color_text("  AWS credentials not configured — see Admin Menu.", fg=YELLOW))
        pause_return(); return

    runlog = _RunLog("AWS S3 — Display Buckets & Sizes")
    try:
        client = _s3_client()
        buckets = client.list_buckets().get("Buckets", [])
    except Exception as e:
        print(color_text(f"  [ERROR] Could not list buckets: {e}", fg=RED))
        runlog.finish(f"FAILED: {e}")
        pause_return(); return

    if not buckets:
        print(color_text("  No buckets found for these credentials.", fg=YELLOW))
        runlog.finish("0 buckets found")
        pause_return(); return

    rows = []
    for b in buckets:
        name = b["Name"]
        total_bytes = 0
        count = 0
        try:
            for obj in _s3_iter_all_objects(client, name, status_label=name):
                total_bytes += obj.get("Size", 0)
                count += 1
            rows.append((name, count, total_bytes, None))
        except Exception as e:
            rows.append((name, 0, 0, str(e)))

    name_w = max(len(r[0]) for r in rows)
    print(f"  {'Bucket':<{name_w}}  {'Objects':>10}  {'Size':>12}")
    print(f"  {'-'*name_w}  {'-'*10}  {'-'*12}")
    for name, count, total_bytes, error in rows:
        if error:
            print(f"  {name:<{name_w}}  {DIM}(could not read: {error}){RESET}")
        else:
            print(f"  {name:<{name_w}}  {count:>10,}  {fmt_size(total_bytes):>12}")

    runlog.finish(f"{len(rows)} bucket(s) listed")
    pause_return()


def search_aws_s3_menu():
    """AWS S3 -> Search AWS S3: filename PREFIX search across every object
    in every bucket the saved credentials can see. S3 has no server-side
    wildcard/glob support (only native prefix matching), so this is a
    startswith() match against each object's basename — not a full glob
    like Find Files in Zip. Slow on large buckets (every object must be
    listed); shows live progress."""
    screen("AWS S3 — Search AWS S3")
    print()

    if not _BOTO3_AVAILABLE:
        print(color_text("  boto3 not installed. Run: pip install boto3", fg=RED))
        pause_return(); return
    if not is_aws_configured():
        print(color_text("  AWS credentials not configured — see Admin Menu.", fg=YELLOW))
        pause_return(); return

    pattern = ask("Filename prefix to search for (no wildcards — e.g. IMG_2026)")
    if not pattern:
        pause_return(); return
    print(f"\n  {YELLOW}Prefix{RESET}: {pattern}")
    print(color_text("  Scanning every object in every bucket — this can take a while…\n", fg=DIM))

    runlog = _RunLog("AWS S3 — Search AWS S3", [("Prefix", pattern)])
    try:
        client = _s3_client()
        buckets = client.list_buckets().get("Buckets", [])
    except Exception as e:
        print(color_text(f"  [ERROR] Could not list buckets: {e}", fg=RED))
        runlog.finish(f"FAILED: {e}")
        pause_return(); return

    pattern_lower = pattern.lower()
    matches = []
    for b in buckets:
        name = b["Name"]
        try:
            for obj in _s3_iter_all_objects(client, name, status_label=name):
                key = obj["Key"]
                if os.path.basename(key).lower().startswith(pattern_lower):
                    matches.append((name, key, obj.get("Size", 0), obj.get("LastModified")))
        except Exception as e:
            print(color_text(f"        {name}: could not scan ({e})", fg=RED))

    if not matches:
        print(color_text("  No matching objects found.", fg=YELLOW))
    else:
        for bucket, key, size, modified in matches:
            mod_str = modified.strftime("%m/%d/%y %-I:%M %p") if modified else ""
            print(f"  {CYAN}{bucket}{RESET}/{key}  {DIM}({fmt_size(size)}, {mod_str}){RESET}")
        print()
        print(color_text(f"  {len(matches)} object(s) found.", fg=BRIGHT_CYAN, style=BOLD))

    runlog.finish(f"{len(matches)} object(s) found")
    if pause_rerun():
        search_aws_s3_menu()


def _aws_s3_wizard_media_files(mode, source, path, exts):
    """File discovery for one AWS S3 wizard phase: every matching file under
    `source` (mode='folder', recursive) or just `path` itself if it matches
    (mode='file'). Returns relative paths (from `source`), same shape
    _find_convertible_videos()/_find_convertible_images() already return."""
    if mode == "folder":
        return _find_convertible_videos(source, True) if exts == VIDEO_CONVERT_EXTS \
            else _find_convertible_images(source, True)
    ext = os.path.splitext(path)[1].lower()
    return [os.path.basename(path)] if ext in exts else []


def aws_s3_optimize_menu():
    """AWS S3 -> Optimize Media, Zip, Log & Upload to AWS S3 — the wizard.
    Gathers all 8 steps up front, then loops on a Review screen + 3…2…1
    countdown: ESC/Q during the countdown re-shows the Review screen with
    the same answers (not re-asked) instead of aborting outright; ESC/Q on
    any of the 8 questions themselves cancels the whole wizard back to the
    AWS S3 menu, same as every other multi-question flow in FM."""
    screen("AWS S3 — Optimize Media, Zip, Log & Upload")
    print()

    # ---- Step 1: Folder or File ----
    mode_options = [
        ("Folder", "Process every matching file found under a folder (recursively)."),
        ("File", "Process just one file."),
    ]
    ch = render_menu("Folder or File?", mode_options)
    if ch == "back":
        return
    mode = "folder" if ch == "1" else "file"
    path = ask_folder("Folder to process") if mode == "folder" else ask_file("File to process")
    if not path:
        pause_return(); return
    path = clean_path(path)
    source = path if mode == "folder" else os.path.dirname(os.path.abspath(path))

    # ---- Step 2: Optimize Video files ----
    do_video = safe_confirm("  Optimize Video files (convert to MP4)?", default=True)
    delete_video = safe_confirm("    Delete originals once converted?", default=False) if do_video else False

    # ---- Step 3: Resize Images ----
    do_resize = safe_confirm("  Resize Images?", default=True)
    resize_h, resize_w, delete_resize = 1200, 1200, False
    if do_resize:
        try:
            resize_h = max(1, int(ask("    Resize Height", default="1200")))
        except (TypeError, ValueError):
            resize_h = 1200
        try:
            resize_w = max(1, int(ask("    Resize Width", default="1200")))
        except (TypeError, ValueError):
            resize_w = 1200
        delete_resize = safe_confirm("    Delete Originals?", default=False)

    # ---- Step 4: Create Media Thumbnails ----
    do_thumbnails = safe_confirm("  Create Media Thumbnails?", default=True)

    # ---- Step 5/6: Zip (folder mode only — a single file isn't "zipped to
    # 1 file" or "zipped by subfolder" in any meaningful sense) ----
    do_zip_one = False
    do_zip_subfolders = False
    if mode == "folder":
        do_zip_one = safe_confirm("  Zip to 1 File?", default=False)
        if not do_zip_one:
            do_zip_subfolders = safe_confirm("  Zip Subfolders?", default=False)

    # ---- Step 7: Log Contents ----
    do_log = safe_confirm("  Log Contents to DocInfo Manager?", default=True)

    # ---- Step 8: AWS S3 Bucket ----
    aws_cfg = get_aws_config()
    bucket = ask("  AWS S3 Bucket", default=aws_cfg.get("bucket", ""))
    if not bucket:
        print(color_text("  Cancelled — a bucket is required.", fg=YELLOW))
        pause_return(); return
    s3_folder = ""
    if safe_confirm("    Folder? (S3 prefix inside the bucket)", default=False):
        s3_folder = ask("    S3 Folder/prefix", default="").strip("/")

    # ---- Review + countdown loop ----
    while True:
        screen("AWS S3 — Review")
        print()
        print(f"  {YELLOW}Mode{RESET}: {'Folder' if mode == 'folder' else 'File'} — {path}")
        video_note = f" (delete originals: {'Yes' if delete_video else 'No'})" if do_video else ""
        print(f"  {YELLOW}Optimize Video{RESET}: {'Yes' if do_video else 'No'}{video_note}")
        resize_note = f" ({resize_h}x{resize_w}, delete originals: {'Yes' if delete_resize else 'No'})" if do_resize else ""
        print(f"  {YELLOW}Resize Images{RESET}: {'Yes' if do_resize else 'No'}{resize_note}")
        print(f"  {YELLOW}Create Thumbnails{RESET}: {'Yes' if do_thumbnails else 'No'}")
        zip_label = "1 File" if do_zip_one else ("Subfolders" if do_zip_subfolders else "None")
        print(f"  {YELLOW}Zip{RESET}: {zip_label}")
        print(f"  {YELLOW}Log Contents{RESET}: {'Yes' if do_log else 'No'}")
        s3_dest = f"s3://{bucket}/{s3_folder}" if s3_folder else f"s3://{bucket}"
        print(f"  {YELLOW}Upload To{RESET}: {s3_dest}")
        print()

        if not safe_confirm("  Run this now?", default=True):
            print(color_text("  Cancelled — nothing done.", fg=YELLOW))
            pause_return(); return

        if _run_countdown(label="Starting"):
            break
        # ESC/Q during the countdown -> loop back to Review with the same
        # answers, per spec (not re-asked, not aborted).

    # ---- Execution ----
    logzip = None
    if do_log:
        logzip, logzip_err = _load_logzip_config()
        if not logzip:
            print(color_text(f"\n  {logzip_err} — Log Contents will be skipped.", fg=YELLOW))
            do_log = False
        elif not ensure_authenticated():
            print(color_text("\n  DocInfo Manager login required — Log Contents will be skipped.", fg=YELLOW))
            do_log = False
            logzip = None

    with _ActivityLog():
        screen("AWS S3 — Optimize Media, Zip, Log & Upload")
        print()
        runlog = _RunLog("AWS S3 Optimize/Zip/Log/Upload",
                          [("Mode", mode), ("Path", path), ("Bucket", bucket), ("S3 Folder", s3_folder or "(root)")])

        # Phase 1: Convert Video
        if do_video:
            print(color_text("\n  Phase: Optimize Video…", fg=BRIGHT_CYAN, style=BOLD))
            if not _MOVIEPY_AVAILABLE:
                print(color_text("    moviepy not installed — skipping. Run: pip install moviepy", fg=RED))
                runlog.action("Optimize Video skipped — moviepy not installed")
            else:
                video_files = _aws_s3_wizard_media_files(mode, source, path, VIDEO_CONVERT_EXTS)
                if video_files:
                    v_ok, v_fail, v_removed, v_kept, _stopped = _convert_video_files_core(
                        video_files, source, source, False, delete_video, 0, runlog)
                    note = f", removed {v_removed} original(s)" if delete_video else ""
                    print(f"    Converted {v_ok}, {v_fail} failed{note}.")
                else:
                    print(color_text("    No convertible video files found.", fg=DIM))

        # Phase 2: Resize Images
        if do_resize:
            print(color_text("\n  Phase: Resize Images…", fg=BRIGHT_CYAN, style=BOLD))
            if not _IMAGE_CONVERT_AVAILABLE:
                print(color_text("    Pillow/pillow-heif not installed — skipping.", fg=RED))
                runlog.action("Resize Images skipped — Pillow/pillow-heif not installed")
            else:
                image_files = _aws_s3_wizard_media_files(mode, source, path, IMAGE_CONVERT_EXTS)
                if image_files:
                    r_ok, r_fail, r_removed, r_kept, _stopped = _convert_image_files_core(
                        image_files, source, source, "jpg", delete_resize, 0, runlog, False,
                        resize_to=(resize_w, resize_h))
                    note = f", removed {r_removed} original(s)" if delete_resize else ""
                    print(f"    Resized {r_ok}, {r_fail} failed{note}.")
                else:
                    print(color_text("    No resizable image files found.", fg=DIM))

        # Phase 3: Thumbnails
        if do_thumbnails:
            print(color_text("\n  Phase: Create Media Thumbnails…", fg=BRIGHT_CYAN, style=BOLD))
            if not _IMAGE_CONVERT_AVAILABLE:
                print(color_text("    Pillow/pillow-heif not installed — skipping.", fg=RED))
            else:
                thumb_files = _aws_s3_wizard_media_files(mode, source, path, IMAGE_CONVERT_EXTS)
                if thumb_files:
                    t_ok, t_fail, _stopped = _create_thumbnails(thumb_files, source, source, runlog, False)
                    print(f"    Created {t_ok} thumbnail(s), {t_fail} failed.")
                else:
                    print(color_text("    No thumbnailable image files found.", fg=DIM))

        # Phase 4: Zip — "Zip to 1 File" and "Zip Subfolders" both build on
        # zip_folder_to_one_file(); "Subfolders" simply calls it once per
        # immediate subfolder rather than reusing the full interactive
        # zip_subfolders() screen (which has its own confirm/countdown — this
        # wizard already ran its own single review+countdown for everything).
        zip_paths = []
        if do_zip_one:
            print(color_text("\n  Phase: Zip to 1 File…", fg=BRIGHT_CYAN, style=BOLD))
            parent = os.path.dirname(source.rstrip("/\\")) or source
            dest_zip = os.path.join(parent, os.path.basename(source.rstrip("/\\")) + ".zip")
            zok, zpath, zcount = zip_folder_to_one_file(source, dest_zip, runlog)
            if zok:
                zip_paths.append(zpath)
                print(f"    Zipped {zcount} file(s) -> {os.path.basename(zpath)}")
            else:
                print(color_text("    Zip FAILED.", fg=RED))
        elif do_zip_subfolders:
            print(color_text("\n  Phase: Zip Subfolders…", fg=BRIGHT_CYAN, style=BOLD))
            subs = immediate_subfolders(source)
            if not subs:
                print(color_text("    No subfolders found.", fg=YELLOW))
            for sub_name in subs:
                sub_path = os.path.join(source, sub_name)
                dest_zip = os.path.join(source, sub_name + ".zip")
                zok, zpath, zcount = zip_folder_to_one_file(sub_path, dest_zip, runlog)
                if zok:
                    zip_paths.append(zpath)
                    print(f"    Zipped {sub_name} ({zcount} file(s)) -> {os.path.basename(zpath)}")
                else:
                    print(color_text(f"    Zip FAILED for {sub_name}.", fg=RED))

        # Phase 5: Log Contents + Upload
        print(color_text("\n  Phase: Log & Upload…", fg=BRIGHT_CYAN, style=BOLD))
        s3_ok = s3_fail = log_ok = log_fail = 0

        if zip_paths:
            # Zipped: upload + log each archive via the existing zip path
            # (api/zipFileLog.php), now carrying the S3 location too.
            for zpath in zip_paths:
                base = os.path.basename(zpath)
                uploaded, s3_key = _s3_push_file(zpath, bucket, s3_folder)
                if uploaded:
                    s3_ok += 1
                else:
                    s3_fail += 1
                if do_log:
                    try:
                        contents = _archive_contents(zpath)
                        load_error = None
                    except (OSError, ValueError, zipfile.BadZipFile, tarfile.TarError) as e:
                        contents, load_error = [], str(e)
                    try:
                        result = _post_zip_log(
                            logzip, zpath, contents, load_failed=bool(load_error),
                            s3_bucket=bucket if uploaded else "", s3_key=s3_key or "",
                            upload_confirmed=uploaded)
                        if str(result.get("success")) == "1":
                            log_ok += 1
                            runlog.action(f"Logged {base} — {result.get('action')}")
                        else:
                            log_fail += 1
                            runlog.action(f"FAILED to log {base}: {result.get('msg')}")
                    except Exception as e:
                        log_fail += 1
                        runlog.action(f"FAILED to log {base}: {e}")
        else:
            # No zip: upload + log every processed media file individually.
            media_exts = VIDEO_CONVERT_EXTS | IMAGE_CONVERT_EXTS | {".mp4", ".jpg", ".jpeg", ".png"}
            if mode == "folder":
                upload_files = []
                for dp, dns, fns in os.walk(source):
                    prune_dirs(dns)
                    for fn in fns:
                        if fn.startswith(".") or is_excluded_file(fn):
                            continue
                        if os.path.splitext(fn)[1].lower() in media_exts:
                            upload_files.append(os.path.relpath(os.path.join(dp, fn), source))
            else:
                upload_files = [os.path.basename(path)] if os.path.splitext(path)[1].lower() in media_exts else []

            for rel in sorted(upload_files):
                full_path = os.path.join(source, rel)
                uploaded, s3_key = _s3_push_file(full_path, bucket, s3_folder)
                if uploaded:
                    s3_ok += 1
                else:
                    s3_fail += 1
                if do_log:
                    try:
                        result = _post_file_log(
                            logzip, full_path, s3_bucket=bucket if uploaded else "",
                            s3_key=s3_key or "", upload_confirmed=uploaded, load_failed=not uploaded)
                        if str(result.get("success")) == "1":
                            log_ok += 1
                            runlog.action(f"Logged {rel} — {result.get('action')}")
                        else:
                            log_fail += 1
                            runlog.action(f"FAILED to log {rel}: {result.get('msg')}")
                    except Exception as e:
                        log_fail += 1
                        runlog.action(f"FAILED to log {rel}: {e}")

        print(f"    Uploaded {s3_ok}, {s3_fail} failed." + (f" Logged {log_ok}, {log_fail} failed." if do_log else ""))

        print()
        summary = f"AWS S3 wizard done — uploaded {s3_ok}, {s3_fail} failed."
        if do_log:
            summary += f" Logged {log_ok}, {log_fail} failed."
        runlog.finish(summary)
        report_result(s3_fail == 0 and log_fail == 0, summary, summary)
        pause_return()


def aws_s3_menu():
    options = [
        ("Display Buckets & Sizes",
         "List every AWS S3 bucket the saved credentials can see, with its "
         "total object count and size. S3 has no cheap per-bucket size API, "
         "so this lists every object in every bucket directly — can take a "
         "while on large buckets."),
        ("Optimize Media, Zip, Log & Upload to AWS S3",
         "A guided wizard: optimize video/images, resize images, create "
         "thumbnails, optionally zip the results, optionally log them to "
         "DocInfo Manager, and upload to an AWS S3 bucket — one folder (or "
         "file) at a time. Ends with a review screen and a 3…2…1 countdown; "
         "ESC during the countdown returns to the review screen with your "
         "answers intact, instead of cancelling outright."),
        ("Search AWS S3",
         "Search for a filename PREFIX across every object in every bucket "
         "the saved credentials can see (e.g. IMG_2026 finds "
         "IMG_2026_final.jpg, IMG_2026-vacation.png, …). S3 has no "
         "server-side wildcard support, so this is prefix matching only — "
         "not a full glob like Find Files in Zip. Slow on large buckets."),
        ("Search DocInfo Manager Records",
         "Search filenames across every record FM has logged to DocInfo "
         "Manager — both zipFileContent (files inside logged zips) and "
         "fileLog (individual files uploaded without zipping) — merged "
         "into one table. Full wildcard matching ('*', '?') is supported "
         "here (a cheap database search, unlike Search AWS S3's live "
         "object listing). Requires a DocInfo Manager login."),
    ]
    while True:
        ch = render_menu("AWS S3", options,
                         intro="Optimize, zip, log, and upload media to AWS S3 — plus bucket/record search.")
        if ch == "back":
            return
        try:
            if ch == "1":
                display_buckets_menu()
            elif ch == "2":
                aws_s3_optimize_menu()
            elif ch == "3":
                search_aws_s3_menu()
            elif ch == "4":
                search_docinfo_records_menu()
        except EscCancelled:
            continue


def about_screen():
    screen("About")
    print()
    print(color_text(f"  {SCRIPT_NAME} {VER}", fg=CYAN, style=BOLD))
    print("  File Manager — a unified interactive + CLI tool that combines the")
    print("  compare, display, eject, find, monitor, remove, sync, zip, optimize")
    print("  media, permissions, and clean-up file-management scripts into a")
    print("  single CB9Lib-based program.")
    print()
    print("  Maintainer: Cloud Box 9 Inc.")
    print(f"  Python: {platform.python_version()}")
    print()
    print("  Copyright © 2026 Cloud Box 9 Inc. All rights reserved.")
    pause_return()


def main_menu():
    options = [
        ("Display  — All drives, folder sizes",
         "• All Drives — size, used, and free space of every mounted drive "
         "(free matches Finder's Available).\n"
         "• Drives in Use — every mounted external drive and whether it's "
         "currently being read from or written to (samples actual disk I/O "
         "for a second, not just an open file). macOS only.\n"
         "• Subfolders (alphabetical or largest-first) — every immediate "
         "subfolder of a chosen folder with its total size "
         "(human-readable and exact bytes) and file count, ending with a "
         "grand total. Good for finding what's eating disk space."),
        ("Find     — Files by combined criteria, folders by name, duplicates",
         "• Find Files — matches one or more criteria AND-ed together "
         "(filename pattern, extension, size over N MB, size under N MB) "
         "— e.g. .mov files under 5 MB.\n"
         "• Find Folders — finds directories by name pattern.\n"
         "• Find Duplicates by Filename — scans one or more folders and "
         "tables files sharing the same name with a size column per "
         "folder.\n"
         "• Find Missing by Filename — compares two folders and tables "
         "the files present in only one of them.\n"
         "• Find & Replace — edits text inside files.\n"
         "• Find & Rename — renames the files themselves (prepend / "
         "append / replace text in the filename).\n"
         "Find & Replace and Find & Rename both dry-run first."),
        ("Compare  — Compare 2 files, or folder contents",
         "Two tools:\n"
         "• Compare 2 Files — shows two text files side by side with a "
         "line-by-line diff.\n"
         "• Compare Folder Contents — compares what's inside two folders "
         "by name and/or size, top-level or recursively; handy for "
         "checking a backup against the original."),
        ("Monitor  — File activity in a folder, real-time",
         "Watch a folder (or a saved profile from fmConfig.json) and report "
         "every created, modified, and deleted file as it happens — on screen "
         "and to fm.log or a CSV. Options: recursive (default yes) and a file "
         "extension filter. Runs until you press [Q/ESC]."),
        ("Eject    — Eject all external drives",
         "List the external drives currently mounted (name, size, mount point) "
         "and eject them all — the same as clicking each drive's eject button. "
         "Asks for confirmation first, shows a per-drive Success/Failed status, "
         "and offers a force eject for drives that won't let go (e.g. Spotlight "
         "or an app is still using them). macOS only (uses diskutil, with a "
         "Finder fallback)."),
        ("Zip      — View zip, log zip file, zip subfolders",
         "• View Zip — lists a zip's contents (sizes, ratios, dates) "
         "without extracting it.\n"
         "• Log Zip File Contents — records a .zip/.tar (or a folder of "
         "them) to the CB9Inventory database.\n"
         "• Zip SubFolders — compresses each subfolder of a target into "
         "its own .zip, with an optional push to AWS S3.\n"
         "• Zip SubFolders & Log — does the same and also logs each "
         "resulting zip to the CB9Inventory database."),
        ("Convert  — Data files between CSV / JSON / XLSX / SQL",
         "Convert a data file to another format. Input: .csv, .json, or .xlsx "
         "(first row / keys = column headers). Output: CSV, JSON (array of "
         "objects), XLSX (needs openpyxl), or SQL (CREATE TABLE with guessed "
         "column types + multi-row INSERTs, table named after the file). The "
         "new file is written next to the source with the same name and the "
         "new extension — collision-safe (name-2.ext, …), never overwrites."),
        ("Optimize Media — Convert video files to MP4",
         "'Convert Video to MP4' converts video files of other formats (.mov, "
         ".avi, .mkv, .wmv, .flv, .m4v) to .mp4 (H.264 + AAC) using moviepy. "
         "Asks for a source folder, an optional destination folder (defaults "
         "to the source folder), whether to include subfolders, and whether "
         "to Clean Up — delete each original file after it converts "
         "successfully AND the new .mp4 passes a validation check (reopens "
         "cleanly, duration matches); a failed check keeps the original "
         "instead of deleting it. Off by default. Collision-safe naming "
         "never overwrites. Requires moviepy (pip install moviepy)."),
        ("Sync     — Push new/updated files between two folders",
         "Sync two folders: push new/updated files A → B, B → A, or two-way "
         "A ↔ B. Options: recursive (default yes), what happens when a file "
         "exists on both sides (newest wins, largest wins, or — one-way syncs "
         "only — choose per file, Ignore/Select with [IA]/[SA] to apply to "
         "the rest), and excluding hidden files. Saved profiles from "
         "fmConfig.json (syncProfiles) run the same way. Nothing "
         "is deleted, and every run previews first — a DRY RUN until you "
         "confirm."),
        ("Remove   — Duplicates, files/folders by name",
         "Delete duplicate files (by name or by exact content), files by name "
         "pattern, or folders by name pattern. Every removal shows a preview and "
         "is a DRY RUN until you confirm — nothing is deleted by accident."),
        ("Clean Up — Junk files, purge old log entries",
         "• Remove Junk Files — finds and deletes .DS_Store, desktop.ini, "
         "and/or *.bak files under a root folder; choose which of the "
         "three to include first (all checked by default).\n"
         "• Purge Old Log Files — trims entries older than N days "
         "(default 90) out of the .log files in ~/Documents/log, saving "
         "a .bak backup of each changed file.\n"
         "Both show a DRY RUN preview and change nothing until you "
         "confirm."),
        ("Mount Shares — Mount SMB shares (saved profiles supported)",
         "Mount one or more SMB shares via Finder's own mount service "
         "(open smb://...) — the only way a plain user can add a new entry "
         "under /Volumes; a direct mkdir/mount_smbfs is refused there.\n"
         "• Mount Manually — enter a server, username, and share name(s) "
         "and mount them right away.\n"
         "• Run a Profile — mounts every share saved in a profile "
         "(fmConfig.json -> mountProfiles) in one go.\n"
         "• Set a Profile — creates or edits those saved profiles (name, "
         "server, username, share list).\n"
         "A saved Keychain login mounts instantly; otherwise Finder pops a "
         "password prompt per share. macOS only."),
        ("Permissions — Apache permissions, saved chmod profiles",
         "• Set Apache Permissions — resets ownership and permissions "
         "under /var/www for Apache (directories 755, files 644, "
         "root:www-data / _www:_www) via sudo. Grayed out when Apache "
         "isn't installed (/var/www doesn't exist).\n"
         "• Run a Profile — applies a saved permission profile "
         "(fmConfig.json → permissionProfiles): scans a folder for files "
         "matching a pattern and changes their permission to a saved "
         "octal value, previewing first — a DRY RUN until you confirm.\n"
         "• Set a Profile — creates or edits those saved profiles "
         "(folder, file name/type pattern, permission, recursive)."),
        ("Environment Vars",
         "List every environment variable currently set for this session — "
         "name and value, sorted alphabetically. Read-only. Not written to "
         "the activity log, since environment variables commonly hold "
         "secrets (API keys, tokens, credentials)."),
        ("Create Random UID — Generate N random UUIDs",
         "Enter how many UUIDs you need and get that many random (version 4) "
         "UUIDs, one per line — e.g. 550e8400-e29b-41d4-a716-446655440000. "
         "Handy for database keys, API tokens, and test data. The list is "
         "also appended to the activity log (fm.log); [R] Run Again "
         "regenerates a fresh batch with the same count."),
    ]
    help_note = (
        "Logging",
        "Commands and their results screens are logged to the FM activity log:\n"
        "• ~/Documents/log/fm.log\n"
        "Each run is appended with a [YYYY-MM-DD HH:MM:SS] timestamp line, "
        "exactly as shown on screen but with colors stripped — including "
        "the folders entered and the full results table.\n"
        "Actions that log to fm.log:\n"
        "• Convert\n"
        "• Eject All External Drives\n"
        "• Find Duplicates by Filename\n"
        "• Find Missing by Filename\n"
        "• Find & Replace\n"
        "• Find & Rename\n"
        "• Create Random UID (reruns via [R] Run Again are logged again)\n"
        "• Monitor File Activity — each event streams in real time, or to "
        "~/Documents/log/fmMonitor.csv when CSV output is chosen\n"
        "• Sync — the preview and what was copied\n"
        "• Permissions → Run a Profile / Set Apache Permissions — start/"
        "finish and every permission changed\n"
        "• Mount Shares → Manual / Run a Profile — every share requested "
        "and whether it mounted immediately or was still pending\n"
        "'Zip → Log Zip File Contents' and 'Zip SubFolders & Log' are "
        "different — they record archives to the CB9Inventory database on "
        "BPA5, not to fm.log.")
    while True:
        # AWS S3 and Admin are appended fresh every redraw (not baked into
        # the static `options` list above) so AWS S3's gray-out state
        # reflects the current boto3/credentials state immediately —
        # e.g. right after setting credentials in Admin Menu and backing
        # out, with no restart needed. Admin is always last (never gated —
        # it's where AWS credentials themselves get set).
        if not _BOTO3_AVAILABLE:
            aws_s3_reason = "boto3 not installed"
        elif not is_aws_configured():
            aws_s3_reason = "AWS credentials not configured — see Admin Menu"
        else:
            aws_s3_reason = None

        menu_options = options + [
            ("AWS S3   — Optimize/zip/log/upload media, bucket & record search",
             "Four tools:\n"
             "• Display Buckets & Sizes — every bucket the saved credentials can "
             "see, with object count and total size.\n"
             "• Optimize Media, Zip, Log & Upload to AWS S3 — a guided wizard: "
             "optimize video/images, resize images, create thumbnails, "
             "optionally zip, optionally log to DocInfo Manager, and upload to "
             "S3 — one folder or file at a time.\n"
             "• Search AWS S3 — filename PREFIX search across every object in "
             "every bucket (S3 has no server-side wildcard support).\n"
             "• Search DocInfo Manager Records — full wildcard filename search "
             "across everything FM has logged (zips and individual files).",
             aws_s3_reason),
            ("Local Scripts — Run a saved local script in a new terminal tab",
             "Arrow-driven list of saved local scripts (fmConfig.json -> "
             "localScripts). [Enter] opens a new tab in whichever terminal "
             "app FM is running in (Terminal.app or iTerm2) and runs the "
             "script's execution string there — FM's own menu keeps running "
             "independently. Adding, editing, and deleting scripts is done "
             "from Admin Menu -> Manage Local Scripts."),
            ("Admin    — DocInfo Manager login/logout, AWS S3 credentials",
             "Login or logout of DocInfo Manager. Logging in is required for "
             "Zip → Log Zip File Contents and Zip SubFolders & Log; the token is saved "
             "encrypted (~/.cb9Auth.enc, valid 90 days) and is independent of any "
             "Show Aliases login. Also set/update/clear the AWS Access Key ID, "
             "Secret Access Key, region, and bucket used by Zip SubFolders' "
             "optional Push to AWS S3 — saved encrypted in the same store, and "
             "manage the Local Scripts list (Manage Local Scripts)."),
        ]
        ch = render_menu("Main Menu", menu_options, is_main=True,
                         help_note=help_note)
        if ch == "quit":
            exit_screen(SCRIPT_NAME, VER)
            return
        if ch == "about":
            about_screen()
            continue
        try:
            if ch == "1":
                display_menu()
            elif ch == "2":
                find_menu()
            elif ch == "3":
                compare_menu()
            elif ch == "4":
                monitor_menu()
            elif ch == "5":
                eject_external_drives()
                pause_return()
            elif ch == "6":
                zip_menu()
            elif ch == "7":
                convert_menu()
            elif ch == "8":
                optimize_media_menu()
            elif ch == "9":
                sync_menu()
            elif ch == "10":
                remove_menu()
            elif ch == "11":
                cleanup_menu()
            elif ch == "12":
                mount_shares_menu()
            elif ch == "13":
                permissions_menu()
            elif ch == "14":
                environment_menu()
            elif ch == "15":
                create_random_uid_menu()
            elif ch == "16":
                aws_s3_menu()
            elif ch == "17":
                local_scripts_menu()
            elif ch == "18":
                admin_menu()
        except EscCancelled:
            continue


# =============================================================================
# CLI
# =============================================================================
def build_parser():
    p = argparse.ArgumentParser(
        prog="fm.py", add_help=True,
        description="File Manager — compare, display, eject, find, monitor, remove, sync, zip, "
                    "convert-video, convert-image, cleanup. Run with no arguments for the interactive menu.")
    sub = p.add_subparsers(dest="cmd")

    c2 = sub.add_parser("compare-2files", help="Compare two files side by side")
    c2.add_argument("a"); c2.add_argument("b")

    cc = sub.add_parser("compare-contents", help="Compare the contents of two folders")
    cc.add_argument("a"); cc.add_argument("b")
    cc.add_argument("--recursive", action="store_true", help="Descend into subfolders (default: top level only)")
    cc.add_argument("--by", choices=["name", "size", "both"], default="both")
    cc.add_argument("--case-sensitive", action="store_true",
                     help="Treat IMG_1.mov and IMG_1.MOV as different files (default: same file)")

    cv = sub.add_parser("convert", help="Convert a data file (.csv/.json/.xlsx) to CSV, JSON, XLSX, or SQL")
    cv.add_argument("file", help="source file (.csv, .json, or .xlsx)")
    cv.add_argument("to", choices=["csv", "json", "xlsx", "sql"],
                    help="output format — written next to the source, same name, new extension")

    uu = sub.add_parser("uuid", help="Generate N random UUIDs, one per line")
    uu.add_argument("count", nargs="?", type=int, default=1,
                    help="how many UUIDs to generate (default 1)")

    sz = sub.add_parser("sizes", help="Display subfolder sizes")
    sz.add_argument("folder", nargs="?", default=".")
    sz.add_argument("--sort", choices=["alpha", "size"], default="alpha")

    sub.add_parser("drives", help="Display size and free space of all mounted drives")

    sub.add_parser("drives-in-use", help="List external drives and whether they're currently being read/written (macOS)")

    fd = sub.add_parser("find", help="Find folders/files (single criterion)")
    fd.add_argument("type", choices=["folder", "name", "ext", "over", "under"])
    fd.add_argument("value")
    fd.add_argument("root", nargs="?", default=".")

    ff = sub.add_parser("find-files", help="Find files by one or more combined criteria (AND)")
    ff.add_argument("root", nargs="?", default=".")
    ff.add_argument("--name", help="filename wildcard pattern (e.g. 'IMG_*.jpg')")
    ff.add_argument("--ext", help="file extension, e.g. mov")
    ff.add_argument("--over", type=float, help="only files larger than N MB")
    ff.add_argument("--under", type=float, help="only files smaller than N MB")

    fdup = sub.add_parser("find-dups", help="Find duplicate filenames across one or more folders (size table)")
    fdup.add_argument("folders", nargs="+")

    ffuz = sub.add_parser("find-fuzzy-dups",
                          help="Find close-name + close-size duplicates (grouped with KEEP/DELETE markers)")
    ffuz.add_argument("folders", nargs="+")

    fmis = sub.add_parser("find-missing", help="Find filenames present in only one of two folders (size table)")
    fmis.add_argument("a"); fmis.add_argument("b")
    fmis.add_argument("--in", dest="mode", choices=["first", "second", "either"], default="either",
                      help="show names in the 1st folder only, the 2nd only, or either (default)")
    fmis.add_argument("--size", action="store_true",
                      help="match by filename AND size (same name with a different size = missing)")

    frp = sub.add_parser("find-replace", help="Find text in files and replace it (dry-run unless --apply)")
    frp.add_argument("root"); frp.add_argument("search"); frp.add_argument("replace")
    frp.add_argument("--ext", help="only files with this extension, e.g. php")
    frp.add_argument("--apply", action="store_true", help="Actually replace (default: dry run)")
    frp.add_argument("--bak", action="store_true", help="save a .bak backup of each modified file")
    frp.add_argument("--yes", action="store_true", help="Skip confirmation when replacing")

    frn = sub.add_parser("find-rename", help="Find files and rename them (dry-run unless --apply)")
    frn.add_argument("root")
    frn_mode = frn.add_mutually_exclusive_group(required=True)
    frn_mode.add_argument("--prepend", metavar="TEXT",
                          help="add TEXT to the start of each filename (skips files already starting with it)")
    frn_mode.add_argument("--append", dest="append_text", metavar="TEXT",
                          help="add TEXT to the end of the name, before the extension (skips files already ending with it)")
    frn_mode.add_argument("--replace", nargs=2, metavar=("FIND", "NEW"),
                          help="replace FIND with NEW inside the filename (literal, case-insensitive; NEW may be '')")
    frn.add_argument("--ext", help="only files with this extension, e.g. mov")
    frn.add_argument("--apply", action="store_true", help="Actually rename (default: dry run)")
    frn.add_argument("--yes", action="store_true", help="Skip confirmation when renaming")

    rm = sub.add_parser("remove", help="Remove items (dry-run unless --delete or --move-to)")
    rm.add_argument("type", choices=["folder", "name", "folder-name", "dup-name", "dup-hash",
                                     "dup-fuzzy", "zero-size"])
    rm.add_argument("args", nargs="+")
    rm_action = rm.add_mutually_exclusive_group()
    rm_action.add_argument("--delete", action="store_true", help="Actually delete (default: dry run)")
    rm_action.add_argument("--move-to", metavar="FOLDER",
                           help="Move instead of deleting (dup-name/dup-hash/dup-fuzzy only) — "
                                "actually moves the files there (default without --delete: dry run)")
    rm.add_argument("--yes", action="store_true", help="Skip confirmation when deleting/moving")

    mo = sub.add_parser("monitor", help="Watch a folder for file activity in real time (Ctrl-C to stop)")
    mo.add_argument("folder", nargs="?", help="Folder to monitor")
    mo.add_argument("--profile", help="run a named monitorProfiles entry from fmConfig.json instead of giving a folder")
    mo.add_argument("--no-recursive", action="store_true", help="top-level files only (default: recursive)")
    mo.add_argument("--ext", help="comma-separated extensions to watch, e.g. jpg,png (default: all files)")
    mo.add_argument("--csv", action="store_true", help="log events to fmMonitor.csv instead of fm.log")

    ej = sub.add_parser("eject", help="Eject all external drives (macOS)")
    ej.add_argument("--list", action="store_true", help="list external drives and exit (no eject)")
    ej.add_argument("--force", action="store_true", help="force-unmount before ejecting (bypasses Spotlight etc.)")
    ej.add_argument("--yes", action="store_true", help="Skip confirmation and eject immediately")

    sy = sub.add_parser("sync", help="Sync: push new/updated files between two folders, one-way or two-way (dry-run unless --copy)")
    sy.add_argument("a", nargs="?", help="Folder A")
    sy.add_argument("b", nargs="?", help="Folder B")
    sy.add_argument("--to", choices=["b", "a", "both"], default="b",
                    help="push direction: 'b' = A→B (default), 'a' = B→A, 'both' = two-way A↔B")
    sy.add_argument("--profile", help="run a named syncProfiles entry from fmConfig.json instead of giving folders")
    sy.add_argument("--conflict", choices=["newest", "largest", "manual"], default="newest",
                    help="when a file exists on both sides: copy if source is newer (default) / larger / "
                         "'manual' to be asked Ignore or Select for each file (one-way sync + a TTY only)")
    sy.add_argument("--no-recursive", action="store_true", help="top-level files only (default: recursive)")
    sy.add_argument("--include-hidden", action="store_true", help="include hidden files (default: excluded)")
    sy.add_argument("--copy", action="store_true", help="Actually copy (default: dry run)")
    sy.add_argument("--yes", action="store_true", help="Skip confirmation when copying")

    zs = sub.add_parser("zip-subfolders", help="Zip each subfolder")
    zs.add_argument("target"); zs.add_argument("dest", nargs="?", default=None)
    zs.add_argument("-r", "--remove", action="store_true", help="Remove source folders after zip")

    zv = sub.add_parser("zip-view", help="View a zip file or browse a folder of zips")
    zv.add_argument("path", nargs="?", default=None)

    zf = sub.add_parser("find-zip", help="Search filenames inside a zip, or every zip under a folder (wildcard, e.g. 'fan*.png')")
    zf.add_argument("path", help="a .zip file, or a folder to recursively search all zips under it")
    zf.add_argument("pattern", nargs="?", default="*", help="filename wildcard pattern (default: *)")

    zl = sub.add_parser("zip-log", help="Log a .zip/.tar (or a folder of them) to CB9Inventory")
    zl.add_argument("target")
    zl.add_argument("-r", "--recursive", action="store_true",
                     help="With a folder TARGET, include archives in subfolders too")

    cv = sub.add_parser("convert-video", help="Convert video files (.mov/.avi/.mkv/.wmv/.flv/.m4v) to .mp4")
    cv.add_argument("source"); cv.add_argument("dest", nargs="?", default=None)
    cv.add_argument("-r", "--recursive", action="store_true", help="Include subfolders")
    cv.add_argument("--flatten", action="store_true",
                     help="With -r and a DEST different from SOURCE, place all converted files "
                          "directly in DEST instead of mirroring the subfolder structure (default: mirror)")
    cv.add_argument("--delete", action="store_true", help="Delete original files after a successful conversion")
    cv.add_argument("--pause", type=float, default=0, help="Seconds to pause between files (drive cooldown, default 0)")

    ci = sub.add_parser("convert-image", help="Convert image files (.heic/.heif/.bmp/.tiff/.webp/.gif) to .jpg or .png")
    ci.add_argument("source"); ci.add_argument("dest", nargs="?", default=None)
    ci.add_argument("-r", "--recursive", action="store_true", help="Include subfolders")
    ci.add_argument("--flatten", action="store_true",
                     help="With -r and a DEST different from SOURCE, place all converted files "
                          "directly in DEST instead of mirroring the subfolder structure (default: mirror)")
    ci.add_argument("--format", choices=list(IMAGE_CONVERT_OUT_FORMATS), default="jpg",
                     help="Output format (default: jpg)")
    ci.add_argument("--delete", action="store_true", help="Delete original files after a successful conversion")
    ci.add_argument("--pause", type=float, default=0, help="Seconds to pause between files (drive cooldown, default 0)")

    cu = sub.add_parser("cleanup", help="Remove junk files or purge old log entries (dry-run unless --delete)")
    cu.add_argument("what", choices=["junk", "logs"])
    cu.add_argument("path", nargs="?", help="junk: root folder (required); logs: log folder (default ~/Documents/log)")
    cu.add_argument("--days", type=int, default=LOG_PURGE_DAYS, help="logs: days of entries to keep (default 90)")
    cu.add_argument("--delete", action="store_true", help="Actually delete/purge (default: dry run)")
    cu.add_argument("--yes", action="store_true", help="Skip confirmation when deleting/purging")

    return p


def run_cli(ns):
    cmd = ns.cmd
    if cmd == "compare-2files":
        compare_two_files(clean_path(ns.a), clean_path(ns.b))
    elif cmd == "convert":
        convert_file(clean_path(ns.file), ns.to)
    elif cmd == "uuid":
        generate_uuids(ns.count)
    elif cmd == "compare-contents":
        compare_folder_contents(clean_path(ns.a), clean_path(ns.b),
                                ns.recursive, ns.by, ns.case_sensitive)
    elif cmd == "sizes":
        display_folder_sizes(clean_path(ns.folder), ns.sort)
    elif cmd == "drives":
        display_all_drives()
    elif cmd == "drives-in-use":
        display_drives_in_use()
    elif cmd == "find":
        root = clean_path(ns.root)
        if ns.type == "folder":
            find_folders(root, ns.value)
        elif ns.type == "name":
            find_files_by_name(root, ns.value)
        elif ns.type == "ext":
            find_files_by_ext(root, ns.value)
        elif ns.type in ("over", "under"):
            find_by_size(root, float(ns.value), over=(ns.type == "over"))
    elif cmd == "find-files":
        find_files_combined(clean_path(ns.root), ns.name, ns.ext, ns.over, ns.under)
    elif cmd == "find-dups":
        find_duplicates_by_filename(ns.folders)
    elif cmd == "find-fuzzy-dups":
        find_duplicates_by_fuzzy_name(ns.folders)
    elif cmd == "find-missing":
        find_missing_by_filename(ns.a, ns.b, ns.mode, ns.size)
    elif cmd == "find-replace":
        find_and_replace(clean_path(ns.root), ns.search, ns.replace, ns.ext,
                         live_requested=ns.apply, assume_yes=ns.yes,
                         backup=ns.bak)
    elif cmd == "find-rename":
        if ns.prepend is not None:
            mode, text, repl = "prepend", ns.prepend, None
        elif ns.append_text is not None:
            mode, text, repl = "append", ns.append_text, None
        else:
            mode, text, repl = "replace", ns.replace[0], ns.replace[1]
        find_and_rename(clean_path(ns.root), mode, text, repl, ns.ext,
                        live_requested=ns.apply, assume_yes=ns.yes)
    elif cmd == "remove":
        move_to = clean_path(ns.move_to) if ns.move_to else None
        if move_to and ns.type not in ("dup-name", "dup-hash", "dup-fuzzy"):
            print(color_text("  --move-to is only supported for dup-name, dup-hash, and dup-fuzzy.", fg=RED))
            return
        live = True if (ns.delete or move_to) else False
        if ns.type == "folder":
            _cli_remove_folder(ns.args[0], live, ns.yes)
        elif ns.type == "name":
            _cli_remove_name(ns.args, live, ns.yes)
        elif ns.type == "folder-name":
            _cli_remove_folder_name(ns.args, live, ns.yes)
        elif ns.type == "dup-name":
            _cli_dupname(ns.args, live, ns.yes, move_to=move_to)
        elif ns.type == "dup-hash":
            _cli_duphash(ns.args, live, ns.yes, move_to=move_to)
        elif ns.type == "dup-fuzzy":
            remove_duplicates_by_fuzzy_name(ns.args, live, assume_yes=ns.yes, move_to=move_to)
        elif ns.type == "zero-size":
            remove_zero_size_files(ns.args, live, assume_yes=ns.yes)
    elif cmd == "monitor":
        if ns.profile:
            profiles, perr = _load_monitor_profiles()
            pr = next((p for p in profiles
                       if str(p.get("name", "")).lower() == ns.profile.lower()), None)
            if perr:
                print(color_text(f"  {perr}", fg=RED))
            elif pr is None:
                names = ", ".join(str(p.get("name", "?")) for p in profiles) or "(none defined)"
                print(color_text(f"  No monitor profile named '{ns.profile}'. Profiles: {names}", fg=RED))
            else:
                monitor_activity(clean_path(str(pr.get("folder", ""))),
                                 recursive=bool(pr.get("recursive", True)),
                                 exts=pr.get("extensions", ""),
                                 output="csv" if (ns.csv or str(pr.get("output", "log")).lower() == "csv") else "log",
                                 profile_name=pr.get("name") or "(unnamed profile)")
        elif not ns.folder:
            print(color_text("  usage: monitor FOLDER  (or: monitor --profile NAME)", fg=RED))
        else:
            monitor_activity(clean_path(ns.folder),
                             recursive=not ns.no_recursive,
                             exts=ns.ext or "",
                             output="csv" if ns.csv else "log")
    elif cmd == "eject":
        if ns.list:
            drives = _external_drives()
            if drives:
                print()
                _print_drive_list(drives)
            else:
                print(color_text("  No external drives found.", fg=YELLOW))
        else:
            eject_external_drives(live_requested=True, assume_yes=ns.yes,
                                  force=ns.force)
    elif cmd == "sync":
        live = True if ns.copy else False
        if ns.profile:
            profiles, perr = _load_sync_profiles()
            pr = next((p for p in profiles
                       if str(p.get("name", "")).lower() == ns.profile.lower()), None)
            if perr:
                print(color_text(f"  {perr}", fg=RED))
            elif pr is None:
                names = ", ".join(str(p.get("name", "?")) for p in profiles) or "(none defined)"
                print(color_text(f"  No sync profile named '{ns.profile}'. Profiles: {names}", fg=RED))
            else:
                _run_sync_profile(pr, live_requested=live, assume_yes=ns.yes)
        elif not ns.a or not ns.b:
            print(color_text("  usage: sync FOLDER_A FOLDER_B  (or: sync --profile NAME)", fg=RED))
        else:
            sync_folders(clean_path(ns.a), clean_path(ns.b),
                         {"b": "a2b", "a": "b2a", "both": "both"}[ns.to],
                         recursive=not ns.no_recursive,
                         conflict=ns.conflict,
                         exclude_hidden=not ns.include_hidden,
                         live_requested=live, assume_yes=ns.yes)
    elif cmd == "zip-subfolders":
        zip_subfolders(ns.target, ns.dest, ns.remove)
    elif cmd == "zip-view":
        zip_view(ns.path)
    elif cmd == "find-zip":
        find_files_in_zip_target(ns.path, ns.pattern)
    elif cmd == "zip-log":
        log_zip_files(ns.target, recursive=ns.recursive)
    elif cmd == "convert-video":
        convert_videos_to_mp4(ns.source, ns.dest, recursive=ns.recursive, remove_after=ns.delete,
                              pause_seconds=ns.pause, flatten=ns.flatten)
    elif cmd == "convert-image":
        convert_images_to_format(ns.source, ns.dest, recursive=ns.recursive, out_format=ns.format,
                                  remove_after=ns.delete, pause_seconds=ns.pause, flatten=ns.flatten)
    elif cmd == "cleanup":
        live = True if ns.delete else False
        if ns.what == "junk":
            if not ns.path:
                print(color_text("  usage: cleanup junk ROOT [--delete] [--yes]", fg=RED))
            else:
                cleanup_junk_files(ns.path, live_requested=live, assume_yes=ns.yes)
        else:
            cleanup_purge_logs(ns.path or LOG_PURGE_DIR, ns.days,
                               live_requested=live, assume_yes=ns.yes)


# CLI remove wrappers reuse the screen functions but pass live_requested and
# assume_yes through the shared _finish_removal path.
def _cli_remove_folder(path, live, yes):
    with _ActivityLog():
        screen("Remove Folder")
        dry_mode = ask_dry_mode(live)
        print(); print_mode_line(dry_mode)
        path = clean_path(path)
        print(f"  {YELLOW}Target{RESET}: {path}\n")
        if not os.path.isdir(path):
            print(color_text(f"  Not a directory: {path}", fg=RED)); return
        total, _ = folder_stats(path)
        _finish_removal([(path, total, True)], live, assume_yes=yes,
                         activity="Remove Folder", fields=[("Target", path)], dry_mode=dry_mode)


def _cli_remove_name(args, live, yes):
    with _ActivityLog():
        if len(args) < 2:
            print(color_text("  usage: remove name PATTERN ROOT", fg=RED)); return
        pattern, root = args[0], clean_path(args[1])
        screen("Remove Files by Name")
        dry_mode = ask_dry_mode(live)
        print(); print_mode_line(dry_mode)
        print(f"  {YELLOW}Root{RESET}: {root}    {YELLOW}Pattern{RESET}: {pattern}\n")
        if not os.path.isdir(root):
            print(color_text(f"  Not a directory: {root}", fg=RED)); return
        items = []
        for dp, dns, fns in os.walk(root):
            for fn in fns:
                if fnmatch.fnmatch(fn, pattern):
                    p = os.path.join(dp, fn)
                    items.append((p, file_size(p), False))
        items.sort(key=lambda t: t[0])
        _finish_removal(items, live, assume_yes=yes, activity="Remove Files by Name",
                         fields=[("Root", root), ("Pattern", pattern)], dry_mode=dry_mode)


def _cli_remove_folder_name(args, live, yes):
    with _ActivityLog():
        if len(args) < 2:
            print(color_text("  usage: remove folder-name PATTERN ROOT", fg=RED)); return
        pattern, root = args[0], clean_path(args[1])
        screen("Remove Folders by Name")
        dry_mode = ask_dry_mode(live)
        print(); print_mode_line(dry_mode)
        print(f"  {YELLOW}Root{RESET}: {root}    {YELLOW}Pattern{RESET}: {pattern}\n")
        if not os.path.isdir(root):
            print(color_text(f"  Not a directory: {root}", fg=RED)); return
        matches = []
        for dp, dns, fns in os.walk(root):
            for d in dns:
                if fnmatch.fnmatch(d, pattern):
                    matches.append(os.path.join(dp, d))
        items = [(m, folder_stats(m)[0], True) for m in _topmost(matches)]
        _finish_removal(items, live, assume_yes=yes, activity="Remove Folders by Name",
                         fields=[("Root", root), ("Pattern", pattern)], dry_mode=dry_mode)


def _cli_dupname(folders, live, yes, move_to=None):
    with _ActivityLog():
        screen("Remove Duplicates (by Name)")
        dry_mode = ask_dry_mode(live)
        print(); print_mode_line(dry_mode)
        files = _gather_files(folders)
        groups = defaultdict(list)
        for p in files:
            groups[os.path.basename(p)].append(p)
        items = []
        keep_map = {}
        for name, paths in sorted(groups.items()):
            if len(paths) > 1:
                sp = sorted(paths)
                for p in sp[1:]:
                    items.append((p, file_size(p), False))
                    keep_map[p] = sp[0]
        _finish_removal(items, live, assume_yes=yes, require_yes=True,
                         activity="Remove Duplicates (by Name)",
                         fields=[("Folders", ', '.join(folders))],
                         allow_move=True, move_to=move_to, dry_mode=dry_mode,
                         keep_map=keep_map)


def _cli_duphash(folders, live, yes, move_to=None):
    with _ActivityLog():
        screen("Remove Duplicates (by Hash)")
        dry_mode = ask_dry_mode(live)
        print(); print_mode_line(dry_mode)
        files = _gather_files(folders)
        by_size = defaultdict(list)
        for p in files:
            by_size[file_size(p)].append(p)
        groups = defaultdict(list)
        for size, paths in by_size.items():
            if len(paths) < 2:
                continue
            for p in paths:
                h = _hash_file(p)
                if h:
                    groups[h].append(p)
        items = []
        keep_map = {}
        for h, paths in groups.items():
            if len(paths) > 1:
                sp = sorted(paths)
                for p in sp[1:]:
                    items.append((p, file_size(p), False))
                    keep_map[p] = sp[0]
        _finish_removal(items, live, assume_yes=yes, require_yes=True,
                         activity="Remove Duplicates (by Hash)",
                         fields=[("Folders", ', '.join(folders))],
                         allow_move=True, move_to=move_to, dry_mode=dry_mode,
                         keep_map=keep_map)


# =============================================================================
# ENTRY POINT
# =============================================================================
def main():
    _check_optional_dependencies()
    _check_config_file()
    if len(sys.argv) > 1:
        global INTERACTIVE
        INTERACTIVE = False
        parser = build_parser()
        ns = parser.parse_args()
        if not ns.cmd:
            parser.print_help()
            return
        try:
            run_cli(ns)
        except KeyboardInterrupt:
            print()
        except EscCancelled:
            print(color_text("\nCancelled.", fg=YELLOW))
        return

    try:
        main_menu()
    except KeyboardInterrupt:
        print()
        exit_screen(SCRIPT_NAME, VER)
    except EscCancelled:
        # A menu function's own try/except should have caught this already —
        # this is a last-resort net so an ESC can never surface as a crash.
        print(color_text("\nCancelled.", fg=YELLOW))
        exit_screen(SCRIPT_NAME, VER)


if __name__ == "__main__":
    main()
